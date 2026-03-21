#!/usr/bin/env python3
"""
Build the DSCR Lead Intelligence demo workbook for Northside Realty.
Premium Excel workbook with Call List, Portfolio Detail, and Market Summary tabs.
"""

import csv
import json
import os
from datetime import datetime, date
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ── Paths ──────────────────────────────────────────────────────────────────
SHOWCASE_CSV = "/Users/stillmindcreative/Documents/dscr_lead_gen/scrape/data/demo/showcase_7ep_wake.csv"
QUALIFIED_CSV = "/Users/stillmindcreative/Documents/dscr_lead_gen/scrape/data/filtered/wake_qualified.csv"
CACHE_PROFILE = "/Users/stillmindcreative/Documents/dscr_lead_gen/scrape/data/demo/attom_7ep_cache/cache_profile.json"
OUTPUT_PATH = "/Users/stillmindcreative/Documents/dscr_lead_gen/sales/demo_tearsheets/demo_workbook_northside.xlsx"

# ── Brand Colors ───────────────────────────────────────────────────────────
OCEAN = "0D9488"
DEEP_OCEAN = "0F766E"
SEAFOAM = "99F6E4"
SEAFOAM_LIGHT = "E0FBF5"
GOLD = "B4873F"
GOLD_LIGHT = "FEF3C7"
CHARCOAL = "1C1917"
STONE = "44403C"
MIST = "E7E5E4"
CLOUD = "F5F5F4"
WHITE = "FAFAF9"
SAGE = "16A34A"
SAGE_LIGHT = "DCFCE7"
AMBER = "D97706"
AMBER_LIGHT = "FEF3C7"
CORAL = "DC2626"
CORAL_LIGHT = "FEE2E2"

# ── Style Objects ──────────────────────────────────────────────────────────
header_fill = PatternFill(start_color=OCEAN, end_color=OCEAN, fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
data_font = Font(name="Calibri", size=10, color=CHARCOAL)
data_font_stone = Font(name="Calibri", size=10, color=STONE)
bold_font = Font(name="Calibri", size=10, bold=True, color=CHARCOAL)
tier1_fill = PatternFill(start_color=SEAFOAM_LIGHT, end_color=SEAFOAM_LIGHT, fill_type="solid")
tier2_fill = PatternFill(start_color=AMBER_LIGHT, end_color=AMBER_LIGHT, fill_type="solid")
cloud_fill = PatternFill(start_color=CLOUD, end_color=CLOUD, fill_type="solid")
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
thin_border = Border(
    left=Side(style="thin", color=MIST),
    right=Side(style="thin", color=MIST),
    top=Side(style="thin", color=MIST),
    bottom=Side(style="thin", color=MIST),
)
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_alignment = Alignment(vertical="center", wrap_text=False)
data_center = Alignment(horizontal="center", vertical="center")
currency_fmt = '$#,##0'
pct_fmt = '0.0%'
number_fmt = '#,##0'

# ── Helpers ────────────────────────────────────────────────────────────────
def safe_float(v, default=0.0):
    try:
        if v is None or v == '' or v == 'nan':
            return default
        return float(v)
    except (ValueError, TypeError):
        return default

def safe_int(v, default=0):
    try:
        if v is None or v == '' or v == 'nan':
            return default
        return int(float(v))
    except (ValueError, TypeError):
        return default

def parse_date(s):
    if not s or s == 'nan':
        return None
    for fmt in ('%Y-%m-%d', '%Y-%m', '%m/%d/%Y', '%Y'):
        try:
            return datetime.strptime(s.strip(), fmt).date()
        except ValueError:
            continue
    return None

def years_between(d1, d2):
    if not d1 or not d2:
        return None
    delta = abs((d2 - d1).days)
    return round(delta / 365.25, 1)

# ── Load Data ──────────────────────────────────────────────────────────────
print("Loading showcase CSV...")
showcase_rows = []
with open(SHOWCASE_CSV) as f:
    reader = csv.DictReader(f)
    for row in reader:
        showcase_rows.append(row)
print(f"  Loaded {len(showcase_rows)} showcase leads")

# Map showcase data by owner name
showcase_by_owner = {}
for row in showcase_rows:
    showcase_by_owner[row['owner_name'].strip()] = row

print("Loading qualified CSV for portfolio data...")
# Owner name variants for expanded matching
ECKENRODE_NAMES = {'ECKENRODE, KYLE BURNS', 'ECKENRODE, KYLE', 'ECKENRODE, KYLE BURNS BRIGGS, KAROLINE OLIVIA'}
KOHLI_NAMES = {'KOHLI, RAVINDER KOHLI, ASHA', 'KOHLI, RAVINDER KOHLI, ASHA K'}

showcase_owner_names = set(showcase_by_owner.keys())

portfolio_rows = defaultdict(list)
with open(QUALIFIED_CSV) as f:
    reader = csv.DictReader(f)
    for row in reader:
        name = row['owner_name_1'].strip()
        if name in showcase_owner_names:
            portfolio_rows[name].append(row)
        elif name in ECKENRODE_NAMES:
            portfolio_rows['ECKENRODE, KYLE BURNS'].append(row)
        elif name in KOHLI_NAMES:
            portfolio_rows['KOHLI, RAVINDER KOHLI, ASHA'].append(row)

print(f"  Portfolio rows loaded for {len(portfolio_rows)} owners")
for name, rows in sorted(portfolio_rows.items(), key=lambda x: -len(x[1])):
    print(f"    {name}: {len(rows)} properties")

print("Loading ATTOM cache for beds/baths/sqft...")
with open(CACHE_PROFILE) as f:
    cache_profile = json.load(f)
print(f"  {len(cache_profile)} cache entries")

# ── Build Investor Profiles ───────────────────────────────────────────────
print("\nBuilding investor profiles...")

TODAY = date(2026, 3, 21)

def get_absentee_status(showcase_row, portfolio_rows_list):
    """Determine absentee status from ATTOM data and mail address."""
    attom_absentee = showcase_row.get('attom_absentee', '')
    mail_addr = showcase_row.get('attom_mail_address', '') or ''

    # Check from portfolio data too
    if not mail_addr and portfolio_rows_list:
        r0 = portfolio_rows_list[0]
        mail_addr = f"{r0.get('mail_street','')} {r0.get('mail_city','')} {r0.get('mail_state','')} {r0.get('mail_zip','')}".strip()

    mail_upper = mail_addr.upper()

    if attom_absentee == 'A':
        # Check if out of state
        if portfolio_rows_list:
            mail_state = portfolio_rows_list[0].get('mail_state', '').strip().upper()
            if mail_state and mail_state != 'NC':
                return 'Out-of-State'
            elif mail_state == 'NC':
                # Check if local (Wake County area)
                mail_city = portfolio_rows_list[0].get('mail_city', '').strip().upper()
                wake_cities = {'RALEIGH', 'CARY', 'APEX', 'HOLLY SPRINGS', 'FUQUAY-VARINA',
                             'FUQUAY VARINA', 'WAKE FOREST', 'GARNER', 'KNIGHTDALE', 'WENDELL',
                             'ZEBULON', 'MORRISVILLE', 'ROLESVILLE'}
                if mail_city in wake_cities:
                    return 'Local'
                return 'In-State'
        # Try parsing mail address for state
        for st in ['TX', 'PA', 'NY', 'NJ', 'FL', 'CA', 'VA', 'MD', 'GA', 'OH', 'IL']:
            if f', {st} ' in mail_upper or mail_upper.endswith(f' {st}'):
                return 'Out-of-State'
        return 'In-State'
    elif attom_absentee == 'O':
        return 'Local'
    return 'Local'

def compute_agent_score(owner_name, showcase_row, port_rows):
    """Compute agent score based on scoring weights."""
    score = 0
    signals = []

    n_properties = len(port_rows)
    is_llc = any(r.get('is_llc') == 'True' for r in port_rows) if port_rows else (showcase_row.get('attom_corporate', '') == 'Y')

    # No homestead (investment property)
    has_homestead = any(r.get('homestead_flag', '') == 'True' for r in port_rows)
    if not has_homestead:
        score += 5

    # Value tier - use ATTOM AVM or assessed
    avm = safe_float(showcase_row.get('attom_avm_value', 0))
    assessed = safe_float(showcase_row.get('attom_assessed_total', 0))
    sample_val = avm if avm > 0 else assessed

    # Average portfolio value
    port_values = []
    for r in port_rows:
        v = safe_float(r.get('just_value', 0))
        if v > 0:
            port_values.append(v)
    avg_val = sum(port_values) / len(port_values) if port_values else sample_val

    if avg_val >= 1000000:
        score += 12
        signals.append("$1M+ avg value")
    elif avg_val >= 500000:
        score += 10
        signals.append("$500K+ avg value")
    elif avg_val >= 150000:
        score += 8

    # New construction
    year_built = safe_int(showcase_row.get('attom_year_built', 0))
    if year_built >= 2023:
        score += 6
        signals.append("new construction")

    # High land ratio
    land_val = safe_float(showcase_row.get('attom_assessed_land', 0))
    total_val = safe_float(showcase_row.get('attom_assessed_total', 0))
    if total_val > 0 and (land_val / total_val) > 0.6:
        score += 5
        signals.append("high land ratio")

    # Absentee
    absentee = get_absentee_status(showcase_row, port_rows)
    if absentee == 'Out-of-State':
        score += 15
        signals.append("out-of-state")
    elif absentee == 'In-State':
        score += 8

    # LLC/Corp
    if is_llc:
        score += 5
        signals.append("LLC")

    # Portfolio size
    if n_properties >= 10:
        score += 25
        signals.append(f"{n_properties} properties")
    elif n_properties >= 5:
        score += 18
        signals.append(f"{n_properties} properties")
    elif n_properties >= 2:
        score += 10
        signals.append(f"{n_properties} properties")

    # Multi-city
    cities = set()
    for r in port_rows:
        c = r.get('prop_city', '').strip()
        if c:
            cities.add(c)
    if len(cities) >= 3:
        score += 8
        signals.append(f"{len(cities)} markets")

    # Recent purchase
    sale_dates = []
    for r in port_rows:
        d = parse_date(r.get('sale_date', ''))
        if d:
            sale_dates.append(d)
    # Also from ATTOM
    attom_sale_date = parse_date(showcase_row.get('attom_last_sale_date', ''))
    if attom_sale_date:
        sale_dates.append(attom_sale_date)

    if sale_dates:
        most_recent = max(sale_dates)
        days_ago = (TODAY - most_recent).days
        if days_ago <= 180:
            score += 15
            signals.append("recent purchase")
        elif days_ago <= 365:
            score += 10
            signals.append("recent purchase")
        elif days_ago <= 730:
            score += 5

    # High velocity (2+ purchases in 24mo)
    recent_purchases = [d for d in sale_dates if (TODAY - d).days <= 730]
    if len(recent_purchases) >= 2:
        score += 12
        signals.append("high velocity")

    # Cash buyer
    cash = showcase_row.get('derived_cash_buyer', '').strip()
    if cash == 'True':
        score += 10
        signals.append("cash buyer")

    # Long hold large portfolio
    if sale_dates and n_properties >= 5:
        oldest = min(sale_dates)
        hold_years = (TODAY - oldest).days / 365.25
        if hold_years >= 5:
            score += 8
            signals.append("long hold")

    # Enrichment: active permits
    permit_count = safe_int(showcase_row.get('attom_permit_count', 0))
    if permit_count > 0:
        score += 6

    permit_value = safe_float(showcase_row.get('attom_total_permit_value', 0))
    if permit_value > 100000:
        score += 5
        signals.append("active developer")

    # REO purchase
    sale_type = showcase_row.get('attom_last_sale_type', '')
    if sale_type and 'REO' in sale_type.upper():
        score += 5
        signals.append("REO buyer")

    # High rent yield
    rent = safe_float(showcase_row.get('attom_rent_estimate', 0))
    if avm > 0 and rent > 0:
        gross_yield = (rent * 12) / avm
        if gross_yield > 0.07:
            score += 4
            signals.append("high yield")

    return min(score, 100), signals

def get_tier(score):
    if score >= 45:
        return "Tier 1 Priority"
    elif score >= 30:
        return "Tier 2 Opportunity"
    elif score >= 15:
        return "Tier 3 Watch"
    else:
        return "Tier 3 Watch"

def get_segment(owner_name, score, signals, n_properties, showcase_row, port_rows):
    """Assign segment based on profile."""
    permit_value = safe_float(showcase_row.get('attom_total_permit_value', 0))
    cash = showcase_row.get('derived_cash_buyer', '').strip() == 'True'
    absentee = get_absentee_status(showcase_row, port_rows)

    # Check for high velocity
    sale_dates = []
    for r in port_rows:
        d = parse_date(r.get('sale_date', ''))
        if d:
            sale_dates.append(d)
    recent_24m = [d for d in sale_dates if (TODAY - d).days <= 730]

    if n_properties >= 10:
        return "Serial Acquirer (10+)"
    if permit_value > 100000:
        return "Active Developer"
    if absentee == 'Out-of-State':
        return "Out-of-State Investor"
    if len(recent_24m) >= 2:
        return "High-Velocity Buyer"
    if n_properties >= 5:
        return "Portfolio Builder (5-9)"
    if cash:
        return "Cash Buyer"
    if n_properties >= 2:
        return "Growing Investor (2-4)"

    # Check REO
    sale_type = showcase_row.get('attom_last_sale_type', '')
    if sale_type and 'REO' in sale_type.upper():
        return "Value Investor"

    return "Growing Investor (2-4)"

def generate_approach(segment, owner_name, n_properties, signals, cities, absentee, showcase_row):
    """Generate personalized recommended approach."""
    name_short = owner_name.split(',')[0].strip() if ',' in owner_name else owner_name.replace(' LLC', '').strip()

    if segment == "Serial Acquirer (10+)":
        city_list = ', '.join(sorted(cities)[:3])
        return f"Your {n_properties}-property Wake County portfolio across {city_list} puts you in the top tier of local investors. I specialize in representing portfolio-scale investors with off-market deal flow and bulk transaction expertise."

    if segment == "Active Developer":
        permit_val = safe_float(showcase_row.get('attom_total_permit_value', 0))
        return f"I noticed your active development activity in Wake County with ${permit_val:,.0f} in recent permits. I work with developers to source land and off-market rehab opportunities before they hit the MLS."

    if segment == "Out-of-State Investor":
        mail = showcase_row.get('attom_mail_address', '') or ''
        # Extract state
        state = ''
        if mail:
            parts = mail.split(',')
            if len(parts) >= 2:
                state_part = parts[-1].strip().split()[0] if parts[-1].strip() else ''
                state = state_part
        return f"Managing {n_properties} Wake County properties from out of state takes serious operational skill. I provide local market intelligence and representation so you can expand your portfolio with confidence — boots on the ground when you need them."

    if segment == "High-Velocity Buyer":
        return f"Your recent acquisition pace in Wake County tells me you're actively deploying capital. I track off-market opportunities and pocket listings daily — let me send you deals that match your buy box before they go public."

    if segment == "Portfolio Builder (5-9)":
        return f"With {n_properties} properties in Wake County, you've built a solid foundation. I work with investors at your stage to identify the next 5 acquisitions that complement your existing portfolio — neighborhoods, price points, and timing."

    if segment == "Cash Buyer":
        return f"Cash buyers are the most sought-after clients in our market. I can connect you with motivated sellers and off-market deals where your ability to close fast gives you a significant competitive advantage."

    if segment == "Value Investor":
        return f"Your track record of finding value in Wake County is impressive. I source distressed and off-market properties before they're widely marketed — exactly the deal flow value investors need."

    if segment == "Growing Investor (2-4)":
        return f"Your {n_properties}-property portfolio in Wake County shows you're serious about building wealth through real estate. I help growing investors identify their next best acquisition based on market timing and neighborhood trajectory."

    return f"I noticed your investment properties in Wake County. I specialize in working with investors to optimize their portfolios through strategic acquisitions and dispositions."


# ── Build Investor Data ────────────────────────────────────────────────────
investors = []

for owner_name in sorted(showcase_by_owner.keys()):
    sc = showcase_by_owner[owner_name]
    port = portfolio_rows.get(owner_name, [])
    n_props = len(port)

    # Portfolio values from wake_qualified (just_value)
    port_values = []
    for r in port:
        v = safe_float(r.get('just_value', 0))
        if v > 0:
            port_values.append(v)

    total_portfolio_value = sum(port_values) if port_values else safe_float(sc.get('attom_assessed_total', 0))
    avg_value = total_portfolio_value / len(port_values) if port_values else total_portfolio_value
    min_value = min(port_values) if port_values else avg_value
    max_value = max(port_values) if port_values else avg_value

    # Cities
    cities = set()
    for r in port:
        c = r.get('prop_city', '').strip()
        if c:
            cities.add(c)

    # Entity type
    is_llc = sc.get('attom_corporate', '') == 'Y' or any(r.get('is_llc') == 'True' for r in port)
    entity_type = 'LLC' if is_llc else 'Individual'
    if 'TRUST' in owner_name.upper():
        entity_type = 'Trust'

    # Sale dates
    sale_dates = []
    for r in port:
        d = parse_date(r.get('sale_date', ''))
        if d:
            sale_dates.append(d)
    attom_sale = parse_date(sc.get('attom_last_sale_date', ''))
    if attom_sale:
        sale_dates.append(attom_sale)

    last_purchase = max(sale_dates) if sale_dates else None

    # Acquisition velocity
    acq_velocity = None
    if sale_dates and len(sale_dates) >= 2:
        oldest = min(sale_dates)
        newest = max(sale_dates)
        span_years = (newest - oldest).days / 365.25
        if span_years > 0:
            acq_velocity = round(len(sale_dates) / span_years, 1)

    # Cash buyer
    cash_buyer = sc.get('derived_cash_buyer', '').strip() == 'True'

    # Absentee
    absentee = get_absentee_status(sc, port)

    # Mailing address
    mail_addr = sc.get('attom_mail_address', '') or ''
    if not mail_addr and port:
        r0 = port[0]
        parts = [r0.get('mail_street', ''), r0.get('mail_city', ''), r0.get('mail_state', ''), r0.get('mail_zip', '')]
        mail_addr = ', '.join(p.strip() for p in parts if p.strip())

    # ATTOM enrichment
    avm = safe_float(sc.get('attom_avm_value', 0))
    rent_est = safe_float(sc.get('attom_rent_estimate', 0))
    gross_yield = (rent_est * 12) / avm if avm > 0 and rent_est > 0 else None

    lender = sc.get('attom_lender_name', '') or ''
    loan_amount = safe_float(sc.get('attom_loan_amount', 0))
    permit_count = safe_int(sc.get('attom_permit_count', 0))
    permit_value = safe_float(sc.get('attom_total_permit_value', 0))
    year_built = safe_int(sc.get('attom_year_built', 0))

    # Score
    score, signal_list = compute_agent_score(owner_name, sc, port)
    tier = get_tier(score)
    segment = get_segment(owner_name, score, signal_list, n_props, sc, port)

    # Key signals string
    key_signals_parts = []
    if cash_buyer:
        key_signals_parts.append("cash buyer")
    if absentee == 'Out-of-State':
        key_signals_parts.append("out-of-state")
    if n_props >= 10:
        key_signals_parts.append(f"{n_props} properties")
    elif n_props >= 5:
        key_signals_parts.append(f"{n_props} properties")
    if last_purchase and (TODAY - last_purchase).days <= 180:
        key_signals_parts.append("recent purchase")
    if permit_value > 100000:
        key_signals_parts.append(f"${permit_value:,.0f} permits")
    if year_built >= 2023:
        key_signals_parts.append("new construction")
    if is_llc:
        key_signals_parts.append("entity-owned")
    if gross_yield and gross_yield > 0.07:
        key_signals_parts.append(f"{gross_yield:.1%} yield")
    key_signals = ", ".join(key_signals_parts)

    approach = generate_approach(segment, owner_name, n_props, signal_list, cities, absentee, sc)

    investors.append({
        'priority': tier,
        'score': score,
        'segment': segment,
        'name': owner_name,
        'entity_type': entity_type,
        'n_properties': n_props,
        'total_value': total_portfolio_value,
        'avg_value': avg_value,
        'value_range': f"${min_value:,.0f} - ${max_value:,.0f}" if min_value != max_value else f"${avg_value:,.0f}",
        'markets': ', '.join(sorted(cities)),
        'last_purchase': last_purchase,
        'acq_velocity': acq_velocity,
        'cash_buyer': 'Yes' if cash_buyer else 'No',
        'absentee': absentee,
        'mail_address': mail_addr,
        'avm': avm,
        'rent_estimate': rent_est,
        'gross_yield': gross_yield,
        'lender': lender if lender else '—',
        'loan_amount': loan_amount,
        'permit_count': permit_count,
        'permit_value': permit_value,
        'year_built': year_built if year_built > 0 else None,
        'key_signals': key_signals,
        'approach': approach,
        # Keep raw data for portfolio detail
        '_showcase': sc,
        '_portfolio': port,
    })

# Sort by score descending
investors.sort(key=lambda x: -x['score'])

print("\nInvestor scores:")
for inv in investors:
    print(f"  {inv['score']:3d}  {inv['priority']:<20s}  {inv['segment']:<30s}  {inv['name']}")

# ── Create Workbook ────────────────────────────────────────────────────────
print("\nBuilding workbook...")
wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════
# TAB 1: CALL LIST
# ═══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Call List"
ws1.sheet_properties.tabColor = OCEAN

# Title row
ws1.merge_cells('A1:AA1')
title_cell = ws1['A1']
title_cell.value = "DSCR Lead Intelligence  |  Wake County Investor Call List  |  Prepared for Northside Realty"
title_cell.font = Font(name="Calibri", size=13, bold=True, color=DEEP_OCEAN)
title_cell.alignment = Alignment(horizontal="left", vertical="center")
title_cell.fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
ws1.row_dimensions[1].height = 32

# Subtitle row
ws1.merge_cells('A2:AA2')
sub_cell = ws1['A2']
sub_cell.value = f"Generated {TODAY.strftime('%B %d, %Y')}  |  15 Investors  |  {sum(inv['n_properties'] for inv in investors)} Properties  |  Sorted by Agent Score (highest priority first)"
sub_cell.font = Font(name="Calibri", size=10, italic=True, color=STONE)
sub_cell.alignment = Alignment(horizontal="left", vertical="center")
sub_cell.fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
ws1.row_dimensions[2].height = 20

# Header row (row 3)
CALL_LIST_HEADERS = [
    "Priority", "Agent Score", "Segment", "Investor Name", "Entity Type",
    "Properties", "Total Portfolio Value", "Avg Property Value", "Property Range",
    "Markets", "Last Purchase", "Acq. Velocity", "Cash Buyer", "Absentee",
    "Mailing Address", "Sample AVM", "Sample Rent Est.", "Gross Yield",
    "Lender", "Loan Amount", "Building Permits", "Permit Value",
    "Year Built", "Key Signals", "Recommended Approach",
    "Phone", "Email"
]

HEADER_ROW = 3
for col_idx, header in enumerate(CALL_LIST_HEADERS, 1):
    cell = ws1.cell(row=HEADER_ROW, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

ws1.row_dimensions[HEADER_ROW].height = 36
ws1.freeze_panes = f"A{HEADER_ROW + 1}"

# Data rows
for row_idx, inv in enumerate(investors, HEADER_ROW + 1):
    row_data = [
        inv['priority'],
        inv['score'],
        inv['segment'],
        inv['name'],
        inv['entity_type'],
        inv['n_properties'],
        inv['total_value'],
        inv['avg_value'],
        inv['value_range'],
        inv['markets'],
        inv['last_purchase'].strftime('%Y-%m-%d') if inv['last_purchase'] else '—',
        f"{inv['acq_velocity']:.1f}/yr" if inv['acq_velocity'] else '—',
        inv['cash_buyer'],
        inv['absentee'],
        inv['mail_address'],
        inv['avm'],
        inv['rent_estimate'],
        inv['gross_yield'],
        inv['lender'],
        inv['loan_amount'],
        inv['permit_count'],
        inv['permit_value'],
        inv['year_built'],
        inv['key_signals'],
        inv['approach'],
        '[skip trace pending]',
        '[skip trace pending]',
    ]

    # Determine row fill
    is_odd = (row_idx - HEADER_ROW) % 2 == 1
    if 'Tier 1' in inv['priority']:
        row_fill = tier1_fill
    elif 'Tier 2' in inv['priority']:
        row_fill = tier2_fill
    else:
        row_fill = cloud_fill if is_odd else white_fill

    for col_idx, value in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        cell.fill = row_fill

        # Column-specific formatting
        col_header = CALL_LIST_HEADERS[col_idx - 1]

        if col_header in ("Total Portfolio Value", "Avg Property Value", "Sample AVM", "Loan Amount", "Permit Value"):
            cell.number_format = currency_fmt
            cell.alignment = data_center
            if value == 0 or value is None:
                cell.value = '—'

        if col_header == "Gross Yield":
            if value and value > 0:
                cell.number_format = pct_fmt
            else:
                cell.value = '—'
            cell.alignment = data_center

        if col_header == "Sample Rent Est.":
            if value and value > 0:
                cell.number_format = currency_fmt
            else:
                cell.value = '—'
            cell.alignment = data_center

        if col_header in ("Agent Score", "Properties", "Building Permits", "Year Built"):
            cell.alignment = data_center
            if col_header == "Year Built" and (not value or value == 0):
                cell.value = '—'

        if col_header in ("Priority", "Cash Buyer", "Absentee", "Entity Type"):
            cell.alignment = data_center

        if col_header == "Priority":
            cell.font = bold_font

        if col_header == "Agent Score":
            cell.font = Font(name="Calibri", size=11, bold=True, color=CHARCOAL)

        if col_header == "Recommended Approach":
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        if col_header == "Key Signals":
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(name="Calibri", size=9, italic=True, color=STONE)

        if col_header in ("Phone", "Email"):
            cell.font = Font(name="Calibri", size=9, italic=True, color=STONE)
            cell.alignment = data_center

    ws1.row_dimensions[row_idx].height = 52

# Conditional formatting for Agent Score column (column B)
score_col = 'B'
score_range = f'{score_col}{HEADER_ROW+1}:{score_col}{HEADER_ROW+len(investors)}'

ws1.conditional_formatting.add(score_range,
    CellIsRule(operator='greaterThan', formula=['60'],
              fill=PatternFill(start_color=SAGE_LIGHT, end_color=SAGE_LIGHT, fill_type="solid"),
              font=Font(name="Calibri", size=11, bold=True, color=SAGE)))

ws1.conditional_formatting.add(score_range,
    CellIsRule(operator='between', formula=['30', '60'],
              fill=PatternFill(start_color=AMBER_LIGHT, end_color=AMBER_LIGHT, fill_type="solid"),
              font=Font(name="Calibri", size=11, bold=True, color=AMBER)))

ws1.conditional_formatting.add(score_range,
    CellIsRule(operator='lessThan', formula=['30'],
              fill=PatternFill(start_color=CORAL_LIGHT, end_color=CORAL_LIGHT, fill_type="solid"),
              font=Font(name="Calibri", size=11, bold=True, color=CORAL)))

# Column widths for Call List
col_widths_1 = {
    1: 18, 2: 13, 3: 26, 4: 36, 5: 14, 6: 12, 7: 20, 8: 18, 9: 24,
    10: 32, 11: 15, 12: 14, 13: 13, 14: 16, 15: 36, 16: 16, 17: 16,
    18: 13, 19: 28, 20: 15, 21: 16, 22: 15, 23: 12, 24: 36, 25: 55,
    26: 20, 27: 20
}
for col, width in col_widths_1.items():
    ws1.column_dimensions[get_column_letter(col)].width = width

# Auto-filter
ws1.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(CALL_LIST_HEADERS))}{HEADER_ROW + len(investors)}"


# ═══════════════════════════════════════════════════════════════════════════
# TAB 2: PORTFOLIO DETAIL
# ═══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Portfolio Detail")
ws2.sheet_properties.tabColor = STONE

# Title
ws2.merge_cells('A1:R1')
t2 = ws2['A1']
t2.value = "DSCR Lead Intelligence  |  Portfolio Detail — One Row Per Property  |  Wake County, NC"
t2.font = Font(name="Calibri", size=13, bold=True, color=DEEP_OCEAN)
t2.alignment = Alignment(horizontal="left", vertical="center")
t2.fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
ws2.row_dimensions[1].height = 32

PORTFOLIO_HEADERS = [
    "Investor Name", "Agent Score", "Property Address", "City",
    "Assessed Value", "Land Value", "Improvement Value", "Land Ratio",
    "Year Built", "Lot Acres", "Sale Date", "Sale Price",
    "Use Code / Description", "AVM", "Rent Estimate",
    "Baths", "SqFt", "Permits"
]

PD_HEADER_ROW = 2
for col_idx, header in enumerate(PORTFOLIO_HEADERS, 1):
    cell = ws2.cell(row=PD_HEADER_ROW, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

ws2.row_dimensions[PD_HEADER_ROW].height = 30
ws2.freeze_panes = f"A{PD_HEADER_ROW + 1}"

# Collect all property rows, sorted by investor score then name
all_props = []
for inv in investors:
    score = inv['score']
    owner = inv['name']
    sc = inv['_showcase']

    for r in inv['_portfolio']:
        prop_addr = r.get('prop_street', '') or ''
        city = r.get('prop_city', '') or ''
        just_val = safe_float(r.get('just_value', 0))
        land_val = safe_float(r.get('land_value', 0))
        imp_val = safe_float(r.get('improvement_value', 0))
        land_ratio = land_val / just_val if just_val > 0 else None
        yr_built = safe_int(r.get('year_built', 0))
        lot_acres = safe_float(r.get('lot_acres', 0))
        sale_date = r.get('sale_date', '') or ''
        sale_price = safe_float(r.get('sale_price', 0))
        use_desc = r.get('use_description', '') or ''
        use_code = r.get('use_code', '') or ''
        use_combined = f"{use_code} — {use_desc}" if use_code and use_desc else use_desc or use_code

        # Check if this property has ATTOM data (is it the showcased property?)
        sc_addr = sc.get('address', '').upper().replace(',', '').strip()
        prop_full = f"{prop_addr} {city} NC".upper().replace(',', '').strip()

        has_attom = False
        avm_val = None
        rent_val = None
        baths_val = None
        sqft_val = None
        permits_val = None

        # Match ATTOM data to showcase address
        if sc_addr and prop_addr:
            # Simple check: does the prop street appear in the showcase address?
            prop_street_clean = prop_addr.upper().strip()
            sc_addr_clean = sc_addr.upper().strip()
            if prop_street_clean and (prop_street_clean in sc_addr_clean or sc_addr_clean.startswith(prop_street_clean.split()[0])):
                has_attom = True
                avm_val = safe_float(sc.get('attom_avm_value', 0))
                rent_val = safe_float(sc.get('attom_rent_estimate', 0))
                baths_val = safe_float(sc.get('attom_baths_full', 0)) or safe_float(sc.get('attom_baths_total', 0))
                sqft_val = safe_float(sc.get('attom_sqft', 0))
                permits_val = safe_int(sc.get('attom_permit_count', 0))

        # Also check from wake_qualified for beds/baths/sqft
        wq_sqft = safe_float(r.get('living_sqft', 0))
        wq_baths = safe_float(r.get('bathrooms', 0))

        if not sqft_val and wq_sqft:
            sqft_val = wq_sqft
        if not baths_val and wq_baths:
            baths_val = wq_baths

        all_props.append({
            'owner': owner,
            'score': score,
            'address': prop_addr,
            'city': city,
            'assessed': just_val,
            'land': land_val,
            'improvement': imp_val,
            'land_ratio': land_ratio,
            'year_built': yr_built,
            'lot_acres': lot_acres,
            'sale_date': sale_date,
            'sale_price': sale_price,
            'use': use_combined,
            'avm': avm_val if has_attom and avm_val else None,
            'rent': rent_val if has_attom and rent_val else None,
            'baths': baths_val if baths_val else None,
            'sqft': sqft_val if sqft_val else None,
            'permits': permits_val if has_attom and permits_val else None,
        })

print(f"\nTotal portfolio properties: {len(all_props)}")

# Write portfolio data
for row_idx, prop in enumerate(all_props, PD_HEADER_ROW + 1):
    row_data = [
        prop['owner'],
        prop['score'],
        prop['address'],
        prop['city'],
        prop['assessed'] if prop['assessed'] else '',
        prop['land'] if prop['land'] else '',
        prop['improvement'] if prop['improvement'] else '',
        prop['land_ratio'],
        prop['year_built'] if prop['year_built'] else '',
        prop['lot_acres'] if prop['lot_acres'] else '',
        prop['sale_date'],
        prop['sale_price'] if prop['sale_price'] else '',
        prop['use'],
        prop['avm'] if prop['avm'] else '',
        prop['rent'] if prop['rent'] else '',
        prop['baths'] if prop['baths'] else '',
        prop['sqft'] if prop['sqft'] else '',
        prop['permits'] if prop['permits'] else '',
    ]

    is_odd = (row_idx - PD_HEADER_ROW) % 2 == 1
    row_fill_pd = cloud_fill if is_odd else white_fill

    for col_idx, value in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=value)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        cell.fill = row_fill_pd

        col_name = PORTFOLIO_HEADERS[col_idx - 1]

        if col_name in ("Assessed Value", "Land Value", "Improvement Value", "AVM", "Rent Estimate", "Sale Price"):
            if value and value != '':
                cell.number_format = currency_fmt
            cell.alignment = data_center

        if col_name == "Land Ratio":
            if value is not None:
                cell.number_format = pct_fmt
            else:
                cell.value = ''
            cell.alignment = data_center

        if col_name in ("Agent Score", "Year Built", "Baths", "SqFt", "Permits"):
            cell.alignment = data_center
            if col_name == "SqFt" and value:
                cell.number_format = number_fmt

        if col_name == "Lot Acres":
            if value:
                cell.number_format = '0.00'
            cell.alignment = data_center

# Column widths for Portfolio Detail
col_widths_2 = {
    1: 36, 2: 13, 3: 32, 4: 18, 5: 16, 6: 14, 7: 18, 8: 12,
    9: 12, 10: 12, 11: 14, 12: 14, 13: 28, 14: 14, 15: 14,
    16: 12, 17: 12, 18: 12
}
for col, width in col_widths_2.items():
    ws2.column_dimensions[get_column_letter(col)].width = width

# Auto-filter
ws2.auto_filter.ref = f"A{PD_HEADER_ROW}:{get_column_letter(len(PORTFOLIO_HEADERS))}{PD_HEADER_ROW + len(all_props)}"

# Conditional formatting for score column
score_range_2 = f'B{PD_HEADER_ROW+1}:B{PD_HEADER_ROW+len(all_props)}'
ws2.conditional_formatting.add(score_range_2,
    CellIsRule(operator='greaterThan', formula=['60'],
              fill=PatternFill(start_color=SAGE_LIGHT, end_color=SAGE_LIGHT, fill_type="solid"),
              font=Font(name="Calibri", size=10, bold=True, color=SAGE)))
ws2.conditional_formatting.add(score_range_2,
    CellIsRule(operator='between', formula=['30', '60'],
              fill=PatternFill(start_color=AMBER_LIGHT, end_color=AMBER_LIGHT, fill_type="solid"),
              font=Font(name="Calibri", size=10, bold=True, color=AMBER)))
ws2.conditional_formatting.add(score_range_2,
    CellIsRule(operator='lessThan', formula=['30'],
              fill=PatternFill(start_color=CORAL_LIGHT, end_color=CORAL_LIGHT, fill_type="solid"),
              font=Font(name="Calibri", size=10, bold=True, color=CORAL)))


# ═══════════════════════════════════════════════════════════════════════════
# TAB 3: MARKET SUMMARY
# ═══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Market Summary")
ws3.sheet_properties.tabColor = GOLD

# Styling helpers
def write_label(ws, row, col, text, merge_end_col=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Calibri", size=10, bold=True, color=STONE)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    cell.fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)

def write_value(ws, row, col, text, fmt=None, merge_end_col=None, bold=False):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Calibri", size=12 if bold else 11, bold=bold, color=CHARCOAL)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    if fmt:
        cell.number_format = fmt
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)

def write_section_header(ws, row, col, text, end_col):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color=OCEAN, end_color=OCEAN, fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = thin_border
    ws.row_dimensions[row].height = 28

# Column widths
for col in range(1, 10):
    ws3.column_dimensions[get_column_letter(col)].width = 20

# Title
ws3.merge_cells('A1:H1')
t3 = ws3['A1']
t3.value = "DSCR Lead Intelligence  |  Wake County Market Summary"
t3.font = Font(name="Calibri", size=16, bold=True, color=DEEP_OCEAN)
t3.alignment = Alignment(horizontal="left", vertical="center")
t3.fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
ws3.row_dimensions[1].height = 40

ws3.merge_cells('A2:H2')
s3 = ws3['A2']
s3.value = f"Prepared for Nick Nicolaysen, Northside Realty  |  {TODAY.strftime('%B %d, %Y')}"
s3.font = Font(name="Calibri", size=11, italic=True, color=STONE)
s3.fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
ws3.row_dimensions[2].height = 24

# ── Overview KPIs ──────────────────────────────────────────────────────────
row = 4
write_section_header(ws3, row, 1, "OVERVIEW", 8)

row = 5
total_props = sum(inv['n_properties'] for inv in investors)
total_value = sum(inv['total_value'] for inv in investors)
avms = [inv['avm'] for inv in investors if inv['avm'] > 0]
avg_avm = sum(avms) / len(avms) if avms else 0

# KPI boxes - row 5-6
kpi_data = [
    ("Total Investors Profiled", "15"),
    ("Total Properties", f"{total_props:,}"),
    ("Total Portfolio Value", f"${total_value:,.0f}"),
    ("Avg Properties / Investor", f"{total_props/15:.1f}"),
]

for i, (label, value) in enumerate(kpi_data):
    col = i * 2 + 1
    cell_l = ws3.cell(row=5, column=col, value=label)
    cell_l.font = Font(name="Calibri", size=9, color=STONE)
    cell_l.alignment = Alignment(horizontal="center", vertical="bottom")
    cell_l.fill = PatternFill(start_color=CLOUD, end_color=CLOUD, fill_type="solid")
    cell_l.border = thin_border
    ws3.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)

    cell_v = ws3.cell(row=6, column=col, value=value)
    cell_v.font = Font(name="Calibri", size=18, bold=True, color=DEEP_OCEAN)
    cell_v.alignment = Alignment(horizontal="center", vertical="top")
    cell_v.fill = PatternFill(start_color=CLOUD, end_color=CLOUD, fill_type="solid")
    cell_v.border = thin_border
    ws3.merge_cells(start_row=6, start_column=col, end_row=6, end_column=col+1)

ws3.row_dimensions[5].height = 22
ws3.row_dimensions[6].height = 34

# Second row of KPIs
kpi_data_2 = [
    ("Avg AVM (Sampled)", f"${avg_avm:,.0f}"),
    ("Avg Portfolio Value", f"${total_value/15:,.0f}"),
    ("Cash Buyers", f"{sum(1 for inv in investors if inv['cash_buyer']=='Yes')}/15"),
    ("Out-of-State", f"{sum(1 for inv in investors if inv['absentee']=='Out-of-State')}/15"),
]

for i, (label, value) in enumerate(kpi_data_2):
    col = i * 2 + 1
    cell_l = ws3.cell(row=7, column=col, value=label)
    cell_l.font = Font(name="Calibri", size=9, color=STONE)
    cell_l.alignment = Alignment(horizontal="center", vertical="bottom")
    cell_l.fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    cell_l.border = thin_border
    ws3.merge_cells(start_row=7, start_column=col, end_row=7, end_column=col+1)

    cell_v = ws3.cell(row=8, column=col, value=value)
    cell_v.font = Font(name="Calibri", size=16, bold=True, color=CHARCOAL)
    cell_v.alignment = Alignment(horizontal="center", vertical="top")
    cell_v.fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    cell_v.border = thin_border
    ws3.merge_cells(start_row=8, start_column=col, end_row=8, end_column=col+1)

ws3.row_dimensions[7].height = 22
ws3.row_dimensions[8].height = 30

# ── Score Distribution ─────────────────────────────────────────────────────
row = 10
write_section_header(ws3, row, 1, "SCORE DISTRIBUTION", 8)

tier_counts = defaultdict(int)
for inv in investors:
    tier_counts[inv['priority']] += 1

tier_table = [
    ("Tier 1 Priority", tier_counts.get("Tier 1 Priority", 0), "Immediate outreach — phone + personalized email", SEAFOAM_LIGHT),
    ("Tier 2 Opportunity", tier_counts.get("Tier 2 Opportunity", 0), "Targeted email sequence with market insights", AMBER_LIGHT),
    ("Tier 3 Watch", tier_counts.get("Tier 3 Watch", 0), "Add to market report drip — nurture", CLOUD),
]

row = 11
# Table headers
for ci, hdr in enumerate(["Tier", "Count", "Recommended Action"], 1):
    c = ws3.cell(row=row, column=ci, value=hdr)
    c.font = Font(name="Calibri", size=10, bold=True, color=STONE)
    c.border = thin_border
    c.fill = PatternFill(start_color=MIST, end_color=MIST, fill_type="solid")
# Merge action col
ws3.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)

for tier_name, count, action, fill_color in tier_table:
    row += 1
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    c1 = ws3.cell(row=row, column=1, value=tier_name)
    c1.font = Font(name="Calibri", size=10, bold=True, color=CHARCOAL)
    c1.fill = fill
    c1.border = thin_border

    c2 = ws3.cell(row=row, column=2, value=count)
    c2.font = Font(name="Calibri", size=14, bold=True, color=DEEP_OCEAN)
    c2.alignment = Alignment(horizontal="center")
    c2.fill = fill
    c2.border = thin_border

    c3 = ws3.cell(row=row, column=3, value=action)
    c3.font = Font(name="Calibri", size=10, color=STONE)
    c3.fill = fill
    c3.border = thin_border
    ws3.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)

# ── Investors by Segment ───────────────────────────────────────────────────
row += 2
write_section_header(ws3, row, 1, "INVESTORS BY SEGMENT", 8)

segment_counts = defaultdict(list)
for inv in investors:
    segment_counts[inv['segment']].append(inv['name'])

row += 1
for ci, hdr in enumerate(["Segment", "Count", "Investors"], 1):
    c = ws3.cell(row=row, column=ci, value=hdr)
    c.font = Font(name="Calibri", size=10, bold=True, color=STONE)
    c.border = thin_border
    c.fill = PatternFill(start_color=MIST, end_color=MIST, fill_type="solid")
ws3.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)

# Sort segments by priority order
seg_order = [
    "Serial Acquirer (10+)", "Active Developer", "Out-of-State Investor",
    "High-Velocity Buyer", "Portfolio Builder (5-9)", "Cash Buyer",
    "Growing Investor (2-4)", "Value Investor", "Long-Hold / Disposition Candidate",
    "General Investor"
]

for seg in seg_order:
    if seg in segment_counts:
        row += 1
        is_odd = (row % 2 == 1)
        fill = cloud_fill if is_odd else white_fill

        c1 = ws3.cell(row=row, column=1, value=seg)
        c1.font = Font(name="Calibri", size=10, bold=True, color=CHARCOAL)
        c1.fill = fill
        c1.border = thin_border

        c2 = ws3.cell(row=row, column=2, value=len(segment_counts[seg]))
        c2.font = Font(name="Calibri", size=11, bold=True, color=DEEP_OCEAN)
        c2.alignment = Alignment(horizontal="center")
        c2.fill = fill
        c2.border = thin_border

        names = ', '.join(segment_counts[seg])
        c3 = ws3.cell(row=row, column=3, value=names)
        c3.font = Font(name="Calibri", size=9, color=STONE)
        c3.fill = fill
        c3.border = thin_border
        c3.alignment = Alignment(wrap_text=True)
        ws3.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
        ws3.row_dimensions[row].height = 28

# ── Investors by City ──────────────────────────────────────────────────────
row += 2
write_section_header(ws3, row, 1, "PROPERTIES BY CITY", 8)

city_counts = defaultdict(int)
for inv in investors:
    for r in inv['_portfolio']:
        city = r.get('prop_city', '').strip()
        if city:
            city_counts[city] += 1

row += 1
for ci, hdr in enumerate(["City", "Properties", "% of Total"], 1):
    c = ws3.cell(row=row, column=ci, value=hdr)
    c.font = Font(name="Calibri", size=10, bold=True, color=STONE)
    c.border = thin_border
    c.fill = PatternFill(start_color=MIST, end_color=MIST, fill_type="solid")

for city, count in sorted(city_counts.items(), key=lambda x: -x[1]):
    row += 1
    is_odd = (row % 2 == 1)
    fill = cloud_fill if is_odd else white_fill

    c1 = ws3.cell(row=row, column=1, value=city)
    c1.font = data_font
    c1.fill = fill
    c1.border = thin_border

    c2 = ws3.cell(row=row, column=2, value=count)
    c2.font = Font(name="Calibri", size=10, bold=True, color=CHARCOAL)
    c2.alignment = Alignment(horizontal="center")
    c2.fill = fill
    c2.border = thin_border

    c3 = ws3.cell(row=row, column=3, value=count/total_props if total_props > 0 else 0)
    c3.number_format = pct_fmt
    c3.font = data_font
    c3.alignment = Alignment(horizontal="center")
    c3.fill = fill
    c3.border = thin_border

# ── Data Sources ───────────────────────────────────────────────────────────
row += 2
write_section_header(ws3, row, 1, "DATA SOURCES", 8)

sources = [
    "County assessor records (Wake County, NC)",
    "NC Secretary of State (entity filings)",
    "ATTOM Property Data (7-endpoint enrichment: AVM, rent estimate, sales history, permits, assessment, ownership, property profile)",
    "Tracerfy skip trace (phone/email append — pending)",
    "Agent scoring model v1.0 (weighted signal scoring optimized for real estate agents)",
]

for src in sources:
    row += 1
    ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws3.cell(row=row, column=1, value=f"  {src}")
    c.font = Font(name="Calibri", size=10, color=STONE)
    c.fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

# ── Footer / Disclaimer ───────────────────────────────────────────────────
row += 2
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
c = ws3.cell(row=row, column=1, value="DSCR Lead Intelligence  |  Confidential — Prepared exclusively for Northside Realty")
c.font = Font(name="Calibri", size=9, italic=True, color=STONE)
c.alignment = Alignment(horizontal="center")

row += 1
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
c = ws3.cell(row=row, column=1, value="All data sourced from public records. No private borrower data (SSN, income, credit) is included or available.")
c.font = Font(name="Calibri", size=8, italic=True, color=MIST)
c.alignment = Alignment(horizontal="center")


# ═══════════════════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════════════════
print(f"\nSaving to {OUTPUT_PATH}...")
wb.save(OUTPUT_PATH)
print("Done!")
print(f"\nWorkbook: {OUTPUT_PATH}")
print(f"  Tab 1: Call List ({len(investors)} investors)")
print(f"  Tab 2: Portfolio Detail ({len(all_props)} properties)")
print(f"  Tab 3: Market Summary")
