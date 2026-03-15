"""
Build DSCR Lead Gen internal cost model workbook.

Generates: sales/internal/cost_model.xlsx
Run:       python sales/internal/build_cost_model.py
"""

from pathlib import Path
from math import ceil
from copy import copy

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent.parent
OUTPUT_PATH = PROJECT_DIR / "sales" / "internal" / "cost_model.xlsx"

# ---------------------------------------------------------------------------
# Color palette
# ---------------------------------------------------------------------------
NAVY = "1a237e"
TEAL = "00796b"
ORANGE = "e65100"
YELLOW_INPUT = "fff9c4"
GREEN_POS = "c8e6c9"
RED_NEG = "ffcdd2"
LIGHT_GRAY = "f5f5f5"
WHITE = "FFFFFF"

FILL_NAVY = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
FILL_TEAL = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
FILL_ORANGE = PatternFill(start_color=ORANGE, end_color=ORANGE, fill_type="solid")
FILL_YELLOW = PatternFill(start_color=YELLOW_INPUT, end_color=YELLOW_INPUT, fill_type="solid")
FILL_GREEN = PatternFill(start_color=GREEN_POS, end_color=GREEN_POS, fill_type="solid")
FILL_RED = PatternFill(start_color=RED_NEG, end_color=RED_NEG, fill_type="solid")
FILL_GRAY = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

FONT_HEADER = Font(name="Calibri", bold=True, color=WHITE, size=11)
FONT_SECTION = Font(name="Calibri", bold=True, color=WHITE, size=12)
FONT_BOLD = Font(name="Calibri", bold=True, size=11)
FONT_NORMAL = Font(name="Calibri", size=11)
FONT_INPUT = Font(name="Calibri", bold=True, size=11, color="333333")

THIN_BORDER = Border(
    left=Side(style="thin", color="cccccc"),
    right=Side(style="thin", color="cccccc"),
    top=Side(style="thin", color="cccccc"),
    bottom=Side(style="thin", color="cccccc"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def style_header_row(ws, row, max_col, fill=FILL_NAVY):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def style_section_row(ws, row, max_col, fill=FILL_TEAL):
    """Merge and style a section header across columns."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1)
    cell.fill = fill
    cell.font = FONT_SECTION
    cell.alignment = ALIGN_LEFT
    for col in range(1, max_col + 1):
        ws.cell(row=row, column=col).fill = fill
        ws.cell(row=row, column=col).border = THIN_BORDER


def style_data_cell(cell, row_idx, is_currency=False, is_pct=False, is_input=False):
    """Style a normal data cell with alternating row colors."""
    cell.font = FONT_INPUT if is_input else FONT_NORMAL
    cell.border = THIN_BORDER
    cell.alignment = ALIGN_CENTER
    if is_input:
        cell.fill = FILL_YELLOW
    elif row_idx % 2 == 0:
        cell.fill = FILL_GRAY
    else:
        cell.fill = FILL_WHITE
    if is_currency:
        cell.number_format = '$#,##0.000'
    if is_pct:
        cell.number_format = '0%'


def set_col_widths(ws, widths):
    """Set column widths from a dict {col_letter: width}."""
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


def write_row(ws, row, values, styles=None):
    """Write a list of values into a row starting at col 1."""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.border = THIN_BORDER
        cell.font = FONT_NORMAL
        cell.alignment = ALIGN_CENTER
    return row


# ---------------------------------------------------------------------------
# Tab 1: Per-Lead Cost Breakdown
# ---------------------------------------------------------------------------
def build_tab1(wb):
    print("  Building Tab 1: Per-Lead Cost Breakdown...")
    ws = wb.active
    ws.title = "Per-Lead Costs"
    max_col = 5

    set_col_widths(ws, {"A": 28, "B": 32, "C": 28, "D": 12, "E": 18})

    # --- Main cost table ---
    r = 1
    style_section_row(ws, r, max_col, FILL_TEAL)
    ws.cell(row=r, column=1, value="Per-Lead Cost Breakdown")

    r = 2
    headers = ["Data Point", "Source", "Cost Structure", "Hit Rate", "Effective $/Lead"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, max_col)

    cost_data = [
        ("Property data", "FDOR NAL / NC OneMap", "FREE", 1.00, 0.000),
        ("Entity resolution", "SunBiz / NC SoS", "FREE (FL) / TBD (NC)", 0.87, 0.000),
        ("Purchase history", "FDOR SDF / County", "FREE", 0.69, 0.000),
        ("Rent estimates", "HUD FMR", "FREE", 1.00, 0.000),
        ("Mortgage/lender (clerk)", "2Captcha solves", "$0.003/solve", 0.94, 0.003),
        ("Mortgage/lender (ATTOM)", "ATTOM Premium Property API", "See ATTOM tiers below", 0.27, None),
        ("Skip trace (phone+email)", "Tracerfy", "$0.02/match (not upload)", 0.45, 0.009),
        ("Email validation", "MillionVerifier", "$0.00245/email", None, 0.001),
        ("Phone type detection", "Twilio v2", "$0.008/lookup", 1.00, 0.008),
        ("DNC scrub (federal)", "FTC registry", "FREE (4 area codes)", 1.00, 0.000),
        ("DNC scrub (comprehensive)", "Tracerfy DNC", "$0.02/phone", 1.00, 0.009),
        ("Wealth signals", "FEC.gov / ProPublica", "FREE", "2-5%", 0.000),
    ]

    for i, (dp, src, cost_str, hit, eff) in enumerate(cost_data):
        r = 3 + i
        ws.cell(row=r, column=1, value=dp)
        ws.cell(row=r, column=2, value=src)
        ws.cell(row=r, column=3, value=cost_str)

        hit_cell = ws.cell(row=r, column=4)
        if isinstance(hit, float):
            hit_cell.value = hit
            hit_cell.number_format = '0%'
        elif hit is None:
            hit_cell.value = "N/A"
        else:
            hit_cell.value = hit

        eff_cell = ws.cell(row=r, column=5)
        if eff is not None:
            eff_cell.value = eff
            eff_cell.number_format = '$#,##0.000'
        else:
            eff_cell.value = "Variable"

        # Alternating row fill
        fill = FILL_GRAY if i % 2 == 0 else FILL_WHITE
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_CENTER if c >= 3 else ALIGN_LEFT

    # --- ATTOM pricing table ---
    r = 3 + len(cost_data) + 2
    attom_start = r
    style_section_row(ws, r, max_col, FILL_TEAL)
    ws.cell(row=r, column=1, value="ATTOM Pricing Tiers")

    r += 1
    attom_headers = ["Calls/Month", "Annual Cost", "Monthly Cost", "Per-Call", ""]
    for c, h in enumerate(attom_headers, 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, max_col)

    attom_data = [
        (5000, 1000, 83.33, 0.017),
        (10000, 1500, 125.00, 0.013),
        (25000, 2500, 208.00, 0.008),
        (50000, 4000, 333.00, 0.007),
        (100000, 6000, 500.00, 0.005),
    ]

    for i, (calls, annual, monthly, per_call) in enumerate(attom_data):
        r += 1
        ws.cell(row=r, column=1, value=calls).number_format = '#,##0'
        ws.cell(row=r, column=2, value=annual).number_format = '$#,##0'
        ws.cell(row=r, column=3, value=monthly).number_format = '$#,##0.00'
        ws.cell(row=r, column=4, value=per_call).number_format = '$#,##0.000'
        fill = FILL_GRAY if i % 2 == 0 else FILL_WHITE
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.font = FONT_NORMAL
            cell.alignment = ALIGN_CENTER

    # --- Totals ---
    r += 2
    style_section_row(ws, r, max_col, FILL_TEAL)
    ws.cell(row=r, column=1, value="Total Per-Lead Cost")

    # Without ATTOM: property(0) + entity(0) + purchase(0) + rent(0) +
    #   clerk(0.003) + skip(0.009) + email(0.001) + phone(0.008) + DNC_free(0) + wealth(0)
    cost_without_attom = 0.003 + 0.009 + 0.001 + 0.008
    # With ATTOM 5K: add ATTOM per-call 0.017 * 0.27 hit rate ≈ effective
    # Actually ATTOM is $0.017/call regardless of hit, so effective = $0.017
    cost_with_attom = cost_without_attom + 0.017

    r += 1
    ws.cell(row=r, column=1, value="Without ATTOM").font = FONT_BOLD
    ws.cell(row=r, column=1).alignment = ALIGN_LEFT
    ws.cell(row=r, column=5, value=cost_without_attom).number_format = '$#,##0.000'
    for c in range(1, max_col + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = FILL_GREEN
        cell.border = THIN_BORDER
    ws.cell(row=r, column=5).font = FONT_BOLD

    r += 1
    ws.cell(row=r, column=1, value="With ATTOM (5K tier)").font = FONT_BOLD
    ws.cell(row=r, column=1).alignment = ALIGN_LEFT
    ws.cell(row=r, column=5, value=cost_with_attom).number_format = '$#,##0.000'
    for c in range(1, max_col + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill = FILL_GREEN
        cell.border = THIN_BORDER
    ws.cell(row=r, column=5).font = FONT_BOLD

    return ws


# ---------------------------------------------------------------------------
# Tab 2: Pricing Scenario Modeling
# ---------------------------------------------------------------------------
def build_tab2(wb):
    print("  Building Tab 2: Pricing Scenario Modeling...")
    ws = wb.create_sheet("Pricing Scenarios")
    max_col = 6

    set_col_widths(ws, {"A": 30, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18})

    # --- Inputs section ---
    r = 1
    style_section_row(ws, r, max_col, FILL_TEAL)
    ws.cell(row=r, column=1, value="Adjustable Inputs (yellow cells are editable)")

    r = 2
    input_labels = [
        ("Pilot price ($)", 500),
        ("Pilot leads", 100),
        ("Starter price ($/mo)", 1500),
        ("Starter leads/mo", 250),
        ("Pro price ($/mo)", 3000),
        ("Pro leads/mo", 750),
        ("Enterprise price ($/mo)", 5000),
        ("Enterprise leads/mo", 2000),
        ("ATTOM tier", "None"),
        ("DNC method", "FTC Free"),
    ]

    # Input cells: labels in A, values in B
    input_rows = {}  # name -> row number for formula references
    for i, (label, default) in enumerate(input_labels):
        row = r + i
        ws.cell(row=row, column=1, value=label).font = FONT_BOLD
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=1).alignment = ALIGN_LEFT
        val_cell = ws.cell(row=row, column=2, value=default)
        val_cell.fill = FILL_YELLOW
        val_cell.font = FONT_INPUT
        val_cell.border = THIN_BORDER
        val_cell.alignment = ALIGN_CENTER
        if isinstance(default, (int, float)) and "price" in label.lower():
            val_cell.number_format = '$#,##0'
        input_rows[label] = row

    # Per-lead cost references (computed, not input)
    r_cost_section = r + len(input_labels) + 1
    ws.cell(row=r_cost_section, column=1, value="Per-lead cost (without ATTOM)").font = FONT_BOLD
    ws.cell(row=r_cost_section, column=1).border = THIN_BORDER
    cost_no_attom_cell = ws.cell(row=r_cost_section, column=2, value=0.021)
    cost_no_attom_cell.number_format = '$#,##0.000'
    cost_no_attom_cell.border = THIN_BORDER
    cost_no_attom_row = r_cost_section

    r_cost_section += 1
    ws.cell(row=r_cost_section, column=1, value="Per-lead cost (with ATTOM 5K)").font = FONT_BOLD
    ws.cell(row=r_cost_section, column=1).border = THIN_BORDER
    cost_attom_cell = ws.cell(row=r_cost_section, column=2, value=0.038)
    cost_attom_cell.number_format = '$#,##0.000'
    cost_attom_cell.border = THIN_BORDER
    cost_attom_row = r_cost_section

    # Per-lead cost that adapts to ATTOM selection
    r_cost_section += 1
    ws.cell(row=r_cost_section, column=1, value="Active per-lead cost").font = FONT_BOLD
    ws.cell(row=r_cost_section, column=1).border = THIN_BORDER
    ws.cell(row=r_cost_section, column=1).fill = FILL_GREEN
    # Formula: if ATTOM tier = "None" use without, else use with
    attom_tier_row = input_rows["ATTOM tier"]
    active_cost_row = r_cost_section
    formula = f'=IF(B{attom_tier_row}="None",B{cost_no_attom_row},B{cost_attom_row})'
    active_cell = ws.cell(row=r_cost_section, column=2, value=formula if isinstance(formula, str) else 0.021)
    active_cell.number_format = '$#,##0.000'
    active_cell.border = THIN_BORDER
    active_cell.fill = FILL_GREEN
    active_cell.font = FONT_BOLD
    # Actually write formula
    ws.cell(row=r_cost_section, column=2).value = formula

    # DNC adjustment note
    r_cost_section += 1
    ws.cell(row=r_cost_section, column=1, value="Note: DNC 'Tracerfy' adds $0.009/lead").font = Font(
        name="Calibri", italic=True, size=10, color="666666"
    )

    # --- Output table ---
    r_out = r_cost_section + 2
    style_section_row(ws, r_out, max_col, FILL_NAVY)
    ws.cell(row=r_out, column=1, value="Tier Economics (Calculated)")

    r_out += 1
    out_headers = ["Metric", "Pilot", "Starter", "Pro", "Enterprise"]
    for c, h in enumerate(out_headers, 1):
        ws.cell(row=r_out, column=c, value=h)
    style_header_row(ws, r_out, max_col - 1)  # 5 cols

    # Row references for tier inputs
    # Pilot: price B2, leads B3
    # Starter: price B4, leads B5
    # Pro: price B6, leads B7
    # Enterprise: price B8, leads B9
    pilot_price_ref = f"B{input_rows['Pilot price ($)']}"
    pilot_leads_ref = f"B{input_rows['Pilot leads']}"
    starter_price_ref = f"B{input_rows['Starter price ($/mo)']}"
    starter_leads_ref = f"B{input_rows['Starter leads/mo']}"
    pro_price_ref = f"B{input_rows['Pro price ($/mo)']}"
    pro_leads_ref = f"B{input_rows['Pro leads/mo']}"
    ent_price_ref = f"B{input_rows['Enterprise price ($/mo)']}"
    ent_leads_ref = f"B{input_rows['Enterprise leads/mo']}"
    active_cost_ref = f"B{active_cost_row}"

    metrics = [
        (
            "Revenue / client / mo",
            f"={pilot_price_ref}",
            f"={starter_price_ref}",
            f"={pro_price_ref}",
            f"={ent_price_ref}",
            '$#,##0',
        ),
        (
            "Leads delivered",
            f"={pilot_leads_ref}",
            f"={starter_leads_ref}",
            f"={pro_leads_ref}",
            f"={ent_leads_ref}",
            '#,##0',
        ),
        (
            "COGS / client / mo",
            f"={pilot_leads_ref}*{active_cost_ref}",
            f"={starter_leads_ref}*{active_cost_ref}",
            f"={pro_leads_ref}*{active_cost_ref}",
            f"={ent_leads_ref}*{active_cost_ref}",
            '$#,##0.00',
        ),
        (
            "Gross margin $",
            f"={pilot_price_ref}-{pilot_leads_ref}*{active_cost_ref}",
            f"={starter_price_ref}-{starter_leads_ref}*{active_cost_ref}",
            f"={pro_price_ref}-{pro_leads_ref}*{active_cost_ref}",
            f"={ent_price_ref}-{ent_leads_ref}*{active_cost_ref}",
            '$#,##0.00',
        ),
    ]

    # We'll also add gross margin % row
    # Need references to the revenue and margin rows
    revenue_row = r_out + 1
    margin_dollar_row = r_out + 4  # after revenue, leads, COGS, margin$

    for i, (metric, *formulas_and_fmt) in enumerate(metrics):
        row = r_out + 1 + i
        ws.cell(row=row, column=1, value=metric).font = FONT_BOLD
        ws.cell(row=row, column=1).alignment = ALIGN_LEFT
        ws.cell(row=row, column=1).border = THIN_BORDER
        fmt = formulas_and_fmt[-1]
        for c, formula in enumerate(formulas_and_fmt[:-1], 2):
            cell = ws.cell(row=row, column=c, value=formula)
            cell.number_format = fmt
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER
            cell.font = FONT_NORMAL
            fill = FILL_GRAY if i % 2 == 0 else FILL_WHITE
            cell.fill = fill

    # Gross margin % row
    gm_pct_row = r_out + 1 + len(metrics)
    ws.cell(row=gm_pct_row, column=1, value="Gross margin %").font = FONT_BOLD
    ws.cell(row=gm_pct_row, column=1).alignment = ALIGN_LEFT
    ws.cell(row=gm_pct_row, column=1).border = THIN_BORDER
    rev_row = r_out + 1  # revenue row
    gm_row = r_out + 4   # gross margin $ row
    for c in range(2, 6):
        col_l = get_column_letter(c)
        cell = ws.cell(
            row=gm_pct_row, column=c,
            value=f"=IF({col_l}{rev_row}=0,0,{col_l}{gm_row}/{col_l}{rev_row})"
        )
        cell.number_format = '0.0%'
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_CENTER
        cell.fill = FILL_GREEN

    # Cost to deliver one pilot batch
    pilot_batch_row = gm_pct_row + 1
    ws.cell(row=pilot_batch_row, column=1, value="Cost to deliver pilot batch").font = FONT_BOLD
    ws.cell(row=pilot_batch_row, column=1).alignment = ALIGN_LEFT
    ws.cell(row=pilot_batch_row, column=1).border = THIN_BORDER
    cell = ws.cell(row=pilot_batch_row, column=2, value=f"={pilot_leads_ref}*{active_cost_ref}")
    cell.number_format = '$#,##0.00'
    cell.border = THIN_BORDER
    cell.fill = FILL_GREEN
    cell.font = FONT_BOLD

    return ws


# ---------------------------------------------------------------------------
# Tab 3: Monthly P&L Projection
# ---------------------------------------------------------------------------
def build_tab3(wb):
    print("  Building Tab 3: Monthly P&L Projection...")
    ws = wb.create_sheet("Monthly P&L")
    max_col = 14  # A (labels) + B-M (months 1-12) + N (total)

    set_col_widths(ws, {"A": 28})
    for c in range(2, max_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 14

    # Shared assumptions (using Starter tier defaults)
    tier_price = 1500
    tier_leads = 250
    per_lead_cost = 0.021  # without ATTOM
    fixed_domain = 1  # $12/yr
    fixed_email = 0
    fixed_attom = 0
    fixed_tools = 20
    fixed_total = fixed_domain + fixed_email + fixed_attom + fixed_tools  # $21

    scenarios = [
        ("Conservative (1 new every 3 months)", [0, 0, 1, 0, 0, 1, 0, 0, 1, 0, 0, 1]),
        ("Moderate (1 new every 2 months)", [0, 1, 0, 1, 0, 1, 0, 1, 0, 1, 0, 1]),
        ("Aggressive (ramp up)", [1, 1, 1, 1, 1, 2, 2, 2, 2, 2, 2, 2]),
    ]

    current_row = 1

    for s_idx, (scenario_name, new_clients_list) in enumerate(scenarios):
        r = current_row

        # Section header
        style_section_row(ws, r, max_col, FILL_TEAL)
        ws.cell(row=r, column=1, value=f"Scenario: {scenario_name}")
        r += 1

        # Assumptions row
        ws.cell(row=r, column=1, value="Tier: Starter ($1,500/mo, 250 leads)").font = Font(
            name="Calibri", italic=True, size=10, color="666666"
        )
        ws.cell(row=r, column=1).border = THIN_BORDER
        r += 1

        # Month headers
        ws.cell(row=r, column=1, value="").border = THIN_BORDER
        for m in range(1, 13):
            ws.cell(row=r, column=m + 1, value=f"Mo {m}")
        ws.cell(row=r, column=14, value="Total")
        style_header_row(ws, r, max_col)
        header_row = r

        # --- New clients (editable / yellow) ---
        r += 1
        new_clients_row = r
        ws.cell(row=r, column=1, value="New clients this month").font = FONT_BOLD
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        for m in range(12):
            cell = ws.cell(row=r, column=m + 2, value=new_clients_list[m])
            cell.fill = FILL_YELLOW
            cell.font = FONT_INPUT
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER
        # Total
        total_col_letter = get_column_letter(14)
        first_data = get_column_letter(2)
        last_data = get_column_letter(13)
        ws.cell(row=r, column=14, value=f"=SUM({first_data}{r}:{last_data}{r})")
        ws.cell(row=r, column=14).border = THIN_BORDER
        ws.cell(row=r, column=14).font = FONT_BOLD

        # --- Cumulative clients (formula) ---
        r += 1
        cum_row = r
        ws.cell(row=r, column=1, value="Cumulative clients").font = FONT_BOLD
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        for m in range(12):
            col = m + 2
            col_l = get_column_letter(col)
            if m == 0:
                formula = f"={col_l}{new_clients_row}"
            else:
                prev_col_l = get_column_letter(col - 1)
                formula = f"={prev_col_l}{cum_row}+{col_l}{new_clients_row}"
            ws.cell(row=r, column=col, value=formula)
            ws.cell(row=r, column=col).border = THIN_BORDER
            ws.cell(row=r, column=col).alignment = ALIGN_CENTER
        # Total col shows final cumulative
        ws.cell(row=r, column=14, value=f"={get_column_letter(13)}{r}")
        ws.cell(row=r, column=14).border = THIN_BORDER
        ws.cell(row=r, column=14).font = FONT_BOLD

        # --- Revenue ---
        r += 1
        rev_row = r
        ws.cell(row=r, column=1, value="Revenue").font = FONT_BOLD
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        for m in range(12):
            col = m + 2
            col_l = get_column_letter(col)
            formula = f"={col_l}{cum_row}*{tier_price}"
            cell = ws.cell(row=r, column=col, value=formula)
            cell.number_format = '$#,##0'
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER
        ws.cell(row=r, column=14, value=f"=SUM({first_data}{r}:{last_data}{r})")
        ws.cell(row=r, column=14).number_format = '$#,##0'
        ws.cell(row=r, column=14).border = THIN_BORDER
        ws.cell(row=r, column=14).font = FONT_BOLD

        # --- COGS ---
        r += 1
        cogs_row = r
        ws.cell(row=r, column=1, value="COGS (lead production)").font = FONT_BOLD
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        for m in range(12):
            col = m + 2
            col_l = get_column_letter(col)
            formula = f"={col_l}{cum_row}*{tier_leads}*{per_lead_cost}"
            cell = ws.cell(row=r, column=col, value=formula)
            cell.number_format = '$#,##0.00'
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER
        ws.cell(row=r, column=14, value=f"=SUM({first_data}{r}:{last_data}{r})")
        ws.cell(row=r, column=14).number_format = '$#,##0.00'
        ws.cell(row=r, column=14).border = THIN_BORDER
        ws.cell(row=r, column=14).font = FONT_BOLD

        # --- Fixed costs breakdown ---
        fixed_items = [
            ("  Domain ($12/yr)", fixed_domain),
            ("  Email infra", fixed_email),
            ("  ATTOM", fixed_attom),
            ("  Tools", fixed_tools),
        ]
        fixed_start = r + 1
        for label, cost in fixed_items:
            r += 1
            ws.cell(row=r, column=1, value=label).font = FONT_NORMAL
            ws.cell(row=r, column=1).border = THIN_BORDER
            ws.cell(row=r, column=1).alignment = ALIGN_LEFT
            for m in range(12):
                cell = ws.cell(row=r, column=m + 2, value=cost)
                cell.number_format = '$#,##0'
                cell.border = THIN_BORDER
                cell.alignment = ALIGN_CENTER
                cell.fill = FILL_GRAY
            ws.cell(row=r, column=14, value=f"=SUM({first_data}{r}:{last_data}{r})")
            ws.cell(row=r, column=14).number_format = '$#,##0'
            ws.cell(row=r, column=14).border = THIN_BORDER
        fixed_end = r

        # --- Total fixed costs ---
        r += 1
        fixed_total_row = r
        ws.cell(row=r, column=1, value="Total fixed costs").font = FONT_BOLD
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        for m in range(12):
            col = m + 2
            col_l = get_column_letter(col)
            formula = f"=SUM({col_l}{fixed_start}:{col_l}{fixed_end})"
            cell = ws.cell(row=r, column=col, value=formula)
            cell.number_format = '$#,##0'
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER
        ws.cell(row=r, column=14, value=f"=SUM({first_data}{r}:{last_data}{r})")
        ws.cell(row=r, column=14).number_format = '$#,##0'
        ws.cell(row=r, column=14).border = THIN_BORDER
        ws.cell(row=r, column=14).font = FONT_BOLD

        # --- Gross profit ---
        r += 1
        gp_row = r
        ws.cell(row=r, column=1, value="Gross profit").font = FONT_BOLD
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        for m in range(12):
            col = m + 2
            col_l = get_column_letter(col)
            formula = f"={col_l}{rev_row}-{col_l}{cogs_row}"
            cell = ws.cell(row=r, column=col, value=formula)
            cell.number_format = '$#,##0.00'
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER
        ws.cell(row=r, column=14, value=f"=SUM({first_data}{r}:{last_data}{r})")
        ws.cell(row=r, column=14).number_format = '$#,##0.00'
        ws.cell(row=r, column=14).border = THIN_BORDER
        ws.cell(row=r, column=14).font = FONT_BOLD

        # --- Net income ---
        r += 1
        ni_row = r
        ws.cell(row=r, column=1, value="Net income").font = FONT_BOLD
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        for m in range(12):
            col = m + 2
            col_l = get_column_letter(col)
            formula = f"={col_l}{gp_row}-{col_l}{fixed_total_row}"
            cell = ws.cell(row=r, column=col, value=formula)
            cell.number_format = '$#,##0.00'
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER
        ws.cell(row=r, column=14, value=f"=SUM({first_data}{r}:{last_data}{r})")
        ws.cell(row=r, column=14).number_format = '$#,##0.00'
        ws.cell(row=r, column=14).border = THIN_BORDER
        ws.cell(row=r, column=14).font = FONT_BOLD

        # --- Cumulative P&L ---
        r += 1
        cum_pl_row = r
        ws.cell(row=r, column=1, value="Cumulative P&L").font = FONT_BOLD
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        for m in range(12):
            col = m + 2
            col_l = get_column_letter(col)
            if m == 0:
                formula = f"={col_l}{ni_row}"
            else:
                prev_col_l = get_column_letter(col - 1)
                formula = f"={prev_col_l}{cum_pl_row}+{col_l}{ni_row}"
            cell = ws.cell(row=r, column=col, value=formula)
            cell.number_format = '$#,##0.00'
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER
        ws.cell(row=r, column=14, value=f"={get_column_letter(13)}{r}")
        ws.cell(row=r, column=14).number_format = '$#,##0.00'
        ws.cell(row=r, column=14).border = THIN_BORDER
        ws.cell(row=r, column=14).font = FONT_BOLD

        # --- Color code net income and cumulative P&L based on computed values ---
        # We need to compute actual values to apply conditional coloring since
        # openpyxl can't do conditional formatting easily with our formulas
        cumulative_clients = 0
        cum_pl = 0.0
        breakeven_found = False
        for m in range(12):
            col = m + 2
            cumulative_clients += new_clients_list[m]
            revenue = cumulative_clients * tier_price
            cogs = cumulative_clients * tier_leads * per_lead_cost
            gp = revenue - cogs
            ni = gp - fixed_total
            cum_pl += ni

            # Color net income row
            ni_cell = ws.cell(row=ni_row, column=col)
            if ni >= 0:
                ni_cell.fill = FILL_GREEN
            else:
                ni_cell.fill = FILL_RED

            # Color cumulative P&L row
            cum_cell = ws.cell(row=cum_pl_row, column=col)
            if cum_pl >= 0:
                cum_cell.fill = FILL_GREEN
                if not breakeven_found:
                    # Highlight breakeven month with bold green
                    cum_cell.font = Font(name="Calibri", bold=True, size=11, color="1b5e20")
                    breakeven_found = True
            else:
                cum_cell.fill = FILL_RED

        # Alternating row colors for non-special rows
        data_rows = [new_clients_row, cum_row, rev_row, cogs_row,
                     fixed_total_row, gp_row]
        for idx, dr in enumerate(data_rows):
            fill = FILL_GRAY if idx % 2 == 0 else FILL_WHITE
            for col in range(2, 15):
                cell = ws.cell(row=dr, column=col)
                if cell.fill == PatternFill():  # only if not already filled
                    cell.fill = fill

        current_row = r + 3  # gap between scenarios

    return ws


# ---------------------------------------------------------------------------
# Tab 4: Unit Economics Summary
# ---------------------------------------------------------------------------
def build_tab4(wb):
    print("  Building Tab 4: Unit Economics Summary...")
    ws = wb.create_sheet("Unit Economics")
    max_col = 3

    set_col_widths(ws, {"A": 45, "B": 22, "C": 22})

    cost_no_attom = 0.021
    cost_with_attom = 0.038

    # Tier definitions
    tiers = {
        "Pilot": (500, 100),
        "Starter": (1500, 250),
        "Pro": (3000, 750),
        "Enterprise": (5000, 2000),
    }

    # --- Production costs ---
    r = 1
    style_section_row(ws, r, max_col, FILL_TEAL)
    ws.cell(row=r, column=1, value="Lead Production Costs")

    r = 2
    headers = ["Metric", "Value", "Notes"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, max_col)

    prod_data = [
        ("Cost per lead (without ATTOM)", cost_no_attom, "Clerk + Tracerfy + MV + Twilio"),
        ("Cost per lead (with ATTOM 5K tier)", cost_with_attom, "Adds $0.017/call ATTOM"),
    ]
    for i, (metric, val, note) in enumerate(prod_data):
        r += 1
        ws.cell(row=r, column=1, value=metric).font = FONT_BOLD
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        ws.cell(row=r, column=2, value=val).number_format = '$#,##0.000'
        ws.cell(row=r, column=3, value=note).font = Font(name="Calibri", size=10, color="666666")
        fill = FILL_GRAY if i % 2 == 0 else FILL_WHITE
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).border = THIN_BORDER

    # --- Revenue per lead ---
    r += 2
    style_section_row(ws, r, max_col, FILL_TEAL)
    ws.cell(row=r, column=1, value="Revenue Per Lead by Tier")

    r += 1
    for c, h in enumerate(["Tier", "Revenue/Lead", "Margin/Lead (no ATTOM)"], 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, max_col)

    for i, (tier_name, (price, leads)) in enumerate(tiers.items()):
        r += 1
        rev_per = price / leads
        margin_per = rev_per - cost_no_attom
        ws.cell(row=r, column=1, value=f"{tier_name} (${price:,} / {leads:,} leads)").font = FONT_BOLD
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        ws.cell(row=r, column=2, value=rev_per).number_format = '$#,##0.00'
        ws.cell(row=r, column=3, value=margin_per).number_format = '$#,##0.00'
        fill = FILL_GRAY if i % 2 == 0 else FILL_WHITE
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.border = THIN_BORDER
            if c == 3:
                cell.fill = FILL_GREEN

    # --- LO Breakeven ---
    r += 2
    style_section_row(ws, r, max_col, FILL_TEAL)
    ws.cell(row=r, column=1, value="LO (Loan Officer) Breakeven")

    r += 1
    for c, h in enumerate(["Metric", "Value", "Notes"], 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, max_col)

    avg_commission = 6000

    lo_data = [
        ("Avg DSCR commission", f"${avg_commission:,}", "Industry average per funded loan"),
        ("Deals to break even on Pilot ($500)", ceil(500 / avg_commission), f"ceil($500 / ${avg_commission:,}) = {ceil(500 / avg_commission)}"),
        ("Deals to break even on Starter ($1,500/mo)", ceil(1500 / avg_commission), f"ceil($1,500 / ${avg_commission:,}) = {ceil(1500 / avg_commission)}"),
        ("Deals to break even on Pro ($3,000/mo)", ceil(3000 / avg_commission), f"ceil($3,000 / ${avg_commission:,}) = {ceil(3000 / avg_commission)}"),
        ("Deals to break even on Enterprise ($5,000/mo)", ceil(5000 / avg_commission), f"ceil($5,000 / ${avg_commission:,}) = {ceil(5000 / avg_commission)}"),
    ]

    for i, (metric, val, note) in enumerate(lo_data):
        r += 1
        ws.cell(row=r, column=1, value=metric).font = FONT_BOLD
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        ws.cell(row=r, column=2, value=val)
        ws.cell(row=r, column=2).alignment = ALIGN_CENTER
        ws.cell(row=r, column=3, value=note).font = Font(name="Calibri", size=10, color="666666")
        fill = FILL_GRAY if i % 2 == 0 else FILL_WHITE
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).border = THIN_BORDER
        if i >= 1:  # deals rows
            ws.cell(row=r, column=2).fill = FILL_GREEN
            ws.cell(row=r, column=2).font = FONT_BOLD

    # --- Our Breakeven ---
    r += 2
    style_section_row(ws, r, max_col, FILL_TEAL)
    ws.cell(row=r, column=1, value="Our Breakeven (Clients Needed)")

    r += 1
    for c, h in enumerate(["Tier", "Clients to Cover Fixed Costs", "Monthly Fixed: $21"], 1):
        ws.cell(row=r, column=c, value=h)
    style_header_row(ws, r, max_col)

    fixed_monthly = 21  # domain $1 + tools $20

    for i, (tier_name, (price, leads)) in enumerate(tiers.items()):
        r += 1
        net_per_client = price - (leads * cost_no_attom) - 0  # no fixed alloc yet
        # Clients needed: fixed_monthly / net_per_client_contribution
        if net_per_client > 0:
            clients_needed = ceil(fixed_monthly / net_per_client)
        else:
            clients_needed = "N/A"

        ws.cell(row=r, column=1, value=tier_name).font = FONT_BOLD
        ws.cell(row=r, column=1).alignment = ALIGN_LEFT
        ws.cell(row=r, column=2, value=clients_needed)
        ws.cell(row=r, column=2).alignment = ALIGN_CENTER
        ws.cell(row=r, column=2).fill = FILL_GREEN
        ws.cell(row=r, column=2).font = FONT_BOLD
        net_str = f"Net ${net_per_client:,.2f}/client/mo after COGS"
        ws.cell(row=r, column=3, value=net_str).font = Font(name="Calibri", size=10, color="666666")
        fill = FILL_GRAY if i % 2 == 0 else FILL_WHITE
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if c != 2:
                ws.cell(row=r, column=c).fill = fill

    # --- Key insight box ---
    r += 2
    style_section_row(ws, r, max_col, FILL_ORANGE)
    ws.cell(row=r, column=1, value="Key Insight")

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=max_col)
    cell = ws.cell(row=r, column=1)
    cell.value = (
        "At $0.021/lead, COGS is negligible. A single Starter client ($1,500/mo) generates "
        "$1,494.75/mo gross margin (99.6%). The LO only needs 1 funded deal from our leads "
        "to more than cover a full month's subscription. This is a near-zero marginal cost business."
    )
    cell.font = Font(name="Calibri", size=11, color="333333")
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    for rr in range(r, r + 3):
        for c in range(1, max_col + 1):
            ws.cell(row=rr, column=c).fill = PatternFill(
                start_color="fff3e0", end_color="fff3e0", fill_type="solid"
            )
            ws.cell(row=rr, column=c).border = THIN_BORDER

    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("Building DSCR Lead Gen cost model...")
    print(f"  Output: {OUTPUT_PATH}")

    wb = openpyxl.Workbook()

    build_tab1(wb)
    build_tab2(wb)
    build_tab3(wb)
    build_tab4(wb)

    # Freeze panes on each sheet
    for ws in wb.worksheets:
        ws.sheet_properties.tabColor = NAVY

    wb.worksheets[0].sheet_properties.tabColor = TEAL
    wb.worksheets[1].sheet_properties.tabColor = "1565c0"
    wb.worksheets[2].sheet_properties.tabColor = "6a1b9a"
    wb.worksheets[3].sheet_properties.tabColor = ORANGE

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_PATH))
    print(f"  Saved: {OUTPUT_PATH}")
    print("Done.")


if __name__ == "__main__":
    main()
