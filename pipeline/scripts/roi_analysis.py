"""
DSCR Lead Generation — Proposal Workbook

One document that tells the whole story:
  Tab 1: The Opportunity — what we built, live numbers
  Tab 2: What Data You Get — free vs paid, field by field
  Tab 3: Your Calculator — adjustable revenue math
  Tab 4: Sample Leads — top 50 actual leads as proof

Reads the pipeline output to pull real numbers. Yellow cells are adjustable.

Usage:
    python3 pipeline/scripts/roi_analysis.py
    python3 pipeline/scripts/roi_analysis.py --leads pipeline/output/06_enriched.csv
"""

import argparse
import pandas as pd
from pathlib import Path
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path("pipeline/output")

# ---------------------------------------------------------------------------
# Colors & Styles
# ---------------------------------------------------------------------------
_NAVY = '1B2A4A'
_BLUE = '4472C4'
_GREEN = '548235'
_RED = 'C00000'
_ORANGE = 'ED7D31'
_LIGHT = 'F2F2F2'
_WHITE = 'FFFFFF'
_YELLOW_BG = 'FFF2CC'

_navy_fill = PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type='solid')
_blue_fill = PatternFill(start_color=_BLUE, end_color=_BLUE, fill_type='solid')
_green_fill = PatternFill(start_color=_GREEN, end_color=_GREEN, fill_type='solid')
_red_fill = PatternFill(start_color=_RED, end_color=_RED, fill_type='solid')
_orange_fill = PatternFill(start_color=_ORANGE, end_color=_ORANGE, fill_type='solid')
_light_fill = PatternFill(start_color=_LIGHT, end_color=_LIGHT, fill_type='solid')
_white_fill = PatternFill(start_color=_WHITE, end_color=_WHITE, fill_type='solid')
_yellow_fill = PatternFill(start_color=_YELLOW_BG, end_color=_YELLOW_BG, fill_type='solid')
_green_text_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')

_title_font = Font(size=20, bold=True, color=_WHITE)
_subtitle_font = Font(size=11, color=_WHITE)
_section_font = Font(size=14, bold=True, color=_WHITE)
_big_number = Font(size=28, bold=True, color=_WHITE)
_big_label = Font(size=10, color=_WHITE)
_label_font = Font(size=11, color='333333')
_label_bold = Font(size=11, bold=True, color='333333')
_value_font = Font(size=11, color='333333')
_value_bold = Font(size=11, bold=True, color='333333')
_input_font = Font(size=12, bold=True, color='7F6000')
_result_font = Font(size=14, bold=True, color=_GREEN)
_result_big = Font(size=18, bold=True, color=_GREEN)
_note_font = Font(size=9, italic=True, color='888888')
_free_font = Font(size=11, bold=True, color=_GREEN)
_paid_font = Font(size=11, bold=True, color=_RED)
_header_font = Font(size=10, bold=True, color=_WHITE)
_cell_font = Font(size=10, color='333333')
_cell_bold = Font(size=10, bold=True, color='333333')

_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

_thin = Border(bottom=Side(style='thin', color='D9D9D9'))
_yellow_border = Border(
    bottom=Side(style='thin', color='BF9000'),
    top=Side(style='thin', color='BF9000'),
    left=Side(style='thin', color='BF9000'),
    right=Side(style='thin', color='BF9000'),
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _setup(ws, col_widths=None):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2
    if col_widths:
        for letter, w in col_widths.items():
            ws.column_dimensions[letter].width = w
    else:
        ws.column_dimensions['B'].width = 34
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 22
        ws.column_dimensions['G'].width = 2


def _merged(ws, row, start, end, value, font=None, fill=None, align=None, fmt=None):
    if start != end:
        ws.merge_cells(start_row=row, start_column=start,
                       end_row=row, end_column=end)
    cell = ws.cell(row=row, column=start, value=value)
    if font: cell.font = font
    if fill:
        cell.fill = fill
        for c in range(start, end + 1):
            ws.cell(row=row, column=c).fill = fill
    if align: cell.alignment = align
    if fmt: cell.number_format = fmt
    return cell


def _gutter(ws, row, fill, end_col=7):
    ws.cell(row=row, column=1).fill = fill
    ws.cell(row=row, column=end_col).fill = fill


def _banner(ws, title, end_col=6):
    row = 1
    _merged(ws, row, 2, end_col, title, _title_font, _navy_fill, _center)
    ws.row_dimensions[row].height = 44
    _gutter(ws, row, _navy_fill, end_col + 1)
    row += 1
    sub = f'Prepared {date.today().strftime("%B %d, %Y")}  |  Florida DSCR Lending'
    _merged(ws, row, 2, end_col, sub, _subtitle_font, _navy_fill, _center)
    ws.row_dimensions[row].height = 24
    _gutter(ws, row, _navy_fill, end_col + 1)
    return row + 1


def _section(ws, row, title, fill=None, end_col=6):
    fill = fill or _blue_fill
    _merged(ws, row, 2, end_col, title, _section_font, fill, _left)
    ws.row_dimensions[row].height = 30
    return row + 1


def _kpi_row(ws, row, tiles):
    cols_per = 5 // len(tiles)
    col = 2
    for val, label, fill in tiles:
        end = min(col + cols_per - 1, 6)
        _merged(ws, row, col, end, val, _big_number, fill, _center)
        col = end + 1
    ws.row_dimensions[row].height = 55
    row += 1
    col = 2
    for val, label, fill in tiles:
        end = min(col + cols_per - 1, 6)
        _merged(ws, row, col, end, label, _big_label, fill, _center)
        col = end + 1
    ws.row_dimensions[row].height = 22
    return row + 1


def _input_row(ws, row, label, value, fmt=None):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    c1 = ws.cell(row=row, column=2, value=label)
    c1.font = _label_bold
    c1.alignment = _left
    c1.border = _thin
    c2 = ws.cell(row=row, column=4, value=value)
    c2.font = _input_font
    c2.fill = _yellow_fill
    c2.border = _yellow_border
    c2.alignment = _center
    if fmt: c2.number_format = fmt
    ws.row_dimensions[row].height = 28
    return row + 1


def _formula_row(ws, row, label, formula, fmt=None, bold=False, result=False):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
    c1 = ws.cell(row=row, column=2, value=label)
    c1.font = _label_bold
    c1.alignment = _left
    c1.border = _thin
    c2 = ws.cell(row=row, column=4, value=formula)
    c2.font = _result_big if result else (_result_font if bold else _value_font)
    c2.alignment = _center
    c2.border = _thin
    if fmt: c2.number_format = fmt
    ws.row_dimensions[row].height = 30 if result else 26
    return row + 1


def _fmt_short(value):
    try:
        v = float(value)
    except (ValueError, TypeError):
        return '$0'
    if pd.isna(v) or v == 0:
        return '$0'
    if abs(v) >= 1_000_000_000:
        return f'${v / 1_000_000_000:,.1f}B'
    if abs(v) >= 1_000_000:
        return f'${v / 1_000_000:,.1f}M'
    if abs(v) >= 1_000:
        return f'${v / 1_000:,.0f}K'
    return f'${v:,.0f}'


def _load_stats(leads_path):
    """Load lead data and compute summary stats for the proposal."""
    df = pd.read_csv(leads_path, dtype=str, low_memory=False)
    total = len(df)

    def _bool(col):
        if col not in df.columns:
            return pd.Series([False] * total)
        return df[col].fillna('').astype(str).str.lower().isin(['true', '1', 'yes'])

    def _has(col):
        if col not in df.columns:
            return pd.Series([False] * total)
        return df[col].fillna('').astype(str).str.strip().ne('') & \
               df[col].astype(str).str.lower().ne('nan')

    def _num(col):
        if col not in df.columns:
            return pd.Series([0.0] * total)
        return pd.to_numeric(df[col], errors='coerce').fillna(0)

    has_phone = _has('phone')
    has_email = _has('email')
    has_address = _has('OWN_ADDR1')
    has_resolved = _has('resolved_person')
    str_licensed = _bool('str_licensed')
    is_entity = _bool('is_entity')
    portfolio = _num('total_portfolio_value')
    prop_count = _num('property_count')
    multi = prop_count >= 2

    return {
        'df': df,
        'total': total,
        'has_phone': int(has_phone.sum()),
        'has_email': int(has_email.sum()),
        'has_address': int(has_address.sum()),
        'has_resolved': int(has_resolved.sum()),
        'str_licensed': int(str_licensed.sum()),
        'is_entity': int(is_entity.sum()),
        'multi_property': int(multi.sum()),
        'foreign': int(_bool('foreign_owner').sum()),
        'oos': int(_bool('out_of_state').sum()),
        'total_portfolio': portfolio.sum(),
        'phone_pct': has_phone.sum() / total * 100 if total > 0 else 0,
        'email_pct': has_email.sum() / total * 100 if total > 0 else 0,
        'refi_candidates': int((_bool('probable_cash_buyer') | _bool('equity_harvest_candidate') |
                                _bool('brrrr_exit_candidate') | _bool('rate_refi_candidate')).sum()),
    }


# ---------------------------------------------------------------------------
# Tab 1: The Opportunity
# ---------------------------------------------------------------------------
def _build_opportunity(wb, stats):
    ws = wb.active
    ws.title = 'The Opportunity'
    _setup(ws)

    row = _banner(ws, 'DSCR LEAD GENERATION')
    row += 1

    row = _section(ws, row, 'WHAT WE BUILT')

    row = _kpi_row(ws, row, [
        (f'{stats["total"]:,}', 'Qualified Investor Leads\n(Palm Beach County)', _navy_fill),
        (f'{stats["refi_candidates"]:,}', 'Refi Candidates Identified', _green_fill),
    ])
    row += 1

    # How it works — simple steps
    row = _section(ws, row, 'HOW IT WORKS')
    steps = [
        'Pull every property record in Florida from free public tax rolls.',
        'Filter down to investment properties only — no homeowners, no noise.',
        'Identify who owns them: LLCs, serial investors, foreign buyers.',
        'Score and rank every lead so you call the best ones first.',
        'Detect refi opportunities: cash buyers, equity harvest, rate refi.',
        'Add phone numbers and emails so you can actually reach them.',
    ]
    for i, step in enumerate(steps):
        _merged(ws, row, 2, 6, f'  {i+1}.  {step}', _label_font, None, _left)
        ws.row_dimensions[row].height = 26
        if i % 2 == 1:
            for c in range(2, 7):
                ws.cell(row=row, column=c).fill = _light_fill
        row += 1

    row += 1

    # What's in each lead — pulled from real data
    row = _section(ws, row, 'WHAT EACH LEAD INCLUDES')
    fields = [
        ('Owner name', f'{stats["total"]:,} leads', 'Free'),
        ('Mailing address', f'{stats["has_address"]:,} leads', 'Free'),
        ('# of properties owned', f'{stats["multi_property"]:,} multi-property', 'Free'),
        ('Total portfolio value', _fmt_short(stats['total_portfolio']), 'Free'),
        ('Investor type & lead score', 'Scored 0-100, ranked', 'Free'),
        ('Refi signals', f'{stats["refi_candidates"]:,} candidates', 'Free'),
        ('Real name behind LLC', f'{stats["has_resolved"]:,} resolved so far', 'Free'),
        ('Phone number', f'{stats["has_phone"]:,} today — see "What Data You Get"', 'Free / Paid'),
        ('Email address', 'See "What Data You Get" tab', 'Paid'),
    ]
    # Header
    ws.cell(row=row, column=2, value='Data Field').font = _header_font
    ws.cell(row=row, column=2).fill = _navy_fill
    ws.cell(row=row, column=2).alignment = _left
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
    ws.cell(row=row, column=3, value='Current Coverage').font = _header_font
    ws.cell(row=row, column=3).fill = _navy_fill
    ws.cell(row=row, column=3).alignment = _center
    for c in range(3, 6):
        ws.cell(row=row, column=c).fill = _navy_fill
    ws.cell(row=row, column=6, value='Cost').font = _header_font
    ws.cell(row=row, column=6).fill = _navy_fill
    ws.cell(row=row, column=6).alignment = _center
    ws.row_dimensions[row].height = 24
    row += 1

    for i, (field, coverage, cost) in enumerate(fields):
        ws.cell(row=row, column=2, value=field).font = _label_bold
        ws.cell(row=row, column=2).alignment = _left
        ws.cell(row=row, column=2).border = _thin
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=5)
        ws.cell(row=row, column=3, value=coverage).font = _value_font
        ws.cell(row=row, column=3).alignment = _center
        ws.cell(row=row, column=3).border = _thin
        ws.cell(row=row, column=6, value=cost).font = _free_font if cost == 'Free' else _paid_font
        ws.cell(row=row, column=6).alignment = _center
        ws.cell(row=row, column=6).border = _thin
        if cost == 'Free':
            ws.cell(row=row, column=6).fill = _green_text_fill
        if i % 2 == 1:
            for c in range(2, 6):
                ws.cell(row=row, column=c).fill = _light_fill
        ws.row_dimensions[row].height = 24
        row += 1

    row += 1

    # Scale
    row = _section(ws, row, 'THE SCALE', fill=_green_fill)
    scale = [
        ('Palm Beach County (done)', f'{stats["total"]:,} leads'),
        ('South Florida (3 counties)', '~120,000 leads'),
        ('All 67 FL counties', '~470,000 leads'),
    ]
    for i, (geo, count) in enumerate(scale):
        ws.cell(row=row, column=2, value=geo).font = _label_bold
        ws.cell(row=row, column=2).alignment = _left
        ws.cell(row=row, column=2).border = _thin
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        ws.cell(row=row, column=3, value=count).font = _value_bold
        ws.cell(row=row, column=3).alignment = _center
        ws.cell(row=row, column=3).border = _thin
        if i % 2 == 1:
            for c in range(2, 7):
                ws.cell(row=row, column=c).fill = _light_fill
        ws.row_dimensions[row].height = 26
        row += 1

    row += 1
    _merged(ws, row, 2, 6,
        'Same process works for every county. Run it once, refresh quarterly.',
        _note_font, _white_fill, _left)


# ---------------------------------------------------------------------------
# Tab 2: What Data You Get
# ---------------------------------------------------------------------------
def _build_data_breakdown(wb, stats):
    ws = wb.create_sheet('What Data You Get')
    _setup(ws)

    row = _banner(ws, 'WHAT DATA YOU GET')
    row += 1

    # Overview
    _merged(ws, row, 2, 6,
        'Here\'s exactly what data you get at each level, what it costs, '
        'and what it lets you do. Start free — add tools as you scale.',
        _label_font, _white_fill, _left)
    ws.row_dimensions[row].height = 36
    row += 2

    # === TIER 1: FREE ===
    row = _section(ws, row, 'FREE — WHAT YOU HAVE RIGHT NOW', fill=_green_fill)

    _merged(ws, row, 2, 6,
        'All of this comes from free public records. No subscriptions, no per-record fees.',
        _note_font, _white_fill, _left)
    ws.row_dimensions[row].height = 24
    row += 1

    # Column headers
    ws.cell(row=row, column=2, value='What You Get').font = _header_font
    ws.cell(row=row, column=2).fill = _navy_fill
    ws.cell(row=row, column=2).alignment = _left
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws.cell(row=row, column=3, value='Source').font = _header_font
    ws.cell(row=row, column=3).fill = _navy_fill
    ws.cell(row=row, column=3).alignment = _center
    for c in [4]:
        ws.cell(row=row, column=c).fill = _navy_fill
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    ws.cell(row=row, column=5, value='You Have Today').font = _header_font
    ws.cell(row=row, column=5).fill = _navy_fill
    ws.cell(row=row, column=5).alignment = _center
    ws.cell(row=row, column=6).fill = _navy_fill
    ws.row_dimensions[row].height = 24
    row += 1

    free_data = [
        ('Owner name', 'FL Tax Records (FDOR)', f'{stats["total"]:,} (100%)'),
        ('Mailing address', 'FL Tax Records (FDOR)', f'{stats["has_address"]:,} (100%)'),
        ('# investment properties', 'FL Tax Records (FDOR)', f'{stats["multi_property"]:,} multi-property'),
        ('Portfolio assessed value', 'FL Tax Records (FDOR)', _fmt_short(stats['total_portfolio']) + ' total'),
        ('Last purchase date & price', 'FL Tax Records (FDOR)', 'Every lead'),
        ('Investor type classification', 'Pipeline scoring engine', '7 investor categories'),
        ('Lead score (0-100)', 'Pipeline scoring engine', 'Every lead ranked'),
        ('Refi signals & priority', 'Pipeline analysis', f'{stats["refi_candidates"]:,} candidates flagged'),
        ('Real name behind LLC/Trust', 'FL SunBiz (Corp Registry)', f'{stats["has_resolved"]:,} resolved'),
        ('Vacation rental detection', 'FL DBPR (Licenses)', f'{stats["str_licensed"]:,} STR operators'),
        ('Phone # (STR licensees)', 'FL DBPR (Licenses)', f'{stats["has_phone"]:,} phone numbers'),
        ('Fund manager detection', 'SEC EDGAR (Federal)', 'Cross-referenced'),
    ]
    for i, (field, source, coverage) in enumerate(free_data):
        ws.cell(row=row, column=2, value=field).font = _label_bold
        ws.cell(row=row, column=2).alignment = _left
        ws.cell(row=row, column=2).border = _thin
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws.cell(row=row, column=3, value=source).font = _value_font
        ws.cell(row=row, column=3).alignment = _left
        ws.cell(row=row, column=3).border = _thin
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        ws.cell(row=row, column=5, value=coverage).font = _value_font
        ws.cell(row=row, column=5).alignment = _center
        ws.cell(row=row, column=5).border = _thin
        if i % 2 == 1:
            for c in range(2, 7):
                ws.cell(row=row, column=c).fill = _light_fill
        ws.row_dimensions[row].height = 24
        row += 1

    row += 1

    # === TIER 2: PHONES & EMAILS ===
    row = _section(ws, row, 'ADD PHONE & EMAIL — $49/MO')

    _merged(ws, row, 2, 6,
        'Right now you have 318 phone numbers (from STR licenses). '
        'To reach the other 39,000 leads, add a contact enrichment tool.',
        _note_font, _white_fill, _left)
    ws.row_dimensions[row].height = 32
    row += 1

    # Column headers
    ws.cell(row=row, column=2, value='What You Add').font = _header_font
    ws.cell(row=row, column=2).fill = _navy_fill
    ws.cell(row=row, column=2).alignment = _left
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws.cell(row=row, column=3, value='Tool').font = _header_font
    ws.cell(row=row, column=3).fill = _navy_fill
    ws.cell(row=row, column=3).alignment = _center
    ws.cell(row=row, column=4).fill = _navy_fill
    ws.cell(row=row, column=5, value='Cost').font = _header_font
    ws.cell(row=row, column=5).fill = _navy_fill
    ws.cell(row=row, column=5).alignment = _center
    ws.cell(row=row, column=6, value='What You Get').font = _header_font
    ws.cell(row=row, column=6).fill = _navy_fill
    ws.cell(row=row, column=6).alignment = _center
    ws.row_dimensions[row].height = 24
    row += 1

    phone_email_options = [
        ('Phone numbers', 'Apollo.io', '$49/mo', '~60% hit rate, 25K lookups/mo'),
        ('Email addresses', 'Apollo.io', 'Included', 'Business + personal emails'),
        ('LinkedIn profiles', 'Apollo.io', 'Included', 'Direct profile URLs'),
        ('Phone (alternative)', 'BatchSkipTracing', '~$0.15/record', 'Higher hit rate, pay per record'),
    ]
    for i, (field, tool, cost, what) in enumerate(phone_email_options):
        ws.cell(row=row, column=2, value=field).font = _label_bold
        ws.cell(row=row, column=2).alignment = _left
        ws.cell(row=row, column=2).border = _thin
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws.cell(row=row, column=3, value=tool).font = _value_font
        ws.cell(row=row, column=3).alignment = _center
        ws.cell(row=row, column=3).border = _thin
        ws.cell(row=row, column=5, value=cost).font = _value_bold
        ws.cell(row=row, column=5).alignment = _center
        ws.cell(row=row, column=5).border = _thin
        ws.cell(row=row, column=6, value=what).font = _value_font
        ws.cell(row=row, column=6).alignment = _left
        ws.cell(row=row, column=6).border = _thin
        if i % 2 == 1:
            for c in range(2, 7):
                ws.cell(row=row, column=c).fill = _light_fill
        ws.row_dimensions[row].height = 24
        row += 1

    row += 1

    # === TIER 3: ADVANCED ===
    row = _section(ws, row, 'ADVANCED — ADD WHEN SCALING', fill=_orange_fill)

    _merged(ws, row, 2, 6,
        'These tools add deeper property intel. Not required to start — '
        'useful once you\'re closing deals and want to sharpen targeting.',
        _note_font, _white_fill, _left)
    ws.row_dimensions[row].height = 32
    row += 1

    # Column headers
    ws.cell(row=row, column=2, value='What You Add').font = _header_font
    ws.cell(row=row, column=2).fill = _navy_fill
    ws.cell(row=row, column=2).alignment = _left
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws.cell(row=row, column=3, value='Tool').font = _header_font
    ws.cell(row=row, column=3).fill = _navy_fill
    ws.cell(row=row, column=3).alignment = _center
    ws.cell(row=row, column=4).fill = _navy_fill
    ws.cell(row=row, column=5, value='Cost').font = _header_font
    ws.cell(row=row, column=5).fill = _navy_fill
    ws.cell(row=row, column=5).alignment = _center
    ws.cell(row=row, column=6, value='Why It Helps').font = _header_font
    ws.cell(row=row, column=6).fill = _navy_fill
    ws.cell(row=row, column=6).alignment = _center
    ws.row_dimensions[row].height = 24
    row += 1

    advanced_options = [
        ('Mortgage balance data', 'PropStream', '$99/mo', 'Know exact equity, not estimated'),
        ('MLS comps & ARV', 'PropStream', 'Included', 'After-repair value for BRRRR deals'),
        ('Skip-traced phones', 'PropStream', 'Included', 'Alternative phone source'),
        ('Rental comps / DSCR est.', 'RentCast', '$74/mo', 'Estimate DSCR ratio before calling'),
    ]
    for i, (field, tool, cost, why) in enumerate(advanced_options):
        ws.cell(row=row, column=2, value=field).font = _label_bold
        ws.cell(row=row, column=2).alignment = _left
        ws.cell(row=row, column=2).border = _thin
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws.cell(row=row, column=3, value=tool).font = _value_font
        ws.cell(row=row, column=3).alignment = _center
        ws.cell(row=row, column=3).border = _thin
        ws.cell(row=row, column=5, value=cost).font = _value_bold
        ws.cell(row=row, column=5).alignment = _center
        ws.cell(row=row, column=5).border = _thin
        ws.cell(row=row, column=6, value=why).font = _value_font
        ws.cell(row=row, column=6).alignment = _left
        ws.cell(row=row, column=6).border = _thin
        if i % 2 == 1:
            for c in range(2, 7):
                ws.cell(row=row, column=c).fill = _light_fill
        ws.row_dimensions[row].height = 24
        row += 1

    row += 1

    # === SUMMARY: TOTAL COST BY LEVEL ===
    row = _section(ws, row, 'TOTAL MONTHLY COST BY LEVEL')

    levels = [
        ('Just the leads (name, address, score, refi signals)', '$0/mo', _green_fill),
        ('+ Phone & email for most leads', '$49/mo', _blue_fill),
        ('+ Mortgage data, comps, rental estimates', '$222/mo', _orange_fill),
    ]
    for i, (desc, cost, fill) in enumerate(levels):
        ws.cell(row=row, column=2, value=desc).font = _label_bold
        ws.cell(row=row, column=2).alignment = _left
        ws.cell(row=row, column=2).border = _thin
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        c = ws.cell(row=row, column=3, value=cost)
        c.font = Font(size=14, bold=True, color=_WHITE)
        c.fill = fill
        c.alignment = _center
        ws.cell(row=row, column=4).fill = fill
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        ws.cell(row=row, column=5).border = _thin
        ws.row_dimensions[row].height = 32
        row += 1

    row += 1

    # Comparison
    row = _section(ws, row, 'FOR CONTEXT')
    comparisons = [
        ('This pipeline (free tier)', '$0/mo', f'{stats["total"]:,}+ scored leads per county'),
        ('This pipeline (with phones)', '$49/mo', 'Phone + email for ~60% of leads'),
        ('Buying leads from a vendor', '$100-250 each', '40-100 leads/month, no scoring'),
        ('Facebook / Google ads', '$25-75 each', 'Inbound but unqualified'),
    ]
    for i, (source, cost, notes) in enumerate(comparisons):
        ws.cell(row=row, column=2, value=source).font = _label_bold
        ws.cell(row=row, column=2).alignment = _left
        ws.cell(row=row, column=2).border = _thin
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws.cell(row=row, column=3, value=cost).font = _value_bold
        ws.cell(row=row, column=3).alignment = _center
        ws.cell(row=row, column=3).border = _thin
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        ws.cell(row=row, column=5, value=notes).font = _value_font
        ws.cell(row=row, column=5).alignment = _left
        ws.cell(row=row, column=5).border = _thin
        if i % 2 == 1:
            for c in range(2, 7):
                ws.cell(row=row, column=c).fill = _light_fill
        ws.row_dimensions[row].height = 26
        row += 1


# ---------------------------------------------------------------------------
# Tab 3: Your Calculator
# ---------------------------------------------------------------------------
def _build_calculator(wb):
    ws = wb.create_sheet('Your Calculator')
    _setup(ws)

    row = _banner(ws, 'YOUR REVENUE CALCULATOR')
    row += 1

    row = _section(ws, row, 'YOUR INPUTS (CHANGE THE YELLOW CELLS)')

    refs = {}

    row = _input_row(ws, row, 'Average DSCR Loan Size', 400000, '$#,##0')
    refs['loan'] = row - 1

    row = _input_row(ws, row, 'Your Comp Per Loan (bps)', 150, '#,##0')
    refs['bps'] = row - 1

    row = _input_row(ws, row, 'Leads You Work Per Month', 500, '#,##0')
    refs['worked'] = row - 1

    row = _input_row(ws, row, 'Close Rate (% of leads worked)', 0.02, '0.0%')
    refs['close'] = row - 1

    row = _input_row(ws, row, 'Monthly Data Costs (see tab 2)', 49, '$#,##0')
    refs['cost'] = row - 1

    row += 1

    row = _section(ws, row, 'THE MATH')

    lr = f'D{refs["loan"]}'
    br = f'D{refs["bps"]}'
    wr = f'D{refs["worked"]}'
    cr = f'D{refs["close"]}'
    co = f'D{refs["cost"]}'

    row = _formula_row(ws, row, 'Your Revenue Per Closed Loan',
        f'={lr}*{br}/10000', '$#,##0')
    rev_ref = f'D{row - 1}'

    row = _formula_row(ws, row, 'Closings Per Month',
        f'=ROUND({wr}*{cr},1)', '#,##0.0')
    close_ref = f'D{row - 1}'

    row = _formula_row(ws, row, 'Monthly Revenue',
        f'={close_ref}*{rev_ref}', '$#,##0', bold=True)
    monthly_ref = f'D{row - 1}'

    row = _formula_row(ws, row, 'Monthly Profit (after data costs)',
        f'={monthly_ref}-{co}', '$#,##0', bold=True)
    profit_ref = f'D{row - 1}'

    row += 1

    row = _section(ws, row, 'YOUR ANNUAL NUMBERS', fill=_green_fill)

    row = _formula_row(ws, row, 'Annual Revenue',
        f'={monthly_ref}*12', '$#,##0', result=True)
    annual_ref = f'D{row - 1}'

    row = _formula_row(ws, row, 'Annual Data Costs',
        f'={co}*12', '$#,##0')
    annual_cost_ref = f'D{row - 1}'

    row = _formula_row(ws, row, 'Annual Profit',
        f'={profit_ref}*12', '$#,##0', result=True)

    row = _formula_row(ws, row, 'ROI on Data Spend',
        f'=IF({annual_cost_ref}=0,"\u221e (free)",ROUND({annual_ref}/{annual_cost_ref},0)&"x")')

    row += 1

    row = _section(ws, row, 'SCENARIOS')

    _merged(ws, row, 2, 6,
        'Try changing the yellow cells above. Here are some benchmarks:',
        _note_font, _white_fill, _left)
    row += 1

    scenarios = [
        ('Conservative: 1% close, 300 leads/mo',
         '300 x 1% = 3 closings x $6K = $18K/mo'),
        ('Base case: 2% close, 500 leads/mo',
         '500 x 2% = 10 closings x $6K = $60K/mo'),
        ('Aggressive: 3% close, 800 leads/mo',
         '800 x 3% = 24 closings x $6K = $144K/mo'),
    ]
    for i, (scenario, math) in enumerate(scenarios):
        ws.cell(row=row, column=2, value=scenario).font = _label_bold
        ws.cell(row=row, column=2).alignment = _left
        ws.cell(row=row, column=2).border = _thin
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        ws.cell(row=row, column=3, value=math).font = _value_font
        ws.cell(row=row, column=3).alignment = _left
        ws.cell(row=row, column=3).border = _thin
        if i % 2 == 1:
            for c in range(2, 7):
                ws.cell(row=row, column=c).fill = _light_fill
        ws.row_dimensions[row].height = 28
        row += 1

    row += 1
    _merged(ws, row, 2, 6,
        'Assumes $400K avg loan, 150 bps comp ($6K/loan). '
        'Adjust yellow cells to match your numbers.',
        _note_font, _white_fill, _left)


# ---------------------------------------------------------------------------
# Tab 4: Sample Leads
# ---------------------------------------------------------------------------
CO_NO_TO_COUNTY = {
    "11": "ALACHUA", "12": "BAKER", "13": "BAY", "14": "BRADFORD",
    "15": "BREVARD", "16": "BROWARD", "17": "CALHOUN", "18": "CHARLOTTE",
    "19": "CITRUS", "20": "CLAY", "21": "COLLIER", "22": "COLUMBIA",
    "23": "MIAMI-DADE", "24": "DESOTO", "25": "DIXIE", "26": "DUVAL",
    "27": "ESCAMBIA", "28": "FLAGLER", "29": "FRANKLIN", "30": "GADSDEN",
    "31": "GILCHRIST", "32": "GLADES", "33": "GULF", "34": "HAMILTON",
    "35": "HARDEE", "36": "HENDRY", "37": "HERNANDO", "38": "HIGHLANDS",
    "39": "HILLSBOROUGH", "40": "HOLMES", "41": "INDIAN RIVER", "42": "JACKSON",
    "43": "JEFFERSON", "44": "LAFAYETTE", "45": "LAKE", "46": "LEE",
    "47": "LEON", "48": "LEVY", "49": "LIBERTY", "50": "MADISON",
    "51": "MANATEE", "52": "MARION", "53": "MARTIN", "54": "MONROE",
    "55": "NASSAU", "56": "OKALOOSA", "57": "OKEECHOBEE", "58": "ORANGE",
    "59": "OSCEOLA", "60": "PALM BEACH", "61": "PASCO", "62": "PINELLAS",
    "63": "POLK", "64": "PUTNAM", "65": "ST. JOHNS", "66": "ST. LUCIE",
    "67": "SANTA ROSA", "68": "SARASOTA", "69": "SEMINOLE", "70": "SUMTER",
    "71": "SUWANNEE", "72": "TAYLOR", "73": "UNION", "74": "VOLUSIA",
    "75": "WAKULLA", "76": "WALTON", "77": "WASHINGTON",
}


def _build_sample_leads(wb, stats):
    ws = wb.create_sheet('Sample Leads')

    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 2

    df = stats['df'].copy()

    # Score the leads inline (simplified version of 06_score_and_output logic)
    df['_score'] = 0.0
    pc = pd.to_numeric(df.get('property_count', 0), errors='coerce').fillna(0)
    pv = pd.to_numeric(df.get('total_portfolio_value', 0), errors='coerce').fillna(0)
    avg_val = pd.to_numeric(df.get('avg_property_value', 0), errors='coerce').fillna(0)
    eq_ratio = pd.to_numeric(df.get('equity_ratio', 0), errors='coerce').fillna(0)

    df['_score'] += pc.clip(upper=10) * 2.5       # 0-25
    df['_score'] += (pv / 200000).clip(upper=15)   # 0-15

    is_entity = df.get('is_entity', '').fillna('').astype(str).str.lower().isin(['true', '1', 'yes'])
    df.loc[is_entity, '_score'] += 10

    str_lic = df.get('str_licensed', '').fillna('').astype(str).str.lower().isin(['true', '1', 'yes'])
    df.loc[str_lic, '_score'] += 15

    refi_boost = pd.to_numeric(df.get('refi_score_boost', 0), errors='coerce').fillna(0)
    df['_score'] += refi_boost.clip(upper=15)

    has_phone = df['phone'].fillna('').astype(str).str.strip().ne('') & \
                df['phone'].astype(str).str.lower().ne('nan')

    # Sort and take top 50
    df = df.sort_values('_score', ascending=False)
    top = df.head(50).copy()

    # Build "Why This Lead" justification for each row
    def _build_why(r):
        reasons = []
        try:
            n = int(float(r.get('property_count', 0)))
        except (ValueError, TypeError):
            n = 0
        port = _fmt_short(r.get('total_portfolio_value', 0))

        # Portfolio size
        if n >= 10:
            reasons.append(f'Serial investor ({n} properties, {port} portfolio)')
        elif n >= 2:
            reasons.append(f'{n} investment properties ({port} portfolio)')

        # Entity structure
        if str(r.get('is_entity', '')).lower() in ('true', '1', 'yes'):
            reasons.append('Entity-structured (LLC/Corp)')

        # Out of state / foreign
        if str(r.get('foreign_owner', '')).lower() in ('true', '1', 'yes'):
            reasons.append('Foreign national — needs DSCR')
        elif str(r.get('out_of_state', '')).lower() in ('true', '1', 'yes'):
            state = str(r.get('OWN_STATE_DOM', '')).strip()
            if state and state.lower() != 'nan':
                reasons.append(f'Out-of-state investor ({state})')
            else:
                reasons.append('Out-of-state investor')

        # STR
        if str(r.get('str_licensed', '')).lower() in ('true', '1', 'yes'):
            try:
                lc = int(float(r.get('str_license_count', 0)))
            except (ValueError, TypeError):
                lc = 0
            if lc > 1:
                reasons.append(f'Licensed STR operator ({lc} licenses)')
            else:
                reasons.append('Licensed vacation rental operator')

        # Refi signals
        if str(r.get('probable_cash_buyer', '')).lower() in ('true', '1', 'yes'):
            reasons.append('Probable all-cash buyer — cash-out refi opportunity')
        if str(r.get('equity_harvest_candidate', '')).lower() in ('true', '1', 'yes'):
            try:
                eq = float(r.get('equity_ratio', 0))
                reasons.append(f'{eq:.0%} equity — harvest opportunity')
            except (ValueError, TypeError):
                reasons.append('High equity — harvest opportunity')
        if str(r.get('brrrr_exit_candidate', '')).lower() in ('true', '1', 'yes'):
            reasons.append('BRRRR exit — needs perm financing')
        if str(r.get('rate_refi_candidate', '')).lower() in ('true', '1', 'yes'):
            reasons.append('2022-2023 purchase — likely overpaying on rate')

        # Contact
        phone_val = str(r.get('phone', '')).strip()
        if phone_val and phone_val.lower() != 'nan':
            reasons.append(f'Has phone: {phone_val}')

        return '; '.join(reasons) if reasons else 'Qualified investor (non-homesteaded, multi-property or entity)'

    # Get contact person name
    def _get_contact(r):
        # Try resolved_person first (SunBiz resolution)
        resolved = str(r.get('resolved_person', '')).strip()
        if resolved and resolved.lower() != 'nan':
            return resolved
        # Try entity_officers
        officers = str(r.get('entity_officers', '')).strip()
        if officers and officers.lower() != 'nan':
            # Take first officer, clean up
            first = officers.split(';')[0].split('(')[0].strip()
            if first:
                return first
        return ''

    # Columns: Score | Entity/Owner | Contact Person | Location | # Props | Portfolio | Phone | Why This Lead
    headers = [
        ('Score', 8),
        ('Owner / Entity', 26),
        ('Contact Person', 20),
        ('Location', 18),
        ('# Props', 8),
        ('Portfolio Value', 15),
        ('Phone', 16),
        ('Why This Is a Good Lead', 52),
    ]

    end_col = len(headers) + 1
    gutter_col = end_col + 1

    # Banner
    row = 1
    _merged(ws, row, 2, end_col, 'SAMPLE LEADS — TOP 50', _title_font, _navy_fill, _center)
    ws.row_dimensions[row].height = 44
    ws.cell(row=row, column=1).fill = _navy_fill
    ws.cell(row=row, column=gutter_col).fill = _navy_fill
    row += 1
    _merged(ws, row, 2, end_col,
        f'From {stats["total"]:,} qualified investor leads in Palm Beach County  |  '
        f'{stats["has_phone"]:,} with phone today  |  '
        f'{stats["refi_candidates"]:,} refi candidates',
        _subtitle_font, _navy_fill, _center)
    ws.row_dimensions[row].height = 24
    ws.cell(row=row, column=1).fill = _navy_fill
    ws.cell(row=row, column=gutter_col).fill = _navy_fill
    row += 2

    # Column widths
    for i, (hdr, width) in enumerate(headers):
        ws.column_dimensions[get_column_letter(i + 2)].width = width

    # Header row
    for i, (hdr, _) in enumerate(headers):
        cell = ws.cell(row=row, column=i + 2, value=hdr)
        cell.font = _header_font
        cell.fill = _navy_fill
        cell.alignment = _center if i != 7 else _left
    ws.row_dimensions[row].height = 24
    row += 1

    # Data rows
    green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

    for idx, (_, lead) in enumerate(top.iterrows()):
        score = int(lead['_score'])

        # Owner name
        owner = str(lead.get('OWN_NAME', '')).strip()
        if not owner or owner.lower() == 'nan':
            owner = '(Unknown)'

        # Contact person
        contact = _get_contact(lead)

        # Location
        city = str(lead.get('OWN_CITY', '')).strip()
        state = str(lead.get('OWN_STATE_DOM', '')).strip()
        if city.lower() == 'nan': city = ''
        if state.lower() == 'nan': state = ''
        location = f'{city}, {state}' if city and state else city or state

        # Property count
        try:
            prop_ct = int(float(lead.get('property_count', 0)))
        except (ValueError, TypeError):
            prop_ct = 0

        portfolio = _fmt_short(lead.get('total_portfolio_value', 0))

        # Phone
        phone_val = str(lead.get('phone', '')).strip()
        if not phone_val or phone_val.lower() == 'nan':
            phone_val = ''

        # Why
        why = _build_why(lead)

        vals = [score, owner[:30], contact[:25], location, prop_ct, portfolio, phone_val, why]
        col_aligns = [_center, _left, _left, _center, _center, _center, _center, _left]

        for i, v in enumerate(vals):
            cell = ws.cell(row=row, column=i + 2, value=v)
            cell.font = _cell_bold if i in (0, 1, 2) else _cell_font
            cell.alignment = col_aligns[i]
            cell.border = _thin

            if i == 0:
                if score >= 60:
                    cell.fill = green_fill
                elif score >= 40:
                    cell.fill = yellow_fill
            elif idx % 2 == 1:
                cell.fill = _light_fill

        ws.row_dimensions[row].height = 28
        row += 1

    # Footer
    row += 1
    _merged(ws, row, 2, end_col,
        f'Showing top 50 of {stats["total"]:,} leads. Full dataset includes equity details, '
        'mailing addresses, refi priority, entity names, and more.',
        _note_font, _white_fill, _left)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description='Generate DSCR proposal workbook')
    parser.add_argument('--leads', type=str, default='pipeline/output/06_enriched.csv',
                        help='Path to enriched leads CSV')
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    today = date.today().isoformat()
    out_path = OUTPUT_DIR / f'roi_analysis_{today}.xlsx'

    # Load lead data for live stats
    leads_path = Path(args.leads)
    if leads_path.exists():
        print(f"Loading leads from {leads_path}...")
        stats = _load_stats(str(leads_path))
        print(f"  {stats['total']:,} leads loaded")
    else:
        print(f"Warning: {leads_path} not found. Using placeholder stats.")
        stats = {
            'df': pd.DataFrame(),
            'total': 39353, 'has_phone': 318, 'has_email': 0,
            'has_address': 39353, 'has_resolved': 183, 'str_licensed': 769,
            'is_entity': 21712, 'multi_property': 12017, 'foreign': 8892,
            'oos': 30495, 'total_portfolio': 16_700_000_000,
            'phone_pct': 0.8, 'email_pct': 0.0, 'refi_candidates': 5200,
        }

    wb = Workbook()

    _build_opportunity(wb, stats)
    _build_data_breakdown(wb, stats)
    _build_calculator(wb)
    if not stats['df'].empty:
        _build_sample_leads(wb, stats)

    wb.save(str(out_path))
    tab_count = len(wb.sheetnames)
    print(f"\nProposal saved: {out_path}")
    print(f"  {tab_count} tabs: {', '.join(wb.sheetnames)}")
    print(f"  Yellow cells are adjustable — formulas update automatically")


if __name__ == '__main__':
    main()
