"""
Module 6: ICP Scoring and Excel Output

Classifies leads into ICP segments, scores by quality,
and produces the final Excel workbook.

Usage:
    python scripts/06_score_and_output.py --input pipeline/output/05_enriched.csv --edgar-input pipeline/output/04_fund_managers.csv --output leads_YYYY-MM-DD.xlsx
"""

import argparse
import pandas as pd
from pathlib import Path
from datetime import datetime, date
import re
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

OUTPUT_DIR = Path("pipeline/output")

# County-to-metro mapping for geographic scoring
SOUTH_FL_COUNTIES = ['PALM BEACH', 'BROWARD', 'MIAMI-DADE']
MAJOR_METRO_COUNTIES = [
    'HILLSBOROUGH', 'ORANGE', 'DUVAL', 'PINELLAS', 'SEMINOLE',
    'OSCEOLA', 'LEE', 'COLLIER', 'SARASOTA', 'MANATEE',
    'VOLUSIA', 'BREVARD', 'PASCO', 'POLK', 'ST. LUCIE', 'MARTIN'
]

# FDOR county code -> name mapping (reverse of codes in 01_fdor_download_filter.py)
CO_NO_TO_COUNTY = {
    "11": "ALACHUA", "12": "BAKER", "13": "BAY", "14": "BRADFORD",
    "15": "BREVARD", "16": "BROWARD", "17": "CALHOUN", "18": "CHARLOTTE",
    "19": "CITRUS", "20": "CLAY", "21": "COLLIER", "22": "COLUMBIA",
    "23": "MIAMI-DADE", "24": "DESOTO", "25": "DIXIE", "26": "DUVAL",
    "27": "ESCAMBIA", "28": "FLAGLER", "29": "FRANKLIN", "30": "GADSDEN",
    "31": "GILCHRIST", "32": "GLADES", "33": "GULF", "34": "HAMILTON",
    "35": "HARDEE", "36": "HENDRY", "37": "HERNANDO", "38": "HIGHLANDS",
    "39": "HILLSBOROUGH", "40": "HOLMES", "41": "INDIAN RIVER", "42": "JACKSON",
    "43": "JEFFERSON", "44": "LAFAYETTE", "45": "LAKE", "46": "LEE",
    "47": "LEON", "48": "LEVY", "49": "LIBERTY", "50": "MADISON",
    "51": "MANATEE", "52": "MARION", "53": "MARTIN", "54": "MONROE",
    "55": "NASSAU", "56": "OKALOOSA", "57": "OKEECHOBEE", "58": "ORANGE",
    "59": "OSCEOLA", "60": "PALM BEACH", "61": "PASCO", "62": "PINELLAS",
    "63": "POLK", "64": "PUTNAM", "65": "SAINT JOHNS", "66": "SAINT LUCIE",
    "67": "SANTA ROSA", "68": "SARASOTA", "69": "SEMINOLE", "70": "SUMTER",
    "71": "SUWANNEE", "72": "TAYLOR", "73": "UNION", "74": "VOLUSIA",
    "75": "WAKULLA", "76": "WALTON", "77": "WASHINGTON",
}


def classify_icp(row: pd.Series) -> tuple:
    """
    Classify a lead into primary ICP segment and tier.
    Returns (icp_primary, icp_secondary, tier)
    """

    def _safe_int(val, default=0):
        try:
            f = float(val)
            return default if pd.isna(f) else int(f)
        except (ValueError, TypeError):
            return default

    property_count = _safe_int(row.get('property_count', 0))
    str_licensed = str(row.get('str_licensed', '')).lower() in ('true', '1', 'yes')
    str_count = _safe_int(row.get('str_license_count', 0))
    foreign_owner = str(row.get('foreign_owner', '')).lower() in ('true', '1', 'yes')
    out_of_state = str(row.get('out_of_state', '')).lower() in ('true', '1', 'yes')
    is_entity = str(row.get('is_entity', '')).lower() in ('true', '1', 'yes')
    entity_count = _safe_int(row.get('entity_count', 0))
    sec_fund = str(row.get('sec_fund_filing', '')).lower() in ('true', '1', 'yes')

    # Get property type info
    prop_types = str(row.get('property_types', '') or row.get('DOR_UC', '')).upper()
    is_multifamily = any(code in prop_types for code in ['03', '08', 'MULTI', 'DUPLEX', 'TRIPLEX', 'FOURPLEX'])

    # Get refinance signals
    probable_cash = str(row.get('probable_cash_buyer', '')).lower() in ('true', '1', 'yes')
    brrrr_exit = str(row.get('brrrr_exit_candidate', '')).lower() in ('true', '1', 'yes')
    equity_harvest = str(row.get('equity_harvest_candidate', '')).lower() in ('true', '1', 'yes')
    rate_refi = str(row.get('rate_refi_candidate', '')).lower() in ('true', '1', 'yes')

    # Get recent purchase info for BRRRR detection
    def _safe_float(val, default=0.0):
        try:
            f = float(val)
            return default if pd.isna(f) else f
        except (ValueError, TypeError):
            return default
    recent_price = _safe_float(row.get('most_recent_price', 0))
    portfolio_value = _safe_float(row.get('total_portfolio_value', 0))

    # Classification logic (priority order)
    icp_primary = 'Single Investment Property'
    icp_secondary = ''
    tier = 3

    if property_count >= 10:
        icp_primary = 'Serial Investor (10+)'
        tier = 1
        if str_licensed:
            icp_secondary = 'STR Operator'
        elif foreign_owner:
            icp_secondary = 'Foreign National'

    elif sec_fund:
        icp_primary = 'Fund Manager / Syndicator'
        tier = 2

    elif str_licensed or str_count > 0:
        icp_primary = 'STR Operator'
        tier = 1
        if property_count >= 5:
            icp_secondary = 'Serial Investor (10+)' if property_count >= 10 else 'Growing Portfolio'

    elif foreign_owner:
        icp_primary = 'Foreign National'
        tier = 1
        if is_entity:
            icp_secondary = 'Entity-Based Investor'

    elif is_entity and entity_count >= 2:
        icp_primary = 'Entity-Based Investor'
        tier = 1
        if property_count >= 5:
            icp_secondary = 'Growing Portfolio'

    elif is_entity and property_count >= 2:
        icp_primary = 'Entity-Based Investor'
        tier = 1

    elif property_count >= 2:
        icp_primary = 'Individual Investor (2-9)'
        tier = 1 if property_count >= 5 else 2
        if is_multifamily:
            icp_secondary = 'Multi-Family Investor'

    elif is_multifamily:
        icp_primary = 'Multi-Family Investor'
        tier = 2

    elif out_of_state:
        icp_primary = 'Out-of-State Investor'
        tier = 2

    else:
        icp_primary = 'Single Investment Property'
        tier = 3

    # Refinance-based secondary tagging (overlay on any primary ICP)
    if not icp_secondary:
        if probable_cash:
            icp_secondary = 'Cash-Out Refi Candidate'
        elif brrrr_exit:
            icp_secondary = 'BRRRR Exit Candidate'
        elif equity_harvest:
            icp_secondary = 'Equity Harvest Candidate'
        elif rate_refi:
            icp_secondary = 'Rate Refi Candidate'

    # Upgrade tier if strong refi signal on otherwise lower-tier lead
    if tier == 3 and (probable_cash or brrrr_exit):
        tier = 2
    if tier == 2 and probable_cash and property_count >= 2:
        tier = 1

    return icp_primary, icp_secondary, tier


def score_lead(row: pd.Series) -> int:
    """
    Score a lead 0-100 based on multiple factors.
    """

    score = 0

    def _safe_int(val, default=0):
        try:
            f = float(val)
            return default if pd.isna(f) else int(f)
        except (ValueError, TypeError):
            return default

    # Property count (0-25)
    pc = _safe_int(row.get('property_count', 0))
    if pc >= 20:
        score += 25
    elif pc >= 10:
        score += 20
    elif pc >= 5:
        score += 15
    elif pc >= 2:
        score += 10
    elif pc >= 1:
        score += 5

    # Recency of last purchase (0-20)
    recent = str(row.get('most_recent_purchase', ''))
    if recent and recent != 'nan':
        try:
            purchase_date = pd.to_datetime(recent)
            days_ago = (datetime.now() - purchase_date).days
            if days_ago < 180:
                score += 20
            elif days_ago < 365:
                score += 15
            elif days_ago < 730:
                score += 10
            elif days_ago < 1095:
                score += 5
        except:
            pass

    # Portfolio value (0-15)
    def _safe_float(val, default=0.0):
        try:
            f = float(val)
            return default if pd.isna(f) else f
        except (ValueError, TypeError):
            return default
    pv = _safe_float(row.get('total_portfolio_value', 0))
    if pv >= 3000000:
        score += 15
    elif pv >= 1000000:
        score += 12
    elif pv >= 500000:
        score += 9
    elif pv >= 200000:
        score += 6
    else:
        score += 3

    # Entity sophistication (0-10)
    is_entity = str(row.get('is_entity', '')).lower() in ('true', '1', 'yes')
    entity_count = _safe_int(row.get('entity_count', 0))
    if entity_count >= 2:
        score += 10
    elif is_entity:
        score += 5

    # STR indicator (0-10)
    str_licensed = str(row.get('str_licensed', '')).lower() in ('true', '1', 'yes')
    if str_licensed:
        score += 10

    # Geographic fit (0-10)
    county = str(row.get('county', '')).upper()
    if any(c in county for c in SOUTH_FL_COUNTIES):
        score += 10
    elif any(c in county for c in MAJOR_METRO_COUNTIES):
        score += 7
    elif county:
        score += 3

    # Contact availability (0-10)
    phone_val = str(row.get('phone', '')).strip()
    email_val = str(row.get('email', '')).strip()
    has_phone = bool(phone_val) and phone_val.lower() != 'nan'
    has_email = bool(email_val) and email_val.lower() != 'nan'
    if has_phone and has_email:
        score += 10
    elif has_phone:
        score += 7
    elif has_email:
        score += 5
    else:
        score += 2

    # Refinance opportunity boost (0-40, from Module 8)
    refi_boost = _safe_int(row.get('refi_score_boost', 0))
    score += refi_boost

    return min(score, 100)


def merge_edgar_data(investor_df: pd.DataFrame, edgar_file: str) -> pd.DataFrame:
    """Merge SEC EDGAR fund manager data into the main lead dataset."""

    if not Path(edgar_file).exists():
        investor_df['sec_fund_filing'] = False
        investor_df['fund_name'] = ''
        investor_df['fund_offering_amount'] = ''
        return investor_df

    edgar_df = pd.read_csv(edgar_file, dtype=str, low_memory=False)

    if edgar_df.empty:
        investor_df['sec_fund_filing'] = False
        investor_df['fund_name'] = ''
        investor_df['fund_offering_amount'] = ''
        return investor_df

    # Create a set of fund manager names for matching
    fund_names = set()
    for col in ['display_name', 'issuer_name', 'gp_name', 'related_persons']:
        if col in edgar_df.columns:
            for val in edgar_df[col].dropna():
                for name in str(val).split(';'):
                    name = name.strip().upper()
                    if name and len(name) > 3:
                        fund_names.add(name)

    # Match against investor owner names and resolved persons
    investor_df['sec_fund_filing'] = False
    investor_df['fund_name'] = ''
    investor_df['fund_offering_amount'] = ''

    name_cols = ['resolved_person', 'owner_name']
    for col in investor_df.columns:
        if 'OWN' in col.upper() and 'NAME' in col.upper():
            name_cols.append(col)

    for idx, row in investor_df.iterrows():
        for col in name_cols:
            if col in row.index:
                name = str(row.get(col, '')).strip().upper()
                if name in fund_names:
                    investor_df.at[idx, 'sec_fund_filing'] = True
                    break

    fund_match_count = investor_df['sec_fund_filing'].sum()
    print(f"SEC EDGAR matches: {fund_match_count}")

    # Also add EDGAR-only leads (fund managers not in property data)
    # These are separate leads worth reaching out to
    edgar_leads = []
    for _, row in edgar_df.iterrows():
        edgar_leads.append({
            'owner_name': row.get('issuer_name', row.get('display_name', '')),
            'resolved_person': row.get('gp_name', ''),
            'phone': row.get('issuer_phone', ''),
            'county': '',
            'property_count': 0,
            'total_portfolio_value': 0,
            'sec_fund_filing': True,
            'fund_name': row.get('issuer_name', row.get('display_name', '')),
            'fund_offering_amount': row.get('offering_amount', ''),
            'is_entity': True,
            'enrichment_source': 'sec_edgar',
        })

    if edgar_leads:
        edgar_additions = pd.DataFrame(edgar_leads)
        investor_df = pd.concat([investor_df, edgar_additions], ignore_index=True)
        print(f"Added {len(edgar_leads)} EDGAR-only fund manager leads")

    return investor_df


def create_excel_output(df: pd.DataFrame, output_file: str):
    """Create multi-tab Excel workbook with formatted, human-readable output."""

    print(f"Creating Excel output: {output_file}")

    # --- Column rename map (internal -> plain English) ---
    COLUMN_RENAME = {
        'lead_id': 'Lead ID',
        'score': 'Lead Score',
        'icp_primary': 'Investor Type',
        'icp_secondary': 'Opportunity Type',
        'tier': 'Priority (1=Hot, 2=Warm, 3=Cold)',
        'owner_name': 'Owner Name',
        'owner_type': 'Owner Structure',
        'resolved_person': 'Contact Person',
        'mailing_address': 'Mailing Address',
        'mailing_city': 'City',
        'mailing_state': 'State',
        'mailing_zip': 'Zip',
        'phone': 'Phone',
        'email': 'Email',
        'enrichment_source': 'Contact Source',
        'property_count': '# Properties',
        'total_portfolio_value': 'Portfolio Value',
        'estimated_equity': 'Est. Equity Per Property',
        'equity_ratio': 'Equity %',
        'max_cashout_75': 'Cash-Out Available (75% LTV)',
        'max_cashout_80': 'Cash-Out Available (80% LTV)',
        'refi_signals': 'Refinance Signals',
        'refi_priority': 'Refi Priority',
        'probable_cash_buyer': 'Paid All Cash?',
        'brrrr_exit_candidate': 'Needs Refi After Rehab?',
        'rate_refi_candidate': 'Overpaying on Rate?',
        'equity_harvest_candidate': 'Has Equity to Tap?',
        'most_recent_purchase': 'Last Purchase Date',
        'most_recent_price': 'Last Purchase Price',
        'county': 'County',
        'property_types': 'Property Types',
        'str_licensed': 'Vacation Rental Licensed?',
        'str_license_count': '# Vacation Rental Licenses',
        'out_of_state': 'Out of State?',
        'foreign_owner': 'Foreign Owner?',
        'entity_count': '# LLCs/Entities Controlled',
        'entity_names': 'Entity Names',
        'sec_fund_filing': 'SEC Fund Filing?',
        'fund_name': 'Fund Name',
        'fund_offering_amount': 'Fund Size',
        'data_sources': 'Data Sources',
        'last_updated': 'Last Updated',
    }

    # Column ordering (internal names)
    output_cols = [
        'lead_id', 'score', 'icp_primary', 'icp_secondary', 'tier',
        'owner_name', 'owner_type', 'resolved_person',
        'mailing_address', 'mailing_city', 'mailing_state', 'mailing_zip',
        'phone', 'email', 'enrichment_source',
        'property_count', 'total_portfolio_value',
        'estimated_equity', 'equity_ratio',
        'max_cashout_75', 'max_cashout_80',
        'refi_signals', 'refi_priority',
        'probable_cash_buyer', 'brrrr_exit_candidate',
        'rate_refi_candidate', 'equity_harvest_candidate',
        'most_recent_purchase', 'most_recent_price',
        'county', 'property_types',
        'str_licensed', 'str_license_count',
        'out_of_state', 'foreign_owner',
        'entity_count', 'entity_names',
        'sec_fund_filing', 'fund_name', 'fund_offering_amount',
        'data_sources', 'last_updated',
    ]

    # Sets of renamed column headers by format type
    DOLLAR_COLS = {
        'Portfolio Value', 'Est. Equity Per Property',
        'Cash-Out Available (75% LTV)', 'Cash-Out Available (80% LTV)',
        'Last Purchase Price', 'Fund Size',
    }
    PCT_COLS = {'Equity %'}
    BOOL_COLS = {
        'Paid All Cash?', 'Needs Refi After Rehab?', 'Overpaying on Rate?',
        'Has Equity to Tap?', 'Vacation Rental Licensed?', 'Out of State?',
        'Foreign Owner?', 'SEC Fund Filing?',
    }
    DATE_COLS = {'Last Purchase Date', 'Last Updated'}

    HEADER_FONT = Font(bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

    # --- Prepare data ---
    df_sorted = df.sort_values('score', ascending=False)
    available_cols = [c for c in output_cols if c in df_sorted.columns]

    def _to_bool_yesno(val):
        if str(val).strip().lower() in ('true', '1', 'yes'):
            return 'Yes'
        return 'No'

    def _tier_label(val):
        mapping = {1: 'Hot', 2: 'Warm', 3: 'Cold'}
        try:
            return mapping.get(int(float(val)), str(val))
        except (ValueError, TypeError):
            return str(val)

    def _to_numeric(val):
        try:
            f = float(val)
            return None if pd.isna(f) else f
        except (ValueError, TypeError):
            return None

    def _prepare_sheet_df(sheet_df):
        """Prepare a DataFrame for writing: rename cols, convert types."""
        out = sheet_df[available_cols].copy()

        # Convert booleans to Yes/No before rename
        for col in available_cols:
            renamed = COLUMN_RENAME.get(col, col)
            if renamed in BOOL_COLS:
                out[col] = out[col].apply(_to_bool_yesno)

        # Convert tier to Hot/Warm/Cold
        if 'tier' in out.columns:
            out['tier'] = out['tier'].apply(_tier_label)

        # Convert dollar/pct columns to numeric
        for col in available_cols:
            renamed = COLUMN_RENAME.get(col, col)
            if renamed in DOLLAR_COLS or renamed in PCT_COLS:
                out[col] = out[col].apply(_to_numeric)

        # Rename columns
        out = out.rename(columns=COLUMN_RENAME)
        return out

    def _format_sheet(ws):
        """Apply formatting to an openpyxl worksheet."""
        # Header formatting + freeze
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.freeze_panes = 'A2'

        # Build column name -> letter mapping
        col_map = {}
        for idx, cell in enumerate(ws[1], 1):
            col_map[cell.value] = idx

        # Number formats
        for col_name, col_idx in col_map.items():
            col_letter = get_column_letter(col_idx)
            if col_name in DOLLAR_COLS:
                for row in range(2, ws.max_row + 1):
                    ws[f'{col_letter}{row}'].number_format = '$#,##0'
            elif col_name in PCT_COLS:
                for row in range(2, ws.max_row + 1):
                    ws[f'{col_letter}{row}'].number_format = '0%'
            elif col_name in DATE_COLS:
                for row in range(2, ws.max_row + 1):
                    ws[f'{col_letter}{row}'].number_format = 'YYYY-MM-DD'

        # Score column: conditional fill
        if 'Lead Score' in col_map:
            score_idx = col_map['Lead Score']
            score_letter = get_column_letter(score_idx)
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{score_letter}{row}']
                try:
                    val = float(cell.value) if cell.value is not None else 0
                except (ValueError, TypeError):
                    val = 0
                if val >= 70:
                    cell.fill = GREEN_FILL
                elif val >= 50:
                    cell.fill = YELLOW_FILL

        # Auto-fit column widths (based on header, with min/max)
        for idx, cell in enumerate(ws[1], 1):
            header_len = len(str(cell.value)) if cell.value else 8
            width = max(10, min(header_len + 4, 35))
            ws.column_dimensions[get_column_letter(idx)].width = width

    # --- Build workbook ---
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:

        # Build Summary tab first
        total = len(df)
        summary_data = []
        summary_data.append({'Metric': 'Total Leads', 'Value': f'{total:,}'})
        summary_data.append({'Metric': '', 'Value': ''})
        summary_data.append({'Metric': 'BY INVESTOR TYPE', 'Value': ''})
        for seg in sorted(df['icp_primary'].unique()):
            count = len(df[df['icp_primary'] == seg])
            avg = df.loc[df['icp_primary'] == seg, 'score'].astype(float).mean()
            summary_data.append({'Metric': f'  {seg}', 'Value': f'{count:,} leads (avg score {avg:.0f})'})

        summary_data.append({'Metric': '', 'Value': ''})
        summary_data.append({'Metric': 'BY PRIORITY', 'Value': ''})
        tier_labels = {1: 'Hot', 2: 'Warm', 3: 'Cold'}
        for tier in sorted(df['tier'].unique()):
            count = len(df[df['tier'] == tier])
            label = tier_labels.get(int(tier), f'Tier {tier}')
            summary_data.append({'Metric': f'  {label}', 'Value': f'{count:,}'})

        summary_data.append({'Metric': '', 'Value': ''})
        summary_data.append({'Metric': 'CONTACT COVERAGE', 'Value': ''})
        has_phone = (df['phone'].fillna('') != '').sum() if 'phone' in df.columns else 0
        has_email = (df['email'].fillna('') != '').sum() if 'email' in df.columns else 0
        has_both = 0
        if 'phone' in df.columns and 'email' in df.columns:
            has_both = ((df['phone'].fillna('') != '') & (df['email'].fillna('') != '')).sum()
        has_either = has_phone + has_email - has_both
        summary_data.append({'Metric': '  Has Phone', 'Value': f'{has_phone:,} ({has_phone/total*100:.1f}%)'})
        summary_data.append({'Metric': '  Has Email', 'Value': f'{has_email:,} ({has_email/total*100:.1f}%)'})
        summary_data.append({'Metric': '  Has Either', 'Value': f'{has_either:,} ({has_either/total*100:.1f}%)'})

        summary_data.append({'Metric': '', 'Value': ''})
        summary_data.append({'Metric': 'INVESTOR PROFILE', 'Value': ''})
        def _safe_bool_count(col_name):
            if col_name not in df.columns:
                return 0
            return (df[col_name].fillna(False).astype(str).str.lower().isin(['true', '1', 'yes'])).sum()

        entity_count = _safe_bool_count('is_entity')
        str_count = _safe_bool_count('str_licensed')
        foreign_count = _safe_bool_count('foreign_owner')
        oos_count = _safe_bool_count('out_of_state')
        summary_data.append({'Metric': '  Entity-Owned', 'Value': f'{entity_count:,} ({entity_count/total*100:.1f}%)'})
        summary_data.append({'Metric': '  STR Licensed', 'Value': f'{str_count:,} ({str_count/total*100:.1f}%)'})
        summary_data.append({'Metric': '  Foreign Owner', 'Value': f'{foreign_count:,} ({foreign_count/total*100:.1f}%)'})
        summary_data.append({'Metric': '  Out-of-State', 'Value': f'{oos_count:,} ({oos_count/total*100:.1f}%)'})

        summary_data.append({'Metric': '', 'Value': ''})
        summary_data.append({'Metric': 'SCORE STATS', 'Value': ''})
        summary_data.append({'Metric': '  Average Score', 'Value': f"{df['score'].astype(float).mean():.1f}"})
        summary_data.append({'Metric': '  Median Score', 'Value': f"{df['score'].astype(float).median():.1f}"})
        score_70 = (df['score'].astype(float) >= 70).sum()
        score_50 = ((df['score'].astype(float) >= 50) & (df['score'].astype(float) < 70)).sum()
        score_lt50 = (df['score'].astype(float) < 50).sum()
        summary_data.append({'Metric': '  Score 70+ (Hot)', 'Value': f'{score_70:,}'})
        summary_data.append({'Metric': '  Score 50-69 (Warm)', 'Value': f'{score_50:,}'})
        summary_data.append({'Metric': '  Score <50', 'Value': f'{score_lt50:,}'})

        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)

        # Segment tabs sorted by average score descending — NO "All Leads" tab
        segment_avgs = df.groupby('icp_primary')['score'].apply(
            lambda s: s.astype(float).mean()
        ).sort_values(ascending=False)

        segment_sheets = []
        for segment in segment_avgs.index:
            sheet_name = segment.replace('/', '-')[:31]
            segment_df = df_sorted[df_sorted['icp_primary'] == segment]
            prepared = _prepare_sheet_df(segment_df)
            prepared.to_excel(writer, sheet_name=sheet_name, index=False)
            segment_sheets.append(sheet_name)

        # Apply formatting to all sheets
        wb = writer.book

        # Format Summary sheet
        ws_summary = wb['Summary']
        for cell in ws_summary[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
        ws_summary.freeze_panes = 'A2'
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 35

        # Format segment sheets
        for sheet_name in segment_sheets:
            _format_sheet(wb[sheet_name])

    print(f"Excel workbook created: {output_file}")
    print(f"  Tabs: Summary + {len(segment_sheets)} segment tabs (no 'All Leads' tab)")


def main():
    parser = argparse.ArgumentParser(description='Score and output leads to Excel')
    parser.add_argument('--input', type=str, default='pipeline/output/05_enriched.csv')
    parser.add_argument('--edgar-input', type=str, default='pipeline/output/04_fund_managers.csv')
    parser.add_argument('--output', type=str,
                        default=f'pipeline/output/leads_{date.today().isoformat()}.xlsx')

    args = parser.parse_args()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Load enriched leads
    print("Loading enriched leads...")
    df = pd.read_csv(args.input, dtype=str, low_memory=False)
    print(f"Loaded {len(df):,} leads")

    # Merge EDGAR data
    print("\nMerging SEC EDGAR fund data...")
    df = merge_edgar_data(df, args.edgar_input)

    # Map CO_NO to county name if needed
    if 'county' not in df.columns and 'CO_NO' in df.columns:
        df['county'] = df['CO_NO'].map(CO_NO_TO_COUNTY).fillna('')
        print(f"Mapped {(df['county'] != '').sum():,} leads to county names")

    # Classify ICP
    print("\nClassifying ICP segments...")
    classifications = df.apply(classify_icp, axis=1)
    df['icp_primary'] = [c[0] for c in classifications]
    df['icp_secondary'] = [c[1] for c in classifications]
    df['tier'] = [c[2] for c in classifications]

    # Score leads
    print("Scoring leads...")
    df['score'] = df.apply(score_lead, axis=1)

    # Generate lead IDs
    df['lead_id'] = [f"DSCR-{i:06d}" for i in range(1, len(df) + 1)]

    # Determine owner type
    def get_owner_type(row):
        if str(row.get('foreign_owner', '')).lower() in ('true', '1', 'yes'):
            return 'Foreign'
        owner = str(row.get('owner_name', '') or row.get(
            [c for c in df.columns if 'OWN' in c.upper() and 'NAME' in c.upper()][0]
            if any('OWN' in c.upper() and 'NAME' in c.upper() for c in df.columns) else '', ''))
        if any(kw in owner.upper() for kw in [' LLC', ' L.L.C', ' INC', ' CORP']):
            return 'LLC'
        if 'TRUST' in owner.upper():
            return 'Trust'
        if any(kw in owner.upper() for kw in [' LP', ' LTD', ' PARTNERS']):
            return 'Partnership'
        return 'Individual'

    df['owner_type'] = df.apply(get_owner_type, axis=1)

    # Add metadata
    df['data_sources'] = 'FDOR'
    if 'str_licensed' in df.columns:
        df.loc[df['str_licensed'].fillna(False).astype(str).str.lower().isin(['true', '1', 'yes']), 'data_sources'] += ',DBPR'
    if 'resolved_person' in df.columns:
        df.loc[df['resolved_person'].fillna('') != '', 'data_sources'] += ',SunBiz'
    if 'sec_fund_filing' in df.columns:
        df.loc[df['sec_fund_filing'].fillna(False).astype(str).str.lower().isin(['true', '1', 'yes']), 'data_sources'] += ',EDGAR'
    if 'enrichment_source' in df.columns:
        df.loc[df['enrichment_source'].fillna('') != '', 'data_sources'] += ',Enrichment'
    df['last_updated'] = date.today().isoformat()

    # Standardize column names for output
    # Prefer OWN_STATE_DOM (2-letter code) over OWN_STATE (full name) for mailing_state
    rename_map = {}
    if 'OWN_STATE_DOM' in df.columns and 'mailing_state' not in df.columns:
        rename_map['OWN_STATE_DOM'] = 'mailing_state'
    for col in df.columns:
        if col in rename_map:
            continue
        if 'OWN' in col.upper() and 'NAME' in col.upper() and col != 'owner_name':
            rename_map[col] = 'owner_name'
        elif ('OWN' in col.upper() or 'MAIL' in col.upper()) and 'ADDR' in col.upper() and 'mailing' not in col.lower():
            rename_map[col] = 'mailing_address'
        elif ('OWN' in col.upper() or 'MAIL' in col.upper()) and 'CITY' in col.upper() and 'mailing' not in col.lower():
            rename_map[col] = 'mailing_city'
        elif ('OWN' in col.upper() or 'MAIL' in col.upper()) and 'STATE' in col.upper() and 'mailing' not in col.lower() and 'mailing_state' not in rename_map.values():
            rename_map[col] = 'mailing_state'
        elif ('OWN' in col.upper() or 'MAIL' in col.upper()) and 'ZIP' in col.upper() and 'mailing' not in col.lower():
            rename_map[col] = 'mailing_zip'

    df = df.rename(columns=rename_map)

    # Print summary
    print(f"\n{'='*60}")
    print(f"FINAL LEAD SUMMARY")
    print(f"{'='*60}")
    print(f"Total leads: {len(df):,}")
    print(f"\nBy ICP Segment:")
    for seg in sorted(df['icp_primary'].unique()):
        count = len(df[df['icp_primary'] == seg])
        print(f"  {seg}: {count:,}")
    print(f"\nBy Tier:")
    for tier in sorted(df['tier'].unique()):
        count = len(df[df['tier'] == tier])
        print(f"  Tier {tier}: {count:,}")
    print(f"\nScore Distribution:")
    print(f"  Mean: {df['score'].astype(float).mean():.1f}")
    print(f"  Median: {df['score'].astype(float).median():.1f}")
    print(f"  Score 70+: {(df['score'].astype(float) >= 70).sum():,}")
    print(f"  Score 50-69: {((df['score'].astype(float) >= 50) & (df['score'].astype(float) < 70)).sum():,}")
    print(f"  Score <50: {(df['score'].astype(float) < 50).sum():,}")

    # Create Excel output
    create_excel_output(df, args.output)


if __name__ == '__main__':
    main()
