"""
Module 6: ICP Scoring and Excel Output

Classifies leads into ICP segments, scores by quality,
and produces the final Excel workbook.

Usage:
    python scripts/06_score_and_output.py --input pipeline/output/05_enriched.csv --edgar-input pipeline/output/04_fund_managers.csv --output leads_YYYY-MM-DD.xlsx
"""

import argparse
import pandas as pd
from pathlib import Path
from datetime import datetime, date
import re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter, column_index_from_string

OUTPUT_DIR = Path("pipeline/output")

# County-to-metro mapping for geographic scoring
SOUTH_FL_COUNTIES = ['PALM BEACH', 'BROWARD', 'MIAMI-DADE']
MAJOR_METRO_COUNTIES = [
    'HILLSBOROUGH', 'ORANGE', 'DUVAL', 'PINELLAS', 'SEMINOLE',
    'OSCEOLA', 'LEE', 'COLLIER', 'SARASOTA', 'MANATEE',
    'VOLUSIA', 'BREVARD', 'PASCO', 'POLK', 'ST. LUCIE', 'MARTIN'
]

# FDOR county code -> name mapping (reverse of codes in 01_fdor_download_filter.py)
CO_NO_TO_COUNTY = {
    "11": "ALACHUA", "12": "BAKER", "13": "BAY", "14": "BRADFORD",
    "15": "BREVARD", "16": "BROWARD", "17": "CALHOUN", "18": "CHARLOTTE",
    "19": "CITRUS", "20": "CLAY", "21": "COLLIER", "22": "COLUMBIA",
    "23": "MIAMI-DADE", "24": "DESOTO", "25": "DIXIE", "26": "DUVAL",
    "27": "ESCAMBIA", "28": "FLAGLER", "29": "FRANKLIN", "30": "GADSDEN",
    "31": "GILCHRIST", "32": "GLADES", "33": "GULF", "34": "HAMILTON",
    "35": "HARDEE", "36": "HENDRY", "37": "HERNANDO", "38": "HIGHLANDS",
    "39": "HILLSBOROUGH", "40": "HOLMES", "41": "INDIAN RIVER", "42": "JACKSON",
    "43": "JEFFERSON", "44": "LAFAYETTE", "45": "LAKE", "46": "LEE",
    "47": "LEON", "48": "LEVY", "49": "LIBERTY", "50": "MADISON",
    "51": "MANATEE", "52": "MARION", "53": "MARTIN", "54": "MONROE",
    "55": "NASSAU", "56": "OKALOOSA", "57": "OKEECHOBEE", "58": "ORANGE",
    "59": "OSCEOLA", "60": "PALM BEACH", "61": "PASCO", "62": "PINELLAS",
    "63": "POLK", "64": "PUTNAM", "65": "SAINT JOHNS", "66": "SAINT LUCIE",
    "67": "SANTA ROSA", "68": "SARASOTA", "69": "SEMINOLE", "70": "SUMTER",
    "71": "SUWANNEE", "72": "TAYLOR", "73": "UNION", "74": "VOLUSIA",
    "75": "WAKULLA", "76": "WALTON", "77": "WASHINGTON",
}


def classify_icp(row: pd.Series) -> tuple:
    """
    Classify a lead into primary ICP segment and tier.
    Returns (icp_primary, icp_secondary, tier)
    """

    def _safe_int(val, default=0):
        try:
            f = float(val)
            return default if pd.isna(f) else int(f)
        except (ValueError, TypeError):
            return default

    property_count = _safe_int(row.get('property_count', 0))
    str_licensed = str(row.get('str_licensed', '')).lower() in ('true', '1', 'yes')
    str_count = _safe_int(row.get('str_license_count', 0))
    foreign_owner = str(row.get('foreign_owner', '')).lower() in ('true', '1', 'yes')
    out_of_state = str(row.get('out_of_state', '')).lower() in ('true', '1', 'yes')
    is_entity = str(row.get('is_entity', '')).lower() in ('true', '1', 'yes')
    entity_count = _safe_int(row.get('entity_count', 0))
    sec_fund = str(row.get('sec_fund_filing', '')).lower() in ('true', '1', 'yes')

    # Get property type info
    prop_types = str(row.get('property_types', '') or row.get('DOR_UC', '')).upper()
    is_multifamily = any(code in prop_types for code in ['03', '08', 'MULTI', 'DUPLEX', 'TRIPLEX', 'FOURPLEX'])

    # Get refinance signals
    probable_cash = str(row.get('probable_cash_buyer', '')).lower() in ('true', '1', 'yes')
    brrrr_exit = str(row.get('brrrr_exit_candidate', '')).lower() in ('true', '1', 'yes')
    equity_harvest = str(row.get('equity_harvest_candidate', '')).lower() in ('true', '1', 'yes')
    rate_refi = str(row.get('rate_refi_candidate', '')).lower() in ('true', '1', 'yes')

    # Get recent purchase info for BRRRR detection
    def _safe_float(val, default=0.0):
        try:
            f = float(val)
            return default if pd.isna(f) else f
        except (ValueError, TypeError):
            return default
    recent_price = _safe_float(row.get('most_recent_price', 0))
    portfolio_value = _safe_float(row.get('total_portfolio_value', 0))

    # Classification logic (priority order)
    icp_primary = 'Single Investment Property'
    icp_secondary = ''
    tier = 3

    if property_count >= 10:
        icp_primary = 'Serial Investor (10+)'
        tier = 1
        if str_licensed:
            icp_secondary = 'STR Operator'
        elif foreign_owner:
            icp_secondary = 'Foreign National'

    elif sec_fund:
        icp_primary = 'Fund Manager / Syndicator'
        tier = 2

    elif str_licensed or str_count > 0:
        icp_primary = 'STR Operator'
        tier = 1
        if property_count >= 5:
            icp_secondary = 'Serial Investor (10+)' if property_count >= 10 else 'Growing Portfolio'

    elif foreign_owner:
        icp_primary = 'Foreign National'
        tier = 1
        if is_entity:
            icp_secondary = 'Entity-Based Investor'

    elif is_entity and entity_count >= 2:
        icp_primary = 'Entity-Based Investor'
        tier = 1
        if property_count >= 5:
            icp_secondary = 'Growing Portfolio'

    elif is_entity and property_count >= 2:
        icp_primary = 'Entity-Based Investor'
        tier = 1

    elif property_count >= 2:
        icp_primary = 'Individual Investor (2-9)'
        tier = 1 if property_count >= 5 else 2
        if is_multifamily:
            icp_secondary = 'Multi-Family Investor'

    elif is_multifamily:
        icp_primary = 'Multi-Family Investor'
        tier = 2

    elif out_of_state:
        icp_primary = 'Out-of-State Investor'
        tier = 2

    else:
        icp_primary = 'Single Investment Property'
        tier = 3

    # Refinance-based secondary tagging (overlay on any primary ICP)
    if not icp_secondary:
        if probable_cash:
            icp_secondary = 'Cash-Out Refi Candidate'
        elif brrrr_exit:
            icp_secondary = 'BRRRR Exit Candidate'
        elif equity_harvest:
            icp_secondary = 'Equity Harvest Candidate'
        elif rate_refi:
            icp_secondary = 'Rate Refi Candidate'

    # Upgrade tier if strong refi signal on otherwise lower-tier lead
    if tier == 3 and (probable_cash or brrrr_exit):
        tier = 2
    if tier == 2 and probable_cash and property_count >= 2:
        tier = 1

    return icp_primary, icp_secondary, tier


def score_lead(row: pd.Series) -> int:
    """
    Score a lead 0-100 based on DSCR qualification factors.

    Scoring philosophy: measures how likely this person is a genuine DSCR
    borrower, NOT how easy they are to reach. Contact availability is tracked
    separately via reachability_score().

    Components (max 100):
      Property count:       0-25  (portfolio scale)
      Recency:              0-15  (active vs dormant investor)
      Portfolio value:       0-15  (deal size / sophistication)
      Entity sophistication: 0-10  (tax planning = repeat borrower)
      STR indicator:        0-15  (strongest DSCR product fit signal)
      Geographic fit:        0-5   (market demand, not lead quality)
      Refi signals:          0-15  (inline, from Module 8)
    """

    score = 0

    def _safe_int(val, default=0):
        try:
            f = float(val)
            return default if pd.isna(f) else int(f)
        except (ValueError, TypeError):
            return default

    def _safe_float(val, default=0.0):
        try:
            f = float(val)
            return default if pd.isna(f) else f
        except (ValueError, TypeError):
            return default

    # Property count (0-25)
    pc = _safe_int(row.get('property_count', 0))
    if pc >= 20:
        score += 25
    elif pc >= 10:
        score += 20
    elif pc >= 5:
        score += 15
    elif pc >= 2:
        score += 10
    elif pc >= 1:
        score += 5

    # Recency of last purchase (0-15)
    recent = str(row.get('most_recent_purchase', ''))
    if recent and recent != 'nan':
        try:
            purchase_date = pd.to_datetime(recent)
            days_ago = (datetime.now() - purchase_date).days
            if days_ago < 180:
                score += 15
            elif days_ago < 365:
                score += 12
            elif days_ago < 730:
                score += 8
            elif days_ago < 1095:
                score += 4
        except:
            pass

    # Portfolio value (0-15)
    pv = _safe_float(row.get('total_portfolio_value', 0))
    if pv >= 3000000:
        score += 15
    elif pv >= 1000000:
        score += 12
    elif pv >= 500000:
        score += 9
    elif pv >= 200000:
        score += 6
    else:
        score += 3

    # Entity sophistication (0-10)
    is_entity = str(row.get('is_entity', '')).lower() in ('true', '1', 'yes')
    entity_count = _safe_int(row.get('entity_count', 0))
    if entity_count >= 2:
        score += 10
    elif is_entity:
        score += 5

    # STR indicator (0-15) — scaled by license count
    str_licensed = str(row.get('str_licensed', '')).lower() in ('true', '1', 'yes')
    str_count = _safe_int(row.get('str_license_count', 0))
    if str_licensed:
        if str_count >= 3:
            score += 15
        elif str_count >= 2:
            score += 12
        else:
            score += 10

    # Geographic fit (0-5) — market demand signal, not lead quality
    county = str(row.get('county', '')).upper()
    if any(c in county for c in SOUTH_FL_COUNTIES):
        score += 5
    elif any(c in county for c in MAJOR_METRO_COUNTIES):
        score += 3
    elif county:
        score += 1

    # Refi signals inline (0-15) — from Module 8, capped and weighted
    refi_boost = _safe_int(row.get('refi_score_boost', 0))
    score += min(refi_boost, 15)

    return min(score, 100)


def reachability_score(row: pd.Series) -> int:
    """
    Separate score (0-10) for how actionable a lead is based on contact data.
    This is NOT a quality signal — it measures outreach readiness.
    """
    phone_val = str(row.get('phone', '')).strip()
    email_val = str(row.get('email', '')).strip()
    has_phone = bool(phone_val) and phone_val.lower() != 'nan'
    has_email = bool(email_val) and email_val.lower() != 'nan'
    if has_phone and has_email:
        return 10
    elif has_phone:
        return 7
    elif has_email:
        return 5
    return 2


def merge_edgar_data(investor_df: pd.DataFrame, edgar_file: str) -> pd.DataFrame:
    """Merge SEC EDGAR fund manager data into the main lead dataset."""

    if not Path(edgar_file).exists():
        investor_df['sec_fund_filing'] = False
        investor_df['fund_name'] = ''
        investor_df['fund_offering_amount'] = ''
        return investor_df

    edgar_df = pd.read_csv(edgar_file, dtype=str, low_memory=False)

    if edgar_df.empty:
        investor_df['sec_fund_filing'] = False
        investor_df['fund_name'] = ''
        investor_df['fund_offering_amount'] = ''
        return investor_df

    # Create a set of fund manager names for matching
    fund_names = set()
    for col in ['display_name', 'issuer_name', 'gp_name', 'related_persons']:
        if col in edgar_df.columns:
            for val in edgar_df[col].dropna():
                for name in str(val).split(';'):
                    name = name.strip().upper()
                    if name and len(name) > 3:
                        fund_names.add(name)

    # Match against investor owner names and resolved persons
    investor_df['sec_fund_filing'] = False
    investor_df['fund_name'] = ''
    investor_df['fund_offering_amount'] = ''

    name_cols = ['resolved_person', 'owner_name']
    for col in investor_df.columns:
        if 'OWN' in col.upper() and 'NAME' in col.upper():
            name_cols.append(col)

    for idx, row in investor_df.iterrows():
        for col in name_cols:
            if col in row.index:
                name = str(row.get(col, '')).strip().upper()
                if name in fund_names:
                    investor_df.at[idx, 'sec_fund_filing'] = True
                    break

    fund_match_count = investor_df['sec_fund_filing'].sum()
    print(f"SEC EDGAR matches: {fund_match_count}")

    # Also add EDGAR-only leads (fund managers not in property data)
    # These are separate leads worth reaching out to
    edgar_leads = []
    for _, row in edgar_df.iterrows():
        edgar_leads.append({
            'owner_name': row.get('issuer_name', row.get('display_name', '')),
            'resolved_person': row.get('gp_name', ''),
            'phone': row.get('issuer_phone', ''),
            'county': '',
            'property_count': 0,
            'total_portfolio_value': 0,
            'sec_fund_filing': True,
            'fund_name': row.get('issuer_name', row.get('display_name', '')),
            'fund_offering_amount': row.get('offering_amount', ''),
            'is_entity': True,
            'enrichment_source': 'sec_edgar',
        })

    if edgar_leads:
        edgar_additions = pd.DataFrame(edgar_leads)
        investor_df = pd.concat([investor_df, edgar_additions], ignore_index=True)
        print(f"Added {len(edgar_leads)} EDGAR-only fund manager leads")

    return investor_df


# ---------------------------------------------------------------------------
# Dashboard helpers
# ---------------------------------------------------------------------------

# Color palette
_NAVY = '1B2A4A'
_ACCENT_BLUE = '4472C4'
_HOT_RED = 'C00000'
_WARM_ORANGE = 'ED7D31'
_MONEY_GREEN = '548235'
_LIGHT_BG = 'F2F2F2'
_WHITE = 'FFFFFF'


def _fmt_dollars_short(value):
    """Format a dollar amount as $1.2M / $4.2B for KPI tiles."""
    try:
        v = float(value)
    except (ValueError, TypeError):
        return '$0'
    if pd.isna(v) or v == 0:
        return '$0'
    if abs(v) >= 1_000_000_000:
        return f'${v / 1_000_000_000:,.1f}B'
    if abs(v) >= 1_000_000:
        return f'${v / 1_000_000:,.1f}M'
    if abs(v) >= 1_000:
        return f'${v / 1_000:,.0f}K'
    return f'${v:,.0f}'


def _write_merged(ws, start_col, end_col, row, value,
                  font=None, fill=None, alignment=None, number_format=None):
    """Write a value into a merged cell range, applying fill to every cell."""
    if start_col != end_col:
        ws.merge_cells(start_row=row, start_column=start_col,
                       end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if number_format:
        cell.number_format = number_format
    # Apply fill to every cell in the range so there are no white gaps
    if fill:
        for c in range(start_col, end_col + 1):
            ws.cell(row=row, column=c).fill = fill


def _build_summary_dashboard(ws, df):
    """Build the executive Summary dashboard onto *ws*."""

    # Pre-compute metrics
    total = len(df)
    scores = df['score'].astype(float)
    avg_score = scores.mean()

    hot_mask = scores >= 60
    warm_mask = (scores >= 40) & (scores < 60)
    cold_mask = scores < 40
    hot_count = int(hot_mask.sum())
    warm_count = int(warm_mask.sum())
    cold_count = int(cold_mask.sum())

    def _safe_float_col(col):
        if col not in df.columns:
            return pd.Series([0.0] * len(df))
        return pd.to_numeric(df[col], errors='coerce').fillna(0)

    portfolio_vals = _safe_float_col('total_portfolio_value')
    total_portfolio = portfolio_vals.sum()

    def _has(col):
        if col not in df.columns:
            return pd.Series([False] * len(df))
        return df[col].fillna('').astype(str).str.strip().ne('') & df[col].astype(str).str.lower().ne('nan')

    has_phone = _has('phone')
    has_email = _has('email')
    phone_count = int(has_phone.sum())
    email_count = int(has_email.sum())
    both_count = int((has_phone & has_email).sum())
    either_count = int((has_phone | has_email).sum())
    none_count = total - either_count

    def _bool_col(col):
        if col not in df.columns:
            return pd.Series([False] * len(df))
        return df[col].fillna('').astype(str).str.lower().isin(['true', '1', 'yes'])

    # Fills
    navy_fill = PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type='solid')
    blue_fill = PatternFill(start_color=_ACCENT_BLUE, end_color=_ACCENT_BLUE, fill_type='solid')
    red_fill = PatternFill(start_color=_HOT_RED, end_color=_HOT_RED, fill_type='solid')
    orange_fill = PatternFill(start_color=_WARM_ORANGE, end_color=_WARM_ORANGE, fill_type='solid')
    green_fill = PatternFill(start_color=_MONEY_GREEN, end_color=_MONEY_GREEN, fill_type='solid')
    light_fill = PatternFill(start_color=_LIGHT_BG, end_color=_LIGHT_BG, fill_type='solid')
    white_fill = PatternFill(start_color=_WHITE, end_color=_WHITE, fill_type='solid')

    # Fonts
    banner_font = Font(size=18, bold=True, color=_WHITE)
    subtitle_font = Font(size=10, color=_WHITE)
    section_font = Font(size=13, bold=True, color=_WHITE)
    kpi_value_font = Font(size=20, bold=True, color=_WHITE)
    kpi_label_font = Font(size=9, color=_WHITE)
    table_header_font = Font(size=10, bold=True, color=_WHITE)
    table_cell_font = Font(size=10, color='333333')
    table_cell_bold = Font(size=10, bold=True, color='333333')
    footer_font = Font(size=8, italic=True, color='999999')

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')

    thin_border = Border(
        bottom=Side(style='thin', color='D9D9D9'),
    )

    # Hide gridlines
    ws.sheet_view.showGridLines = False

    # Column widths: A=gutter(2), B-G=content(~20 each), H=gutter(2)
    ws.column_dimensions['A'].width = 2
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 22
    ws.column_dimensions['H'].width = 2

    row = 1  # row cursor

    # ===== SECTION 1: Banner =====
    _write_merged(ws, 2, 7, row, 'DSCR LEAD GENERATION DASHBOARD',
                  font=banner_font, fill=navy_fill, alignment=center)
    ws.row_dimensions[row].height = 40
    # Gutter fills
    ws.cell(row=row, column=1).fill = navy_fill
    ws.cell(row=row, column=8).fill = navy_fill
    row += 1

    subtitle = f'Generated {date.today().strftime("%B %d, %Y")}  |  CrossCountry Mortgage  |  Florida Investment Properties'
    _write_merged(ws, 2, 7, row, subtitle,
                  font=subtitle_font, fill=navy_fill, alignment=center)
    ws.row_dimensions[row].height = 22
    ws.cell(row=row, column=1).fill = navy_fill
    ws.cell(row=row, column=8).fill = navy_fill
    row += 1

    # Spacer
    row += 1

    # ===== SECTION 2: Pipeline at a Glance — 6 KPI tiles =====
    _write_merged(ws, 2, 7, row, 'PIPELINE AT A GLANCE',
                  font=section_font, fill=blue_fill, alignment=left)
    ws.row_dimensions[row].height = 28
    row += 1

    # Row of 3 KPI tiles: Total Leads | Hot Leads | Warm Leads
    kpi_tiles_row1 = [
        (f'{total:,}', 'Total Leads', navy_fill),
        (f'{hot_count:,}', 'Hot Leads (60+)', red_fill),
        (f'{warm_count:,}', 'Warm Leads (40-59)', orange_fill),
    ]
    for i, (val, label, fill) in enumerate(kpi_tiles_row1):
        col_start = 2 + i * 2
        col_end = col_start + 1
        _write_merged(ws, col_start, col_end, row, val,
                      font=kpi_value_font, fill=fill, alignment=center)
    ws.row_dimensions[row].height = 45
    row += 1

    for i, (val, label, fill) in enumerate(kpi_tiles_row1):
        col_start = 2 + i * 2
        col_end = col_start + 1
        _write_merged(ws, col_start, col_end, row, label,
                      font=kpi_label_font, fill=fill, alignment=center)
    ws.row_dimensions[row].height = 20
    row += 1

    # Row of 3 KPI tiles: Avg Score | Contactable | Total Portfolio
    kpi_tiles_row2 = [
        (f'{avg_score:.0f}', 'Avg Lead Score', blue_fill),
        (f'{phone_count:,}', 'Contactable (Phone)', green_fill),
        (_fmt_dollars_short(total_portfolio), 'Total Portfolio Value', green_fill),
    ]
    for i, (val, label, fill) in enumerate(kpi_tiles_row2):
        col_start = 2 + i * 2
        col_end = col_start + 1
        _write_merged(ws, col_start, col_end, row, val,
                      font=kpi_value_font, fill=fill, alignment=center)
    ws.row_dimensions[row].height = 45
    row += 1

    for i, (val, label, fill) in enumerate(kpi_tiles_row2):
        col_start = 2 + i * 2
        col_end = col_start + 1
        _write_merged(ws, col_start, col_end, row, label,
                      font=kpi_label_font, fill=fill, alignment=center)
    ws.row_dimensions[row].height = 20
    row += 1

    # Spacer
    row += 1

    # ===== SECTION 3: Score Distribution =====
    _write_merged(ws, 2, 7, row, 'SCORE DISTRIBUTION',
                  font=section_font, fill=blue_fill, alignment=left)
    ws.row_dimensions[row].height = 28
    row += 1

    # Header row
    score_headers = ['Band', '# Leads', '% of Total', 'Avg Properties', 'Avg Portfolio $', 'Phone %']
    for i, hdr in enumerate(score_headers):
        ws.cell(row=row, column=2 + i, value=hdr).font = table_header_font
        ws.cell(row=row, column=2 + i).fill = navy_fill
        ws.cell(row=row, column=2 + i).alignment = center
    ws.row_dimensions[row].height = 24
    row += 1

    prop_counts = _safe_float_col('property_count')
    # Data rows: Hot, Warm, Cold
    bands = [
        ('Hot (60+)', hot_mask, red_fill),
        ('Warm (40-59)', warm_mask, orange_fill),
        ('Cold (<40)', cold_mask, light_fill),
    ]
    for band_label, mask, band_fill in bands:
        n = int(mask.sum())
        pct = f'{n / total * 100:.1f}%' if total > 0 else '0%'
        avg_props = f'{prop_counts[mask].mean():.1f}' if n > 0 else '0'
        avg_port = _fmt_dollars_short(portfolio_vals[mask].mean()) if n > 0 else '$0'
        phone_pct = f'{has_phone[mask].sum() / n * 100:.0f}%' if n > 0 else '0%'
        vals = [band_label, f'{n:,}', pct, avg_props, avg_port, phone_pct]
        for i, v in enumerate(vals):
            cell = ws.cell(row=row, column=2 + i, value=v)
            cell.font = table_cell_bold if i == 0 else table_cell_font
            cell.alignment = center if i > 0 else left
            cell.border = thin_border
            if band_label.startswith('Cold'):
                cell.fill = light_fill
        ws.row_dimensions[row].height = 22
        row += 1

    # Spacer
    row += 1

    # ===== SECTION 4: Investor Segments =====
    _write_merged(ws, 2, 7, row, 'INVESTOR SEGMENTS',
                  font=section_font, fill=blue_fill, alignment=left)
    ws.row_dimensions[row].height = 28
    row += 1

    seg_headers = ['Segment', '# Leads', 'Avg Score', 'Tier 1 %', 'Avg Portfolio $', 'Phone %']
    for i, hdr in enumerate(seg_headers):
        ws.cell(row=row, column=2 + i, value=hdr).font = table_header_font
        ws.cell(row=row, column=2 + i).fill = navy_fill
        ws.cell(row=row, column=2 + i).alignment = center
    ws.row_dimensions[row].height = 24
    row += 1

    seg_stats = df.groupby('icp_primary').agg(
        count=('score', 'size'),
        avg_score=('score', lambda s: s.astype(float).mean()),
    ).sort_values('avg_score', ascending=False)

    for idx_num, (seg_name, seg_row) in enumerate(seg_stats.iterrows()):
        seg_mask = df['icp_primary'] == seg_name
        n = int(seg_row['count'])
        tier1_pct = f"{(df.loc[seg_mask, 'tier'].astype(int) == 1).sum() / n * 100:.0f}%" if n > 0 else '0%'
        avg_port = _fmt_dollars_short(portfolio_vals[seg_mask].mean())
        phone_pct = f'{has_phone[seg_mask].sum() / n * 100:.0f}%' if n > 0 else '0%'
        vals = [seg_name, f'{n:,}', f'{seg_row["avg_score"]:.0f}', tier1_pct, avg_port, phone_pct]
        for i, v in enumerate(vals):
            cell = ws.cell(row=row, column=2 + i, value=v)
            cell.font = table_cell_bold if i == 0 else table_cell_font
            cell.alignment = center if i > 0 else left
            cell.border = thin_border
            if idx_num % 2 == 1:
                cell.fill = light_fill
        ws.row_dimensions[row].height = 22
        row += 1

    # Spacer
    row += 1

    # ===== SECTION 5: Refinance Opportunities =====
    _write_merged(ws, 2, 7, row, 'REFINANCE OPPORTUNITIES',
                  font=section_font, fill=green_fill, alignment=left)
    ws.row_dimensions[row].height = 28
    row += 1

    # Refi KPI tiles (3 across)
    cash_buyer = _bool_col('probable_cash_buyer')
    equity_harv = _bool_col('equity_harvest_candidate')
    brrrr = _bool_col('brrrr_exit_candidate')
    rate_refi = _bool_col('rate_refi_candidate')
    any_refi = cash_buyer | equity_harv | brrrr | rate_refi
    refi_total = int(any_refi.sum())

    cashout_75 = _safe_float_col('max_cashout_75')
    est_cashout = cashout_75[any_refi].sum()

    high_refi = int((any_refi & hot_mask).sum())

    refi_kpis = [
        (f'{refi_total:,}', 'Refi Candidates', green_fill),
        (_fmt_dollars_short(est_cashout), 'Est. Cash-Out Potential', green_fill),
        (f'{high_refi:,}', 'High-Priority Refi', red_fill),
    ]
    for i, (val, label, fill) in enumerate(refi_kpis):
        col_start = 2 + i * 2
        col_end = col_start + 1
        _write_merged(ws, col_start, col_end, row, val,
                      font=kpi_value_font, fill=fill, alignment=center)
    ws.row_dimensions[row].height = 45
    row += 1

    for i, (val, label, fill) in enumerate(refi_kpis):
        col_start = 2 + i * 2
        col_end = col_start + 1
        _write_merged(ws, col_start, col_end, row, label,
                      font=kpi_label_font, fill=fill, alignment=center)
    ws.row_dimensions[row].height = 20
    row += 1

    # Spacer row
    row += 1

    # Refi breakdown table
    refi_headers = ['Refi Type', '# Leads', 'Avg Score', 'Avg Portfolio $']
    for i, hdr in enumerate(refi_headers):
        ws.cell(row=row, column=2 + i, value=hdr).font = table_header_font
        ws.cell(row=row, column=2 + i).fill = navy_fill
        ws.cell(row=row, column=2 + i).alignment = center
    ws.row_dimensions[row].height = 24
    row += 1

    refi_types = [
        ('Cash Buyer (Leverage Up)', cash_buyer),
        ('Equity Harvest (30%+ Equity)', equity_harv),
        ('BRRRR Exit (Hard Money Out)', brrrr),
        ('Rate Refi (2022-23 Vintage)', rate_refi),
    ]
    for idx_num, (rtype, rmask) in enumerate(refi_types):
        n = int(rmask.sum())
        avg_sc = f'{scores[rmask].mean():.0f}' if n > 0 else '-'
        avg_port = _fmt_dollars_short(portfolio_vals[rmask].mean()) if n > 0 else '$0'
        vals = [rtype, f'{n:,}', avg_sc, avg_port]
        for i, v in enumerate(vals):
            cell = ws.cell(row=row, column=2 + i, value=v)
            cell.font = table_cell_bold if i == 0 else table_cell_font
            cell.alignment = center if i > 0 else left
            cell.border = thin_border
            if idx_num % 2 == 1:
                cell.fill = light_fill
        ws.row_dimensions[row].height = 22
        row += 1

    # Spacer
    row += 1

    # ===== SECTION 6: Contact & Outreach Readiness =====
    _write_merged(ws, 2, 7, row, 'CONTACT & OUTREACH READINESS',
                  font=section_font, fill=blue_fill, alignment=left)
    ws.row_dimensions[row].height = 28
    row += 1

    # Left block (B-D): contact coverage
    contact_hdr = ['Contact Type', '# Leads', '% of Total']
    for i, hdr in enumerate(contact_hdr):
        ws.cell(row=row, column=2 + i, value=hdr).font = table_header_font
        ws.cell(row=row, column=2 + i).fill = navy_fill
        ws.cell(row=row, column=2 + i).alignment = center
    # Right block (E-G): actionable combos
    action_hdr = ['Actionable Segment', '# Leads', '']
    for i, hdr in enumerate(action_hdr):
        ws.cell(row=row, column=5 + i, value=hdr).font = table_header_font
        ws.cell(row=row, column=5 + i).fill = navy_fill
        ws.cell(row=row, column=5 + i).alignment = center
    ws.row_dimensions[row].height = 24
    row += 1

    # Actionable combos
    str_licensed = _bool_col('str_licensed')
    hot_w_phone = int((hot_mask & has_phone).sum())
    warm_w_phone = int((warm_mask & has_phone).sum())
    str_w_phone = int((str_licensed & has_phone).sum())
    refi_w_phone = int((any_refi & has_phone).sum())

    contact_rows = [
        ('Has Phone', phone_count, f'{phone_count / total * 100:.1f}%' if total > 0 else '0%'),
        ('Has Email', email_count, f'{email_count / total * 100:.1f}%' if total > 0 else '0%'),
        ('Has Both', both_count, f'{both_count / total * 100:.1f}%' if total > 0 else '0%'),
        ('Has Either', either_count, f'{either_count / total * 100:.1f}%' if total > 0 else '0%'),
        ('No Contact Info', none_count, f'{none_count / total * 100:.1f}%' if total > 0 else '0%'),
    ]
    action_rows = [
        ('Score 60+ w/ Phone', hot_w_phone),
        ('Score 40+ w/ Phone', warm_w_phone + hot_w_phone),
        ('STR Operator w/ Phone', str_w_phone),
        ('Refi Candidate w/ Phone', refi_w_phone),
        ('', ''),  # pad to match left
    ]

    for idx_num, ((ct_label, ct_n, ct_pct), (act_label, act_n)) in enumerate(zip(contact_rows, action_rows)):
        # Left block
        ws.cell(row=row, column=2, value=ct_label).font = table_cell_bold
        ws.cell(row=row, column=2).alignment = left
        ws.cell(row=row, column=3, value=f'{ct_n:,}' if isinstance(ct_n, int) else ct_n).font = table_cell_font
        ws.cell(row=row, column=3).alignment = center
        ws.cell(row=row, column=4, value=ct_pct).font = table_cell_font
        ws.cell(row=row, column=4).alignment = center
        # Right block
        ws.cell(row=row, column=5, value=act_label).font = table_cell_bold
        ws.cell(row=row, column=5).alignment = left
        ws.cell(row=row, column=6, value=f'{act_n:,}' if isinstance(act_n, int) and act_label else '').font = table_cell_font
        ws.cell(row=row, column=6).alignment = center
        # Alternating fill
        for c in range(2, 8):
            ws.cell(row=row, column=c).border = thin_border
            if idx_num % 2 == 1:
                ws.cell(row=row, column=c).fill = light_fill
        ws.row_dimensions[row].height = 22
        row += 1

    # Spacer
    row += 1

    # ===== SECTION 7: Top 10 Leads Preview =====
    _write_merged(ws, 2, 7, row, 'TOP 10 LEADS PREVIEW',
                  font=section_font, fill=navy_fill, alignment=left)
    ws.row_dimensions[row].height = 28
    row += 1

    top10_headers = ['Score', 'Owner', 'Investor Type', '# Props', 'Portfolio Value', 'Refi Signal']
    for i, hdr in enumerate(top10_headers):
        ws.cell(row=row, column=2 + i, value=hdr).font = table_header_font
        ws.cell(row=row, column=2 + i).fill = navy_fill
        ws.cell(row=row, column=2 + i).alignment = center
    ws.row_dimensions[row].height = 24
    row += 1

    top10 = df.nlargest(10, 'score')
    for idx_num, (_, lead) in enumerate(top10.iterrows()):
        owner = str(lead.get('owner_name', '') or lead.get('resolved_person', '')).strip()
        if not owner or owner.lower() == 'nan':
            owner = str(lead.get('resolved_person', '')).strip()
        if owner.lower() == 'nan':
            owner = '(Unknown)'

        refi_signals = []
        if str(lead.get('probable_cash_buyer', '')).lower() in ('true', '1', 'yes'):
            refi_signals.append('Cash Buyer')
        if str(lead.get('equity_harvest_candidate', '')).lower() in ('true', '1', 'yes'):
            refi_signals.append('Equity Harvest')
        if str(lead.get('brrrr_exit_candidate', '')).lower() in ('true', '1', 'yes'):
            refi_signals.append('BRRRR Exit')
        if str(lead.get('rate_refi_candidate', '')).lower() in ('true', '1', 'yes'):
            refi_signals.append('Rate Refi')
        refi_str = ', '.join(refi_signals) if refi_signals else '-'

        prop_ct = 0
        try:
            prop_ct = int(float(lead.get('property_count', 0)))
        except (ValueError, TypeError):
            pass

        vals = [
            int(float(lead.get('score', 0))),
            owner[:30],
            str(lead.get('icp_primary', '')),
            prop_ct,
            _fmt_dollars_short(lead.get('total_portfolio_value', 0)),
            refi_str,
        ]
        for i, v in enumerate(vals):
            cell = ws.cell(row=row, column=2 + i, value=v)
            cell.font = table_cell_bold if i <= 1 else table_cell_font
            cell.alignment = center if i != 1 and i != 2 and i != 5 else left
            cell.border = thin_border
            if idx_num % 2 == 1:
                cell.fill = light_fill
        ws.row_dimensions[row].height = 22
        row += 1

    # Spacer + Footer
    row += 1
    footer = 'Data sources: FDOR (FL Dept of Revenue) | SunBiz | DBPR | SEC EDGAR | Contact Enrichment'
    _write_merged(ws, 2, 7, row, footer,
                  font=footer_font, fill=None, alignment=center)


def create_excel_output(df: pd.DataFrame, output_file: str):
    """Create multi-tab Excel workbook with formatted, human-readable output."""

    print(f"Creating Excel output: {output_file}")

    # --- Column rename map (internal -> plain English) ---
    COLUMN_RENAME = {
        'lead_id': 'Lead ID',
        'score': 'Lead Score',
        'icp_primary': 'Investor Type',
        'icp_secondary': 'Opportunity Type',
        'tier': 'Priority Tier',
        'reachability': 'Reachability (0-10)',
        'owner_name': 'Owner Name',
        'owner_type': 'Owner Structure',
        'resolved_person': 'Contact Person',
        'mailing_address': 'Mailing Address',
        'mailing_city': 'City',
        'mailing_state': 'State',
        'mailing_zip': 'Zip',
        'phone': 'Phone',
        'email': 'Email',
        'enrichment_source': 'Contact Source',
        'property_count': '# Properties',
        'total_portfolio_value': 'Portfolio Value',
        'estimated_equity': 'Est. Equity Per Property',
        'equity_ratio': 'Equity %',
        'max_cashout_75': 'Cash-Out Available (75% LTV)',
        'max_cashout_80': 'Cash-Out Available (80% LTV)',
        'refi_signals': 'Refinance Signals',
        'refi_priority': 'Refi Priority',
        'probable_cash_buyer': 'Paid All Cash?',
        'brrrr_exit_candidate': 'Needs Refi After Rehab?',
        'rate_refi_candidate': 'Overpaying on Rate?',
        'equity_harvest_candidate': 'Has Equity to Tap?',
        'most_recent_purchase': 'Last Purchase Date',
        'most_recent_price': 'Last Purchase Price',
        'county': 'County',
        'property_types': 'Property Types',
        'str_licensed': 'Vacation Rental Licensed?',
        'str_license_count': '# Vacation Rental Licenses',
        'out_of_state': 'Out of State?',
        'foreign_owner': 'Foreign Owner?',
        'entity_count': '# LLCs/Entities Controlled',
        'entity_names': 'Entity Names',
        'sec_fund_filing': 'SEC Fund Filing?',
        'fund_name': 'Fund Name',
        'fund_offering_amount': 'Fund Size',
        'data_sources': 'Data Sources',
        'last_updated': 'Last Updated',
    }

    # Column ordering (internal names)
    output_cols = [
        'lead_id', 'score', 'reachability', 'icp_primary', 'icp_secondary', 'tier',
        'owner_name', 'owner_type', 'resolved_person',
        'mailing_address', 'mailing_city', 'mailing_state', 'mailing_zip',
        'phone', 'email', 'enrichment_source',
        'property_count', 'total_portfolio_value',
        'estimated_equity', 'equity_ratio',
        'max_cashout_75', 'max_cashout_80',
        'refi_signals', 'refi_priority',
        'probable_cash_buyer', 'brrrr_exit_candidate',
        'rate_refi_candidate', 'equity_harvest_candidate',
        'most_recent_purchase', 'most_recent_price',
        'county', 'property_types',
        'str_licensed', 'str_license_count',
        'out_of_state', 'foreign_owner',
        'entity_count', 'entity_names',
        'sec_fund_filing', 'fund_name', 'fund_offering_amount',
        'data_sources', 'last_updated',
    ]

    # Sets of renamed column headers by format type
    DOLLAR_COLS = {
        'Portfolio Value', 'Est. Equity Per Property',
        'Cash-Out Available (75% LTV)', 'Cash-Out Available (80% LTV)',
        'Last Purchase Price', 'Fund Size',
    }
    PCT_COLS = {'Equity %'}
    BOOL_COLS = {
        'Paid All Cash?', 'Needs Refi After Rehab?', 'Overpaying on Rate?',
        'Has Equity to Tap?', 'Vacation Rental Licensed?', 'Out of State?',
        'Foreign Owner?', 'SEC Fund Filing?',
    }
    DATE_COLS = {'Last Purchase Date', 'Last Updated'}

    HEADER_FONT = Font(bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    YELLOW_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')

    # --- Prepare data ---
    df_sorted = df.sort_values('score', ascending=False)
    available_cols = [c for c in output_cols if c in df_sorted.columns]

    def _to_bool_yesno(val):
        if str(val).strip().lower() in ('true', '1', 'yes'):
            return 'Yes'
        return 'No'

    def _tier_label(val):
        mapping = {1: 'Hot', 2: 'Warm', 3: 'Cold'}
        try:
            return mapping.get(int(float(val)), str(val))
        except (ValueError, TypeError):
            return str(val)

    def _to_numeric(val):
        try:
            f = float(val)
            return None if pd.isna(f) else f
        except (ValueError, TypeError):
            return None

    def _prepare_sheet_df(sheet_df):
        """Prepare a DataFrame for writing: rename cols, convert types."""
        out = sheet_df[available_cols].copy()

        # Convert booleans to Yes/No before rename
        for col in available_cols:
            renamed = COLUMN_RENAME.get(col, col)
            if renamed in BOOL_COLS:
                out[col] = out[col].apply(_to_bool_yesno)

        # Convert tier to Hot/Warm/Cold
        if 'tier' in out.columns:
            out['tier'] = out['tier'].apply(_tier_label)

        # Convert dollar/pct columns to numeric
        for col in available_cols:
            renamed = COLUMN_RENAME.get(col, col)
            if renamed in DOLLAR_COLS or renamed in PCT_COLS:
                out[col] = out[col].apply(_to_numeric)

        # Rename columns
        out = out.rename(columns=COLUMN_RENAME)
        return out

    def _format_sheet(ws):
        """Apply formatting to an openpyxl worksheet."""
        # Header formatting + freeze
        for cell in ws[1]:
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        ws.freeze_panes = 'A2'

        # Build column name -> letter mapping
        col_map = {}
        for idx, cell in enumerate(ws[1], 1):
            col_map[cell.value] = idx

        # Number formats
        for col_name, col_idx in col_map.items():
            col_letter = get_column_letter(col_idx)
            if col_name in DOLLAR_COLS:
                for row in range(2, ws.max_row + 1):
                    ws[f'{col_letter}{row}'].number_format = '$#,##0'
            elif col_name in PCT_COLS:
                for row in range(2, ws.max_row + 1):
                    ws[f'{col_letter}{row}'].number_format = '0%'
            elif col_name in DATE_COLS:
                for row in range(2, ws.max_row + 1):
                    ws[f'{col_letter}{row}'].number_format = 'YYYY-MM-DD'

        # Score column: conditional fill
        if 'Lead Score' in col_map:
            score_idx = col_map['Lead Score']
            score_letter = get_column_letter(score_idx)
            for row in range(2, ws.max_row + 1):
                cell = ws[f'{score_letter}{row}']
                try:
                    val = float(cell.value) if cell.value is not None else 0
                except (ValueError, TypeError):
                    val = 0
                if val >= 60:
                    cell.fill = GREEN_FILL
                elif val >= 40:
                    cell.fill = YELLOW_FILL

        # Auto-fit column widths (based on header, with min/max)
        for idx, cell in enumerate(ws[1], 1):
            header_len = len(str(cell.value)) if cell.value else 8
            width = max(10, min(header_len + 4, 35))
            ws.column_dimensions[get_column_letter(idx)].width = width

    # --- Build workbook ---
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:

        # Segment tabs sorted by average score descending
        segment_avgs = df.groupby('icp_primary')['score'].apply(
            lambda s: s.astype(float).mean()
        ).sort_values(ascending=False)

        segment_sheets = []
        for segment in segment_avgs.index:
            sheet_name = segment.replace('/', '-')[:31]
            segment_df = df_sorted[df_sorted['icp_primary'] == segment]
            prepared = _prepare_sheet_df(segment_df)
            prepared.to_excel(writer, sheet_name=sheet_name, index=False)
            segment_sheets.append(sheet_name)

        wb = writer.book

        # Build Summary dashboard as first tab
        ws_summary = wb.create_sheet('Summary', 0)
        _build_summary_dashboard(ws_summary, df)

        # Format segment sheets
        for sheet_name in segment_sheets:
            _format_sheet(wb[sheet_name])

    print(f"Excel workbook created: {output_file}")
    print(f"  Tabs: Summary + {len(segment_sheets)} segment tabs (no 'All Leads' tab)")


def main():
    parser = argparse.ArgumentParser(description='Score and output leads to Excel')
    parser.add_argument('--input', type=str, default='pipeline/output/05_enriched.csv')
    parser.add_argument('--edgar-input', type=str, default='pipeline/output/04_fund_managers.csv')
    parser.add_argument('--output', type=str,
                        default=f'pipeline/output/leads_{date.today().isoformat()}.xlsx')

    args = parser.parse_args()
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Load enriched leads
    print("Loading enriched leads...")
    df = pd.read_csv(args.input, dtype=str, low_memory=False)
    print(f"Loaded {len(df):,} leads")

    # Merge EDGAR data
    print("\nMerging SEC EDGAR fund data...")
    df = merge_edgar_data(df, args.edgar_input)

    # Map CO_NO to county name if needed
    if 'county' not in df.columns and 'CO_NO' in df.columns:
        df['county'] = df['CO_NO'].map(CO_NO_TO_COUNTY).fillna('')
        print(f"Mapped {(df['county'] != '').sum():,} leads to county names")

    # Classify ICP
    print("\nClassifying ICP segments...")
    classifications = df.apply(classify_icp, axis=1)
    df['icp_primary'] = [c[0] for c in classifications]
    df['icp_secondary'] = [c[1] for c in classifications]
    df['tier'] = [c[2] for c in classifications]

    # Score leads
    print("Scoring leads...")
    df['score'] = df.apply(score_lead, axis=1)
    df['reachability'] = df.apply(reachability_score, axis=1)

    # Generate lead IDs
    df['lead_id'] = [f"DSCR-{i:06d}" for i in range(1, len(df) + 1)]

    # Determine owner type
    def get_owner_type(row):
        if str(row.get('foreign_owner', '')).lower() in ('true', '1', 'yes'):
            return 'Foreign'
        owner = str(row.get('owner_name', '') or row.get(
            [c for c in df.columns if 'OWN' in c.upper() and 'NAME' in c.upper()][0]
            if any('OWN' in c.upper() and 'NAME' in c.upper() for c in df.columns) else '', ''))
        if any(kw in owner.upper() for kw in [' LLC', ' L.L.C', ' INC', ' CORP']):
            return 'LLC'
        if 'TRUST' in owner.upper():
            return 'Trust'
        if any(kw in owner.upper() for kw in [' LP', ' LTD', ' PARTNERS']):
            return 'Partnership'
        return 'Individual'

    df['owner_type'] = df.apply(get_owner_type, axis=1)

    # Add metadata
    df['data_sources'] = 'FDOR'
    if 'str_licensed' in df.columns:
        df.loc[df['str_licensed'].fillna(False).astype(str).str.lower().isin(['true', '1', 'yes']), 'data_sources'] += ',DBPR'
    if 'resolved_person' in df.columns:
        df.loc[df['resolved_person'].fillna('') != '', 'data_sources'] += ',SunBiz'
    if 'sec_fund_filing' in df.columns:
        df.loc[df['sec_fund_filing'].fillna(False).astype(str).str.lower().isin(['true', '1', 'yes']), 'data_sources'] += ',EDGAR'
    if 'enrichment_source' in df.columns:
        df.loc[df['enrichment_source'].fillna('') != '', 'data_sources'] += ',Enrichment'
    df['last_updated'] = date.today().isoformat()

    # Standardize column names for output
    # Prefer OWN_STATE_DOM (2-letter code) over OWN_STATE (full name) for mailing_state
    rename_map = {}
    if 'OWN_STATE_DOM' in df.columns and 'mailing_state' not in df.columns:
        rename_map['OWN_STATE_DOM'] = 'mailing_state'
    for col in df.columns:
        if col in rename_map:
            continue
        if 'OWN' in col.upper() and 'NAME' in col.upper() and col != 'owner_name':
            rename_map[col] = 'owner_name'
        elif ('OWN' in col.upper() or 'MAIL' in col.upper()) and 'ADDR' in col.upper() and 'mailing' not in col.lower():
            rename_map[col] = 'mailing_address'
        elif ('OWN' in col.upper() or 'MAIL' in col.upper()) and 'CITY' in col.upper() and 'mailing' not in col.lower():
            rename_map[col] = 'mailing_city'
        elif ('OWN' in col.upper() or 'MAIL' in col.upper()) and 'STATE' in col.upper() and 'mailing' not in col.lower() and 'mailing_state' not in rename_map.values():
            rename_map[col] = 'mailing_state'
        elif ('OWN' in col.upper() or 'MAIL' in col.upper()) and 'ZIP' in col.upper() and 'mailing' not in col.lower():
            rename_map[col] = 'mailing_zip'

    df = df.rename(columns=rename_map)

    # Print summary
    print(f"\n{'='*60}")
    print(f"FINAL LEAD SUMMARY")
    print(f"{'='*60}")
    print(f"Total leads: {len(df):,}")
    print(f"\nBy ICP Segment:")
    for seg in sorted(df['icp_primary'].unique()):
        count = len(df[df['icp_primary'] == seg])
        print(f"  {seg}: {count:,}")
    print(f"\nBy Tier:")
    for tier in sorted(df['tier'].unique()):
        count = len(df[df['tier'] == tier])
        print(f"  Tier {tier}: {count:,}")
    print(f"\nScore Distribution:")
    print(f"  Mean: {df['score'].astype(float).mean():.1f}")
    print(f"  Median: {df['score'].astype(float).median():.1f}")
    print(f"  Hot (60+):   {(df['score'].astype(float) >= 60).sum():,}")
    print(f"  Warm (40-59): {((df['score'].astype(float) >= 40) & (df['score'].astype(float) < 60)).sum():,}")
    print(f"  Cold (<40):  {(df['score'].astype(float) < 40).sum():,}")
    print(f"\nReachability:")
    reach = df['reachability'].astype(int)
    print(f"  Phone+Email (10): {(reach == 10).sum():,}")
    print(f"  Phone only (7):   {(reach == 7).sum():,}")
    print(f"  Email only (5):   {(reach == 5).sum():,}")
    print(f"  Address only (2): {(reach == 2).sum():,}")

    # Create Excel output
    create_excel_output(df, args.output)


if __name__ == '__main__':
    main()
