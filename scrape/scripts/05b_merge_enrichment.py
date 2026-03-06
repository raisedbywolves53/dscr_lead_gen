"""
Step 5b: Merge Enrichment Results
==================================

After you've done manual research (research_tracker.xlsx) and/or
uploaded to Datazapp, this script merges everything into a single
enriched file ready for validation.

Input sources (uses whatever exists):
  1. data/enriched/top_leads_enriched.csv (from script 05)
  2. data/enriched/research_tracker.xlsx (your manual lookups — yellow columns)
  3. data/enriched/tracerfy_results.csv (Tracerfy skip trace — primary source)
  4. data/enriched/tracerfy_dnc_results.csv (Tracerfy DNC scrub results)
  5. data/enriched/datazapp_results.csv (Datazapp skip trace output)
  6. data/enriched/wiza_results.csv (Wiza LinkedIn lookups)
  7. data/enriched/apollo_results.csv (Apollo.io API — already merged in step 10)

Output:
  data/enriched/merged_enriched.csv

Usage:
    python scripts/05b_merge_enrichment.py
"""

import re
from pathlib import Path

import pandas as pd

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
ENRICHED_DIR = PROJECT_DIR / "data" / "enriched"

LEADS_CSV = ENRICHED_DIR / "top_leads_enriched.csv"
TRACKER_XLSX = ENRICHED_DIR / "research_tracker.xlsx"
TRACERFY_CSV = ENRICHED_DIR / "tracerfy_results.csv"
TRACERFY_DNC_CSV = ENRICHED_DIR / "tracerfy_dnc_results.csv"
DATAZAPP_CSV = ENRICHED_DIR / "datazapp_results.csv"
WIZA_CSV = ENRICHED_DIR / "wiza_results.csv"
APOLLO_CSV = ENRICHED_DIR / "apollo_results.csv"
OUTPUT_CSV = ENRICHED_DIR / "merged_enriched.csv"


def normalize_phone(phone) -> str:
    """Normalize phone to 10 digits."""
    phone = re.sub(r"\D", "", str(phone))
    if len(phone) == 11 and phone.startswith("1"):
        phone = phone[1:]
    return phone if len(phone) == 10 else ""


def load_tracker_results() -> dict:
    """
    Load manual research results from the research tracker Excel.
    Returns dict: owner_name → {phone, email, source, notes}
    """
    if not TRACKER_XLSX.exists() or not HAS_OPENPYXL:
        return {}

    print("  Loading research_tracker.xlsx...")
    wb = load_workbook(TRACKER_XLSX, data_only=True)
    ws = wb.active

    results = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        # Columns: Rank, Score, LLC/Owner, Resolved Person, Properties,
        #          Portfolio Value, Mailing Address, Existing Phone, Existing Email,
        #          TPS Link, FPS Link, LinkedIn, FOUND:Phone, FOUND:Email, FOUND:Source, Notes
        if len(row) < 16:
            continue

        owner = str(row[2].value or "").strip()
        found_phone = normalize_phone(row[12].value or "")
        found_email = str(row[13].value or "").strip()
        found_source = str(row[14].value or "").strip()
        notes = str(row[15].value or "").strip()

        if owner and (found_phone or found_email):
            results[owner.upper()] = {
                "phone": found_phone,
                "email": found_email,
                "source": found_source or "manual_research",
                "notes": notes,
            }

    print(f"  Found {len(results)} manual research results.")
    return results


def load_tracerfy_results() -> dict:
    """
    Load Tracerfy skip trace results.
    Returns dict: (first+last+zip) → {phone, phone_type, email, all_phones, all_emails}

    Tracerfy output (normalized by script 08) has columns:
    tracerfy_first, tracerfy_last, tracerfy_zip, tracerfy_phone_1,
    tracerfy_phone_1_type, tracerfy_email_1, tracerfy_match, etc.
    """
    if not TRACERFY_CSV.exists():
        return {}

    print("  Loading tracerfy_results.csv...")
    tf = pd.read_csv(TRACERFY_CSV, dtype=str)

    results = {}
    for _, row in tf.iterrows():
        if str(row.get("tracerfy_match", "")).lower() != "yes":
            continue

        first = str(row.get("tracerfy_first", "")).strip().upper()
        last = str(row.get("tracerfy_last", "")).strip().upper()
        zipcode = str(row.get("tracerfy_zip", "")).strip()[:5]

        key = f"{first}|{last}|{zipcode}"

        phone = normalize_phone(row.get("tracerfy_phone_1", ""))
        phone2 = normalize_phone(row.get("tracerfy_phone_2", ""))
        phone_type = str(row.get("tracerfy_phone_1_type", "")).strip()
        phone2_type = str(row.get("tracerfy_phone_2_type", "")).strip()
        email = str(row.get("tracerfy_email_1", "")).strip()
        email2 = str(row.get("tracerfy_email_2", "")).strip()

        if email.upper() in ("NAN", "NONE", ""):
            email = ""
        if email2.upper() in ("NAN", "NONE", ""):
            email2 = ""
        if phone_type.upper() in ("NAN", "NONE", ""):
            phone_type = ""
        if phone2_type.upper() in ("NAN", "NONE", ""):
            phone2_type = ""

        if phone or email:
            results[key] = {
                "phone": phone,
                "phone_type": phone_type,
                "phone2": phone2,
                "phone2_type": phone2_type,
                "email": email,
                "email2": email2,
                "source": "tracerfy",
            }

    print(f"  Found {len(results)} Tracerfy matches.")
    return results


def load_tracerfy_dnc() -> set:
    """
    Load Tracerfy DNC scrub results.
    Returns set of phone numbers flagged as DNC/litigator.
    """
    if not TRACERFY_DNC_CSV.exists():
        return set()

    print("  Loading tracerfy_dnc_results.csv...")
    dnc = pd.read_csv(TRACERFY_DNC_CSV, dtype=str)

    flagged = set()
    for _, row in dnc.iterrows():
        is_clean = str(row.get("is_clean", "")).lower()
        if is_clean not in ("true", "1", "yes"):
            phone = normalize_phone(row.get("phone", ""))
            if phone:
                flagged.add(phone)

    print(f"  Found {len(flagged)} DNC-flagged phones.")
    return flagged


def load_datazapp_results() -> dict:
    """
    Load Datazapp skip trace results.
    Returns dict: (first+last+zip) → {phone, email}

    Datazapp output columns vary but typically include:
    First Name, Last Name, Address, City, State, Zip,
    Cell Phone, Landline, Email
    """
    if not DATAZAPP_CSV.exists():
        return {}

    print("  Loading datazapp_results.csv...")
    dz = pd.read_csv(DATAZAPP_CSV, dtype=str)

    # Normalize column names (Datazapp output varies)
    col_map = {}
    for col in dz.columns:
        cl = col.lower().strip()
        if "first" in cl and "name" in cl:
            col_map[col] = "first_name"
        elif "last" in cl and "name" in cl:
            col_map[col] = "last_name"
        elif "cell" in cl or ("phone" in cl and "land" not in cl):
            col_map[col] = "cell_phone"
        elif "land" in cl:
            col_map[col] = "landline"
        elif "email" in cl:
            col_map[col] = "email"
        elif "zip" in cl:
            col_map[col] = "zip"
    dz = dz.rename(columns=col_map)

    results = {}
    for _, row in dz.iterrows():
        first = str(row.get("first_name", "")).strip().upper()
        last = str(row.get("last_name", "")).strip().upper()
        zipcode = str(row.get("zip", "")).strip()[:5]

        key = f"{first}|{last}|{zipcode}"

        phone = normalize_phone(row.get("cell_phone", ""))
        if not phone:
            phone = normalize_phone(row.get("landline", ""))
        email = str(row.get("email", "")).strip()
        if email.upper() in ("NAN", "NONE", ""):
            email = ""

        if phone or email:
            results[key] = {
                "phone": phone,
                "phone_type": "mobile" if normalize_phone(row.get("cell_phone", "")) else "landline",
                "email": email,
                "source": "datazapp",
            }

    print(f"  Found {len(results)} Datazapp matches.")
    return results


def load_wiza_results() -> dict:
    """
    Load Wiza LinkedIn lookup results.
    Returns dict: owner_name_upper → {phone, email, linkedin, source}

    Wiza CSV export typically has columns:
    First Name, Last Name, Email, Phone, LinkedIn URL, Company, Title
    """
    if not WIZA_CSV.exists():
        return {}

    print("  Loading wiza_results.csv...")
    wz = pd.read_csv(WIZA_CSV, dtype=str)

    # Normalize column names
    col_map = {}
    for col in wz.columns:
        cl = col.lower().strip()
        if "first" in cl and "name" in cl:
            col_map[col] = "first_name"
        elif "last" in cl and "name" in cl:
            col_map[col] = "last_name"
        elif cl == "email" or "email" in cl:
            col_map[col] = "email"
        elif "phone" in cl or "mobile" in cl:
            col_map[col] = "phone"
        elif "linkedin" in cl:
            col_map[col] = "linkedin"
    wz = wz.rename(columns=col_map)

    results = {}
    for _, row in wz.iterrows():
        first = str(row.get("first_name", "")).strip().upper()
        last = str(row.get("last_name", "")).strip().upper()
        if not first or not last:
            continue

        # Key by "LAST, FIRST" to match OWN_NAME / resolved_person format
        key = f"{last}, {first}"

        phone = normalize_phone(row.get("phone", ""))
        email = str(row.get("email", "")).strip()
        linkedin = str(row.get("linkedin", "")).strip()

        if email.upper() in ("NAN", "NONE", ""):
            email = ""
        if linkedin.upper() in ("NAN", "NONE", ""):
            linkedin = ""

        if phone or email:
            results[key] = {
                "phone": phone,
                "email": email,
                "linkedin": linkedin,
                "source": "wiza",
            }

    print(f"  Found {len(results)} Wiza matches.")
    return results


def load_apollo_contacts() -> dict:
    """
    Load Apollo results that have actual contact data (email/phone).
    Apollo results are already in the base CSV from script 10, but if
    apollo_results.csv has data that wasn't merged, pick it up here.
    Returns dict: owner_name_upper → {phone, email, linkedin}
    """
    if not APOLLO_CSV.exists():
        return {}

    print("  Loading apollo_results.csv...")
    ap = pd.read_csv(APOLLO_CSV, dtype=str)

    results = {}
    for _, row in ap.iterrows():
        owner = str(row.get("OWN_NAME", "")).strip().upper()
        if not owner:
            continue

        email = str(row.get("apollo_email", "")).strip()
        phone = normalize_phone(row.get("apollo_mobile", "") or row.get("apollo_phone", ""))
        linkedin = str(row.get("apollo_linkedin", "")).strip()

        if email.upper() in ("NAN", "NONE", ""):
            email = ""
        if linkedin.upper() in ("NAN", "NONE", ""):
            linkedin = ""

        if phone or email:
            results[owner] = {
                "phone": phone,
                "email": email,
                "linkedin": linkedin,
                "source": "apollo",
            }

    matched = len(results)
    print(f"  Found {matched} Apollo contacts with data.")
    return results


def main():
    if not LEADS_CSV.exists():
        print(f"\nBase enriched file not found: {LEADS_CSV}")
        print("Run script 05 first: python scripts/05_enrich_contacts.py --limit 25")
        return

    print("\n  Merging enrichment sources...\n")

    # Load base data
    df = pd.read_csv(LEADS_CSV, dtype=str)
    print(f"  Base records: {len(df):,}")

    # Ensure contact columns exist
    for col in ["phone_1", "phone_1_source", "phone_1_type",
                 "phone_2", "phone_2_source", "phone_2_type",
                 "email_1", "email_1_source", "email_2", "email_2_source",
                 "enrichment_sources"]:
        if col not in df.columns:
            df[col] = ""

    # Pre-fill from existing phone/email columns
    for idx, row in df.iterrows():
        existing_phone = str(row.get("phone", "")).strip()
        if existing_phone and existing_phone.upper() not in ("NAN", "NONE", ""):
            df.at[idx, "phone_1"] = normalize_phone(existing_phone)
            df.at[idx, "phone_1_source"] = str(row.get("enrichment_source", "existing"))

        existing_email = str(row.get("email", "")).strip()
        if existing_email and existing_email.upper() not in ("NAN", "NONE", "") and "@" in existing_email:
            df.at[idx, "email_1"] = existing_email
            df.at[idx, "email_1_source"] = str(row.get("enrichment_source", "existing"))

    # Merge manual research results
    tracker_results = load_tracker_results()
    tracker_merged = 0
    for idx, row in df.iterrows():
        owner = str(row.get("OWN_NAME", "")).strip().upper()
        if owner in tracker_results:
            res = tracker_results[owner]
            if res["phone"] and not df.at[idx, "phone_1"]:
                df.at[idx, "phone_1"] = res["phone"]
                df.at[idx, "phone_1_source"] = res["source"]
            elif res["phone"]:
                df.at[idx, "phone_2"] = res["phone"]
                df.at[idx, "phone_2_source"] = res["source"]

            if res["email"] and not df.at[idx, "email_1"]:
                df.at[idx, "email_1"] = res["email"]
                df.at[idx, "email_1_source"] = res["source"]
            elif res["email"]:
                df.at[idx, "email_2"] = res["email"]
                df.at[idx, "email_2_source"] = res["source"]

            tracker_merged += 1

    if tracker_merged:
        print(f"  Merged {tracker_merged} manual research results.")

    # Merge Tracerfy results (primary skip trace source)
    tracerfy_results = load_tracerfy_results()
    tracerfy_dnc = load_tracerfy_dnc()
    tf_merged = 0
    if tracerfy_results:
        for idx, row in df.iterrows():
            # Build lookup key from resolved person or owner name
            resolved = str(row.get("resolved_person", "")).strip()
            owner = str(row.get("OWN_NAME", "")).strip()
            zipcode = str(row.get("OWN_ZIPCD", "")).strip()[:5]

            person_name = resolved if resolved.upper() not in ("", "NAN", "NONE") else owner
            # Strip trailing "&" (joint ownership marker)
            person_name = person_name.rstrip("& ").strip()
            if "," in person_name:
                # "LAST, FIRST" format
                parts = person_name.split(",", 1)
                last = parts[0].strip().upper()
                first = parts[1].strip().split()[0].upper() if len(parts) > 1 and parts[1].strip() else ""
            else:
                # FDOR format: "LAST FIRST" or "LAST FIRST MIDDLE"
                parts = person_name.split()
                if len(parts) >= 2:
                    last = parts[0].upper()
                    first = parts[1].upper()
                elif len(parts) == 1:
                    last = parts[0].upper()
                    first = ""
                else:
                    last = ""
                    first = ""

            key = f"{first}|{last}|{zipcode}"
            if key in tracerfy_results:
                res = tracerfy_results[key]

                # Phone 1
                if res["phone"] and not df.at[idx, "phone_1"]:
                    df.at[idx, "phone_1"] = res["phone"]
                    df.at[idx, "phone_1_source"] = "tracerfy"
                    df.at[idx, "phone_1_type"] = res.get("phone_type", "")
                elif res["phone"]:
                    if not df.at[idx, "phone_2"]:
                        df.at[idx, "phone_2"] = res["phone"]
                        df.at[idx, "phone_2_source"] = "tracerfy"
                        df.at[idx, "phone_2_type"] = res.get("phone_type", "")

                # Phone 2 from Tracerfy
                if res.get("phone2") and not df.at[idx, "phone_2"]:
                    df.at[idx, "phone_2"] = res["phone2"]
                    df.at[idx, "phone_2_source"] = "tracerfy"
                    df.at[idx, "phone_2_type"] = res.get("phone2_type", "")

                # Email 1
                if res["email"] and not df.at[idx, "email_1"]:
                    df.at[idx, "email_1"] = res["email"]
                    df.at[idx, "email_1_source"] = "tracerfy"
                elif res["email"]:
                    if not df.at[idx, "email_2"]:
                        df.at[idx, "email_2"] = res["email"]
                        df.at[idx, "email_2_source"] = "tracerfy"

                # Email 2 from Tracerfy
                if res.get("email2") and not df.at[idx, "email_2"]:
                    df.at[idx, "email_2"] = res["email2"]
                    df.at[idx, "email_2_source"] = "tracerfy"

                tf_merged += 1

        print(f"  Merged {tf_merged} Tracerfy results.")

    # Apply Tracerfy DNC flags
    dnc_flagged_count = 0
    if tracerfy_dnc:
        if "phone_1_dnc" not in df.columns:
            df["phone_1_dnc"] = ""
        if "phone_2_dnc" not in df.columns:
            df["phone_2_dnc"] = ""

        for idx, row in df.iterrows():
            p1 = normalize_phone(row.get("phone_1", ""))
            p2 = normalize_phone(row.get("phone_2", ""))
            if p1 and p1 in tracerfy_dnc:
                df.at[idx, "phone_1_dnc"] = "True"
                dnc_flagged_count += 1
            if p2 and p2 in tracerfy_dnc:
                df.at[idx, "phone_2_dnc"] = "True"

        if dnc_flagged_count:
            print(f"  DNC flagged: {dnc_flagged_count} phone numbers.")

    # Merge Datazapp results
    datazapp_results = load_datazapp_results()
    dz_merged = 0
    if datazapp_results:
        for idx, row in df.iterrows():
            # Build lookup key from resolved person or owner name
            resolved = str(row.get("resolved_person", "")).strip()
            owner = str(row.get("OWN_NAME", "")).strip()
            zipcode = str(row.get("OWN_ZIPCD", "")).strip()[:5]

            # Try resolved person first
            person_name = resolved if resolved.upper() not in ("", "NAN", "NONE") else owner
            person_name = person_name.rstrip("& ").strip()
            if "," in person_name:
                parts = person_name.split(",", 1)
                last = parts[0].strip().upper()
                first = parts[1].strip().split()[0].upper() if len(parts) > 1 else ""
            else:
                # FDOR format: "LAST FIRST"
                parts = person_name.split()
                if len(parts) >= 2:
                    last = parts[0].upper()
                    first = parts[1].upper()
                elif len(parts) == 1:
                    last = parts[0].upper()
                    first = ""
                else:
                    last = ""
                    first = ""

            key = f"{first}|{last}|{zipcode}"
            if key in datazapp_results:
                res = datazapp_results[key]
                if res["phone"] and not df.at[idx, "phone_1"]:
                    df.at[idx, "phone_1"] = res["phone"]
                    df.at[idx, "phone_1_source"] = "datazapp"
                    df.at[idx, "phone_1_type"] = res.get("phone_type", "")
                elif res["phone"]:
                    df.at[idx, "phone_2"] = res["phone"]
                    df.at[idx, "phone_2_source"] = "datazapp"
                    df.at[idx, "phone_2_type"] = res.get("phone_type", "")

                if res["email"] and not df.at[idx, "email_1"]:
                    df.at[idx, "email_1"] = res["email"]
                    df.at[idx, "email_1_source"] = "datazapp"
                elif res["email"]:
                    df.at[idx, "email_2"] = res["email"]
                    df.at[idx, "email_2_source"] = "datazapp"

                dz_merged += 1

        print(f"  Merged {dz_merged} Datazapp results.")

    # Merge Wiza results
    wiza_results = load_wiza_results()
    wiza_merged = 0
    if wiza_results:
        for idx, row in df.iterrows():
            # Try matching by resolved person name or owner name
            resolved = str(row.get("resolved_person", "")).strip().upper()
            owner = str(row.get("OWN_NAME", "")).strip().upper()

            res = None
            if resolved and resolved not in ("NAN", "NONE", ""):
                res = wiza_results.get(resolved)
            if not res and owner:
                res = wiza_results.get(owner)

            if res:
                if res["phone"] and not df.at[idx, "phone_1"]:
                    df.at[idx, "phone_1"] = res["phone"]
                    df.at[idx, "phone_1_source"] = "wiza"
                    df.at[idx, "phone_1_type"] = "mobile"
                elif res["phone"]:
                    df.at[idx, "phone_2"] = res["phone"]
                    df.at[idx, "phone_2_source"] = "wiza"
                    df.at[idx, "phone_2_type"] = "mobile"

                if res["email"] and not df.at[idx, "email_1"]:
                    df.at[idx, "email_1"] = res["email"]
                    df.at[idx, "email_1_source"] = "wiza"
                elif res["email"]:
                    df.at[idx, "email_2"] = res["email"]
                    df.at[idx, "email_2_source"] = "wiza"

                # Store LinkedIn if we got one
                if res.get("linkedin") and "apollo_linkedin" in df.columns:
                    if not str(df.at[idx, "apollo_linkedin"]).strip() or str(df.at[idx, "apollo_linkedin"]).upper() in ("NAN", "NONE", ""):
                        df.at[idx, "apollo_linkedin"] = res["linkedin"]

                wiza_merged += 1

        print(f"  Merged {wiza_merged} Wiza results.")

    # Merge Apollo contacts (for any that have actual data)
    apollo_results = load_apollo_contacts()
    apollo_merged = 0
    if apollo_results:
        for idx, row in df.iterrows():
            owner = str(row.get("OWN_NAME", "")).strip().upper()
            if owner not in apollo_results:
                continue

            res = apollo_results[owner]
            if res["phone"] and not df.at[idx, "phone_1"]:
                df.at[idx, "phone_1"] = res["phone"]
                df.at[idx, "phone_1_source"] = "apollo"
            elif res["phone"] and not df.at[idx, "phone_2"]:
                df.at[idx, "phone_2"] = res["phone"]
                df.at[idx, "phone_2_source"] = "apollo"

            if res["email"] and not df.at[idx, "email_1"]:
                df.at[idx, "email_1"] = res["email"]
                df.at[idx, "email_1_source"] = "apollo"
            elif res["email"] and not df.at[idx, "email_2"]:
                df.at[idx, "email_2"] = res["email"]
                df.at[idx, "email_2_source"] = "apollo"

            apollo_merged += 1

        print(f"  Merged {apollo_merged} Apollo results.")

    # Build enrichment_sources summary
    for idx, row in df.iterrows():
        sources = set()
        for src_col in ["phone_1_source", "phone_2_source", "email_1_source", "email_2_source"]:
            val = str(row.get(src_col, "")).strip()
            if val and val.upper() not in ("NAN", "NONE", ""):
                sources.add(val)
        df.at[idx, "enrichment_sources"] = ", ".join(sorted(sources))

    # Save
    df.to_csv(OUTPUT_CSV, index=False)

    # Summary
    has_phone = df["phone_1"].astype(str).str.len().ge(10).sum()
    has_email = df["email_1"].astype(str).str.contains("@", na=False).sum()
    total = len(df)

    print()
    print("=" * 60)
    print("  MERGE SUMMARY")
    print("=" * 60)
    print(f"  Total records:       {total}")
    print(f"  Has phone:           {has_phone}  ({has_phone/total*100:.0f}%)")
    print(f"  Has email:           {has_email}  ({has_email/total*100:.0f}%)")
    print(f"  From Tracerfy:       {tf_merged}")
    print(f"  From manual:         {tracker_merged}")
    print(f"  From Datazapp:       {dz_merged}")
    print(f"  From Wiza:           {wiza_merged}")
    print(f"  From Apollo:         {apollo_merged}")
    if tracerfy_dnc:
        print(f"  DNC flagged:         {dnc_flagged_count}")
    print(f"\n  Saved: {OUTPUT_CSV}")
    print()
    print("  NEXT STEPS:")
    print(f"  python scripts/06_validate_contacts.py --county merged")
    print(f"  python scripts/07_export_campaign_ready.py --county merged")
    print()


if __name__ == "__main__":
    main()
