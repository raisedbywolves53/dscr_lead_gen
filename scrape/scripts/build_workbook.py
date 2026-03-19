"""
DSCR Unified Excel Workbook — Pipeline + Dashboard + Investor Detail Tabs
=========================================================================

Consolidates PDF dossiers, call sheet, and tracking into one Excel workbook.

Tab Structure:
  1. PIPELINE — Daily driver call sheet with conditional formatting + data bars
  2-N. Per-investor detail tabs — 3-act dossier layout (replaces PDFs)

Usage:
    python scripts/build_workbook.py --input data/enriched/fl_client_samples.csv
    python scripts/build_workbook.py --input data/enriched/fl_client_samples.csv --redacted
    python scripts/build_workbook.py --input data/enriched/fl_client_samples.csv --output data/workbooks/custom.xlsx
"""

import argparse
import math
import re
import sys
from pathlib import Path

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import (
    CellIsRule, FormulaRule, DataBarRule,
)

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent

# ── Color Palette (openpyxl hex, no '#') ─────────────────────────────
NAVY = "0A2342"
TEAL = "1A6B6A"
ACCENT = "0066B3"
GREEN = "228B22"
AMBER = "CC8800"
RED = "BF3131"
ALT_ROW = "F8F9FB"
LIGHT_BG = "EDF1F7"
LIGHT_BLUE = "DCEAFC"
WHITE = "FFFFFF"
BLACK = "000000"
LIGHT_GREEN = "C8E6C9"
LIGHT_YELLOW = "FFF9C4"
LIGHT_RED = "FFCDD2"
MED_GRAY = "999999"
DARK_GRAY = "595959"

# ── Fills ─────────────────────────────────────────────────────────────
navy_fill = PatternFill("solid", fgColor=NAVY)
teal_fill = PatternFill("solid", fgColor=TEAL)
alt_fill = PatternFill("solid", fgColor=ALT_ROW)
light_bg_fill = PatternFill("solid", fgColor=LIGHT_BG)
light_blue_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
green_fill = PatternFill("solid", fgColor=GREEN)
amber_fill = PatternFill("solid", fgColor=AMBER)
gray_fill = PatternFill("solid", fgColor="D0D0D0")
white_fill = PatternFill("solid", fgColor=WHITE)
cond_green = PatternFill("solid", fgColor=LIGHT_GREEN)
cond_yellow = PatternFill("solid", fgColor=LIGHT_YELLOW)
cond_red = PatternFill("solid", fgColor=LIGHT_RED)
cond_blue = PatternFill("solid", fgColor=LIGHT_BLUE)

# ── Fonts ─────────────────────────────────────────────────────────────
white_bold_14 = Font(bold=True, color=WHITE, size=14, name="Calibri")
white_bold_11 = Font(bold=True, color=WHITE, size=11, name="Calibri")
white_bold_10 = Font(bold=True, color=WHITE, size=10, name="Calibri")
navy_bold_20 = Font(bold=True, color=NAVY, size=20, name="Calibri")
navy_bold_16 = Font(bold=True, color=WHITE, size=16, name="Calibri")
navy_bold_12 = Font(bold=True, color=NAVY, size=12, name="Calibri")
navy_bold_10 = Font(bold=True, color=NAVY, size=10, name="Calibri")
navy_bold_9 = Font(bold=True, color=NAVY, size=9, name="Calibri")
teal_10 = Font(color=TEAL, size=10, name="Calibri")
teal_bold_10 = Font(bold=True, color=TEAL, size=10, name="Calibri")
dark_gray_7 = Font(color=DARK_GRAY, size=7, name="Calibri")
dark_gray_8 = Font(color=DARK_GRAY, size=8, name="Calibri")
normal_10 = Font(size=10, name="Calibri")
normal_9 = Font(size=9, name="Calibri")
bold_10 = Font(bold=True, size=10, name="Calibri")
bold_9 = Font(bold=True, size=9, name="Calibri")
green_bold_10 = Font(bold=True, color=GREEN, size=10, name="Calibri")
red_bold_10 = Font(bold=True, color=RED, size=10, name="Calibri")

# ── Alignments ────────────────────────────────────────────────────────
wrap_top = Alignment(wrap_text=True, vertical="top")
wrap_center = Alignment(wrap_text=True, vertical="center", horizontal="center")
center_center = Alignment(vertical="center", horizontal="center")
left_center = Alignment(vertical="center", horizontal="left")
right_center = Alignment(vertical="center", horizontal="right")

# ── Borders ───────────────────────────────────────────────────────────
thin_border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# ── FL County Codes ───────────────────────────────────────────────────
FL_COUNTY_MAP = {
    "1": "Alachua", "5": "Brevard", "6": "Broward", "13": "Dade",
    "16": "Duval", "29": "Hillsborough", "48": "Orange",
    "50": "Palm Beach", "52": "Pinellas", "60": "Palm Beach",
}

NC_COUNTY_MAP = {
    "wake": "Wake", "mecklenburg": "Mecklenburg", "durham": "Durham",
    "guilford": "Guilford", "forsyth": "Forsyth",
}


# ════════════════════════════════════════════════════════════════════
# DATA HELPERS (ported from build_dossier_pdf.py / build_google_sheets.py)
# ════════════════════════════════════════════════════════════════════

def safe_float(val, default=None):
    try:
        v = float(val)
        return v if not np.isnan(v) else default
    except (ValueError, TypeError):
        return default


def safe_int(val, default=0):
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default


def s(val):
    """Safe string — NaN/None to empty string."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return ""
    v = str(val)
    return "" if v == "nan" else v


def fv(val):
    """Format value — return None if empty/nan."""
    v = str(val).strip()
    if not v or v.upper() in ("NAN", "NONE", "", "0", "0.0", "N/A", "--"):
        return None
    return v


def fc(val):
    """Format currency — return display string or None."""
    try:
        v = str(val).replace(",", "").replace("$", "").strip()
        if not v or v.upper() in ("NAN", "NONE", ""):
            return None
        f = float(v)
        if f == 0 or math.isnan(f):
            return None
        if f >= 1_000_000:
            return f"${f/1_000_000:.1f}M"
        return f"${f:,.0f}"
    except (ValueError, TypeError):
        return None


def fphone(val):
    digits = re.sub(r"\D", "", str(val))
    if len(digits) == 11 and digits[0] == "1":
        digits = digits[1:]
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    return s(val) if fv(val) else ""


def clean_name(val):
    v = str(val).strip().rstrip("& ").rstrip(",").strip()
    return v if v and v.upper() not in ("NAN", "NONE") else None


def display_name(raw_name):
    if not raw_name:
        return "Unknown"
    v = str(raw_name).strip().rstrip("& ").rstrip(",").strip()
    if not v or v.upper() in ("NAN", "NONE"):
        return "Unknown"
    entity_markers = [
        "LLC", "L.L.C", "INC", "CORP", "TRUST", "LTD", "LP", "L.P.",
        "REALTY", "PROPERTIES", "INVESTMENTS", "HOLDINGS", "PARTNERS",
        "GROUP", "VENTURES", "CAPITAL", "MANAGEMENT", "ASSOCIATES",
        "ENTERPRISES", "DEVELOPMENT", "FOUNDATION",
    ]
    upper = v.upper()
    if any(m in upper for m in entity_markers):
        result = v.title()
        for m in ["Llc", "L.L.C", "Inc", "Corp", "Ltd", "Lp", "L.P."]:
            result = result.replace(m, m.upper())
        return result
    parts = v.split()
    if not parts:
        return "Unknown"
    if len(parts) == 1:
        return parts[0].title()
    last = parts[0].title()
    first_middle = " ".join(parts[1:]).title()
    fm_parts = first_middle.split()
    if len(fm_parts) > 1:
        result = fm_parts[0]
        for p in fm_parts[1:]:
            result += f" {p}." if len(p) == 1 else f" {p}"
        first_middle = result
    return f"{first_middle} {last}" if first_middle else last


def get_county_state(row):
    state = fv(row.get("OWN_STATE_DOM", "")) or fv(row.get("OWN_STATE", ""))
    co_no = str(row.get("CO_NO", "")).strip()
    county = None
    if state and state.upper() == "FL" and co_no:
        county = FL_COUNTY_MAP.get(co_no)
    if not county:
        county_raw = fv(row.get("county", "")) or fv(row.get("county_name", ""))
        if county_raw:
            county = county_raw.title()
    if county and state:
        return f"{county} County, {state.upper()}"
    elif state:
        return state.upper()
    return ""


def generate_financing_angle(row):
    rate = safe_float(row.get("attom_interest_rate"))
    cash_buyer = s(row.get("probable_cash_buyer")) == "True"
    hard_money = s(row.get("est_hard_money")) == "True"
    equity = safe_float(row.get("estimated_equity"), 0)
    equity_ratio = safe_float(row.get("equity_ratio"), 0)
    purchases_12 = safe_int(row.get("purchases_last_12mo"))
    flip_count = safe_int(row.get("flip_count"))
    dscr = safe_float(row.get("est_dscr"))
    props = safe_int(row.get("property_count", row.get("props", 0)))
    value = safe_float(row.get("total_portfolio_value"), 0)
    payment = safe_float(row.get("est_monthly_payment"))

    if rate and rate > 8:
        if payment and payment > 0:
            savings = payment * (1 - 6.25 / rate)
            return f"Currently at {rate:.1f}% - save ${savings:,.0f}/mo switching to 6.25% DSCR"
        return f"Currently at {rate:.1f}% - significant savings at 6.25% DSCR"
    if cash_buyer and equity > 500000:
        return f"Cash buyer - unlock ${equity:,.0f} equity with cash-out refi at 6.25%"
    if cash_buyer:
        return "Cash buyer - unlock equity with cash-out refi at 6.25%"
    if hard_money:
        return "Hard money loan - refi into long-term DSCR at 6.25%"
    if equity_ratio > 0.6 and equity > 0:
        return f"{equity_ratio*100:.0f}% equity - ${equity:,.0f} available for cash-out"
    if purchases_12 > 0:
        return f"Active buyer ({purchases_12} in 12mo) - needs ongoing DSCR financing"
    if flip_count > 2:
        return "Flipper profile - bridge-to-DSCR exit strategy"
    if dscr is not None and dscr >= 1.0:
        return f"DSCR-ready at {dscr:.2f}x - quick approval"
    v = f"${value/1_000_000:.1f}M" if value >= 1_000_000 else f"${value/1_000:.0f}K"
    return f"{props}-property portfolio, {v} - portfolio consolidation opportunity"


def build_talking_points(row):
    pts = []
    props = safe_int(row.get("props", row.get("property_count", 1)), 1)
    pval = safe_float(row.get("total_portfolio_value"), 0)
    is_ent = s(row.get("is_entity")).upper() == "TRUE"
    oos = s(row.get("out_of_state")).upper() == "TRUE"
    lender = fv(row.get("attom_lender_name", row.get("best_lender", "")))
    rate = fv(row.get("attom_rate_type", ""))
    cashout = safe_float(row.get("max_cashout_75", row.get("portfolio_cashout_75", 0)), 0)
    refi = s(row.get("refi_priority")).lower()
    brrrr = s(row.get("brrrr_exit_candidate")).upper() == "TRUE"
    equity_harv = s(row.get("equity_harvest_candidate")).upper() == "TRUE"
    p12 = safe_int(row.get("purchases_last_12mo"))
    rent = safe_float(row.get("est_annual_rent"), 0)

    if props >= 10:
        pts.append(f"Institutional-scale investor with {props} properties and a ${pval:,.0f} portfolio.")
    elif props >= 5:
        pts.append(f"Established portfolio landlord with {props} properties worth ${pval:,.0f}.")
    elif props >= 2:
        pts.append(f"Active investor scaling a {props}-property portfolio valued at ${pval:,.0f}.")
    if is_ent:
        pts.append("Entity-structured ownership signals sophistication and comfort with non-QM lending products.")
    if oos:
        pts.append("Out-of-state investor managing remotely -- may prefer streamlined DSCR process.")
    if lender and rate:
        pts.append(f"Currently financed through {lender} ({rate.lower()}).")
    elif lender:
        pts.append(f"Currently financed through {lender}.")
    if cashout > 100000:
        pts.append(f"Cash-out refi potential: up to ${cashout:,.0f} available at 75% LTV.")
    if brrrr:
        pts.append("Recent below-market acquisition indicates BRRRR strategy -- may need DSCR exit financing.")
    if equity_harv:
        pts.append("Long-held properties with accumulated equity -- prime candidate for cash-out refi.")
    if p12 >= 2:
        pts.append(f"Highly active: {p12} acquisitions in the past 12 months.")
    elif p12 == 1:
        pts.append("Recent acquisition indicates active investment posture.")
    if rent > 50000:
        pts.append(f"Est. ${rent:,.0f}/year rental income supports strong DSCR qualification.")
    if refi == "high":
        pts.append("Flagged as HIGH refinance priority.")
    return " ".join(pts) if pts else "Confirmed investment property owner -- DSCR lending conversation opportunity."


def calculate_priority(row):
    """Return (label, sort_key) tuple."""
    score = safe_float(row.get("score", row.get("_score", 0)), 0)
    tier = s(row.get("selling_tier", "")).lower()
    if "hot" in tier or "1" in tier or score >= 50:
        return ("HOT", 0)
    elif "warm" in tier or "2" in tier or score >= 30:
        return ("WARM", 1)
    return ("NURTURE", 2)


def redact(val):
    if not val:
        return ""
    return "X" * min(len(str(val)), 12)


def tab_name(index, name_raw):
    """Build a safe tab name: '01 FirstName LastName' (max 31 chars)."""
    name = re.sub(r"[^\w\s-]", "", str(name_raw))[:26].strip()
    return f"{index:02d} {name}"[:31]


# ════════════════════════════════════════════════════════════════════
# TAB 1: PIPELINE
# ════════════════════════════════════════════════════════════════════

PIPELINE_HEADERS = [
    "Priority", "Score", "Investor", "Phone", "Type", "Properties",
    "Portfolio $", "Equity", "DSCR", "Cash-Out 75%", "ICP Segment",
    "Financing Angle", "Email", "Call Status", "Call Date",
    "Follow-Up", "Notes",
]

PIPELINE_WIDTHS = [12, 8, 28, 16, 10, 10, 16, 14, 8, 14, 20, 40, 26, 14, 12, 12, 36]


def build_pipeline_tab(wb, df, investor_tabs, is_redacted=False):
    ws = wb.active
    ws.title = "PIPELINE"
    ws.sheet_properties.tabColor = GREEN

    num_cols = len(PIPELINE_HEADERS)

    # ── Row 1: Banner ─────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    banner = ws.cell(row=1, column=1, value="DSCR INVESTOR PIPELINE")
    banner.fill = navy_fill
    banner.font = white_bold_14
    banner.alignment = center_center
    ws.row_dimensions[1].height = 32

    # ── Row 2: Column headers ─────────────────────────────────────
    for c, hdr in enumerate(PIPELINE_HEADERS, 1):
        cell = ws.cell(row=2, column=c, value=hdr)
        cell.fill = navy_fill
        cell.font = white_bold_10
        cell.alignment = wrap_center
        cell.border = thin_border
    ws.row_dimensions[2].height = 28

    # Column widths
    for c, w in enumerate(PIPELINE_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ── Sort data ─────────────────────────────────────────────────
    rows_data = []
    for _, row in df.iterrows():
        priority_label, priority_sort = calculate_priority(row)
        score = safe_float(row.get("score", row.get("_score", 0)), 0)
        rows_data.append((priority_sort, -score, row, priority_label))
    rows_data.sort(key=lambda x: (x[0], x[1]))

    # ── Data rows ─────────────────────────────────────────────────
    for i, (_, _, row, priority_label) in enumerate(rows_data):
        r = i + 3  # data starts at row 3
        ws.row_dimensions[r].height = 36

        owner_raw = clean_name(row.get("OWN_NAME", "")) or "Unknown"
        owner = display_name(owner_raw)
        score = safe_float(row.get("score", row.get("_score", 0)), 0)
        phone1 = fphone(row.get("phone_1", ""))
        phone1_type = s(row.get("phone_1_type", "")).capitalize()
        props = safe_int(row.get("props", row.get("property_count", 0)))
        portfolio = safe_float(row.get("total_portfolio_value"), 0)
        equity = safe_float(row.get("estimated_equity"), 0)
        dscr = safe_float(row.get("est_dscr"))
        cashout = safe_float(row.get("max_cashout_75", row.get("portfolio_cashout_75", 0)), 0)
        segment = s(row.get("_icp", row.get("selling_segment", "")))
        angle = generate_financing_angle(row)
        email1 = s(row.get("email_1", ""))

        if is_redacted:
            phone1 = redact(phone1)
            email1 = redact(email1)
            owner = redact(owner)

        values = [
            priority_label, score, owner,
            phone1, phone1_type, props,
            portfolio, equity,
            dscr if dscr is not None else "",
            cashout if cashout else "",
            segment, angle, email1,
            "", "", "", "",  # Call Status, Call Date, Follow-Up, Notes
        ]

        is_alt = i % 2 == 1
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = wrap_top
            cell.font = normal_9
            cell.border = thin_border
            if is_alt:
                cell.fill = alt_fill

        # ── Priority fill (col A) ────────────────────────────────
        pri_cell = ws.cell(row=r, column=1)
        pri_cell.alignment = center_center
        pri_cell.font = bold_10
        if priority_label == "HOT":
            pri_cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
        elif priority_label == "WARM":
            pri_cell.fill = PatternFill("solid", fgColor="FFE0A0")
        else:
            pri_cell.fill = PatternFill("solid", fgColor="D8D8D8")

        # ── Score fill (col B) ────────────────────────────────────
        score_cell = ws.cell(row=r, column=2)
        score_cell.alignment = center_center
        score_cell.font = bold_10
        if score >= 50:
            score_cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
        elif score >= 30:
            score_cell.fill = PatternFill("solid", fgColor=LIGHT_YELLOW)

        # ── Investor name bold + hyperlink to detail tab (col C) ──
        inv_cell = ws.cell(row=r, column=3)
        inv_cell.font = Font(bold=True, size=10, name="Calibri", color=ACCENT)
        tab = investor_tabs.get(i)
        if tab and not is_redacted:
            safe_tab = tab.replace("'", "''")
            inv_cell.hyperlink = f"#'{safe_tab}'!A1"

        # ── Phone type fill (col E) ───────────────────────────────
        type_cell = ws.cell(row=r, column=5)
        if phone1_type.lower() == "mobile":
            type_cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)

        # ── Currency formatting ───────────────────────────────────
        for c in [7, 8, 10]:  # Portfolio $, Equity, Cash-Out
            ws.cell(row=r, column=c).number_format = "$#,##0"
            ws.cell(row=r, column=c).alignment = right_center

        # ── DSCR formatting (col I) ──────────────────────────────
        dscr_cell = ws.cell(row=r, column=9)
        if dscr is not None:
            dscr_cell.number_format = "0.00"
            if dscr >= 1.0:
                dscr_cell.font = green_bold_10
            else:
                dscr_cell.font = red_bold_10
        dscr_cell.alignment = center_center

        # ── Equity color (col H) ─────────────────────────────────
        eq_cell = ws.cell(row=r, column=8)
        if equity >= 200000:
            eq_cell.font = green_bold_10
        elif equity < 0:
            eq_cell.font = red_bold_10

    nr = len(rows_data) + 2  # last data row

    # ── Data bars ─────────────────────────────────────────────────
    # Score (col B)
    ws.conditional_formatting.add(
        f"B3:B{nr}",
        DataBarRule(start_type="num", start_value=0,
                    end_type="num", end_value=100,
                    color=GREEN),
    )
    # Properties (col F)
    ws.conditional_formatting.add(
        f"F3:F{nr}",
        DataBarRule(start_type="min", end_type="max", color=TEAL),
    )
    # Portfolio $ (col G)
    ws.conditional_formatting.add(
        f"G3:G{nr}",
        DataBarRule(start_type="min", end_type="max", color=NAVY),
    )
    # Cash-Out (col J)
    ws.conditional_formatting.add(
        f"J3:J{nr}",
        DataBarRule(start_type="min", end_type="max", color=ACCENT),
    )

    # ── Call Status dropdown (col N) ──────────────────────────────
    dv = DataValidation(
        type="list",
        formula1='"Not Called,VM Left,Connected,Interested,Scheduled,DNC"',
        allow_blank=True,
    )
    dv.error = "Select from dropdown"
    ws.add_data_validation(dv)
    dv.add(f"N3:N{nr}")

    # ── Call Status conditional formatting ────────────────────────
    for label, fill in [
        ("Connected", cond_blue), ("Interested", cond_green),
        ("Scheduled", PatternFill("solid", fgColor="81C784")),
        ("DNC", cond_red),
    ]:
        ws.conditional_formatting.add(
            f"N3:N{nr}",
            CellIsRule(operator="equal", formula=[f'"{label}"'], fill=fill),
        )

    # ── Date formatting (cols O, P) ──────────────────────────────
    for c in [15, 16]:
        for r in range(3, nr + 1):
            ws.cell(row=r, column=c).number_format = "MMM DD, YYYY"

    # ── Notes wrap (col Q) ────────────────────────────────────────
    for r in range(3, nr + 1):
        ws.cell(row=r, column=17).alignment = wrap_top

    # ── Freeze panes + auto-filter ────────────────────────────────
    ws.freeze_panes = "D3"
    ws.auto_filter.ref = f"A2:{get_column_letter(num_cols)}{nr}"

    return ws



# ════════════════════════════════════════════════════════════════════
# TABS 3-N: PER-INVESTOR DETAIL TABS
# ════════════════════════════════════════════════════════════════════

def build_investor_tab(wb, row, index, is_redacted=False):
    owner_raw = clean_name(row.get("OWN_NAME", "")) or "Unknown"
    owner = display_name(owner_raw)
    tname = tab_name(index + 1, owner)
    ws = wb.create_sheet(tname)

    score = safe_float(row.get("score", row.get("_score", 0)), 0)
    segment = s(row.get("selling_segment", row.get("_icp", "Investor")))
    tier = s(row.get("selling_tier", ""))
    props = safe_int(row.get("props", row.get("property_count", 1)), 1)
    county_state = get_county_state(row)

    priority_label, _ = calculate_priority(row)

    # Tab color by tier
    if priority_label == "HOT":
        ws.sheet_properties.tabColor = GREEN
    elif priority_label == "WARM":
        ws.sheet_properties.tabColor = AMBER
    else:
        ws.sheet_properties.tabColor = MED_GRAY

    # Hide gridlines
    ws.sheet_view.showGridLines = False

    r = lambda v: redact(v) if is_redacted else v

    # ════════════════════════════════════════════════════════════
    # ACT 1: "WHO IS THIS?" (Rows 1-7)
    # ════════════════════════════════════════════════════════════

    # Row 1: Navy banner with name + score
    ws.merge_cells("A1:J1")
    name_cell = ws.cell(row=1, column=1, value=r(owner))
    name_cell.fill = navy_fill
    name_cell.font = navy_bold_16
    name_cell.alignment = left_center

    # Score badge in col K-L
    ws.merge_cells("K1:L1")
    score_cell = ws.cell(row=1, column=11, value=f"SCORE: {int(score)}")
    score_cell.alignment = center_center
    score_cell.font = Font(bold=True, size=14, color=WHITE, name="Calibri")
    if score >= 50:
        score_cell.fill = PatternFill("solid", fgColor=GREEN)
    elif score >= 30:
        score_cell.fill = PatternFill("solid", fgColor=AMBER)
    else:
        score_cell.fill = PatternFill("solid", fgColor=MED_GRAY)
    ws.row_dimensions[1].height = 36

    # Row 2: Teal subtitle
    ws.merge_cells("A2:L2")
    subtitle = f"{segment}  |  {props} Properties  |  {county_state}  |  {priority_label}"
    sub_cell = ws.cell(row=2, column=1, value=subtitle)
    sub_cell.font = teal_bold_10
    sub_cell.alignment = left_center
    ws.row_dimensions[2].height = 22

    # Row 3: blank spacer
    ws.row_dimensions[3].height = 8

    # Rows 4-6: KPI cards (4 cards across 3 rows: label, value, spacer)
    portfolio_val = safe_float(row.get("total_portfolio_value"), 0)
    equity_val = safe_float(row.get("estimated_equity"), 0)
    equity_ratio = safe_float(row.get("equity_ratio"), 0)
    dscr_val = safe_float(row.get("est_dscr"))

    kpi_data = [
        ("PORTFOLIO VALUE", fc(portfolio_val) or "--"),
        ("TOTAL EQUITY", fc(equity_val) or "--"),
        ("EQUITY RATIO", f"{equity_ratio*100:.0f}%" if equity_ratio else "--"),
        ("EST. DSCR", f"{dscr_val:.2f}" if dscr_val else "--"),
    ]

    # 4 cards: cols A-C, D-F, G-I, J-L
    card_cols = [(1, 3), (4, 6), (7, 9), (10, 12)]
    for ci, ((label, value), (cs, ce)) in enumerate(zip(kpi_data, card_cols)):
        # Label row (row 4)
        ws.merge_cells(start_row=4, start_column=cs, end_row=4, end_column=ce)
        lbl = ws.cell(row=4, column=cs, value=label)
        lbl.font = dark_gray_7
        lbl.alignment = center_center
        lbl.fill = light_bg_fill

        # Value row (row 5)
        ws.merge_cells(start_row=5, start_column=cs, end_row=5, end_column=ce)
        val_cell = ws.cell(row=5, column=cs, value=value)
        val_cell.font = navy_bold_20
        val_cell.alignment = center_center
        val_cell.fill = light_bg_fill

        # Color-code equity and DSCR
        if ci == 1 and equity_val > 0:  # Equity
            if equity_val >= 200000:
                val_cell.font = Font(bold=True, color=GREEN, size=20, name="Calibri")
            elif equity_val < 0:
                val_cell.font = Font(bold=True, color=RED, size=20, name="Calibri")
        if ci == 3 and dscr_val is not None:  # DSCR
            if dscr_val >= 1.0:
                val_cell.font = Font(bold=True, color=GREEN, size=20, name="Calibri")
            else:
                val_cell.font = Font(bold=True, color=RED, size=20, name="Calibri")

        # Teal accent top border
        for c in range(cs, ce + 1):
            ws.cell(row=4, column=c).border = Border(top=Side(style="medium", color=TEAL))

    ws.row_dimensions[4].height = 16
    ws.row_dimensions[5].height = 32
    ws.row_dimensions[6].height = 8  # spacer

    # ════════════════════════════════════════════════════════════
    # ACT 2: "WHAT DO THEY OWN?" (Rows 7+)
    # ════════════════════════════════════════════════════════════

    cur_row = 8

    # ── LEFT PANEL: Contact + Entity + Signals (cols A-F) ─────
    # CONTACT section label
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=6)
    ws.cell(row=cur_row, column=1, value="CONTACT").font = navy_bold_9
    ws.cell(row=cur_row, column=1).border = Border(bottom=Side(style="medium", color=TEAL))
    cur_left = cur_row + 1

    phone1 = fphone(row.get("phone_1", ""))
    phone1_type = s(row.get("phone_1_type", "")).capitalize()
    phone2 = fphone(row.get("phone_2", ""))
    email1 = s(row.get("email_1", ""))
    email2 = s(row.get("email_2", ""))
    mail_addr = s(row.get("OWN_ADDR1", row.get("mail_street", "")))
    mail_city = s(row.get("OWN_CITY", ""))
    mail_state = s(row.get("OWN_STATE_DOM", row.get("OWN_STATE", "")))
    mail_zip = s(row.get("OWN_ZIPCD", ""))

    def _info_row(ws, r, label, value, col_start=1, col_end=6):
        if not value:
            return r
        ws.cell(row=r, column=col_start, value=label).font = dark_gray_8
        ws.merge_cells(start_row=r, start_column=col_start + 2, end_row=r, end_column=col_end)
        ws.cell(row=r, column=col_start + 2, value=r_val(value)).font = bold_9
        return r + 1

    def r_val(v):
        return redact(v) if is_redacted else v

    if phone1:
        ws.cell(row=cur_left, column=1, value="Phone").font = dark_gray_8
        ws.merge_cells(start_row=cur_left, start_column=3, end_row=cur_left, end_column=5)
        ws.cell(row=cur_left, column=3, value=r_val(phone1)).font = bold_9
        # Type badge
        type_cell = ws.cell(row=cur_left, column=6, value=phone1_type)
        if phone1_type.lower() == "mobile":
            type_cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
            type_cell.font = Font(bold=True, size=8, color=GREEN, name="Calibri")
        else:
            type_cell.font = dark_gray_8
        type_cell.alignment = center_center
        cur_left += 1

    if phone2 and phone2 != phone1:
        ws.cell(row=cur_left, column=1, value="Phone 2").font = dark_gray_8
        ws.merge_cells(start_row=cur_left, start_column=3, end_row=cur_left, end_column=6)
        ws.cell(row=cur_left, column=3, value=r_val(phone2)).font = bold_9
        cur_left += 1

    if email1:
        ws.cell(row=cur_left, column=1, value="Email").font = dark_gray_8
        ws.merge_cells(start_row=cur_left, start_column=3, end_row=cur_left, end_column=6)
        ws.cell(row=cur_left, column=3, value=r_val(email1)).font = normal_9
        cur_left += 1

    if email2 and email2 != email1:
        ws.cell(row=cur_left, column=1, value="Email 2").font = dark_gray_8
        ws.merge_cells(start_row=cur_left, start_column=3, end_row=cur_left, end_column=6)
        ws.cell(row=cur_left, column=3, value=r_val(email2)).font = normal_9
        cur_left += 1

    if mail_addr:
        ws.cell(row=cur_left, column=1, value="Address").font = dark_gray_8
        ws.merge_cells(start_row=cur_left, start_column=3, end_row=cur_left, end_column=6)
        city_line = ", ".join(filter(None, [mail_city, mail_state]))
        if mail_zip:
            city_line += f" {mail_zip}"
        ws.cell(row=cur_left, column=3, value=r_val(f"{mail_addr}, {city_line}")).font = normal_9
        cur_left += 1

    cur_left += 1  # spacer

    # Entity details
    is_entity = s(row.get("is_entity")).upper() == "TRUE"
    if is_entity:
        ws.merge_cells(start_row=cur_left, start_column=1, end_row=cur_left, end_column=6)
        ws.cell(row=cur_left, column=1, value="ENTITY DETAILS").font = navy_bold_9
        ws.cell(row=cur_left, column=1).border = Border(bottom=Side(style="medium", color=TEAL))
        cur_left += 1

        agent = fv(row.get("registered_agent_name", row.get("registered_agent", "")))
        officers = fv(row.get("entity_officers", row.get("officer_names", "")))
        entity_status = fv(row.get("entity_status", row.get("sunbiz_status", "")))

        for label, value in [("Agent", agent), ("Officers", officers), ("Status", entity_status)]:
            if value:
                ws.cell(row=cur_left, column=1, value=label).font = dark_gray_8
                ws.merge_cells(start_row=cur_left, start_column=3, end_row=cur_left, end_column=6)
                ws.cell(row=cur_left, column=3, value=r_val(value)).font = normal_9
                cur_left += 1
        cur_left += 1

    # Signals
    signals = []
    if s(row.get("brrrr_exit_candidate")).upper() == "TRUE":
        signals.append(("BRRRR EXIT", AMBER))
    if s(row.get("equity_harvest_candidate")).upper() == "TRUE":
        signals.append(("EQUITY HARVEST", GREEN))
    if s(row.get("rate_refi_candidate")).upper() == "TRUE":
        signals.append(("RATE REFI", ACCENT))
    if s(row.get("out_of_state")).upper() == "TRUE":
        signals.append(("OUT-OF-STATE", MED_GRAY))
    if s(row.get("str_licensed")).upper() == "TRUE":
        signals.append(("STR LICENSED", "6A32C8"))
    refi_priority = fv(row.get("refi_priority", ""))
    if refi_priority and refi_priority.lower() == "high":
        signals.append(("REFI PRIORITY", RED))

    if signals:
        ws.merge_cells(start_row=cur_left, start_column=1, end_row=cur_left, end_column=6)
        ws.cell(row=cur_left, column=1, value="OPPORTUNITY SIGNALS").font = navy_bold_9
        ws.cell(row=cur_left, column=1).border = Border(bottom=Side(style="medium", color=TEAL))
        cur_left += 1

        for ci, (sig_text, sig_color) in enumerate(signals):
            col = (ci % 3) * 2 + 1
            cell = ws.cell(row=cur_left + ci // 3, column=col, value=sig_text)
            cell.fill = PatternFill("solid", fgColor=sig_color)
            cell.font = Font(bold=True, size=7, color=WHITE, name="Calibri")
            cell.alignment = center_center
        cur_left += (len(signals) - 1) // 3 + 2

    # ── RIGHT PANEL: Financing + DSCR (cols G-L) ─────────────
    cur_right = cur_row

    lender = fv(row.get("attom_lender_name", row.get("best_lender", "")))
    loan_amt = fc(row.get("attom_loan_amount", ""))
    rate_type = fv(row.get("attom_rate_type", ""))
    interest_rate = fv(row.get("attom_interest_rate", row.get("est_interest_rate", "")))
    loan_date = fv(row.get("attom_loan_date", ""))
    due_date = fv(row.get("attom_due_date", row.get("est_maturity_date", "")))
    cashout = fc(row.get("max_cashout_75", row.get("portfolio_cashout_75", "")))
    attom_prop = fv(row.get("attom_property_address", ""))

    # Label financing section — ATTOM only matched one property
    has_financing = any([lender, loan_amt, rate_type, interest_rate])
    if has_financing and attom_prop and props > 1:
        fin_label = f"MORTGAGE DETAIL (ATTOM match: 1 of {props})"
    else:
        fin_label = "FINANCING INTELLIGENCE"

    ws.merge_cells(start_row=cur_right, start_column=7, end_row=cur_right, end_column=12)
    ws.cell(row=cur_right, column=7, value=fin_label).font = navy_bold_9
    ws.cell(row=cur_right, column=7).border = Border(bottom=Side(style="medium", color=TEAL))
    cur_right += 1

    # Show which property this mortgage belongs to
    if has_financing and attom_prop:
        ws.cell(row=cur_right, column=7, value="Property").font = dark_gray_8
        ws.merge_cells(start_row=cur_right, start_column=9, end_row=cur_right, end_column=12)
        ws.cell(row=cur_right, column=9, value=attom_prop.title()[:45]).font = Font(italic=True, size=8, color=DARK_GRAY, name="Calibri")
        cur_right += 1

    for label, value in [
        ("Lender", lender), ("Loan Amount", loan_amt),
        ("Rate Type", rate_type), ("Interest Rate", f"{interest_rate}%" if interest_rate else None),
        ("Loan Date", loan_date), ("Maturity", due_date),
        ("Cash-Out (75%)", cashout),
    ]:
        if value:
            ws.cell(row=cur_right, column=7, value=label).font = dark_gray_8
            ws.merge_cells(start_row=cur_right, start_column=9, end_row=cur_right, end_column=12)
            ws.cell(row=cur_right, column=9, value=value).font = bold_9
            cur_right += 1

    # Note about unmatched properties
    if has_financing and props > 1:
        ws.merge_cells(start_row=cur_right, start_column=7, end_row=cur_right, end_column=12)
        ws.cell(row=cur_right, column=7,
                value=f"Remaining {props - 1} properties: no ATTOM match yet").font = Font(italic=True, size=7, color=MED_GRAY, name="Calibri")
        cur_right += 1

    cur_right += 1

    # DSCR Analysis
    ws.merge_cells(start_row=cur_right, start_column=7, end_row=cur_right, end_column=12)
    ws.cell(row=cur_right, column=7, value="DSCR ANALYSIS").font = navy_bold_9
    ws.cell(row=cur_right, column=7).border = Border(bottom=Side(style="medium", color=TEAL))
    cur_right += 1

    dscr_display = f"{dscr_val:.2f}" if dscr_val else "--"
    ws.merge_cells(start_row=cur_right, start_column=8, end_row=cur_right + 1, end_column=11)
    dscr_cell = ws.cell(row=cur_right, column=8, value=dscr_display)
    dscr_cell.alignment = center_center
    if dscr_val and dscr_val >= 1.0:
        dscr_cell.font = Font(bold=True, size=28, color=GREEN, name="Calibri")
    elif dscr_val:
        dscr_cell.font = Font(bold=True, size=28, color=RED, name="Calibri")
    else:
        dscr_cell.font = Font(bold=True, size=28, color=MED_GRAY, name="Calibri")
    cur_right += 2

    rent = fc(row.get("est_annual_rent", ""))
    noi = fc(row.get("est_noi", ""))
    if rent:
        ws.cell(row=cur_right, column=7, value="Annual Rent").font = dark_gray_8
        ws.cell(row=cur_right, column=9, value=rent).font = normal_9
        cur_right += 1
    if noi:
        ws.cell(row=cur_right, column=7, value="NOI").font = dark_gray_8
        ws.cell(row=cur_right, column=9, value=noi).font = normal_9
        cur_right += 1

    # ── Equity vs Debt bar (full width) ───────────────────────
    bar_row = max(cur_left, cur_right) + 1
    equity_num = safe_float(row.get("estimated_equity"), 0)
    loan_num = safe_float(row.get("attom_loan_amount"), 0)
    total = equity_num + loan_num

    if total > 0:
        ws.merge_cells(start_row=bar_row, start_column=1, end_row=bar_row, end_column=12)
        ws.cell(row=bar_row, column=1, value="EQUITY vs DEBT").font = navy_bold_9
        ws.cell(row=bar_row, column=1).border = Border(bottom=Side(style="medium", color=TEAL))
        bar_row += 1

        eq_cols = max(1, round(12 * equity_num / total))
        dt_cols = 12 - eq_cols

        if eq_cols > 0:
            ws.merge_cells(start_row=bar_row, start_column=1, end_row=bar_row, end_column=eq_cols)
            eq_cell = ws.cell(row=bar_row, column=1, value=f"Equity: {fc(equity_num) or '--'}  ({equity_num/total*100:.0f}%)")
            eq_cell.fill = PatternFill("solid", fgColor=GREEN)
            eq_cell.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
            eq_cell.alignment = center_center

        if dt_cols > 0 and loan_num > 0:
            ws.merge_cells(start_row=bar_row, start_column=eq_cols + 1, end_row=bar_row, end_column=12)
            dt_cell = ws.cell(row=bar_row, column=eq_cols + 1, value=f"Debt: {fc(loan_num) or '--'}  ({loan_num/total*100:.0f}%)")
            dt_cell.fill = PatternFill("solid", fgColor=RED)
            dt_cell.font = Font(bold=True, size=10, color=WHITE, name="Calibri")
            dt_cell.alignment = center_center

        ws.row_dimensions[bar_row].height = 28
        bar_row += 1

        # Cash-out callout
        cashout_num = safe_float(row.get("max_cashout_75", row.get("portfolio_cashout_75", 0)), 0)
        if cashout_num > 0:
            ws.merge_cells(start_row=bar_row, start_column=1, end_row=bar_row, end_column=12)
            ws.cell(row=bar_row, column=1,
                    value=f"Cash-out potential (75% LTV): {fc(cashout_num)}").font = navy_bold_10
            ws.cell(row=bar_row, column=1).alignment = center_center
            bar_row += 1

        bar_row += 1  # spacer

    # ── Property Portfolio Table ──────────────────────────────
    table_row = bar_row
    ws.merge_cells(start_row=table_row, start_column=1, end_row=table_row, end_column=12)
    ws.cell(row=table_row, column=1,
            value=f"PROPERTY PORTFOLIO ({props} Properties)").font = navy_bold_9
    ws.cell(row=table_row, column=1).border = Border(bottom=Side(style="medium", color=TEAL))
    table_row += 1

    # Header
    prop_headers = ["#", "Address", "", "", "Est. Value", "Lender", "", "Rate", "Loan", "", "Equity", ""]
    prop_header_labels = {1: "#", 2: "Address", 5: "Est. Value", 6: "Lender", 8: "Rate", 9: "Loan", 11: "Equity"}
    for c, val in prop_header_labels.items():
        cell = ws.cell(row=table_row, column=c, value=val)
        cell.fill = navy_fill
        cell.font = white_bold_10
        cell.alignment = center_center
    # Fill remaining header cells with navy
    for c in range(1, 13):
        ws.cell(row=table_row, column=c).fill = navy_fill
    # Merge address cols
    ws.merge_cells(start_row=table_row, start_column=2, end_row=table_row, end_column=4)
    ws.merge_cells(start_row=table_row, start_column=6, end_row=table_row, end_column=7)
    ws.merge_cells(start_row=table_row, start_column=9, end_row=table_row, end_column=10)
    ws.merge_cells(start_row=table_row, start_column=11, end_row=table_row, end_column=12)
    table_row += 1

    # Parse addresses
    phy_addr_raw = str(row.get("PHY_ADDR1", "")).strip()
    addresses = []
    if phy_addr_raw and phy_addr_raw.upper() not in ("NAN", "NONE", ""):
        addresses = [a.strip().title()[:40] for a in phy_addr_raw.split("|") if a.strip()]

    attom_addr = str(row.get("attom_property_address", "")).strip().upper()
    avg_val_num = safe_float(row.get("avg_property_value"), 0)

    show_addrs = addresses[:10]
    overflow = max(0, props - len(show_addrs))
    total_value = 0
    total_loan_table = 0
    total_equity_table = 0

    for idx, addr in enumerate(show_addrs):
        is_alt = idx % 2 == 1

        # ATTOM match check
        is_attom = False
        if attom_addr:
            addr_upper = addr.upper().replace(",", "").strip()
            if addr_upper in attom_addr or attom_addr in addr_upper:
                is_attom = True

        prop_val = avg_val_num
        prop_lender = fv(row.get("attom_lender_name", "")) if is_attom else "--"
        prop_rate = fv(row.get("attom_rate_type", "")) if is_attom else "--"
        prop_loan = safe_float(row.get("attom_loan_amount"), 0) if is_attom else 0
        prop_equity = prop_val - prop_loan if prop_val > 0 else prop_val

        total_value += prop_val
        total_loan_table += prop_loan
        total_equity_table += prop_equity

        r_idx = table_row + idx
        fill = alt_fill if is_alt else white_fill

        ws.cell(row=r_idx, column=1, value=idx + 1).font = normal_9
        ws.cell(row=r_idx, column=1).fill = fill
        ws.cell(row=r_idx, column=1).alignment = center_center

        ws.merge_cells(start_row=r_idx, start_column=2, end_row=r_idx, end_column=4)
        ws.cell(row=r_idx, column=2, value=r_val(addr)).font = normal_9
        ws.cell(row=r_idx, column=2).fill = fill

        ws.cell(row=r_idx, column=5, value=fc(prop_val) or "--").font = normal_9
        ws.cell(row=r_idx, column=5).fill = fill

        ws.merge_cells(start_row=r_idx, start_column=6, end_row=r_idx, end_column=7)
        ws.cell(row=r_idx, column=6, value=prop_lender or "--").font = normal_9
        ws.cell(row=r_idx, column=6).fill = fill

        ws.cell(row=r_idx, column=8, value=prop_rate or "--").font = normal_9
        ws.cell(row=r_idx, column=8).fill = fill

        ws.merge_cells(start_row=r_idx, start_column=9, end_row=r_idx, end_column=10)
        ws.cell(row=r_idx, column=9, value=fc(prop_loan) if prop_loan > 0 else "--").font = normal_9
        ws.cell(row=r_idx, column=9).fill = fill

        ws.merge_cells(start_row=r_idx, start_column=11, end_row=r_idx, end_column=12)
        ws.cell(row=r_idx, column=11, value=fc(prop_equity) if prop_equity > 0 else "--").font = normal_9
        ws.cell(row=r_idx, column=11).fill = fill

    table_row += len(show_addrs)

    # Overflow
    if overflow > 0:
        ws.merge_cells(start_row=table_row, start_column=1, end_row=table_row, end_column=12)
        ws.cell(row=table_row, column=1,
                value=f"  ...and {overflow} more properties").font = Font(italic=True, size=9, color=DARK_GRAY, name="Calibri")
        ws.cell(row=table_row, column=1).fill = alt_fill
        table_row += 1

    # Totals row
    portfolio_val_display = fc(row.get("total_portfolio_value", "")) or fc(total_value) or "--"
    portfolio_equity_display = fc(row.get("estimated_equity", "")) or fc(total_equity_table) or "--"
    portfolio_loan_display = fc(total_loan_table) if total_loan_table > 0 else "--"

    for c in range(1, 13):
        ws.cell(row=table_row, column=c).fill = light_blue_fill
    ws.cell(row=table_row, column=1).font = navy_bold_10
    ws.merge_cells(start_row=table_row, start_column=2, end_row=table_row, end_column=4)
    ws.cell(row=table_row, column=2, value=f"TOTAL  |  {props} Properties").font = navy_bold_10

    ws.cell(row=table_row, column=5, value=portfolio_val_display).font = navy_bold_10
    ws.merge_cells(start_row=table_row, start_column=9, end_row=table_row, end_column=10)
    ws.cell(row=table_row, column=9, value=portfolio_loan_display).font = navy_bold_10
    ws.merge_cells(start_row=table_row, start_column=11, end_row=table_row, end_column=12)
    ws.cell(row=table_row, column=11, value=portfolio_equity_display).font = navy_bold_10

    table_row += 2

    # ════════════════════════════════════════════════════════════
    # ACT 3: "HOW TO WIN THE BUSINESS"
    # ════════════════════════════════════════════════════════════

    ws.merge_cells(start_row=table_row, start_column=1, end_row=table_row, end_column=12)
    ws.cell(row=table_row, column=1, value="HOW TO WIN THE BUSINESS").font = navy_bold_10
    for c in range(1, 13):
        ws.cell(row=table_row, column=c).border = Border(
            top=Side(style="medium", color=TEAL),
            bottom=Side(style="thin", color=TEAL),
        )
    ws.cell(row=table_row, column=1).fill = PatternFill("solid", fgColor="F5F7FA")
    table_row += 1

    talking = build_talking_points(row)
    ws.merge_cells(start_row=table_row, start_column=1, end_row=table_row + 1, end_column=12)
    tp_cell = ws.cell(row=table_row, column=1, value=talking)
    tp_cell.font = normal_9
    tp_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[table_row].height = 48
    table_row += 2

    # Acquisition summary
    p12 = fv(row.get("purchases_last_12mo", ""))
    p36 = fv(row.get("purchases_last_36mo", ""))
    avg_purchase = fc(row.get("avg_purchase_price", row.get("avg_sale_price", "")))
    flip_count = fv(row.get("flip_count", ""))
    acq_parts = []
    if p12:
        acq_parts.append(f"12mo: {p12}")
    if p36:
        acq_parts.append(f"36mo: {p36}")
    if avg_purchase:
        acq_parts.append(f"Avg: {avg_purchase}")
    if flip_count and flip_count != "0":
        acq_parts.append(f"Flips: {flip_count}")
    if acq_parts:
        ws.merge_cells(start_row=table_row, start_column=1, end_row=table_row, end_column=12)
        ws.cell(row=table_row, column=1,
                value="Purchases: " + " | ".join(acq_parts)).font = Font(bold=True, size=8, color=DARK_GRAY, name="Calibri")
        table_row += 1

    # Footer
    table_row += 1
    ws.merge_cells(start_row=table_row, start_column=1, end_row=table_row, end_column=12)
    footer = "Confidential" if not is_redacted else "SAMPLE"
    ws.cell(row=table_row, column=1,
            value=f"{footer}  |  Still Mind Creative  |  Source: Public Records + Enrichment APIs").font = Font(size=7, color=MED_GRAY, name="Calibri")
    ws.cell(row=table_row, column=1).alignment = center_center

    # ── Column widths ─────────────────────────────────────────
    col_widths = [10, 12, 10, 10, 14, 12, 10, 10, 12, 10, 12, 10]
    for c, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    return tname


# ════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="Build unified DSCR Excel workbook")
    parser.add_argument("--input", type=str, required=True)
    parser.add_argument("--output", type=str, default=None)
    parser.add_argument("--redacted", action="store_true")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.is_absolute():
        input_path = PROJECT_DIR / args.input
    if not input_path.exists():
        print(f"Input not found: {input_path}")
        return

    output_path = Path(args.output) if args.output else PROJECT_DIR / "data" / "workbooks" / "dscr_workbook.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)

    df = pd.read_csv(input_path, dtype=str)
    print(f"Loaded {len(df)} leads from {input_path.name}")

    wb = Workbook()

    # ── Build investor detail tabs first (need names for hyperlinks) ──
    investor_tabs = {}
    total = len(df)
    detail_cutoff = 50

    # Sort by score descending for tab ordering
    df["_sort_score"] = pd.to_numeric(df.get("score", df.get("_score", 0)), errors="coerce").fillna(0)
    df_sorted = df.sort_values("_sort_score", ascending=False).reset_index(drop=True)

    print(f"\nBuilding investor detail tabs...")
    for i, (_, row) in enumerate(df_sorted.iterrows()):
        priority_label, _ = calculate_priority(row)
        # >50 leads: only HOT + WARM get detail tabs
        if total > detail_cutoff and priority_label == "NURTURE":
            continue
        tname = build_investor_tab(wb, row, i, is_redacted=args.redacted)
        investor_tabs[i] = tname
        print(f"  [{i+1}/{total}] {tname}")

    # ── Build pipeline tab (needs investor_tabs for hyperlinks) ───
    # Re-sort for pipeline display (by priority then score)
    print(f"\nBuilding PIPELINE tab...")

    # Map pipeline row index → investor tab name
    # Pipeline sorts by priority then score; detail tabs sort by score only
    # We need to map each lead to its detail tab by matching OWN_NAME
    pipeline_investor_tabs = {}
    name_to_tab = {}
    for i, (_, row) in enumerate(df_sorted.iterrows()):
        owner = clean_name(row.get("OWN_NAME", "")) or f"lead_{i}"
        if i in investor_tabs:
            name_to_tab[owner] = investor_tabs[i]

    # Build pipeline rows in pipeline sort order and map
    rows_data = []
    for _, row in df.iterrows():
        priority_label, priority_sort = calculate_priority(row)
        score = safe_float(row.get("score", row.get("_score", 0)), 0)
        rows_data.append((priority_sort, -score, row))
    rows_data.sort(key=lambda x: (x[0], x[1]))

    for pi, (_, _, row) in enumerate(rows_data):
        owner = clean_name(row.get("OWN_NAME", "")) or f"lead_{pi}"
        if owner in name_to_tab:
            pipeline_investor_tabs[pi] = name_to_tab[owner]

    build_pipeline_tab(wb, df, pipeline_investor_tabs, is_redacted=args.redacted)

    # ── Save ──────────────────────────────────────────────────────
    wb.save(str(output_path))
    size_kb = output_path.stat().st_size / 1024
    print(f"\nDone! {output_path.name} ({size_kb:.0f} KB)")
    print(f"  PIPELINE: {len(df)} leads")
    print(f"  Detail tabs: {len(investor_tabs)}")
    print(f"  Output: {output_path}")


if __name__ == "__main__":
    main()
