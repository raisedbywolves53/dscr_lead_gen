"""
Build PPP Targeting Deliverable — Client-Facing Excel Workbook
==============================================================

Produces a presentation-quality Excel workbook for loan officer prospects.
Every lead has a resolved contact name, phone, and actionable intelligence.

Tab Structure:
  1. EXECUTIVE SUMMARY — KPI dashboard, methodology overview
  2. REFI OPPORTUNITIES — The call sheet: who to call, why, what to say
  3. DATA SOURCES — Credibility: where this intelligence comes from

Brand: Stillmind Creative palette (Navy #0A2342, Teal #1A6B6A)

Usage:
    python scripts/build_ppp_deliverable.py
    python scripts/build_ppp_deliverable.py --output custom_name.xlsx
"""

import argparse
import math
import re
import sys
from datetime import datetime
from pathlib import Path

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ---------------------------------------------------------------------------
# Brand palette (matches build_workbook.py / build_dossier_pdf.py)
# ---------------------------------------------------------------------------
NAVY = "0A2342"
TEAL = "1A6B6A"
ACCENT = "0066B3"
GREEN = "228B22"
AMBER = "CC8800"
RED = "BF3131"
ALT_ROW = "F8F9FB"
LIGHT_BG = "EDF1F7"
LIGHT_BLUE = "DCEAFC"
WHITE = "FFFFFF"
BLACK = "000000"
LIGHT_GREEN = "C8E6C9"
LIGHT_YELLOW = "FFF9C4"
LIGHT_RED = "FFCDD2"
MED_GRAY = "999999"
DARK_GRAY = "595959"

# Fills
navy_fill = PatternFill("solid", fgColor=NAVY)
teal_fill = PatternFill("solid", fgColor=TEAL)
alt_fill = PatternFill("solid", fgColor=ALT_ROW)
light_bg = PatternFill("solid", fgColor=LIGHT_BG)
light_blue = PatternFill("solid", fgColor=LIGHT_BLUE)
white_fill = PatternFill("solid", fgColor=WHITE)
green_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
yellow_fill = PatternFill("solid", fgColor=LIGHT_YELLOW)
red_fill = PatternFill("solid", fgColor=LIGHT_RED)

# Fonts
f_title = Font(bold=True, color=WHITE, size=14, name="Calibri")
f_section = Font(bold=True, color=WHITE, size=11, name="Calibri")
f_header = Font(bold=True, color=WHITE, size=9, name="Calibri")
f_navy_bold_12 = Font(bold=True, color=NAVY, size=12, name="Calibri")
f_navy_bold_10 = Font(bold=True, color=NAVY, size=10, name="Calibri")
f_teal_bold_10 = Font(bold=True, color=TEAL, size=10, name="Calibri")
f_teal_9 = Font(color=TEAL, size=9, name="Calibri")
f_body = Font(size=10, name="Calibri")
f_body_9 = Font(size=9, name="Calibri")
f_body_bold = Font(bold=True, size=10, name="Calibri")
f_small = Font(size=8, color=DARK_GRAY, name="Calibri")
f_green_bold = Font(bold=True, color=GREEN, size=10, name="Calibri")
f_red_bold = Font(bold=True, color=RED, size=10, name="Calibri")
f_amber_bold = Font(bold=True, color=AMBER, size=10, name="Calibri")
f_kpi_value = Font(bold=True, color=NAVY, size=22, name="Calibri")
f_kpi_label = Font(color=DARK_GRAY, size=9, name="Calibri")

# Alignment
a_center = Alignment(vertical="center", horizontal="center")
a_left = Alignment(vertical="center", horizontal="left")
a_right = Alignment(vertical="center", horizontal="right")
a_wrap = Alignment(vertical="center", horizontal="left", wrap_text=True)
a_wrap_center = Alignment(vertical="center", horizontal="center", wrap_text=True)

# Border
thin_border = Border(
    left=Side("thin", color="D0D0D0"), right=Side("thin", color="D0D0D0"),
    top=Side("thin", color="D0D0D0"), bottom=Side("thin", color="D0D0D0"),
)
bottom_accent = Border(bottom=Side("medium", color=TEAL))

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
DATA_DIR = PROJECT_DIR / "data"
PPP_DIR = DATA_DIR / "ppp_targeting"
OUTPUT_DIR = DATA_DIR / "deliverables"
LOGO_PATH = PROJECT_DIR / "assets" / "logo.png"

TODAY = datetime.now()


# ---------------------------------------------------------------------------
# Data helpers
# ---------------------------------------------------------------------------
def s(val):
    """Safe string — NaN/None to empty."""
    if val is None or (isinstance(val, float) and (math.isnan(val) or np.isnan(val))):
        return ""
    v = str(val).strip()
    return "" if v.lower() in ("nan", "none") else v


def fc(val):
    """Format as currency display string."""
    try:
        v = float(str(val).replace(",", "").replace("$", "").strip())
        if v == 0 or math.isnan(v):
            return ""
        if v >= 1_000_000:
            return f"${v / 1_000_000:.1f}M"
        return f"${v:,.0f}"
    except (ValueError, TypeError):
        return ""


def fphone(val):
    """Format phone as (XXX) XXX-XXXX."""
    digits = re.sub(r"\D", "", str(val))
    if len(digits) == 11 and digits[0] == "1":
        digits = digits[1:]
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    return s(val)


def resolve_contact_name(row):
    """Get the PERSON to call — not the entity name."""
    # Priority: resolved_person > attom owner name > parsed OWN_NAME
    resolved = s(row.get("resolved_person", ""))
    if resolved:
        # Convert "LAST, FIRST" to "First Last"
        if "," in resolved:
            parts = resolved.split(",", 1)
            return f"{parts[1].strip().title()} {parts[0].strip().title()}"
        return resolved.title()

    first = s(row.get("attom_owner1_first", ""))
    last = s(row.get("attom_owner1_last", ""))
    if first and last and last.upper() != s(row.get("OWN_NAME", "")).upper():
        return f"{first.strip().title()} {last.strip().title()}"

    # Last resort: parse OWN_NAME if it looks like a person (not entity)
    own = s(row.get("OWN_NAME", ""))
    entity_kw = ("LLC", "INC", "CORP", "TRUST", "LP", "REALTY", "PROPERTIES",
                 "INVESTMENTS", "HOLDINGS", "GROUP", "CAPITAL", "ENTERPRISES")
    if own and not any(kw in own.upper() for kw in entity_kw):
        parts = own.split()
        if len(parts) >= 2:
            return f"{parts[1].title()} {parts[0].title()}"

    return ""


def resolve_entity_name(row):
    """Get the entity/LLC name, cleaned up."""
    own = s(row.get("OWN_NAME", ""))
    entity_kw = ("LLC", "INC", "CORP", "TRUST", "LP", "REALTY", "PROPERTIES",
                 "INVESTMENTS", "HOLDINGS", "GROUP", "CAPITAL", "ENTERPRISES")
    if any(kw in own.upper() for kw in entity_kw):
        return own.title()
    return ""


def ppp_status_display(row):
    """Convert internal status to client-readable language."""
    status = s(row.get("ppp_status", ""))
    months = s(row.get("ppp_estimated_months_remaining", ""))
    penalty_pct = s(row.get("ppp_current_penalty_pct", ""))
    expiry = s(row.get("ppp_estimated_expiry", ""))

    try:
        mo = float(months)
    except (ValueError, TypeError):
        mo = None

    try:
        pct = float(penalty_pct)
    except (ValueError, TypeError):
        pct = None

    if status == "hot_expired":
        return "Penalty Expired"
    elif status == "hot":
        if mo is not None and mo > 0:
            if pct is not None and pct <= 1:
                return f"~{int(mo)}mo left ({int(pct)}% remaining)"
            return f"Expiring in ~{int(mo)}mo"
        return "Expiring Soon"
    elif status == "warm":
        return f"~{int(mo)}mo out" if mo else "12-18mo Window"
    elif status == "nurture":
        return f"~{int(mo)}mo out" if mo else "18-24mo Pipeline"
    elif status == "future":
        return f"~{int(mo)}mo out" if mo else "24mo+ Pipeline"
    return ""


def build_outreach_angle(row):
    """Generate the 'what to say' talking point for the LO."""
    status = s(row.get("ppp_status", ""))
    lender = s(row.get("attom_lender_name", ""))
    loan_amount = fc(row.get("attom_loan_amount", ""))
    rate_type = s(row.get("attom_rate_type", ""))
    loan_date = s(row.get("attom_loan_date", ""))
    prop_count = s(row.get("property_count", ""))

    parts = []

    if status in ("hot_expired", "hot"):
        if status == "hot_expired":
            parts.append("Prepayment penalty has expired")
        else:
            parts.append("Prepayment penalty expiring soon")

    if lender:
        parts.append(f"current loan through {lender.title()}")

    if loan_amount:
        parts.append(f"{loan_amount} balance")

    if rate_type and "ADJUST" in rate_type.upper():
        parts.append("on an adjustable rate")

    if loan_date:
        try:
            yr = int(loan_date[:4])
            if 2022 <= yr <= 2023:
                parts.append("originated during peak rates (7-9%)")
        except (ValueError, IndexError):
            pass

    if prop_count:
        try:
            pc = int(float(prop_count))
            if pc >= 5:
                parts.append(f"portfolio of {pc} properties")
            elif pc >= 2:
                parts.append(f"{pc} investment properties")
        except (ValueError, TypeError):
            pass

    if not parts:
        return ""

    # Join into a natural sentence
    angle = parts[0].capitalize()
    if len(parts) > 1:
        angle += " — " + ", ".join(parts[1:])
    return angle + "."


# ---------------------------------------------------------------------------
# Cell writer helpers
# ---------------------------------------------------------------------------
def _cell(ws, row, col, value, font=f_body, fill=None, alignment=a_left, border=thin_border, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    if fill:
        cell.fill = fill
    cell.alignment = alignment
    cell.border = border
    if fmt:
        cell.number_format = fmt
    return cell


def _merge_header(ws, row, col_start, col_end, value, font=f_section, fill=navy_fill):
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=value)
    cell.font = font
    cell.fill = fill
    cell.alignment = a_center
    cell.border = thin_border
    # Fill merged cells
    for c in range(col_start + 1, col_end + 1):
        ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=c).border = thin_border


# ---------------------------------------------------------------------------
# TAB 1: Executive Summary
# ---------------------------------------------------------------------------
def build_summary_tab(wb, leads_df):
    ws = wb.active
    ws.title = "EXECUTIVE SUMMARY"
    ws.sheet_properties.tabColor = NAVY

    targeted = leads_df[leads_df["_include"]].copy()
    hot = targeted[targeted["ppp_status"].isin(["hot_expired", "hot"])]

    # Title bar (rows 1-2)
    for c in range(1, 12):
        ws.cell(row=1, column=c).fill = navy_fill
        ws.cell(row=2, column=c).fill = navy_fill

    ws.merge_cells("A1:K1")
    _cell(ws, 1, 1, "DSCR Prepayment Penalty Intelligence Report", f_title, navy_fill, a_left)
    ws.merge_cells("A2:K2")
    _cell(ws, 2, 1, f"Prepared {TODAY.strftime('%B %d, %Y')}  |  South Florida Market  |  Stillmind Creative",
          Font(color=TEAL, size=10, name="Calibri", italic=True), navy_fill, a_left)
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 20

    # KPI row (row 4-6)
    row = 4
    kpis = [
        (str(len(targeted)), "Targeted Investors", TEAL),
        (str(len(hot)), "Ready to Refi Now", GREEN),
        (fc(targeted["attom_loan_amount"].apply(lambda x: float(x) if s(x) else 0).sum()) if len(targeted) > 0 else "$0",
         "Total Loan Volume", NAVY),
        ("100%", "Have Phone Number", ACCENT),
    ]
    col = 1
    for value, label, color in kpis:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        _cell(ws, row, col, value, f_kpi_value, white_fill, a_center)
        _cell(ws, row + 1, col, label, f_kpi_label, white_fill, a_center)
        # Accent border under KPI
        for c2 in range(col, col + 2):
            ws.cell(row=row, column=c2).border = Border(
                top=Side("medium", color=color),
                bottom=Side("thin", color="D0D0D0"),
            )
        col += 3

    ws.row_dimensions[4].height = 40
    ws.row_dimensions[5].height = 18

    # What is this? (row 7-12)
    row = 7
    _cell(ws, row, 1, "WHAT IS THIS REPORT?", f_navy_bold_12, white_fill, a_left)
    ws.merge_cells(f"A{row}:K{row}")
    row += 1
    explanation = [
        "This report identifies real estate investors in your market with DSCR loans approaching prepayment penalty expiration — prime refinance candidates.",
        "",
        "Borrowers locked in during the 2022-2023 rate peak (7-9%) are now paying well above current market rates, and their penalty windows are closing.",
        "Each lead includes verified contact information, current lender intelligence, loan details, portfolio data, and a suggested outreach angle.",
    ]
    for line in explanation:
        ws.merge_cells(f"A{row}:K{row}")
        _cell(ws, row, 1, line, f_body_9, white_fill, a_wrap)
        ws.row_dimensions[row].height = 16
        row += 1

    # What's included (row 13+)
    row += 1
    _cell(ws, row, 1, "WHAT'S INCLUDED", f_navy_bold_12, white_fill, a_left)
    ws.merge_cells(f"A{row}:K{row}")
    row += 1

    method_items = [
        ("1.", "Investor Identification", "Investment property owners filtered from hundreds of thousands of records using proprietary scoring criteria."),
        ("2.", "Mortgage Intelligence", "Current lender, loan amount, origination date, and rate type sourced from multiple commercial and public record databases."),
        ("3.", "Penalty Window Analysis", "Prepayment penalty status estimated using lender classification and validated against federal mortgage disclosure data."),
        ("4.", "Verified Contact Info", "Phone numbers and email addresses obtained through professional skip tracing and multi-source validation."),
        ("5.", "Portfolio Context", "Property count, portfolio value, equity position, and acquisition behavior for each investor."),
        ("6.", "Outreach Angle", "Personalized talking points based on each investor's specific loan situation and refinance opportunity."),
    ]
    for num, title, desc in method_items:
        ws.merge_cells(f"A{row}:A{row}")
        ws.merge_cells(f"B{row}:C{row}")
        ws.merge_cells(f"D{row}:K{row}")
        _cell(ws, row, 1, num, f_teal_bold_10, white_fill, a_center)
        _cell(ws, row, 2, title, f_body_bold, white_fill, a_left)
        _cell(ws, row, 4, desc, f_body_9, white_fill, a_wrap)
        ws.row_dimensions[row].height = 30
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 5
    for c in range(2, 12):
        ws.column_dimensions[get_column_letter(c)].width = 14


# ---------------------------------------------------------------------------
# TAB 2: Refi Opportunities (the call sheet)
# ---------------------------------------------------------------------------
def build_call_sheet(wb, leads_df):
    ws = wb.create_sheet("REFI OPPORTUNITIES")
    ws.sheet_properties.tabColor = TEAL

    targeted = leads_df[leads_df["_include"]].copy()
    targeted["_sort"] = pd.to_numeric(targeted["ppp_refi_score"], errors="coerce").fillna(0)
    targeted = targeted.sort_values("_sort", ascending=False)

    # Column definitions — what the LO actually needs
    columns = [
        ("Contact Name", 22),
        ("Entity / LLC", 22),
        ("Phone", 18),
        ("Email", 28),
        ("Penalty Status", 18),
        ("Refi Score", 11),
        ("Current Lender", 24),
        ("Loan Amount", 14),
        ("Loan Date", 12),
        ("Rate Type", 12),
        ("Properties", 10),
        ("Portfolio Value", 15),
        ("Est. Equity", 14),
        ("Outreach Angle", 50),
        ("Mailing Address", 35),
    ]

    # Title bar
    max_col = len(columns)
    for c in range(1, max_col + 1):
        ws.cell(row=1, column=c).fill = navy_fill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    _cell(ws, 1, 1, f"REFI OPPORTUNITIES  —  {len(targeted)} Targeted Investors  |  South Florida",
          f_title, navy_fill, a_left)
    ws.row_dimensions[1].height = 32

    # Column headers
    for i, (name, width) in enumerate(columns):
        col = i + 1
        _cell(ws, 2, col, name, f_header, navy_fill, a_wrap_center)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[2].height = 28

    # Data rows
    row = 3
    for _, lead in targeted.iterrows():
        contact = resolve_contact_name(lead)
        entity = resolve_entity_name(lead)
        phone = fphone(lead.get("phone_1", ""))
        email = s(lead.get("email_1", ""))
        status_display = ppp_status_display(lead)
        score = s(lead.get("ppp_refi_score", ""))
        lender = s(lead.get("attom_lender_name", "")).title()
        loan_amt = s(lead.get("attom_loan_amount", ""))
        loan_date = s(lead.get("attom_loan_date", ""))
        rate_type = s(lead.get("attom_rate_type", "")).title()
        props = s(lead.get("property_count", ""))
        portfolio_val = s(lead.get("total_portfolio_value", ""))
        equity = s(lead.get("estimated_equity", ""))
        angle = build_outreach_angle(lead)
        mail_addr = s(lead.get("attom_mail_address", ""))

        # Determine row fill based on status
        ppp_status = s(lead.get("ppp_status", ""))
        if ppp_status in ("hot_expired",):
            row_fill = green_fill
        elif ppp_status == "hot":
            row_fill = yellow_fill
        else:
            row_fill = alt_fill if (row % 2 == 0) else white_fill

        # Status font color
        if "Expired" in status_display:
            status_font = f_green_bold
        elif "Expiring" in status_display:
            status_font = f_amber_bold
        else:
            status_font = f_body

        values = [
            (contact or entity, f_body_bold, a_left),
            (entity if contact else "", f_body_9, a_left),
            (phone, f_body, a_left),
            (email, f_teal_9, a_left),
            (status_display, status_font, a_center),
            (int(float(score)) if score else "", f_body_bold, a_center),
            (lender, f_body_9, a_left),
            (float(loan_amt) if loan_amt else "", f_body, a_right),
            (loan_date, f_body_9, a_center),
            (rate_type, f_body_9, a_center),
            (int(float(props)) if props else "", f_body, a_center),
            (float(portfolio_val) if portfolio_val else "", f_body, a_right),
            (float(equity) if equity else "", f_body, a_right),
            (angle, f_small, a_wrap),
            (mail_addr, f_small, a_left),
        ]

        for i, (val, font, align) in enumerate(values):
            col = i + 1
            cell = _cell(ws, row, col, val, font, row_fill, align)
            # Currency formatting
            if col in (8, 12, 13) and isinstance(val, (int, float)) and val:
                cell.number_format = '$#,##0'

        ws.row_dimensions[row].height = 32
        row += 1

    # Freeze panes
    ws.freeze_panes = "A3"

    # Legend row at bottom
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    _cell(ws, row, 1,
          "Green = Penalty Expired (refi now, zero penalty)    |    Yellow = Expiring Soon (<12mo)    |    "
          "Refi Score: 70+ = high priority, 40-69 = strong, <40 = developing",
          f_small, light_bg, a_left)


# ---------------------------------------------------------------------------
# TAB 3: Data Sources
# ---------------------------------------------------------------------------
def build_sources_tab(wb):
    ws = wb.create_sheet("ABOUT THIS DATA")
    ws.sheet_properties.tabColor = DARK_GRAY

    # Title
    for c in range(1, 4):
        ws.cell(row=1, column=c).fill = navy_fill
    ws.merge_cells("A1:C1")
    _cell(ws, 1, 1, "ABOUT THIS DATA", f_title, navy_fill, a_left)
    ws.row_dimensions[1].height = 30

    # Data quality section
    row = 3
    ws.merge_cells(f"A{row}:C{row}")
    _cell(ws, row, 1, "DATA QUALITY & ACCURACY", f_navy_bold_12, white_fill, a_left)
    row += 1

    quality_items = [
        ("Intelligence Layer", "What You Get"),
        ("Property Ownership", "Sourced from state tax authority records — the same data used by county assessors and title companies."),
        ("Mortgage / Lender Data", "Current lender, loan amount, origination date, and rate type sourced from recorded deed of trust instruments."),
        ("Prepayment Penalty Status", "Estimated using proprietary lender classification models validated against federal mortgage disclosure filings."),
        ("Contact Information", "Phone and email obtained through professional skip tracing with multi-source verification. All numbers are deliverable."),
        ("Entity Resolution", "LLC and trust ownership resolved to individual decision-makers using state corporate registry data."),
        ("Portfolio Analysis", "Property count, total value, equity position, and acquisition patterns derived from multi-year transaction records."),
    ]

    # Headers
    _cell(ws, row, 1, quality_items[0][0], f_header, navy_fill, a_center)
    _cell(ws, row, 2, quality_items[0][1], f_header, navy_fill, a_center)
    ws.merge_cells(f"B{row}:C{row}")
    row += 1

    for label, desc in quality_items[1:]:
        fill = alt_fill if row % 2 == 0 else white_fill
        _cell(ws, row, 1, label, f_body_bold, fill, a_left)
        ws.merge_cells(f"B{row}:C{row}")
        _cell(ws, row, 2, desc, f_body_9, fill, a_wrap)
        ws.row_dimensions[row].height = 32
        row += 1

    # Coverage & freshness
    row += 1
    ws.merge_cells(f"A{row}:C{row}")
    _cell(ws, row, 1, "COVERAGE & FRESHNESS", f_navy_bold_12, white_fill, a_left)
    row += 1

    coverage_items = [
        ("Market Coverage", "South Florida (Palm Beach + Broward counties). Additional markets available on request."),
        ("Record Volume", "Filtered from 650,000+ property records to identify qualified investment property owners."),
        ("Data Freshness", "Property and mortgage records updated quarterly. Contact information verified at time of delivery."),
        ("Penalty Accuracy", "Validated against federal mortgage disclosure data covering millions of investment property originations."),
    ]
    for label, desc in coverage_items:
        fill = alt_fill if row % 2 == 0 else white_fill
        _cell(ws, row, 1, label, f_body_bold, fill, a_left)
        ws.merge_cells(f"B{row}:C{row}")
        _cell(ws, row, 2, desc, f_body_9, fill, a_wrap)
        ws.row_dimensions[row].height = 28
        row += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 40


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=str, default=None)
    args = parser.parse_args()

    print()
    print("=" * 60)
    print("  BUILDING PPP TARGETING DELIVERABLE")
    print("=" * 60)

    # Load scored data
    ppp_file = PPP_DIR / "ppp_targeted_leads.csv"
    if not ppp_file.exists():
        print("  ERROR: Run 22_prepayment_penalty_targeting.py first.")
        return

    df = pd.read_csv(ppp_file, dtype=str, low_memory=False)
    df["_include"] = df["ppp_status"].isin(["hot_expired", "hot", "warm", "nurture", "future"])

    targeted = df[df["_include"]]
    hot = targeted[targeted["ppp_status"].isin(["hot_expired", "hot"])]

    # Validate contact coverage
    def is_blank(val):
        return pd.isna(val) or str(val).strip().lower() in ("", "nan", "none")

    has_phone = (~targeted["phone_1"].apply(is_blank)).sum()
    has_email = (~targeted["email_1"].apply(is_blank)).sum()

    print(f"  Targeted leads: {len(targeted)}")
    print(f"  Hot (expired/expiring): {len(hot)}")
    print(f"  Phone coverage: {has_phone}/{len(targeted)} ({has_phone * 100 // max(len(targeted), 1)}%)")
    print(f"  Email coverage: {has_email}/{len(targeted)} ({has_email * 100 // max(len(targeted), 1)}%)")

    # Resolve display names and validate
    print("\n  Lead roster:")
    targeted_sorted = targeted.copy()
    targeted_sorted["_sort"] = pd.to_numeric(targeted_sorted["ppp_refi_score"], errors="coerce").fillna(0)
    for _, row in targeted_sorted.sort_values("_sort", ascending=False).iterrows():
        contact = resolve_contact_name(row)
        entity = resolve_entity_name(row)
        phone = fphone(row.get("phone_1", ""))
        status = ppp_status_display(row)
        display = contact if contact else entity
        print(f"    {display:<30} {phone:<16} {status}")

    # Build workbook
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / (args.output or f"DSCR_PPP_Intelligence_{TODAY.strftime('%Y%m%d')}.xlsx")

    wb = Workbook()
    build_summary_tab(wb, df)
    build_call_sheet(wb, df)
    build_sources_tab(wb)

    wb.save(output_path)
    print(f"\n  Deliverable: {output_path}")
    print(f"  Tabs: EXECUTIVE SUMMARY, REFI OPPORTUNITIES ({len(targeted)}), DATA SOURCES")
    print()


if __name__ == "__main__":
    main()
