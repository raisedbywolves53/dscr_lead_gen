#!/usr/bin/env python3
"""
build_demo_crm.py — Build a demo CRM Excel workbook for Northside Realty.

Agent/Broker channel: NO mortgage data (RESPA compliance).
No lender names, loan amounts, interest rates.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# ── Config ──────────────────────────────────────────────────────────────────
INPUT = Path(__file__).resolve().parent.parent / "data" / "filtered" / "wake_qualified.csv"
OUTPUT = Path(__file__).resolve().parent.parent.parent / "sales" / "demo_tearsheets" / "demo_crm_northside.xlsx"

RALEIGH_CITIES = [
    "RALEIGH", "CARY", "WAKE FOREST", "APEX", "HOLLY SPRINGS",
    "FUQUAY-VARINA", "GARNER", "KNIGHTDALE", "ROLESVILLE", "WENDELL",
]

INSTITUTIONAL_NAMES = [
    "OPENDOOR", "INVITATION HOMES", "AMERICAN HOMES 4 RENT",
    "PROGRESS RESIDENTIAL", "CERBERUS", "PRETIUM", "FIRSTKEY",
    "TRICON", "AMHERST", "STARWOOD", "COLONY",
]

INSTITUTIONAL_PATTERNS = ["TRUST I", "TRUST II", "TRUST III", "TRUST IV"]

TOP_N = 100

# ── Styles ──────────────────────────────────────────────────────────────────
TEAL = PatternFill(start_color="008080", end_color="008080", fill_type="solid")
LIGHT_GRAY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(size=10)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_CENTER = Alignment(wrap_text=True, vertical="top", horizontal="center")


def is_institutional(name: str) -> bool:
    """Return True if owner name matches institutional buyer patterns."""
    upper = name.upper()
    for inst in INSTITUTIONAL_NAMES:
        if inst in upper:
            return True
    for pat in INSTITUTIONAL_PATTERNS:
        if pat in upper:
            return True
    return False


def fmt_currency(val) -> str:
    """Format a numeric string as currency."""
    try:
        v = float(val)
        if v == 0:
            return ""
        if v >= 1_000_000:
            return f"${v / 1_000_000:.1f}M"
        if v >= 1_000:
            return f"${v / 1_000:.0f}K"
        return f"${v:,.0f}"
    except (ValueError, TypeError):
        return ""


def build_lead_data(df: pd.DataFrame) -> pd.DataFrame:
    """Filter, group, and shape leads for the demo CRM tab."""

    # Filter to Raleigh-area cities
    df = df[df["prop_city"].str.upper().isin(RALEIGH_CITIES)].copy()
    print(f"  After city filter: {len(df):,} rows")

    # Exclude institutional buyers
    df = df[~df["owner_name_1"].apply(is_institutional)].copy()
    print(f"  After institutional exclusion: {len(df):,} rows")

    # Convert score to numeric for sorting
    df["icp_score_num"] = pd.to_numeric(df["icp_score"], errors="coerce").fillna(0)

    # Fill NaN in sale_date before groupby to avoid mixed str/float comparison
    df["sale_date"] = df["sale_date"].fillna("")

    # Group by owner — aggregate portfolio info
    grouped = df.groupby("owner_name_1", as_index=False).agg(
        properties=("parcel_id", "count"),
        cities=("prop_city", lambda x: ", ".join(sorted(set(x.str.title())))),
        segment=("icp_segment", "first"),
        score=("icp_score_num", "max"),
        tier=("icp_tier", "first"),
        signals=("icp_signals", "first"),
        portfolio_value=("just_value", lambda x: pd.to_numeric(x, errors="coerce").sum()),
        last_sale=("sale_date", lambda x: max((v for v in x if v), default="")),
        cash_buyer=("is_cash_buyer", lambda x: "Yes" if any(v == "True" for v in x) else "No"),
    )

    # Only multi-property investors (2+)
    grouped = grouped[grouped["properties"] >= 2].copy()
    print(f"  Multi-property investors (2+): {len(grouped):,}")

    # Sort by score descending, take top N
    grouped = grouped.sort_values("score", ascending=False).head(TOP_N).reset_index(drop=True)

    # Format columns for display
    grouped["portfolio_value"] = grouped["portfolio_value"].apply(fmt_currency)
    grouped["last_sale"] = grouped["last_sale"].fillna("")

    # Clean up signals — shorten for readability
    def clean_signals(s):
        if pd.isna(s) or s == "":
            return ""
        parts = [p.strip().replace("_", " ").title() for p in str(s).split(",")]
        return "; ".join(parts[:4])  # max 4 signals for readability

    grouped["signals"] = grouped["signals"].apply(clean_signals)

    # Rename for display
    result = grouped.rename(columns={
        "owner_name_1": "Investor Name",
        "properties": "Properties",
        "cities": "Cities",
        "segment": "Segment",
        "score": "Score",
        "tier": "Tier",
        "signals": "Key Signals",
        "portfolio_value": "Portfolio Est. Value",
        "last_sale": "Last Activity",
        "cash_buyer": "Cash Buyer",
    })

    return result[["Investor Name", "Properties", "Cities", "Segment", "Score",
                    "Tier", "Key Signals", "Portfolio Est. Value", "Last Activity", "Cash Buyer"]]


def write_lead_feed_tab(ws, data: pd.DataFrame):
    """Write the Lead Feed tab with formatting."""
    headers = list(data.columns)

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = TEAL
        cell.alignment = WRAP_CENTER
        cell.border = THIN_BORDER

    # Write data rows
    for row_idx, (_, row) in enumerate(data.iterrows(), 2):
        fill = LIGHT_GRAY if row_idx % 2 == 0 else WHITE
        for col_idx, header in enumerate(headers, 1):
            val = row[header]
            if header == "Score":
                val = int(val) if pd.notna(val) else 0
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.alignment = WRAP
            cell.border = THIN_BORDER

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-width columns (with sensible max)
    col_widths = {
        "Investor Name": 30,
        "Properties": 12,
        "Cities": 28,
        "Segment": 24,
        "Score": 8,
        "Tier": 18,
        "Key Signals": 45,
        "Portfolio Est. Value": 18,
        "Last Activity": 14,
        "Cash Buyer": 12,
    }
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header, 15)


def write_info_tab(ws):
    """Write the 'What You Get' info tab."""
    SECTION_FONT = Font(bold=True, size=13, color="008080")
    SUBSECTION_FONT = Font(bold=True, size=11)
    NORMAL_FONT = Font(size=10)
    BOLD_FONT = Font(bold=True, size=10)

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 70

    content = [
        (None, None),
        (SECTION_FONT, "DSCR Investor Intelligence — Monthly Lead Feed"),
        (None, None),
        (SUBSECTION_FONT, "What's Included"),
        (NORMAL_FONT, "  - Scored & segmented investor leads refreshed monthly"),
        (NORMAL_FONT, "  - Multi-property owner identification & grouping"),
        (NORMAL_FONT, "  - ICP scoring (0-100) across 11 investor segments"),
        (NORMAL_FONT, "  - Portfolio size, estimated value, cash buyer flags"),
        (NORMAL_FONT, "  - City & market segmentation for territory planning"),
        (NORMAL_FONT, "  - RESPA-compliant: no mortgage/lender data included"),
        (None, None),
        (SUBSECTION_FONT, "What's NOT Included (Available in Full Dossier)"),
        (NORMAL_FONT, "  - Skip-traced phone & email (validated)"),
        (NORMAL_FONT, "  - Property-level AVM & rental estimates"),
        (NORMAL_FONT, "  - Purchase history & transaction timeline"),
        (NORMAL_FONT, "  - Entity resolution (LLC → beneficial owner)"),
        (NORMAL_FONT, "  - Wealth signals (FEC donations, nonprofit boards)"),
        (NORMAL_FONT, "  - Network mapping (shared LLCs, repeat agents)"),
        (NORMAL_FONT, "  - PDF tear sheet per lead"),
        (None, None),
        (SUBSECTION_FONT, "Pricing"),
        (None, None),
        (BOLD_FONT, "  RE Agent"),
        (NORMAL_FONT, "    $300–500/mo  or  $10–12/lead"),
        (NORMAL_FONT, "    Monthly CSV feed, scored & segmented."),
        (None, None),
        (BOLD_FONT, "  RE Broker / Brokerage"),
        (NORMAL_FONT, "    $800–1,500/mo"),
        (NORMAL_FONT, "    Office-wide distribution, territory segmentation,"),
        (NORMAL_FONT, "    monthly market report, priority support."),
        (None, None),
        (BOLD_FONT, "  Full Dossier Add-On"),
        (NORMAL_FONT, "    $60–100 per lead (on top of feed)"),
        (NORMAL_FONT, "    PDF tear sheet with all enrichment layers."),
        (None, None),
        (SUBSECTION_FONT, "Data Freshness"),
        (NORMAL_FONT, "  - Property records: updated monthly from county assessor"),
        (NORMAL_FONT, "  - Scoring: recalculated each refresh cycle"),
        (NORMAL_FONT, "  - Contact enrichment: validated at time of dossier build"),
        (None, None),
        (SUBSECTION_FONT, "Coverage"),
        (NORMAL_FONT, "  - Currently available: Wake County NC, Palm Beach & Broward FL"),
        (NORMAL_FONT, "  - Deployable to any U.S. county within 2 weeks"),
        (None, None),
        (None, None),
        (Font(italic=True, size=9, color="888888"),
         "Prepared by Still Mind Creative  |  stillmindcreative.com"),
    ]

    for row_idx, (font, text) in enumerate(content, 1):
        if text is None:
            continue
        cell = ws.cell(row=row_idx, column=2, value=text)
        if font:
            cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def main():
    print(f"Loading {INPUT}...")
    df = pd.read_csv(INPUT, dtype=str)
    print(f"  Total rows: {len(df):,}")

    lead_data = build_lead_data(df)
    print(f"  Final lead count: {len(lead_data)}")

    # Build workbook
    wb = Workbook()

    # Tab 1: Lead Feed
    ws1 = wb.active
    ws1.title = "Lead Feed — Sample"
    write_lead_feed_tab(ws1, lead_data)

    # Tab 2: What You Get
    ws2 = wb.create_sheet("What You Get")
    write_info_tab(ws2)

    # Save
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"\nSaved: {OUTPUT}")
    print(f"  Tab 1: {len(lead_data)} leads")
    print(f"  Tab 2: Pricing & info sheet")


if __name__ == "__main__":
    main()
