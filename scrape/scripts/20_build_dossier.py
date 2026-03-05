"""
Step 20: Build Investor Dossier
================================

Assembles ALL intelligence sources into comprehensive investor profiles.
This is the final output — a multi-tab Excel workbook with everything
we know about each investor.

What this script does:
  1. Reads enriched contacts (script 05/10)
  2. Reads financing data (script 11) — mortgages, lenders, LTV
  3. Reads purchase history (script 12) — acquisition timeline
  4. Reads rent estimates (script 13) — portfolio NOI/DSCR
  5. Reads wealth signals (script 14) — FEC, 990, SunBiz reverse
  6. Reads network map (script 15) — co-investors, shared services
  7. Reads life events (script 16) — divorce, liens, lis pendens
  8. Merges into single investor profile per lead
  9. Calculates final opportunity score (0-100)
  10. Outputs multi-tab Excel dossier + JSON for programmatic access

Usage:
    python scripts/20_build_dossier.py
    python scripts/20_build_dossier.py --limit 25
    python scripts/20_build_dossier.py --input data/enriched/top_leads_enriched.csv
"""

import argparse
import json
from datetime import datetime
from pathlib import Path

import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not installed. Run: pip install openpyxl")

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent.parent / ".env")
except ImportError:
    pass

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
ENRICHED_DIR = PROJECT_DIR / "data" / "enriched"
FINANCING_DIR = PROJECT_DIR / "data" / "financing"
HISTORY_DIR = PROJECT_DIR / "data" / "history"
SIGNALS_DIR = PROJECT_DIR / "data" / "signals"
DOSSIER_DIR = PROJECT_DIR / "data" / "dossiers"

DEFAULT_INPUT = ENRICHED_DIR / "top_leads_enriched.csv"


# ---------------------------------------------------------------------------
# Data loading — each source is optional
# ---------------------------------------------------------------------------

def load_if_exists(path: Path, label: str) -> pd.DataFrame:
    """Load a CSV if it exists, return empty DataFrame if not."""
    if path.exists():
        df = pd.read_csv(path, dtype=str, low_memory=False)
        print(f"  Loaded {label}: {len(df)} records")
        return df
    else:
        print(f"  {label}: not available yet")
        return pd.DataFrame()


def find_best_input() -> Path:
    """Find the best available input file."""
    candidates = [
        ENRICHED_DIR / "apollo_results.csv",
        ENRICHED_DIR / "merged_enriched.csv",
        ENRICHED_DIR / "top_leads_enriched.csv",
        DEFAULT_INPUT,
    ]
    for p in candidates:
        if p.exists():
            return p
    return DEFAULT_INPUT


# ---------------------------------------------------------------------------
# Merge intelligence sources
# ---------------------------------------------------------------------------

def merge_source(master: pd.DataFrame, source: pd.DataFrame,
                 join_col: str, cols_to_add: list, prefix: str = "") -> pd.DataFrame:
    """
    Merge columns from a source DataFrame into the master.
    Only adds columns that don't already exist (avoids duplicates).
    """
    if source.empty:
        return master

    # Find matching column
    if join_col not in source.columns:
        return master

    for col in cols_to_add:
        if col not in source.columns:
            continue
        target_col = f"{prefix}{col}" if prefix else col
        if target_col in master.columns:
            # Fill blanks from source
            source_map = source.set_index(join_col)[col].to_dict()
            for idx, row in master.iterrows():
                key = str(row.get(join_col, ""))
                if key in source_map:
                    existing = str(master.at[idx, target_col])
                    if not existing or existing.upper() in ("", "NAN", "NONE"):
                        master.at[idx, target_col] = str(source_map[key])
        else:
            # Add new column
            source_map = source.set_index(join_col)[col].to_dict()
            master[target_col] = master[join_col].map(lambda x: source_map.get(str(x), ""))

    return master


# ---------------------------------------------------------------------------
# Opportunity scoring
# ---------------------------------------------------------------------------

def calculate_opportunity_score(row: pd.Series) -> int:
    """
    Final opportunity score (0-100) combining all intelligence.

    Scoring breakdown:
      Contact quality:     0-15
      Portfolio signals:   0-25
      Financing signals:   0-25
      Behavior signals:    0-15
      Wealth/network:      0-10
      Urgency signals:     0-10
    """
    score = 0

    def safe_float(val, default=0):
        try:
            return float(str(val))
        except (ValueError, TypeError):
            return default

    def safe_int(val, default=0):
        try:
            return int(float(str(val)))
        except (ValueError, TypeError):
            return default

    def is_true(val):
        return str(val).strip().lower() in ("true", "1", "yes")

    # --- Contact quality (0-15) ---
    resolved = str(row.get("resolved_person", ""))
    if resolved and resolved.upper() not in ("", "NAN", "NONE"):
        score += 3
    phone = str(row.get("phone", "")) or str(row.get("apollo_phone", ""))
    if phone and phone.upper() not in ("", "NAN", "NONE"):
        score += 5
    email = str(row.get("email", "")) or str(row.get("apollo_email", ""))
    if email and email.upper() not in ("", "NAN", "NONE"):
        score += 4
    linkedin = str(row.get("apollo_linkedin", ""))
    if linkedin and linkedin.upper() not in ("", "NAN", "NONE"):
        score += 3

    # --- Portfolio signals (0-25) ---
    prop_count = safe_int(row.get("property_count"), 1)
    if prop_count >= 10:
        score += 15
    elif prop_count >= 5:
        score += 10
    elif prop_count >= 2:
        score += 5

    portfolio_val = safe_float(row.get("total_portfolio_value"))
    if portfolio_val >= 5_000_000:
        score += 10
    elif portfolio_val >= 2_000_000:
        score += 7
    elif portfolio_val >= 500_000:
        score += 4

    # --- Financing signals (0-25) ---
    if is_true(row.get("probable_cash_buyer")):
        score += 10  # Cash buyer = big DSCR refi opportunity

    if is_true(row.get("rate_refi_candidate")):
        score += 8

    if is_true(row.get("brrrr_exit_candidate")):
        score += 10

    if is_true(row.get("equity_harvest_candidate")):
        score += 8

    # Hard money exposure (from financing data)
    hard_money = safe_int(row.get("hard_money_count"))
    if hard_money > 0:
        score += 7

    # Maturing loans
    maturing = safe_int(row.get("loans_maturing_24mo"))
    if maturing > 0:
        score += 5

    # Cap financing at 25
    financing_score = min(score - 15 - 25, 25)  # Approximate

    # --- Behavior signals (0-15) ---
    purchases_12mo = safe_int(row.get("purchases_last_12mo"))
    purchases_36mo = safe_int(row.get("purchases_last_36mo"))
    if purchases_12mo >= 3:
        score += 10
    elif purchases_12mo >= 1:
        score += 7
    elif purchases_36mo >= 3:
        score += 5

    if is_true(row.get("str_licensed")):
        score += 3

    if is_true(row.get("out_of_state")):
        score += 2

    # --- Wealth / Network (0-10) ---
    fec_total = safe_float(row.get("fec_total_donated"))
    if fec_total >= 10000:
        score += 5
    elif fec_total >= 1000:
        score += 3

    sunbiz_count = safe_int(row.get("sunbiz_entity_count"))
    if sunbiz_count >= 5:
        score += 5
    elif sunbiz_count >= 3:
        score += 3

    network_score = safe_int(row.get("network_score"))
    score += min(network_score, 5)

    # --- Urgency signals (0-10) ---
    life_urgency = safe_int(row.get("life_event_urgency_max"))
    if life_urgency >= 4:
        score += 10
    elif life_urgency >= 3:
        score += 7
    elif life_urgency >= 2:
        score += 4

    return min(score, 100)


# ---------------------------------------------------------------------------
# Excel dossier generation
# ---------------------------------------------------------------------------

def build_excel_dossier(df: pd.DataFrame, output_path: Path):
    """Build multi-tab Excel workbook with investor dossiers."""
    if not HAS_OPENPYXL:
        print("  WARNING: openpyxl not available. Skipping Excel generation.")
        return

    wb = Workbook()

    # Styles
    header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    score_high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    score_med_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    score_low_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    wrap = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def style_header(ws, headers):
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = wrap
            cell.border = thin_border
        ws.freeze_panes = "A2"

    def write_row(ws, row_idx, values):
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = wrap
            cell.border = thin_border

    def safe_str(val):
        s = str(val) if val is not None else ""
        return "" if s.upper() in ("NAN", "NONE") else s

    # =======================================================================
    # Tab 1: Investor Summary
    # =======================================================================
    ws1 = wb.active
    ws1.title = "Investor Summary"

    headers1 = [
        "Rank", "Opportunity Score", "Investor Name", "Decision Maker",
        "LLC / Entity", "Phone", "Email", "LinkedIn",
        "Properties", "Portfolio Value", "Mailing Address",
        "Primary Market", "ICP Segment", "Key Signals",
    ]
    style_header(ws1, headers1)

    df_sorted = df.sort_values("opportunity_score", ascending=False)

    for rank, (_, row) in enumerate(df_sorted.iterrows(), 1):
        opp_score = int(float(str(row.get("opportunity_score", 0))))
        phone = safe_str(row.get("phone")) or safe_str(row.get("apollo_phone"))
        email = safe_str(row.get("email")) or safe_str(row.get("apollo_email"))
        linkedin = safe_str(row.get("apollo_linkedin"))

        # Build key signals string
        signals = []
        if str(row.get("probable_cash_buyer", "")).lower() in ("true", "1"):
            signals.append("Cash Buyer")
        if str(row.get("equity_harvest_candidate", "")).lower() in ("true", "1"):
            signals.append("Equity Harvest")
        if str(row.get("rate_refi_candidate", "")).lower() in ("true", "1"):
            signals.append("Rate Refi")
        if str(row.get("brrrr_exit_candidate", "")).lower() in ("true", "1"):
            signals.append("BRRRR Exit")
        if str(row.get("str_licensed", "")).lower() in ("true", "1"):
            signals.append("STR Licensed")
        if str(row.get("out_of_state", "")).lower() in ("true", "1"):
            signals.append("Out-of-State")
        life_events = safe_str(row.get("life_event_types"))
        if life_events:
            signals.append(f"Life: {life_events}")

        values = [
            rank,
            opp_score,
            safe_str(row.get("OWN_NAME")),
            safe_str(row.get("resolved_person")),
            safe_str(row.get("OWN_NAME")) if str(row.get("is_entity", "")).lower() in ("true", "1") else "",
            phone,
            email,
            linkedin,
            safe_str(row.get("property_count")),
            safe_str(row.get("total_portfolio_value")),
            f"{safe_str(row.get('OWN_ADDR1'))}, {safe_str(row.get('OWN_CITY'))}, {safe_str(row.get('OWN_STATE_DOM'))} {safe_str(row.get('OWN_ZIPCD'))[:5]}",
            safe_str(row.get("OWN_CITY")),
            safe_str(row.get("icp_segment")) or safe_str(row.get("refi_priority")),
            ", ".join(signals),
        ]
        write_row(ws1, rank + 1, values)

        # Color-code score
        score_cell = ws1.cell(row=rank + 1, column=2)
        if opp_score >= 60:
            score_cell.fill = score_high_fill
        elif opp_score >= 40:
            score_cell.fill = score_med_fill
        else:
            score_cell.fill = score_low_fill

    # Column widths
    widths1 = [5, 10, 35, 25, 35, 15, 30, 35, 8, 15, 40, 20, 20, 40]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[ws1.cell(row=1, column=i).column_letter].width = w

    # =======================================================================
    # Tab 2: Entity & Contact Detail
    # =======================================================================
    ws2 = wb.create_sheet("Entity & Contact")

    headers2 = [
        "Owner / Entity", "Decision Maker", "Entity Officers",
        "Registered Agent", "Entity Status",
        "Phone", "Email", "LinkedIn",
        "Employer / Title", "Twitter", "Facebook",
        "Mailing Address", "City", "State", "ZIP",
    ]
    style_header(ws2, headers2)

    for rank, (_, row) in enumerate(df_sorted.iterrows(), 1):
        values = [
            safe_str(row.get("OWN_NAME")),
            safe_str(row.get("resolved_person")),
            safe_str(row.get("entity_officers")),
            safe_str(row.get("registered_agent")),
            safe_str(row.get("entity_status")),
            safe_str(row.get("phone")) or safe_str(row.get("apollo_phone")),
            safe_str(row.get("email")) or safe_str(row.get("apollo_email")),
            safe_str(row.get("apollo_linkedin")),
            f"{safe_str(row.get('apollo_title'))} @ {safe_str(row.get('apollo_employer'))}".strip(" @"),
            safe_str(row.get("apollo_twitter")),
            safe_str(row.get("apollo_facebook")),
            safe_str(row.get("OWN_ADDR1")),
            safe_str(row.get("OWN_CITY")),
            safe_str(row.get("OWN_STATE_DOM")),
            safe_str(row.get("OWN_ZIPCD"))[:5],
        ]
        write_row(ws2, rank + 1, values)

    widths2 = [35, 25, 50, 30, 12, 15, 30, 35, 30, 25, 25, 35, 20, 5, 8]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[ws2.cell(row=1, column=i).column_letter].width = w

    # =======================================================================
    # Tab 3: Portfolio & Financing
    # =======================================================================
    ws3 = wb.create_sheet("Portfolio & Financing")

    headers3 = [
        "Owner", "Properties", "Portfolio Value", "Property Types",
        "Estimated Equity", "Equity Ratio",
        "Cash Buyer?", "Rate Refi?", "BRRRR Exit?", "Equity Harvest?",
        "Est Monthly Rent", "Est Annual NOI",
        "Refi Signals", "Refi Priority",
    ]
    style_header(ws3, headers3)

    for rank, (_, row) in enumerate(df_sorted.iterrows(), 1):
        values = [
            safe_str(row.get("OWN_NAME")),
            safe_str(row.get("property_count")),
            safe_str(row.get("total_portfolio_value")),
            safe_str(row.get("property_types")),
            safe_str(row.get("estimated_equity")),
            safe_str(row.get("equity_ratio")),
            safe_str(row.get("probable_cash_buyer")),
            safe_str(row.get("rate_refi_candidate")),
            safe_str(row.get("brrrr_exit_candidate")),
            safe_str(row.get("equity_harvest_candidate")),
            safe_str(row.get("est_monthly_rent")),
            safe_str(row.get("est_annual_noi")),
            safe_str(row.get("refi_signals")),
            safe_str(row.get("refi_priority")),
        ]
        write_row(ws3, rank + 1, values)

    widths3 = [35, 8, 15, 25, 15, 10, 10, 10, 10, 10, 15, 15, 40, 15]
    for i, w in enumerate(widths3, 1):
        ws3.column_dimensions[ws3.cell(row=1, column=i).column_letter].width = w

    # =======================================================================
    # Tab 4: Wealth & Network
    # =======================================================================
    ws4 = wb.create_sheet("Wealth & Network")

    headers4 = [
        "Owner", "FEC Donations Total", "FEC Donation Count", "FEC Recipients",
        "SunBiz Entities", "Entity Count",
        "Nonprofit Orgs",
        "Co-Investors", "Connected Leads", "Shared Lenders",
        "Connection Count", "Network Score",
    ]
    style_header(ws4, headers4)

    for rank, (_, row) in enumerate(df_sorted.iterrows(), 1):
        values = [
            safe_str(row.get("OWN_NAME")),
            safe_str(row.get("fec_total_donated")),
            safe_str(row.get("fec_donation_count")),
            safe_str(row.get("fec_recipients")),
            safe_str(row.get("sunbiz_entities")),
            safe_str(row.get("sunbiz_entity_count")),
            safe_str(row.get("nonprofit_orgs_found")),
            safe_str(row.get("co_investors")),
            safe_str(row.get("connected_leads")),
            safe_str(row.get("shared_lenders")),
            safe_str(row.get("connection_count")),
            safe_str(row.get("network_score")),
        ]
        write_row(ws4, rank + 1, values)

    widths4 = [35, 15, 10, 40, 50, 10, 40, 40, 40, 30, 10, 10]
    for i, w in enumerate(widths4, 1):
        ws4.column_dimensions[ws4.cell(row=1, column=i).column_letter].width = w

    # =======================================================================
    # Tab 5: Opportunity Signals
    # =======================================================================
    ws5 = wb.create_sheet("Opportunity Signals")

    headers5 = [
        "Owner", "Opportunity Score",
        "Life Events", "Life Event Urgency", "Life Event Details",
        "Purchases Last 12mo", "Purchases Last 36mo",
        "Avg Purchase Price", "Flip Count", "Hold Count",
        "STR Licensed?", "Out of State?",
        "Recommended Approach",
    ]
    style_header(ws5, headers5)

    for rank, (_, row) in enumerate(df_sorted.iterrows(), 1):
        # Determine recommended approach
        approach = []
        if str(row.get("probable_cash_buyer", "")).lower() in ("true", "1"):
            approach.append("Cash-out refi pitch")
        if str(row.get("rate_refi_candidate", "")).lower() in ("true", "1"):
            approach.append("Rate reduction refi")
        if str(row.get("brrrr_exit_candidate", "")).lower() in ("true", "1"):
            approach.append("BRRRR exit to DSCR")
        if str(row.get("equity_harvest_candidate", "")).lower() in ("true", "1"):
            approach.append("Equity harvest / portfolio refi")
        p12 = 0
        try:
            p12 = int(float(str(row.get("purchases_last_12mo", 0))))
        except (ValueError, TypeError):
            pass
        if p12 >= 1:
            approach.append("New acquisition DSCR")
        if not approach:
            approach.append("General DSCR education")

        values = [
            safe_str(row.get("OWN_NAME")),
            safe_str(row.get("opportunity_score")),
            safe_str(row.get("life_event_types")),
            safe_str(row.get("life_event_urgency_max")),
            safe_str(row.get("life_event_details")),
            safe_str(row.get("purchases_last_12mo")),
            safe_str(row.get("purchases_last_36mo")),
            safe_str(row.get("avg_purchase_price")),
            safe_str(row.get("flip_count")),
            safe_str(row.get("hold_count")),
            safe_str(row.get("str_licensed")),
            safe_str(row.get("out_of_state")),
            " | ".join(approach),
        ]
        write_row(ws5, rank + 1, values)

        score_cell = ws5.cell(row=rank + 1, column=2)
        opp_score = int(float(str(row.get("opportunity_score", 0))))
        if opp_score >= 60:
            score_cell.fill = score_high_fill
        elif opp_score >= 40:
            score_cell.fill = score_med_fill
        else:
            score_cell.fill = score_low_fill

    widths5 = [35, 10, 20, 10, 50, 10, 10, 15, 8, 8, 10, 10, 50]
    for i, w in enumerate(widths5, 1):
        ws5.column_dimensions[ws5.cell(row=1, column=i).column_letter].width = w

    # Save
    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Build investor dossier (Step 20)")
    parser.add_argument("--input", type=str, default="",
                        help="Input CSV (auto-detects best available)")
    parser.add_argument("--limit", type=int, default=0,
                        help="Limit to top N leads by opportunity score")
    args = parser.parse_args()

    DOSSIER_DIR.mkdir(parents=True, exist_ok=True)

    # -----------------------------------------------------------------------
    # 1. Load all data sources
    # -----------------------------------------------------------------------
    print("\n  Loading intelligence sources...")
    print("  " + "-" * 50)

    # Primary leads
    if args.input:
        input_path = Path(args.input)
    else:
        input_path = find_best_input()

    if not input_path.exists():
        print(f"  ERROR: Input file not found: {input_path}")
        return

    print(f"  Primary input: {input_path}")
    df = pd.read_csv(input_path, dtype=str, low_memory=False)
    print(f"  Leads loaded: {len(df)}")

    # Apollo enrichment
    apollo = load_if_exists(ENRICHED_DIR / "apollo_results.csv", "Apollo enrichment")

    # Financing
    financing = pd.DataFrame()
    for county in ["palm_beach", "broward"]:
        f = load_if_exists(FINANCING_DIR / f"{county}_mortgages.csv", f"Financing ({county})")
        if not f.empty:
            financing = pd.concat([financing, f])

    # Purchase history
    history = load_if_exists(HISTORY_DIR / "purchase_history.csv", "Purchase history")

    # Rent estimates
    rents = load_if_exists(ENRICHED_DIR / "rent_estimates.csv", "Rent estimates")

    # Wealth signals
    wealth = load_if_exists(SIGNALS_DIR / "wealth_signals.csv", "Wealth signals")

    # Network map
    network = load_if_exists(SIGNALS_DIR / "network_map.csv", "Network map")

    # Life events
    events = load_if_exists(SIGNALS_DIR / "life_events.csv", "Life events")

    # -----------------------------------------------------------------------
    # 2. Merge all sources
    # -----------------------------------------------------------------------
    print("\n  Merging intelligence sources...")

    join_col = "OWN_NAME"

    # Apollo columns
    apollo_cols = ["apollo_email", "apollo_phone", "apollo_mobile", "apollo_linkedin",
                   "apollo_title", "apollo_employer", "apollo_twitter", "apollo_facebook"]
    df = merge_source(df, apollo, join_col, apollo_cols)

    # Purchase history columns
    history_cols = ["total_acquisitions", "total_dispositions", "purchases_last_12mo",
                    "purchases_last_36mo", "avg_purchase_price", "flip_count", "hold_count",
                    "avg_hold_period_months", "purchase_frequency_months"]
    df = merge_source(df, history, join_col, history_cols)

    # Rent estimate columns
    rent_cols = ["est_monthly_rent", "est_annual_rent", "est_noi", "rent_to_value_ratio"]
    df = merge_source(df, rents, join_col, rent_cols)

    # Wealth signal columns
    wealth_cols = ["fec_total_donated", "fec_donation_count", "fec_recipients",
                   "nonprofit_orgs_found", "sunbiz_entity_count", "sunbiz_entities",
                   "wealth_signal_score"]
    df = merge_source(df, wealth, join_col, wealth_cols)

    # Network columns
    network_cols = ["connection_count", "co_investors", "shared_lenders",
                    "connected_leads", "network_score"]
    df = merge_source(df, network, join_col, network_cols)

    # Life event columns
    event_cols = ["life_event_count", "life_event_types", "life_event_urgency_max",
                  "life_event_details", "life_event_score"]
    df = merge_source(df, events, join_col, event_cols)

    # -----------------------------------------------------------------------
    # 3. Calculate opportunity score
    # -----------------------------------------------------------------------
    print("  Calculating opportunity scores...")
    df["opportunity_score"] = df.apply(calculate_opportunity_score, axis=1)

    # Sort by score
    df = df.sort_values("opportunity_score", ascending=False)

    # Limit if requested
    if args.limit > 0:
        df = df.head(args.limit)

    # -----------------------------------------------------------------------
    # 4. Generate Excel dossier
    # -----------------------------------------------------------------------
    today = datetime.now().strftime("%Y-%m-%d")
    excel_path = DOSSIER_DIR / f"investor_dossiers_{today}.xlsx"
    print(f"\n  Building Excel dossier: {excel_path}")
    build_excel_dossier(df, excel_path)

    # -----------------------------------------------------------------------
    # 5. Save JSON for programmatic access
    # -----------------------------------------------------------------------
    json_path = DOSSIER_DIR / f"investor_dossiers_{today}.json"
    records = df.to_dict(orient="records")
    with open(json_path, "w") as f:
        json.dump(records, f, indent=2, default=str)
    print(f"  JSON output: {json_path}")

    # -----------------------------------------------------------------------
    # 6. Save merged CSV
    # -----------------------------------------------------------------------
    csv_path = DOSSIER_DIR / f"investor_dossiers_{today}.csv"
    df.to_csv(csv_path, index=False)
    print(f"  CSV output: {csv_path}")

    # -----------------------------------------------------------------------
    # Summary
    # -----------------------------------------------------------------------
    print()
    print("=" * 60)
    print("  INVESTOR DOSSIER SUMMARY")
    print("=" * 60)
    print(f"  Total profiles:     {len(df)}")

    scores = df["opportunity_score"].astype(int)
    print(f"  Score range:        {scores.min()} — {scores.max()}")
    print(f"  Score 60+ (hot):    {(scores >= 60).sum()}")
    print(f"  Score 40-59 (warm): {((scores >= 40) & (scores < 60)).sum()}")
    print(f"  Score <40 (cool):   {(scores < 40).sum()}")
    print()

    # Data completeness
    def pct(col):
        filled = df[col].fillna("").apply(lambda x: x.upper() not in ("", "NAN", "NONE")).sum()
        return f"{filled}/{len(df)} ({filled/len(df)*100:.0f}%)" if len(df) > 0 else "0/0"

    print("  DATA COMPLETENESS:")
    print(f"    Resolved person:   {pct('resolved_person')}")

    phone_col = "phone" if "phone" in df.columns else "apollo_phone"
    email_col = "email" if "email" in df.columns else "apollo_email"
    print(f"    Phone:             {pct(phone_col)}")
    print(f"    Email:             {pct(email_col)}")

    for col, label in [
        ("apollo_linkedin", "LinkedIn"),
        ("purchases_last_12mo", "Purchase history"),
        ("est_monthly_rent", "Rent estimate"),
        ("fec_total_donated", "FEC donations"),
        ("sunbiz_entity_count", "SunBiz entities"),
        ("connection_count", "Network connections"),
        ("life_event_count", "Life events"),
    ]:
        if col in df.columns:
            print(f"    {label + ':':20s} {pct(col)}")

    print()
    print(f"  Top 5 leads:")
    for rank, (_, row) in enumerate(df.head(5).iterrows(), 1):
        name = str(row.get("resolved_person", "")) or str(row.get("OWN_NAME", ""))
        score = int(float(str(row.get("opportunity_score", 0))))
        props = str(row.get("property_count", "?"))
        print(f"    {rank}. {name[:35]:35s} Score: {score:3d}  Props: {props}")
    print()
    print(f"  FILES CREATED:")
    print(f"    {excel_path}")
    print(f"    {json_path}")
    print(f"    {csv_path}")
    print()


if __name__ == "__main__":
    main()
