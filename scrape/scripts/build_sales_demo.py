"""
Build Sales Demo Package — DSCR Lead Gen
==========================================

Creates a professional demo spreadsheet to pitch loan officers.
Uses real FL data (sanitized) as proof-of-concept and NC market stats
to show what's available in their market.

4 tabs:
  1. "System Proof (FL)"     — 25 sanitized FL leads showing signal depth
  2. "Your Market (NC)"      — NC aggregate stats, tier breakdown, top zips
  3. "Sample Leads (NC)"     — 15 NC leads with signals but NO contact info
  4. "What You Get"          — Pricing, deliverables, comparison

Usage:
    python scripts/build_sales_demo.py
    python scripts/build_sales_demo.py --xlsx-only
"""

import argparse
import hashlib
import os
import sys
from pathlib import Path

import pandas as pd
import numpy as np

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
PROJECT_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_DIR / "data"
FL_INPUT = DATA_DIR / "mvp" / "pilot_500_master.csv"
NC_WAKE = DATA_DIR / "filtered" / "wake_qualified.csv"
NC_MECK = DATA_DIR / "filtered" / "mecklenburg_qualified.csv"
OUTPUT_DIR = DATA_DIR / "demo"
XLSX_OUTPUT = OUTPUT_DIR / "dscr_sales_demo.xlsx"
TOKEN_FILE = PROJECT_DIR / "google_token.json"

# Enriched showcase data (from enrich_showcase_leads.py)
SHOWCASE_ENRICHED = OUTPUT_DIR / "showcase_enriched.csv"
SHOWCASE_PROPS = OUTPUT_DIR / "showcase_properties.csv"

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",
]

# ---------------------------------------------------------------------------
# Color palette for openpyxl
# ---------------------------------------------------------------------------
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

NAVY_FILL = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
TEAL_FILL = PatternFill(start_color="009688", end_color="009688", fill_type="solid")
GREEN_FILL = PatternFill(start_color="2e7d32", end_color="2e7d32", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="e65100", end_color="e65100", fill_type="solid")
PURPLE_FILL = PatternFill(start_color="4a148c", end_color="4a148c", fill_type="solid")
LIGHT_GRAY_FILL = PatternFill(start_color="f5f5f5", end_color="f5f5f5", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="c8e6c9", end_color="c8e6c9", fill_type="solid")
LIGHT_YELLOW_FILL = PatternFill(start_color="fff9c4", end_color="fff9c4", fill_type="solid")
LIGHT_RED_FILL = PatternFill(start_color="ffcdd2", end_color="ffcdd2", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="bbdefb", end_color="bbdefb", fill_type="solid")
WHITE_FILL = PatternFill(start_color="ffffff", end_color="ffffff", fill_type="solid")

HEADER_FONT = Font(name="Arial", size=11, bold=True, color="ffffff")
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="1a237e")
SUBTITLE_FONT = Font(name="Arial", size=11, bold=True, color="333333")
BODY_FONT = Font(name="Arial", size=10, color="333333")
BOLD_FONT = Font(name="Arial", size=10, bold=True, color="333333")
METRIC_FONT = Font(name="Arial", size=24, bold=True, color="1a237e")
METRIC_LABEL_FONT = Font(name="Arial", size=10, color="666666")
LINK_FONT = Font(name="Arial", size=10, color="1565c0", underline="single")

THIN_BORDER = Border(
    left=Side(style="thin", color="e0e0e0"),
    right=Side(style="thin", color="e0e0e0"),
    top=Side(style="thin", color="e0e0e0"),
    bottom=Side(style="thin", color="e0e0e0"),
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def sanitize_name(name: str, idx: int) -> str:
    """Replace real name with anonymous label."""
    if not name or str(name).strip() == "" or str(name).lower() == "nan":
        return f"Investor {idx + 1}"
    name = str(name).strip().upper()
    # Keep entity type indicators but anonymize
    for kw in ["LLC", "INC", "CORP", "TRUST", "LP", "HOLDINGS", "INVESTMENTS",
                "PROPERTIES", "GROUP", "CAPITAL", "VENTURES", "REALTY"]:
        if kw in name:
            # Hash to get consistent anonymization
            h = hashlib.md5(name.encode()).hexdigest()[:4].upper()
            return f"{kw} Entity {h}"
    return f"Investor {idx + 1}"


def fmt_currency(val):
    """Format as currency string."""
    try:
        v = float(val)
        if v >= 1_000_000:
            return f"${v/1_000_000:.1f}M"
        elif v >= 1_000:
            return f"${v/1_000:.0f}K"
        else:
            return f"${v:,.0f}"
    except (ValueError, TypeError):
        return ""


def fmt_number(val):
    """Format with commas."""
    try:
        return f"{int(float(val)):,}"
    except (ValueError, TypeError):
        return ""


def apply_header_row(ws, row, fill, num_cols):
    """Apply header formatting to a row."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def apply_data_rows(ws, start_row, end_row, num_cols):
    """Apply alternating row formatting."""
    for row in range(start_row, end_row + 1):
        fill = LIGHT_GRAY_FILL if (row - start_row) % 2 == 0 else WHITE_FILL
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER


def write_section_header(ws, row, col, text, fill, span=3):
    """Write a colored section header spanning multiple columns.

    Merges FIRST, then writes only to the top-left cell to avoid
    Excel repair warnings about content inside merged regions.
    """
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + span - 1)
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = HEADER_FONT
    cell.fill = fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER


# ---------------------------------------------------------------------------
# Showcase leads — hand-picked for maximum completeness
# ---------------------------------------------------------------------------
SHOWCASE_LEADS = [
    "DEMIRAY HOLDINGS INC",         # 26 props, score 70, mobile+email, hard money 12%, maturity urgent, 10 buys/yr
    "MCDOUGALL LIVING TRUST",       # 12 props, score 60, mobile+email, BofA, 12 buys/yr, 100% equity
    "STEBBINS LIDIA B",             # 6 props, score 60, mobile+email, 8.3% rate, high refi, BRRRR, 8 buys/yr
    "MSNO PROPERTIES LLC",          # 5 props, score 65, mobile+email, hard money 12%, maturity urgent, 95% equity
    "JSF ENTERPRISES LLC",          # 5 props, score 65, mobile+email, hard money 12%, maturity urgent, 97% equity
]

ANON_NAMES = [
    "Apex Property Group Inc",
    "Lakeside Family Trust",
    "Lidia S. (Individual Investor)",
    "Meridian Realty Holdings LLC",
    "Coastal Equity Ventures LLC",
]


def safe_val(row, col, default=""):
    """Get a field value, returning default for empty/nan."""
    v = str(row.get(col, "")).strip()
    return default if v in ("", "nan", "None") else v


def write_dossier_field(ws, row, col, label, value, label_fill=None):
    """Write a label:value pair into the sheet."""
    lbl_cell = ws.cell(row=row, column=col, value=label)
    lbl_cell.font = BOLD_FONT
    lbl_cell.alignment = Alignment(vertical="center")
    lbl_cell.border = THIN_BORDER
    if label_fill:
        lbl_cell.fill = label_fill

    val_cell = ws.cell(row=row, column=col + 1, value=value)
    val_cell.font = BODY_FONT
    val_cell.alignment = Alignment(vertical="center", wrap_text=True)
    val_cell.border = THIN_BORDER
    return row


# ---------------------------------------------------------------------------
# Tab 1: Investor Dossiers (FL) — deep-dive showcase (UPGRADED)
# ---------------------------------------------------------------------------
def build_fl_proof(wb):
    """Build deep dossiers for 5 showcase FL investors.

    If enriched data exists (from enrich_showcase_leads.py), uses:
      - Per-property detail tables with individual values
      - All phone numbers (up to 8)
      - All emails (up to 5)
      - Enhanced talking points with specific $ figures
      - Human-readable property types
    """
    print("  Building Tab 1: Investor Dossiers (FL)...")

    df = pd.read_csv(FL_INPUT, dtype=str)

    # Load enriched data if available
    enriched = None
    props_detail = None
    if SHOWCASE_ENRICHED.exists():
        enriched = pd.read_csv(SHOWCASE_ENRICHED, dtype=str)
        print("    Using enriched showcase data")
    if SHOWCASE_PROPS.exists():
        props_detail = pd.read_csv(SHOWCASE_PROPS, dtype=str)
        print(f"    Loaded {len(props_detail)} per-property records")

    ws = wb.create_sheet("Investor Dossiers (FL)", 0)

    # Title
    ws.merge_cells("A1:F1")
    ws["A1"].value = "DSCR Lead Intelligence — Full Investor Dossiers"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="1a237e")
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:F2")
    ws["A2"].value = (
        "5 real investor profiles from Palm Beach County, FL (names anonymized). "
        "Each profile shows the full depth of intelligence our system produces — "
        "contact data, portfolio analysis, financing details, refi signals, and personalized talking points."
    )
    ws["A2"].font = BODY_FONT
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 45

    r = 4  # current row

    for idx, (real_name, anon_name) in enumerate(zip(SHOWCASE_LEADS, ANON_NAMES)):
        match = df[df["OWN_NAME"] == real_name]
        if len(match) == 0:
            print(f"    WARNING: {real_name} not found, skipping")
            continue
        lead = match.iloc[0]

        # Check for enriched version of this lead
        e_lead = None
        if enriched is not None:
            e_match = enriched[enriched["OWN_NAME"] == real_name]
            if len(e_match) > 0:
                e_lead = e_match.iloc[0]

        # --- Dossier Header ---
        write_section_header(ws, r, 1, f"INVESTOR {idx + 1}:  {anon_name}", NAVY_FILL, span=6)
        ws.row_dimensions[r].height = 30
        r += 1

        # === IDENTITY & CONTACT ===
        write_section_header(ws, r, 1, "IDENTITY & CONTACT", PatternFill(start_color="1565c0", end_color="1565c0", fill_type="solid"), span=6)
        r += 1

        # Collect all phones for display
        phones_display = []
        if e_lead is not None:
            all_phones = str(e_lead.get("all_phones", "")).strip()
            if all_phones and all_phones != "nan":
                phones_display = [p.strip() for p in all_phones.split("|") if p.strip()]
        if not phones_display:
            p1 = safe_val(lead, "phone_1", "")
            if p1:
                phones_display = [p1]

        # Collect all emails for display
        emails_display = []
        if e_lead is not None:
            all_emails = str(e_lead.get("all_emails", "")).strip()
            if all_emails and all_emails != "nan":
                emails_display = [e.strip() for e in all_emails.split("|") if e.strip()]
        if not emails_display:
            e1 = safe_val(lead, "email_1", "")
            if e1:
                emails_display = [e1]

        phone_str = phones_display[0] if phones_display else "Not yet enriched"
        email_str = emails_display[0] if emails_display else "Not yet enriched"
        phone_count_str = f" (+{len(phones_display)-1} more)" if len(phones_display) > 1 else ""
        email_count_str = f" (+{len(emails_display)-1} more)" if len(emails_display) > 1 else ""

        fields_left = [
            ("Investor Type", safe_val(lead, "_icp")),
            ("ICP Score", f'{safe_val(lead, "_score")} / 100'),
            ("Entity Status", safe_val(lead, "sunbiz_status", "N/A")),
            ("Registered Agent", safe_val(lead, "registered_agent_name", "N/A")),
            ("Mailing Address", f'{safe_val(lead, "OWN_ADDR1")}, {safe_val(lead, "OWN_CITY")}, {safe_val(lead, "OWN_STATE")} {safe_val(lead, "OWN_ZIPCD")}'),
        ]
        fields_right = [
            ("Decision Maker", safe_val(lead, "contact_name", safe_val(lead, "resolved_person", "N/A"))),
            ("Phone", f"{phone_str}{phone_count_str}"),
            ("Phone Type", safe_val(lead, "phone_1_type", "—")),
            ("Email", f"{email_str}{email_count_str}"),
            ("Absentee?", "Yes — out of state" if safe_val(lead, "out_of_state") == "True" else "Yes — in state" if safe_val(lead, "is_absentee") == "True" else "No"),
        ]

        for i, ((ll, lv), (rl, rv)) in enumerate(zip(fields_left, fields_right)):
            bg = LIGHT_BLUE_FILL if i % 2 == 0 else WHITE_FILL
            write_dossier_field(ws, r, 1, ll, lv, bg)
            ws.cell(row=r, column=2).fill = bg
            write_dossier_field(ws, r, 4, rl, rv, bg)
            ws.cell(row=r, column=5).fill = bg
            if rl == "Phone" and "Not yet" not in rv:
                ws.cell(row=r, column=5).fill = LIGHT_GREEN_FILL
            if rl == "Email" and "Not yet" not in rv:
                ws.cell(row=r, column=5).fill = LIGHT_GREEN_FILL
            r += 1

        # Show additional phones/emails if enriched
        if len(phones_display) > 1:
            write_dossier_field(ws, r, 4, "All Phones", " | ".join(phones_display), LIGHT_GREEN_FILL)
            ws.cell(row=r, column=5).fill = LIGHT_GREEN_FILL
            r += 1
        if len(emails_display) > 1:
            write_dossier_field(ws, r, 4, "All Emails", " | ".join(emails_display), LIGHT_GREEN_FILL)
            ws.cell(row=r, column=5).fill = LIGHT_GREEN_FILL
            r += 1

        # LinkedIn placeholder
        write_dossier_field(ws, r, 4, "LinkedIn", "[Manual lookup required]", WHITE_FILL)
        r += 1

        r += 1

        # === PORTFOLIO (UPGRADED with per-property detail) ===
        write_section_header(ws, r, 1, "PORTFOLIO", PatternFill(start_color="00796b", end_color="00796b", fill_type="solid"), span=6)
        r += 1

        # Use enriched counts if available
        if e_lead is not None:
            prop_count = safe_val(e_lead, "prop_count_verified", safe_val(lead, "property_count", "1"))
            total_val = safe_val(e_lead, "total_value_verified", safe_val(lead, "total_portfolio_value"))
            total_eq = safe_val(e_lead, "total_equity_verified", safe_val(lead, "est_portfolio_equity", safe_val(lead, "estimated_equity")))
            eq_pct = safe_val(e_lead, "equity_pct_verified", safe_val(lead, "est_equity_pct", safe_val(lead, "equity_ratio", "0")))
        else:
            prop_count = safe_val(lead, "property_count", "1")
            total_val = safe_val(lead, "total_portfolio_value")
            total_eq = safe_val(lead, "est_portfolio_equity", safe_val(lead, "estimated_equity"))
            eq_pct = safe_val(lead, "est_equity_pct", safe_val(lead, "equity_ratio", "0"))

        avg_val = safe_val(lead, "avg_property_value")

        portfolio_fields = [
            ("Total Properties", prop_count),
            ("Total Portfolio Value", fmt_currency(total_val)),
            ("Avg Property Value", fmt_currency(avg_val)),
            ("Property Types", safe_val(lead, "property_types")),
            ("Estimated Portfolio Equity", fmt_currency(total_eq)),
            ("Equity %", f'{float(eq_pct):.1f}%' if eq_pct and str(eq_pct) not in ("", "nan", "0") else "—"),
            ("Max Cash-Out (75% LTV)", fmt_currency(safe_val(lead, "portfolio_cashout_75", safe_val(lead, "max_cashout_75")))),
        ]
        for i, (lbl, val) in enumerate(portfolio_fields):
            bg = LIGHT_GRAY_FILL if i % 2 == 0 else WHITE_FILL
            write_dossier_field(ws, r, 1, lbl, val, bg)
            ws.cell(row=r, column=2).fill = bg
            r += 1

        # Per-property detail table (UPGRADED)
        r += 1
        ws.cell(row=r, column=1, value="Per-Property Detail").font = Font(name="Arial", size=10, bold=True, color="00796b")
        r += 1

        # Check if we have enriched per-property data
        owner_props = None
        if props_detail is not None:
            owner_props = props_detail[props_detail["owner_real"] == real_name]

        if owner_props is not None and len(owner_props) > 0:
            # UPGRADED: Full per-property table with mortgage data
            prop_headers = ["#", "Address", "Est. Value", "Type", "Lender", "Rate", "Est. Equity"]
            for i, h in enumerate(prop_headers, 1):
                ws.cell(row=r, column=i, value=h)
            apply_header_row(ws, r, TEAL_FILL, len(prop_headers))
            r += 1

            for pi, (_, prop) in enumerate(owner_props.iterrows()):
                ws.cell(row=r, column=1, value=pi + 1).font = BODY_FONT
                ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=r, column=2, value=str(prop.get("address", ""))).font = BODY_FONT
                ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
                ws.cell(row=r, column=3, value=fmt_currency(prop.get("est_value", 0))).font = BODY_FONT
                ws.cell(row=r, column=4, value=str(prop.get("property_type", ""))).font = BODY_FONT
                ws.cell(row=r, column=5, value=str(prop.get("lender", ""))).font = BODY_FONT
                rate_val = str(prop.get("est_rate", "")).strip()
                ws.cell(row=r, column=6, value=f"{rate_val}%" if rate_val and rate_val != "nan" else "—").font = BODY_FONT
                ws.cell(row=r, column=7, value=fmt_currency(prop.get("est_equity", 0))).font = BODY_FONT
                bg = LIGHT_GRAY_FILL if pi % 2 == 0 else WHITE_FILL
                for c in range(1, len(prop_headers) + 1):
                    ws.cell(row=r, column=c).fill = bg
                    ws.cell(row=r, column=c).border = THIN_BORDER
                r += 1
        else:
            # Fallback: old-style address list
            addr_headers = ["#", "Address", "Est. Value"]
            for i, h in enumerate(addr_headers, 1):
                ws.cell(row=r, column=i, value=h)
            apply_header_row(ws, r, TEAL_FILL, 3)
            r += 1

            addresses = str(lead.get("PHY_ADDR1", "")).split(" | ")
            try:
                per_prop_val = float(total_val) / max(len(addresses), 1)
            except (ValueError, TypeError):
                per_prop_val = 0

            for pi, addr in enumerate(addresses):
                addr = addr.strip()
                if not addr:
                    continue
                ws.cell(row=r, column=1, value=pi + 1).font = BODY_FONT
                ws.cell(row=r, column=1).alignment = Alignment(horizontal="center")
                ws.cell(row=r, column=2, value=f"{addr}, Palm Beach County, FL").font = BODY_FONT
                ws.cell(row=r, column=3, value=fmt_currency(per_prop_val)).font = BODY_FONT
                bg = LIGHT_GRAY_FILL if pi % 2 == 0 else WHITE_FILL
                for c in range(1, 4):
                    ws.cell(row=r, column=c).fill = bg
                    ws.cell(row=r, column=c).border = THIN_BORDER
                r += 1

        r += 1

        # === FINANCING ===
        write_section_header(ws, r, 1, "FINANCING & REFI SIGNALS", PatternFill(start_color="e65100", end_color="e65100", fill_type="solid"), span=6)
        r += 1

        fin_fields = [
            ("Current Lender", safe_val(lead, "clean_lender", "Unknown")),
            ("Lender Type", safe_val(lead, "best_lender_type", "—")),
            ("Loan Type", safe_val(lead, "est_loan_type", "—")),
            ("Est. Interest Rate", f'{safe_val(lead, "est_interest_rate", "—")}%' if safe_val(lead, "est_interest_rate") else "—"),
            ("Est. Remaining Balance", fmt_currency(safe_val(lead, "est_remaining_balance"))),
            ("Est. Maturity Date", safe_val(lead, "est_maturity_date", "—")),
            ("Months to Maturity", safe_val(lead, "est_months_to_maturity", "—")),
            ("Maturity Urgent?", "YES" if safe_val(lead, "est_maturity_urgent") == "True" else "No"),
            ("Mortgage Count (Clerk)", safe_val(lead, "clerk_all_mtg_count", "—")),
            ("Refi Priority", safe_val(lead, "refi_priority", "—")),
            ("Refi Signals", safe_val(lead, "est_refi_signals", safe_val(lead, "refi_signals", "—"))),
            ("BRRRR Exit Candidate?", "YES" if safe_val(lead, "brrrr_exit_candidate") == "True" else "No"),
            ("Equity Harvest Candidate?", "YES" if safe_val(lead, "equity_harvest_candidate") == "True" else "No"),
            ("Est. Annual Debt Service", fmt_currency(safe_val(lead, "est_annual_debt_service"))),
        ]
        for i, (lbl, val) in enumerate(fin_fields):
            bg = LIGHT_GRAY_FILL if i % 2 == 0 else WHITE_FILL
            write_dossier_field(ws, r, 1, lbl, val, bg)
            ws.cell(row=r, column=2).fill = bg
            if "YES" in str(val):
                ws.cell(row=r, column=2).fill = LIGHT_GREEN_FILL
                ws.cell(row=r, column=2).font = Font(name="Arial", size=10, bold=True, color="2e7d32")
            if lbl == "Refi Priority" and val == "High":
                ws.cell(row=r, column=2).fill = LIGHT_GREEN_FILL
                ws.cell(row=r, column=2).font = Font(name="Arial", size=10, bold=True, color="2e7d32")
            r += 1

        r += 1

        # === ACQUISITION BEHAVIOR ===
        write_section_header(ws, r, 1, "ACQUISITION BEHAVIOR", PatternFill(start_color="4a148c", end_color="4a148c", fill_type="solid"), span=6)
        r += 1

        acq_fields = [
            ("Total Acquisitions", safe_val(lead, "total_acquisitions", "—")),
            ("Purchases Last 12 Months", safe_val(lead, "purchases_last_12mo", "0")),
            ("Purchases Last 36 Months", safe_val(lead, "purchases_last_36mo", "0")),
            ("Avg Purchase Price", fmt_currency(safe_val(lead, "avg_purchase_price"))),
            ("Most Recent Purchase", safe_val(lead, "most_recent_purchase_date", safe_val(lead, "most_recent_purchase", "—"))),
            ("Most Recent Price", fmt_currency(safe_val(lead, "most_recent_purchase_price"))),
            ("Purchase Frequency", f'Every {safe_val(lead, "purchase_frequency_months", "—")} months' if safe_val(lead, "purchase_frequency_months") else "—"),
            ("Cash Purchase %", f'{safe_val(lead, "cash_purchase_pct", "0")}%'),
        ]
        for i, (lbl, val) in enumerate(acq_fields):
            bg = LIGHT_GRAY_FILL if i % 2 == 0 else WHITE_FILL
            write_dossier_field(ws, r, 1, lbl, val, bg)
            ws.cell(row=r, column=2).fill = bg
            r += 1

        r += 1

        # === RENTAL & DSCR ===
        write_section_header(ws, r, 1, "RENTAL INCOME & DSCR", PatternFill(start_color="2e7d32", end_color="2e7d32", fill_type="solid"), span=6)
        r += 1

        rent_fields = [
            ("Est. Monthly Rent", fmt_currency(safe_val(lead, "est_monthly_rent"))),
            ("Est. Annual Rent", fmt_currency(safe_val(lead, "est_annual_rent"))),
            ("Est. NOI (60% expense ratio)", fmt_currency(safe_val(lead, "est_noi"))),
            ("Est. DSCR", safe_val(lead, "est_dscr", "—")),
        ]
        for i, (lbl, val) in enumerate(rent_fields):
            bg = LIGHT_GRAY_FILL if i % 2 == 0 else WHITE_FILL
            write_dossier_field(ws, r, 1, lbl, val, bg)
            ws.cell(row=r, column=2).fill = bg
            r += 1

        r += 1

        # === TALKING POINTS (UPGRADED) ===
        write_section_header(ws, r, 1, "TALKING POINTS (AI-GENERATED)", PatternFill(start_color="616161", end_color="616161", fill_type="solid"), span=6)
        r += 1

        # Use enhanced talking points if available
        if e_lead is not None:
            tp = safe_val(e_lead, "enhanced_talking_points", "")
        else:
            tp = ""
        if not tp:
            tp = safe_val(lead, "talking_points", "No talking points generated")

        ws.merge_cells(start_row=r, start_column=1, end_row=r + 2, end_column=6)
        tp_cell = ws.cell(row=r, column=1, value=tp)
        tp_cell.font = Font(name="Arial", size=10, italic=True, color="333333")
        tp_cell.alignment = Alignment(wrap_text=True, vertical="top")
        tp_cell.border = THIN_BORDER
        ws.row_dimensions[r].height = 50
        ws.row_dimensions[r + 1].height = 50
        ws.row_dimensions[r + 2].height = 50
        r += 4

        # Separator between dossiers
        if idx < len(SHOWCASE_LEADS) - 1:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            ws.cell(row=r, column=1).fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
            ws.row_dimensions[r].height = 4
            r += 2

    # --- Aggregate stats footer ---
    r += 1
    write_section_header(ws, r, 1, "FLORIDA DEPLOYMENT RESULTS", NAVY_FILL, span=6)
    r += 1
    stats = [
        ("Total Leads Scored", "7,537"),
        ("Phone Numbers Found", "3,143 (42% hit rate)"),
        ("Emails Found", "2,592 (34% hit rate)"),
        ("Data Sources Used", "FDOR, SunBiz, County Clerk, Tracerfy, ATTOM, HUD FMR"),
        ("Total Enrichment Cost", "$71.50"),
        ("Cost Per Lead", "$0.009"),
    ]
    for i, (lbl, val) in enumerate(stats):
        bg = LIGHT_BLUE_FILL if i % 2 == 0 else WHITE_FILL
        write_dossier_field(ws, r, 1, lbl, val, bg)
        ws.cell(row=r, column=2).fill = bg
        r += 1

    # Column widths (wider to accommodate per-property table)
    widths = [28, 40, 18, 28, 40, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    print(f"    {len(SHOWCASE_LEADS)} investor dossiers written")
    return ws


# ---------------------------------------------------------------------------
# Tab 2: Your Market (NC) — aggregate stats
# ---------------------------------------------------------------------------
def build_nc_market(wb):
    """Build NC market stats tab with aggregate numbers."""
    print("  Building Tab 2: Your Market (NC)...")

    # Load both counties
    dfs = []
    for path, county in [(NC_WAKE, "Wake"), (NC_MECK, "Mecklenburg")]:
        if path.exists():
            df = pd.read_csv(path, dtype=str)
            df["_county"] = county
            dfs.append(df)
            print(f"    Loaded {county}: {len(df):,} leads")

    if not dfs:
        print("    ERROR: No NC data found. Run pipeline steps 01-03 first.")
        return None

    nc = pd.concat(dfs, ignore_index=True)

    # Compute stats
    nc["icp_score_num"] = pd.to_numeric(nc["icp_score"], errors="coerce")
    nc["just_value_num"] = pd.to_numeric(nc["just_value"], errors="coerce")
    nc["portfolio_count_num"] = pd.to_numeric(nc["portfolio_count"], errors="coerce")

    total = len(nc)
    tier1 = len(nc[nc["icp_tier"].str.contains("Hot", na=False)])
    tier2 = len(nc[nc["icp_tier"].str.contains("Warm", na=False)])
    tier3 = len(nc[nc["icp_tier"].str.contains("Nurture", na=False)])
    llc_owned = len(nc[nc["is_llc"] == "True"])
    out_of_state = len(nc[nc["is_absentee"] == "True"])
    portfolio_5 = len(nc[nc["portfolio_count_num"] >= 5])
    avg_value = nc["just_value_num"].mean()
    median_value = nc["just_value_num"].median()
    total_value = nc["just_value_num"].sum()

    # Tier breakdown by county
    wake = nc[nc["_county"] == "Wake"]
    meck = nc[nc["_county"] == "Mecklenburg"]

    # Top zips by Tier 1 count
    tier1_df = nc[nc["icp_tier"].str.contains("Hot", na=False)]
    top_zips = tier1_df.groupby("prop_zip").size().sort_values(ascending=False).head(10)

    # ICP segment breakdown
    segments = nc["icp_segment"].value_counts().head(8)

    # Also need raw + all_scored counts for the funnel
    import subprocess
    raw_total = 0
    residential_total = 0
    for raw_path, all_path in [
        (DATA_DIR / "raw" / "wake_raw.csv", DATA_DIR / "filtered" / "wake_all_scored.csv"),
        (DATA_DIR / "raw" / "mecklenburg_raw.csv", DATA_DIR / "filtered" / "mecklenburg_all_scored.csv"),
    ]:
        if raw_path.exists():
            n = int(subprocess.run(["wc", "-l", str(raw_path)], capture_output=True, text=True).stdout.split()[0]) - 1
            raw_total += n
        if all_path.exists():
            n = int(subprocess.run(["wc", "-l", str(all_path)], capture_output=True, text=True).stdout.split()[0]) - 1
            residential_total += n

    ws = wb.create_sheet("Your Market (NC)", 1)

    # Title
    ws.merge_cells("A1:L1")
    ws["A1"].value = "RALEIGH / CHARLOTTE — Investor Market Intelligence"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="1a237e")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:L2")
    ws["A2"].value = "Wake County (Raleigh) + Mecklenburg County (Charlotte) — every number below comes from public record data, not estimates"
    ws["A2"].font = BODY_FONT
    ws.row_dimensions[2].height = 25

    # =====================================================================
    # DATA SOURCE & METHODOLOGY (rows 4-25ish)
    # =====================================================================
    r = 4
    write_section_header(ws, r, 1, "DATA SOURCE & METHODOLOGY", NAVY_FILL, span=8)
    r += 1

    # Source
    src_label = ws.cell(row=r, column=1, value="Data Source")
    src_label.font = BOLD_FONT
    src_label.border = THIN_BORDER
    src_label.fill = LIGHT_BLUE_FILL
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    src_val = ws.cell(row=r, column=2, value="NC OneMap — North Carolina's official statewide parcel database, maintained by the NC Center for Geographic Information & Analysis. Free, public record.")
    src_val.font = BODY_FONT
    src_val.alignment = Alignment(wrap_text=True)
    src_val.border = THIN_BORDER
    src_val.fill = LIGHT_BLUE_FILL
    ws.row_dimensions[r].height = 35
    r += 1

    api_label = ws.cell(row=r, column=1, value="API Endpoint")
    api_label.font = BOLD_FONT
    api_label.border = THIN_BORDER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    api_val = ws.cell(row=r, column=2, value="services.nconemap.gov/secure/rest/services/NC1Map_Parcels/FeatureServer/1")
    api_val.font = Font(name="Arial", size=9, color="666666")
    api_val.border = THIN_BORDER
    r += 1

    date_label = ws.cell(row=r, column=1, value="Date Pulled")
    date_label.font = BOLD_FONT
    date_label.border = THIN_BORDER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    date_val = ws.cell(row=r, column=2, value="March 2026 — data refreshes as counties update their tax rolls")
    date_val.font = BODY_FONT
    date_val.border = THIN_BORDER
    r += 2

    # --- DATA FUNNEL ---
    write_section_header(ws, r, 1, "HOW WE GOT THESE NUMBERS", PatternFill(start_color="1565c0", end_color="1565c0", fill_type="solid"), span=8)
    r += 1

    funnel_headers = ["Step", "Description", "Records", "What Happened"]
    for i, h in enumerate(funnel_headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, PatternFill(start_color="1565c0", end_color="1565c0", fill_type="solid"), 4)
    r += 1

    discarded = residential_total - total

    funnel_rows = [
        ("1. Download", "Pull every parcel record from NC OneMap for Wake + Mecklenburg counties",
         fmt_number(raw_total), "Raw download via ArcGIS API — every parcel in both counties"),
        ("2. Filter Residential", "Remove commercial, industrial, agricultural, vacant land, government parcels",
         fmt_number(residential_total), f"Kept residential use codes only (SFR, condo, multi-family, mobile home). Removed {fmt_number(raw_total - residential_total)} non-residential"),
        ("3. Score (12 signals)", "Score each parcel against 12 investor signals using NC-specific weights",
         fmt_number(residential_total), "Every residential parcel gets scored 0-100"),
        ("4. Qualify (score ≥ 10)", "Remove parcels with zero investor signals (owner-occupied primary homes with no other flags)",
         fmt_number(total), f"Removed {fmt_number(discarded)} parcels that matched zero investor signals"),
    ]

    for step, desc, count, detail in funnel_rows:
        ws.cell(row=r, column=1, value=step).font = BOLD_FONT
        ws.cell(row=r, column=2, value=desc).font = BODY_FONT
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=3, value=count).font = Font(name="Arial", size=11, bold=True, color="1a237e")
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=4, value=detail).font = Font(name="Arial", size=9, color="666666")
        ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True)
        bg = LIGHT_GRAY_FILL if (r % 2 == 0) else WHITE_FILL
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = bg
            ws.cell(row=r, column=c).border = THIN_BORDER
        ws.row_dimensions[r].height = 40
        r += 1

    r += 1

    # --- SCORING SIGNALS EXPLAINED ---
    write_section_header(ws, r, 1, "THE 12 SCORING SIGNALS", PatternFill(start_color="4a148c", end_color="4a148c", fill_type="solid"), span=8)
    r += 1

    signal_headers = ["Signal", "Points", "What It Means", "Why It Matters for DSCR"]
    for i, h in enumerate(signal_headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, PURPLE_FILL, 4)
    r += 1

    signals = [
        ("Out-of-State Owner", "15", "Mailing address outside NC", "Non-local = almost certainly an investor, not owner-occupant"),
        ("In-State Absentee", "10", "NC address but different zip than property", "Local investor — owns the property but doesn't live there"),
        ("LLC / Corp / Trust", "10", "Entity ownership (not a person's name)", "Sophisticated investor, already structured for investment lending"),
        ("Portfolio 5+ Properties", "20", "Owner name appears on 5+ parcels in the county", "High-value target — serial investor, likely needs ongoing financing"),
        ("Portfolio 2-4 Properties", "10", "Owner name appears on 2-4 parcels", "Growing investor — building a portfolio, may need DSCR for next deal"),
        ("Value $150K-$500K", "10", "County-assessed value in DSCR sweet spot", "Most common DSCR loan size — high volume, high close rate"),
        ("Value $500K-$1M", "8", "Higher-value property", "Larger loan amounts but fewer investors qualify"),
        ("Multi-Family (2-4 units)", "10", "Duplex, triplex, or quad", "Higher rental income = better DSCR ratios, easier to qualify"),
        ("STR-Eligible Zip", "5", "Property in a known vacation/tourist zip", "Short-term rental strategy viable — different DSCR calculation"),
        ("Cash Purchase (no mtg)", "15", "Bought with cash, no recorded mortgage", "Prime candidate for DSCR cash-out refinance"),
        ("Purchased < 12 months", "10", "Recent acquisition", "Active buyer — likely needs financing for next deal"),
        ("Purchased 12-24 months", "5", "Moderately recent purchase", "Still active, may be ready for refi or next acquisition"),
    ]

    for i, (signal, pts, what, why) in enumerate(signals):
        ws.cell(row=r, column=1, value=signal).font = BOLD_FONT
        ws.cell(row=r, column=2, value=pts).font = Font(name="Arial", size=11, bold=True, color="1a237e")
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value=what).font = BODY_FONT
        ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=4, value=why).font = Font(name="Arial", size=9, color="333333")
        ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True)
        bg = LIGHT_GRAY_FILL if i % 2 == 0 else WHITE_FILL
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = bg
            ws.cell(row=r, column=c).border = THIN_BORDER
        ws.row_dimensions[r].height = 32
        r += 1

    r += 1

    # --- TIER DEFINITIONS ---
    write_section_header(ws, r, 1, "WHAT THE TIERS MEAN", PatternFill(start_color="e65100", end_color="e65100", fill_type="solid"), span=8)
    r += 1

    tier_headers = ["Tier", "Score Range", "What It Means", "Example"]
    for i, h in enumerate(tier_headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, ORANGE_FILL, 4)
    r += 1

    tiers = [
        ("Tier 1 — Hot", "40+", "Matched 3+ strong investor signals. Out-of-state LLC with a portfolio, or cash buyer with recent purchases.",
         "LLC in Texas owns 8 properties in Charlotte, all bought in last 2 years"),
        ("Tier 2 — Warm", "25-39", "Matched 2+ signals. Clear investor profile but fewer urgency indicators.",
         "In-state absentee with 3 properties valued $150K-$500K"),
        ("Tier 3 — Nurture", "10-24", "Matched 1-2 signals. Possible investor but needs more research before outreach.",
         "Single property in STR-eligible zip, no other signals"),
        ("Discarded", "0-9", "No meaningful investor signals. Likely owner-occupied primary residence.",
         "Person living in their own home, no LLC, no portfolio, no absentee flags"),
    ]

    fills = [LIGHT_GREEN_FILL, LIGHT_YELLOW_FILL, LIGHT_BLUE_FILL, LIGHT_RED_FILL]
    for i, ((tier, score_range, meaning, example), fill) in enumerate(zip(tiers, fills)):
        ws.cell(row=r, column=1, value=tier).font = BOLD_FONT
        ws.cell(row=r, column=1).fill = fill
        ws.cell(row=r, column=2, value=score_range).font = Font(name="Arial", size=11, bold=True, color="1a237e")
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value=meaning).font = BODY_FONT
        ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True)
        ws.cell(row=r, column=4, value=example).font = Font(name="Arial", size=9, italic=True, color="666666")
        ws.cell(row=r, column=4).alignment = Alignment(wrap_text=True)
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = THIN_BORDER
        ws.row_dimensions[r].height = 45
        r += 1

    r += 2

    # =====================================================================
    # NOW THE MARKET STATS (same as before, but repositioned below methodology)
    # =====================================================================

    # --- KPI Cards ---
    write_section_header(ws, r, 1, "MARKET RESULTS", NAVY_FILL, span=12)
    r += 1
    kpi_row = r
    kpis = [
        (fmt_number(total), "Total Qualified Leads"),
        (fmt_number(tier1), "Tier 1 — Hot Leads"),
        (fmt_number(tier2), "Tier 2 — Warm Leads"),
        (fmt_number(llc_owned), "LLC/Entity Owned"),
        (fmt_number(out_of_state), "Out-of-State Investors"),
        (fmt_number(portfolio_5), "Portfolio 5+ Properties"),
    ]

    for i, (value, label) in enumerate(kpis):
        col = (i * 2) + 1
        # Merge 2 columns per KPI — merge first, then write only to top-left
        ws.merge_cells(start_row=kpi_row, start_column=col, end_row=kpi_row, end_column=col + 1)
        ws.merge_cells(start_row=kpi_row + 1, start_column=col, end_row=kpi_row + 1, end_column=col + 1)
        cell_val = ws.cell(row=kpi_row, column=col, value=value)
        cell_val.font = METRIC_FONT
        cell_val.alignment = Alignment(horizontal="center")
        cell_val.fill = LIGHT_BLUE_FILL
        cell_val.border = THIN_BORDER
        cell_lbl = ws.cell(row=kpi_row + 1, column=col, value=label)
        cell_lbl.font = METRIC_LABEL_FONT
        cell_lbl.alignment = Alignment(horizontal="center")
        cell_lbl.fill = LIGHT_BLUE_FILL
        cell_lbl.border = THIN_BORDER

    ws.row_dimensions[kpi_row].height = 40
    ws.row_dimensions[kpi_row + 1].height = 22
    r = kpi_row + 3

    # --- County Breakdown ---
    write_section_header(ws, r, 1, "COUNTY BREAKDOWN", TEAL_FILL, span=6)
    r += 1
    county_headers = ["County", "Total Leads", "Tier 1 (Hot)", "Tier 2 (Warm)", "LLC-Owned", "Out-of-State"]
    for i, h in enumerate(county_headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, TEAL_FILL, len(county_headers))

    for county_name, county_df in [("Wake (Raleigh)", wake), ("Mecklenburg (Charlotte)", meck)]:
        r += 1
        ws.cell(row=r, column=1, value=county_name)
        ws.cell(row=r, column=2, value=fmt_number(len(county_df)))
        ws.cell(row=r, column=3, value=fmt_number(len(county_df[county_df["icp_tier"].str.contains("Hot", na=False)])))
        ws.cell(row=r, column=4, value=fmt_number(len(county_df[county_df["icp_tier"].str.contains("Warm", na=False)])))
        ws.cell(row=r, column=5, value=fmt_number(len(county_df[county_df["is_llc"] == "True"])))
        ws.cell(row=r, column=6, value=fmt_number(len(county_df[county_df["is_absentee"] == "True"])))
    apply_data_rows(ws, r - 1, r, len(county_headers))

    # Combined row
    r += 1
    ws.cell(row=r, column=1, value="COMBINED").font = BOLD_FONT
    ws.cell(row=r, column=2, value=fmt_number(total)).font = BOLD_FONT
    ws.cell(row=r, column=3, value=fmt_number(tier1)).font = BOLD_FONT
    ws.cell(row=r, column=4, value=fmt_number(tier2)).font = BOLD_FONT
    ws.cell(row=r, column=5, value=fmt_number(llc_owned)).font = BOLD_FONT
    ws.cell(row=r, column=6, value=fmt_number(out_of_state)).font = BOLD_FONT
    for c in range(1, 7):
        ws.cell(row=r, column=c).fill = LIGHT_GREEN_FILL
        ws.cell(row=r, column=c).border = THIN_BORDER

    # --- ICP Segments (row r+3) ---
    r += 3
    write_section_header(ws, r, 1, "INVESTOR SEGMENTS", PURPLE_FILL, span=3)
    r += 1
    ws.cell(row=r, column=1, value="Segment")
    ws.cell(row=r, column=2, value="Count")
    ws.cell(row=r, column=3, value="% of Total")
    apply_header_row(ws, r, PURPLE_FILL, 3)

    for seg_name, seg_count in segments.items():
        r += 1
        ws.cell(row=r, column=1, value=seg_name).font = BODY_FONT
        ws.cell(row=r, column=2, value=fmt_number(seg_count)).font = BODY_FONT
        ws.cell(row=r, column=3, value=f"{seg_count/total*100:.1f}%").font = BODY_FONT
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = THIN_BORDER

    # --- Top Zip Codes (next to segments) ---
    seg_start = r - len(segments)
    write_section_header(ws, seg_start - 1, 5, "TOP ZIP CODES (Tier 1 Leads)", ORANGE_FILL, span=3)
    ws.cell(row=seg_start, column=5, value="Zip Code")
    ws.cell(row=seg_start, column=6, value="Hot Leads")
    ws.cell(row=seg_start, column=7, value="% of Tier 1")
    apply_header_row(ws, seg_start, ORANGE_FILL, 3)
    # Fix: apply to correct columns
    for col in range(5, 8):
        cell = ws.cell(row=seg_start, column=col)
        cell.font = HEADER_FONT
        cell.fill = ORANGE_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    zip_row = seg_start + 1
    for zip_code, zip_count in top_zips.items():
        ws.cell(row=zip_row, column=5, value=str(zip_code)).font = BODY_FONT
        ws.cell(row=zip_row, column=6, value=fmt_number(zip_count)).font = BODY_FONT
        ws.cell(row=zip_row, column=7, value=f"{zip_count/tier1*100:.1f}%").font = BODY_FONT
        for c in range(5, 8):
            ws.cell(row=zip_row, column=c).border = THIN_BORDER
        zip_row += 1

    # --- Property Value Stats (row r+3) ---
    r += 3
    write_section_header(ws, r, 1, "PROPERTY VALUE DISTRIBUTION", GREEN_FILL, span=4)
    r += 1
    value_stats = [
        ("Average Property Value", fmt_currency(avg_value)),
        ("Median Property Value", fmt_currency(median_value)),
        ("Total Portfolio Value (all leads)", fmt_currency(total_value)),
        ("DSCR Sweet Spot ($150K–$500K)", fmt_number(len(nc[(nc["just_value_num"] >= 150000) & (nc["just_value_num"] <= 500000)]))),
        ("Premium ($500K–$1M)", fmt_number(len(nc[(nc["just_value_num"] > 500000) & (nc["just_value_num"] <= 1000000)]))),
    ]
    for metric, value in value_stats:
        ws.cell(row=r, column=1, value=metric).font = BODY_FONT
        ws.cell(row=r, column=2, value=value).font = BOLD_FONT
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    # Column widths
    col_widths = [28, 40, 16, 45, 16, 16, 16, 16, 16, 16, 16, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    print(f"    {total:,} leads analyzed, stats written")
    return ws


# ---------------------------------------------------------------------------
# Tab 3: Sample Leads (NC) — 15 leads, no contact info
# ---------------------------------------------------------------------------
def build_nc_sample(wb):
    """Build NC sample leads — property + signals, NO contact info."""
    print("  Building Tab 3: Sample Leads (NC)...")

    dfs = []
    for path, county in [(NC_WAKE, "Wake"), (NC_MECK, "Mecklenburg")]:
        if path.exists():
            df = pd.read_csv(path, dtype=str)
            df["_county"] = county
            dfs.append(df)

    if not dfs:
        return None

    nc = pd.concat(dfs, ignore_index=True)
    nc["icp_score_num"] = pd.to_numeric(nc["icp_score"], errors="coerce")

    # Pick 15 diverse Tier 1 leads — mix of segments and counties
    tier1 = nc[nc["icp_tier"].str.contains("Hot", na=False)].copy()

    selected = []
    # Get a mix from each county and segment
    for county in ["Wake", "Mecklenburg"]:
        county_df = tier1[tier1["_county"] == county]
        segments = county_df["icp_segment"].unique()
        for seg in segments:
            seg_df = county_df[county_df["icp_segment"] == seg]
            if len(seg_df) > 0 and len(selected) < 15:
                selected.append(seg_df.sample(n=1, random_state=42))
    if len(selected) < 15:
        remaining = tier1[~tier1.index.isin(pd.concat(selected).index)]
        extra = remaining.sample(n=min(15 - len(selected), len(remaining)), random_state=42)
        selected.append(extra)

    sample = pd.concat(selected).head(15).reset_index(drop=True)

    ws = wb.create_sheet("Sample Leads (NC)", 2)

    # Title
    ws.merge_cells("A1:K1")
    ws["A1"].value = "NC Sample Leads — What Your Data Looks Like"
    ws["A1"].font = TITLE_FONT
    ws.row_dimensions[1].height = 35

    ws.merge_cells("A2:K2")
    ws["A2"].value = "15 Tier 1 leads from Wake + Mecklenburg. Contact info (phone, email, LinkedIn) added after enrichment — that's what you're buying."
    ws["A2"].font = BODY_FONT
    ws.row_dimensions[2].height = 30

    # Headers
    headers = [
        "County", "Investor Type", "Score", "Property Value",
        "Use Type", "Zip Code", "LLC?", "Out-of-State?",
        "Portfolio Size", "ICP Segment", "Signals Matched"
    ]
    hr = 4
    for col, h in enumerate(headers, 1):
        ws.cell(row=hr, column=col, value=h)
    apply_header_row(ws, hr, NAVY_FILL, len(headers))

    for i, (_, row) in enumerate(sample.iterrows()):
        r = hr + 1 + i
        ws.cell(row=r, column=1, value=row.get("_county", ""))
        # Anonymize but show entity type
        name = str(row.get("owner_name_1", ""))
        ws.cell(row=r, column=2, value=sanitize_name(name, i))
        ws.cell(row=r, column=3, value=row.get("icp_score", ""))
        ws.cell(row=r, column=4, value=fmt_currency(row.get("just_value", "")))
        ws.cell(row=r, column=5, value=str(row.get("use_description", ""))[:25])
        ws.cell(row=r, column=6, value=str(row.get("prop_zip", ""))[:5])

        is_llc = str(row.get("is_llc", "")) == "True"
        is_abs = str(row.get("is_absentee", "")) == "True"
        llc_cell = ws.cell(row=r, column=7, value="Yes" if is_llc else "No")
        abs_cell = ws.cell(row=r, column=8, value="Yes" if is_abs else "No")
        if is_llc:
            llc_cell.fill = LIGHT_GREEN_FILL
        if is_abs:
            abs_cell.fill = LIGHT_GREEN_FILL

        ws.cell(row=r, column=9, value=row.get("portfolio_count", ""))
        ws.cell(row=r, column=10, value=row.get("icp_segment", ""))

        # Signals — show these to prove the intelligence
        signals = str(row.get("icp_signals", ""))
        ws.cell(row=r, column=11, value=signals if signals != "nan" else "")

    end_row = hr + len(sample)
    apply_data_rows(ws, hr + 1, end_row, len(headers))

    # Re-apply green fills
    for i, (_, row) in enumerate(sample.iterrows()):
        r = hr + 1 + i
        if str(row.get("is_llc", "")) == "True":
            ws.cell(row=r, column=7).fill = LIGHT_GREEN_FILL
        if str(row.get("is_absentee", "")) == "True":
            ws.cell(row=r, column=8).fill = LIGHT_GREEN_FILL

    # Locked row at bottom
    lr = end_row + 2
    ws.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=11)
    lock_cell = ws.cell(row=lr, column=1, value=(
        "Contact data (phone, email, LinkedIn) is added via skip trace enrichment. "
        "That's the deliverable — scored leads with verified contact info, ready for outreach."
    ))
    lock_cell.font = Font(name="Arial", size=10, bold=True, color="e65100")
    lock_cell.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[lr].height = 35

    # Column widths
    widths = [14, 22, 8, 14, 16, 10, 8, 12, 12, 24, 45]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    print(f"    {len(sample)} sample leads written")
    return ws


# ---------------------------------------------------------------------------
# Tab 4: What You Get — deliverables and pricing
# ---------------------------------------------------------------------------
def build_pricing(wb):
    """Build two-program pricing tab."""
    print("  Building Tab 4: What You Get...")

    ws = wb.create_sheet("What You Get", 3)
    SPAN = 7  # total columns used

    RED_FILL = PatternFill(start_color="b71c1c", end_color="b71c1c", fill_type="solid")
    DARK_TEAL_FILL = PatternFill(start_color="00695c", end_color="00695c", fill_type="solid")
    DARK_BLUE_FILL = PatternFill(start_color="0d47a1", end_color="0d47a1", fill_type="solid")

    # Title
    ws.merge_cells(f"A1:G1")
    ws["A1"].value = "DSCR Lead Intelligence — Two Ways to Win"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="1a237e")
    ws.row_dimensions[1].height = 40

    # =====================================================================
    # THE PROBLEM
    # =====================================================================
    r = 3
    write_section_header(ws, r, 1, "THE PROBLEM", RED_FILL, span=SPAN)
    r += 1
    problems = [
        "You're cold-calling random homeowners who don't own investment property",
        "Platforms like PropStream / BatchData cost $700+/mo and give you the same list everyone else has",
        "You have no idea which investors actually need DSCR financing right now",
        "You waste hours dialing wrong numbers and dead-end leads",
        "Even when you find investors, you don't know their current lender, equity position, or refi urgency",
    ]
    for p in problems:
        ws.cell(row=r, column=1, value="X").font = Font(name="Arial", size=10, bold=True, color="b71c1c")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=SPAN)
        ws.cell(row=r, column=2, value=p).font = BODY_FONT
        r += 1

    # =====================================================================
    # THE SOLUTION — Two Programs
    # =====================================================================
    r += 1
    write_section_header(ws, r, 1, "TWO PROGRAMS — CHOOSE YOUR LEVEL", NAVY_FILL, span=SPAN)
    r += 2

    # --- PROGRAM 1: Deal Intelligence ---
    write_section_header(ws, r, 1, "PROGRAM 1:  DEAL INTELLIGENCE PLATFORM", DARK_TEAL_FILL, span=SPAN)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=SPAN)
    p1_desc = ws.cell(row=r, column=1, value="We deliver the investor dossiers. You make the calls.")
    p1_desc.font = Font(name="Arial", size=11, italic=True, color="00695c")
    ws.row_dimensions[r].height = 25
    r += 2

    # What you get
    ws.cell(row=r, column=1, value="What's Included:").font = Font(name="Arial", size=11, bold=True, color="00695c")
    r += 1
    p1_features = [
        "Scored investor dossiers — portfolio size, property values, equity positions, LLC details",
        "Verified contact data — mobile phone numbers + validated email addresses",
        "Financing intelligence — current lender, interest rate, loan maturity, refi urgency flags",
        "Personalized talking points — AI-generated pitch script tailored to each investor's situation",
        "ICP scoring (0-100) — every lead ranked by DSCR opportunity strength",
        "Monthly data refresh — new investors, new transactions, updated scores",
        "Professional Google Sheet / CRM delivery — ready to work from day one",
    ]
    for feat in p1_features:
        ws.cell(row=r, column=1, value="-->").font = Font(name="Arial", size=10, bold=True, color="00695c")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=SPAN)
        ws.cell(row=r, column=2, value=feat).font = BODY_FONT
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 22
        r += 1

    r += 1

    # Program 1 tiers
    p1_headers = ["Tier", "Monthly", "Dossiers / Month", "Counties", "Refresh Cycle", "Best For"]
    for i, h in enumerate(p1_headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, DARK_TEAL_FILL, len(p1_headers))
    r += 1

    p1_tiers = [
        ("Starter", "$1,500/mo", "250 Tier 1", "2 counties", "Monthly", "Solo LO testing the system"),
        ("Pro", "$3,000/mo", "750 Tier 1 + Tier 2", "5 counties", "Bi-weekly", "Active LO or small team"),
        ("Enterprise", "$5,000+/mo", "Unlimited", "Full state", "Weekly", "Teams, brokerages, lender desks"),
    ]
    for tier_name, price, dossiers, counties, refresh, best_for in p1_tiers:
        ws.cell(row=r, column=1, value=tier_name).font = BOLD_FONT
        ws.cell(row=r, column=2, value=price).font = Font(name="Arial", size=11, bold=True, color="00695c")
        ws.cell(row=r, column=3, value=dossiers).font = BODY_FONT
        ws.cell(row=r, column=4, value=counties).font = BODY_FONT
        ws.cell(row=r, column=5, value=refresh).font = BODY_FONT
        ws.cell(row=r, column=6, value=best_for).font = Font(name="Arial", size=9, color="666666")
        for c in range(1, len(p1_headers) + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER
        ws.row_dimensions[r].height = 28
        r += 1

    r += 2

    # --- PROGRAM 2: Done-For-You Outbound ---
    write_section_header(ws, r, 1, "PROGRAM 2:  DONE-FOR-YOU OUTBOUND ENGINE", DARK_BLUE_FILL, span=SPAN)
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=SPAN)
    p2_desc = ws.cell(row=r, column=1, value="We run the campaigns. You show up when investors respond.")
    p2_desc.font = Font(name="Arial", size=11, italic=True, color="0d47a1")
    ws.row_dimensions[r].height = 25
    r += 2

    # What you get
    ws.cell(row=r, column=1, value="Everything in Program 1, plus:").font = Font(name="Arial", size=11, bold=True, color="0d47a1")
    r += 1
    p2_features = [
        "Multi-step email sequences — personalized outreach sent on your behalf to scored investors",
        "Direct mail campaigns — targeted postcards/letters to highest-value leads",
        "LinkedIn outreach — connection requests + messages to investor decision-makers",
        "Response management — warm handoffs when investors reply, with full context",
        "Campaign analytics — open rates, response rates, meetings booked, pipeline tracking",
        "A/B testing — subject lines, messaging angles, channel mix optimization",
        "You never touch a cold lead — only talk to investors who raised their hand",
    ]
    for feat in p2_features:
        ws.cell(row=r, column=1, value="-->").font = Font(name="Arial", size=10, bold=True, color="0d47a1")
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=SPAN)
        ws.cell(row=r, column=2, value=feat).font = BODY_FONT
        ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 22
        r += 1

    r += 1

    # Program 2 tiers
    p2_headers = ["Tier", "Monthly", "Leads / Month", "Channels", "Direct Mail", "Best For"]
    for i, h in enumerate(p2_headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, DARK_BLUE_FILL, len(p2_headers))
    r += 1

    p2_tiers = [
        ("Launch", "$3,500/mo", "500", "Email + Direct Mail", "500 pieces/mo", "LO who wants hands-off lead gen"),
        ("Growth", "$5,000/mo", "1,000", "Email + Mail + LinkedIn", "750 pieces/mo", "Growing team, multiple markets"),
        ("Scale", "$7,500+/mo", "2,000+", "Full multi-channel", "1,000+ pieces/mo", "Brokerage or lender desk"),
    ]
    for tier_name, price, leads, channels, mail, best_for in p2_tiers:
        ws.cell(row=r, column=1, value=tier_name).font = BOLD_FONT
        ws.cell(row=r, column=2, value=price).font = Font(name="Arial", size=11, bold=True, color="0d47a1")
        ws.cell(row=r, column=3, value=leads).font = BODY_FONT
        ws.cell(row=r, column=4, value=channels).font = BODY_FONT
        ws.cell(row=r, column=5, value=mail).font = BODY_FONT
        ws.cell(row=r, column=6, value=best_for).font = Font(name="Arial", size=9, color="666666")
        for c in range(1, len(p2_headers) + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER
        ws.row_dimensions[r].height = 28
        r += 1

    # =====================================================================
    # THE MATH — Why this is a no-brainer
    # =====================================================================
    r += 2
    write_section_header(ws, r, 1, "THE MATH — WHY THIS PAYS FOR ITSELF", PatternFill(start_color="1b5e20", end_color="1b5e20", fill_type="solid"), span=SPAN)
    r += 1

    math_rows = [
        ("Average DSCR loan size", "$400,000", ""),
        ("Your commission (1.5%)", "$6,000 per funded deal", ""),
        ("One repeat investor (5+ loans over time)", "$30,000+ lifetime value", ""),
        ("", "", ""),
        ("Program 1 Starter cost", "$1,500/mo", ""),
        ("Deals needed to break even", "1 deal per quarter", "$6,000 commission vs $4,500 in fees"),
        ("", "", ""),
        ("Program 2 Launch cost", "$3,500/mo", ""),
        ("Deals needed to break even", "1 deal every 2 months", "$6,000 commission vs $7,000 in fees"),
        ("After breakeven, every deal is pure upside", "", "And DSCR investors come back — they buy 3-10+ properties"),
    ]
    for label, value, note in math_rows:
        if not label and not value:
            r += 1
            continue
        lbl_cell = ws.cell(row=r, column=1, value=label)
        lbl_cell.font = BOLD_FONT if label else BODY_FONT
        lbl_cell.border = THIN_BORDER
        val_cell = ws.cell(row=r, column=2, value=value)
        val_cell.font = Font(name="Arial", size=11, bold=True, color="1b5e20") if value else BODY_FONT
        val_cell.border = THIN_BORDER
        if note:
            ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=SPAN)
            ws.cell(row=r, column=3, value=note).font = Font(name="Arial", size=9, italic=True, color="666666")
        r += 1

    # =====================================================================
    # VS. THE ALTERNATIVES
    # =====================================================================
    r += 2
    write_section_header(ws, r, 1, "VS. WHAT YOU'RE PAYING NOW", ORANGE_FILL, span=SPAN)
    r += 1

    comp_headers = ["", "Our System", "PropStream", "Zillow", "Lead Gen Agency", "DIY Cold Call"]
    for i, h in enumerate(comp_headers, 1):
        ws.cell(row=r, column=i, value=h)
    apply_header_row(ws, r, ORANGE_FILL, len(comp_headers))
    r += 1

    comparisons = [
        ("DSCR-specific scoring", "Yes", "No", "No", "No", "No"),
        ("Investor-only leads", "Yes", "Partial", "No", "Partial", "No"),
        ("Portfolio & equity intel", "Yes", "No", "No", "No", "No"),
        ("Current lender + refi signals", "Yes", "No", "No", "No", "No"),
        ("Verified mobile phone", "Included", "$$ extra", "Shared", "Shared", "Manual"),
        ("Personalized talking points", "Yes", "No", "No", "No", "No"),
        ("Done-for-you outreach", "Program 2", "No", "No", "Yes", "No"),
        ("Exclusive to you", "Yes", "No", "No", "Sometimes", "Yes"),
        ("Custom to your market", "Yes", "Generic", "Generic", "Generic", "N/A"),
        ("Monthly cost", "From $1,500", "$199+", "$1,000-4,000", "$1,500-5,000", "Your time"),
        ("Cost per funded deal", "< $1,500", "Unknown", "$3,750+", "$2,000+", "Infinite hours"),
    ]
    for row_data in comparisons:
        for i, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=i, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if i == 2 and val in ("Yes", "Included", "Program 2"):
                cell.fill = LIGHT_GREEN_FILL
                cell.font = Font(name="Arial", size=10, bold=True, color="1b5e20")
            elif val == "No":
                cell.fill = LIGHT_RED_FILL
        ws.cell(row=r, column=1).font = BOLD_FONT
        ws.row_dimensions[r].height = 22
        r += 1

    # =====================================================================
    # PILOT OFFER
    # =====================================================================
    r += 2
    write_section_header(ws, r, 1, "START HERE — $500 PILOT", NAVY_FILL, span=SPAN)
    r += 1

    pilot_lines = [
        ("What you get:", "100 Tier 1 investor dossiers in your target market — fully scored, skip-traced, contact-verified"),
        ("Delivered as:", "Professional Google Sheet with dossiers, scoring breakdown, and talking points"),
        ("Timeline:", "5 business days from kickoff"),
        ("Risk:", "Zero. If the data isn't worth 10x what you paid, walk away."),
        ("What happens next:", "If the leads work, we set you up on a monthly program. If they don't, you're out $500."),
    ]
    for label, value in pilot_lines:
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=1).fill = LIGHT_BLUE_FILL
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=SPAN)
        val_cell = ws.cell(row=r, column=2, value=value)
        val_cell.font = BODY_FONT
        val_cell.alignment = Alignment(wrap_text=True)
        val_cell.border = THIN_BORDER
        val_cell.fill = LIGHT_BLUE_FILL
        ws.row_dimensions[r].height = 28
        r += 1

    # CTA
    r += 2
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=SPAN)
    cta_cell = ws.cell(row=r, column=1, value="Ready to see what's in your market? Let's talk.")
    cta_cell.font = Font(name="Arial", size=14, bold=True, color="1a237e")
    cta_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[r].height = 45

    # Column widths
    widths = [28, 22, 18, 18, 18, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    print("    Pricing tab written")
    return ws


# ---------------------------------------------------------------------------
# Google Sheets upload
# ---------------------------------------------------------------------------
def rgb_gs(hex_color):
    """Hex → Google Sheets API RGB dict (0-1 floats)."""
    h = hex_color.lstrip('#')
    return {'red': int(h[0:2], 16) / 255, 'green': int(h[2:4], 16) / 255, 'blue': int(h[4:6], 16) / 255}


def get_google_creds():
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request

    creds = None
    if TOKEN_FILE.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            oauth_path = os.getenv("GOOGLE_OAUTH_CREDENTIALS", "")
            if not oauth_path or not Path(oauth_path).exists():
                print("  ERROR: Set GOOGLE_OAUTH_CREDENTIALS to your OAuth JSON path")
                return None
            flow = InstalledAppFlow.from_client_secrets_file(oauth_path, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, "w") as f:
            f.write(creds.to_json())
    return creds


def upload_to_sheets(xlsx_path):
    """Upload XLSX to Google Sheets with formatting."""
    try:
        from googleapiclient.discovery import build as build_service
        from googleapiclient.http import MediaFileUpload
    except ImportError:
        print("  Google API libraries not installed. Run: pip install google-api-python-client google-auth-oauthlib")
        return None

    creds = get_google_creds()
    if not creds:
        return None

    drive = build_service("drive", "v3", credentials=creds)
    sheets = build_service("sheets", "v4", credentials=creds)

    # Upload as Google Sheet
    file_metadata = {
        "name": "DSCR Lead Intelligence — Sales Demo",
        "mimeType": "application/vnd.google-apps.spreadsheet",
    }
    media = MediaFileUpload(
        str(xlsx_path),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    file = drive.files().create(body=file_metadata, media_body=media, fields="id").execute()
    sheet_id = file["id"]
    sheet_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/edit"

    print(f"  Uploaded to Google Sheets: {sheet_url}")

    # Make it viewable by anyone with link
    drive.permissions().create(
        fileId=sheet_id,
        body={"type": "anyone", "role": "reader"},
    ).execute()
    print("  Set to 'Anyone with the link can view'")

    return sheet_url


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Build DSCR Sales Demo spreadsheet")
    parser.add_argument("--xlsx-only", action="store_true", help="Build XLSX only, skip Google Sheets upload")
    args = parser.parse_args()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Check inputs
    if not FL_INPUT.exists():
        print(f"  ERROR: FL data not found: {FL_INPUT}")
        print("  Run the FL pipeline first to generate pilot data.")
        sys.exit(1)

    nc_exists = NC_WAKE.exists() or NC_MECK.exists()
    if not nc_exists:
        print(f"  ERROR: NC data not found. Run NC pipeline steps 01-03 first.")
        sys.exit(1)

    print("\n" + "=" * 60)
    print("DSCR SALES DEMO BUILDER")
    print("=" * 60)

    from openpyxl import Workbook
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_fl_proof(wb)
    build_nc_market(wb)
    build_nc_sample(wb)
    build_pricing(wb)

    # Save XLSX (initial)
    wb.save(str(XLSX_OUTPUT))

    # Post-save fixup: strip style/value attributes from cells inside merged ranges
    # directly in the XML. openpyxl leaks theme styling into these cells which causes
    # Excel to show a "found a problem with some content" repair warning.
    import zipfile, shutil, re as _re
    import xml.etree.ElementTree as ET
    from openpyxl.utils import range_boundaries
    from openpyxl.utils.cell import get_column_letter as _gcl

    tmp_path = str(XLSX_OUTPUT) + ".tmp"
    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ET.register_namespace("", ns)
    # Also register common spreadsheet namespaces to avoid ns0/ns1 prefixes
    ET.register_namespace("r", "http://schemas.openxmlformats.org/officeDocument/2006/relationships")
    ET.register_namespace("mc", "http://schemas.openxmlformats.org/markup-compatibility/2006")
    ET.register_namespace("x14ac", "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac")

    with zipfile.ZipFile(str(XLSX_OUTPUT), "r") as zin, \
         zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if _re.match(r"xl/worksheets/sheet\d+\.xml", item.filename):
                tree = ET.parse(zin.open(item.filename))
                root = tree.getroot()
                # Build set of all cell refs inside merged ranges (excluding top-left)
                merged_interior = set()
                for mc_el in root.iter(f"{{{ns}}}mergeCell"):
                    ref = mc_el.get("ref", "")
                    if ":" not in ref:
                        continue
                    min_col, min_row, max_col, max_row = range_boundaries(ref)
                    for r in range(min_row, max_row + 1):
                        for c in range(min_col, max_col + 1):
                            if r == min_row and c == min_col:
                                continue
                            merged_interior.add(f"{_gcl(c)}{r}")
                # Strip style and value from cells inside merged ranges
                for row_el in root.iter(f"{{{ns}}}row"):
                    for c_el in list(row_el):
                        ref = c_el.get("r", "")
                        if ref in merged_interior:
                            # Remove style index and any child elements (value, formula)
                            if "s" in c_el.attrib:
                                del c_el.attrib["s"]
                            if "t" in c_el.attrib:
                                del c_el.attrib["t"]
                            for child in list(c_el):
                                c_el.remove(child)
                data = ET.tostring(root, xml_declaration=True, encoding="UTF-8")
            zout.writestr(item, data)
    shutil.move(tmp_path, str(XLSX_OUTPUT))
    size_kb = XLSX_OUTPUT.stat().st_size / 1024
    print(f"\n  SAVED: {XLSX_OUTPUT.name} ({size_kb:.0f} KB)")

    # Upload to Google Sheets
    if not args.xlsx_only:
        print("\n  Uploading to Google Sheets...")
        url = upload_to_sheets(XLSX_OUTPUT)
        if url:
            print(f"\n  DEMO SHEET: {url}")
    else:
        print("  (Skipping Google Sheets upload — use without --xlsx-only to upload)")

    print("\n" + "=" * 60)
    print("DONE — Sales demo package built!")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
