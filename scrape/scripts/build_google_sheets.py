"""
Build Professional Google Sheets MVP — DSCR Lead Gen
=====================================================

4 sheets with full formatting:
  1. "Frank's Call Sheet" — Daily driver with priority badges, financing angles
  2. "Battlecards" — Full dossier with colored section headers
  3. "Performance" — Pipeline tracking with dropdowns
  4. "Dashboard" — KPIs, ICP breakdown, charts

Uses OAuth2 (desktop app) for Google Sheets API access.
First run opens browser for Google sign-in, then caches token locally.

Usage:
    python scripts/build_google_sheets.py
    python scripts/build_google_sheets.py --xlsx-only
"""

import os
import sys
import re
import argparse
from pathlib import Path

import pandas as pd
import numpy as np

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
PROJECT_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = PROJECT_DIR / "data" / "mvp"
INPUT_FILE = DATA_DIR / "pilot_500_master.csv"
XLSX_OUTPUT = DATA_DIR / "dscr_mvp_sheets.xlsx"
TOKEN_FILE = PROJECT_DIR / "google_token.json"

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",
]

# ---------------------------------------------------------------------------
# Color palette — normalized 0-1 RGB for Google Sheets API
# ---------------------------------------------------------------------------
def rgb(hex_color):
    h = hex_color.lstrip('#')
    return {'red': int(h[0:2], 16) / 255, 'green': int(h[2:4], 16) / 255, 'blue': int(h[4:6], 16) / 255}

NAVY = rgb('#1a237e')
WHITE = rgb('#ffffff')
BLACK = rgb('#000000')
LIGHT_GRAY = rgb('#f5f5f5')
MED_GRAY = rgb('#e0e0e0')
DARK_GRAY = rgb('#9e9e9e')
GREEN = rgb('#4caf50')
LIGHT_GREEN = rgb('#c8e6c9')
SCHED_GREEN = rgb('#81c784')
YELLOW = rgb('#ffeb3b')
LIGHT_YELLOW = rgb('#fff9c4')
RED = rgb('#f44336')
LIGHT_RED = rgb('#ffcdd2')
ORANGE = rgb('#ff9800')
TEAL = rgb('#009688')
PURPLE = rgb('#673ab7')
BLUE = rgb('#2196f3')
LIGHT_BLUE = rgb('#bbdefb')

# Section header colors for Battlecards
BC_IDENTITY = rgb('#1565c0')
BC_PORTFOLIO = rgb('#00796b')
BC_FINANCING = rgb('#e65100')
BC_ACQUISITION = rgb('#4a148c')
BC_REFI = rgb('#b71c1c')
BC_WEALTH = rgb('#2e7d32')
BC_OUTREACH = rgb('#616161')

# Section header colors for Performance
PF_LEAD = rgb('#546e7a')
PF_OUTREACH = rgb('#1565c0')
PF_FUNNEL = rgb('#2e7d32')
PF_LEARNING = rgb('#f9a825')


# ---------------------------------------------------------------------------
# Google Auth
# ---------------------------------------------------------------------------
def get_google_creds():
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request

    creds = None
    if TOKEN_FILE.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            oauth_path = os.getenv("GOOGLE_OAUTH_CREDENTIALS", "")
            if not oauth_path or not Path(oauth_path).exists():
                print("ERROR: Set GOOGLE_OAUTH_CREDENTIALS in .env to your OAuth client JSON")
                sys.exit(1)
            flow = InstalledAppFlow.from_client_secrets_file(oauth_path, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, "w") as f:
            f.write(creds.to_json())
    return creds


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def safe_float(val, default=None):
    try:
        v = float(val)
        return v if not np.isnan(v) else default
    except (ValueError, TypeError):
        return default

def safe_int(val, default=0):
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return default

def fmt_phone(phone):
    if not phone or pd.isna(phone):
        return ""
    digits = re.sub(r'\D', '', str(phone))
    if len(digits) == 11 and digits[0] == '1':
        digits = digits[1:]
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    return str(phone)

def sanitize_row(row):
    return ['' if v is None else v for v in row]

def check(val):
    return '\u2713' if str(val) == 'True' else ''

def county_name(co_no):
    return {'60': 'Palm Beach', '16': 'Broward'}.get(str(co_no or ''), '')

def s(val):
    """Safe string — convert NaN/None to empty string."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return ''
    v = str(val)
    return '' if v == 'nan' else v


# ---------------------------------------------------------------------------
# Financing Angle Generator
# ---------------------------------------------------------------------------
def generate_financing_angle(row):
    rate = safe_float(row.get('attom_interest_rate'))
    cash_buyer = s(row.get('probable_cash_buyer')) == 'True'
    hard_money = s(row.get('est_hard_money')) == 'True'
    equity = safe_float(row.get('estimated_equity'), 0)
    equity_ratio = safe_float(row.get('equity_ratio'), 0)
    purchases_12 = safe_int(row.get('purchases_last_12mo'))
    flip_count = safe_int(row.get('flip_count'))
    dscr = safe_float(row.get('est_dscr'))
    props = safe_int(row.get('property_count'))
    value = safe_float(row.get('total_portfolio_value'), 0)
    payment = safe_float(row.get('est_monthly_payment'))

    if rate and rate > 8:
        if payment and payment > 0:
            savings = payment * (1 - 6.25 / rate)
            return f"Currently at {rate:.1f}% - save ${savings:,.0f}/mo switching to 6.25% DSCR"
        return f"Currently at {rate:.1f}% - significant savings at 6.25% DSCR"

    if cash_buyer and equity > 500000:
        return f"Cash buyer - unlock ${equity:,.0f} equity with cash-out refi at 6.25%"
    if cash_buyer:
        return "Cash buyer - unlock equity with cash-out refi at 6.25%"

    if hard_money:
        return "Hard money loan - refi into long-term DSCR at 6.25%"

    if equity_ratio > 0.6 and equity > 0:
        return f"{equity_ratio*100:.0f}% equity - ${equity:,.0f} available for cash-out"

    if purchases_12 > 0:
        return f"Active buyer ({purchases_12} in 12mo) - needs ongoing DSCR financing"

    if flip_count > 2:
        return "Flipper profile - bridge-to-DSCR exit strategy"

    if dscr is not None and dscr >= 1.0:
        return f"DSCR-ready at {dscr:.2f}x - quick approval"

    v = f"${value/1_000_000:.1f}M" if value >= 1_000_000 else f"${value/1_000:.0f}K"
    return f"{props}-property portfolio, {v} - portfolio consolidation opportunity"


# ---------------------------------------------------------------------------
# Priority Calculator
# ---------------------------------------------------------------------------
def calculate_priority(row):
    rate = safe_float(row.get('attom_interest_rate'))
    hard_money = s(row.get('est_hard_money')) == 'True'
    cash_buyer = s(row.get('probable_cash_buyer')) == 'True'
    equity = safe_float(row.get('estimated_equity'), 0)
    refi_priority = s(row.get('refi_priority'))
    purchases_12 = safe_int(row.get('purchases_last_12mo'))

    if (rate and rate > 8) or hard_money or (cash_buyer and equity > 500000):
        return '\U0001f534 HIGH'
    if refi_priority in ('High', 'Medium') or purchases_12 > 0 or equity > 200000:
        return '\U0001f7e0 MEDIUM'
    return '\U0001f7e1 STANDARD'

PRIORITY_ORDER = {'\U0001f534 HIGH': 0, '\U0001f7e0 MEDIUM': 1, '\U0001f7e1 STANDARD': 2}


# ---------------------------------------------------------------------------
# Load data
# ---------------------------------------------------------------------------
def load_master_data():
    if not INPUT_FILE.exists():
        print(f"ERROR: {INPUT_FILE} not found")
        sys.exit(1)
    df = pd.read_csv(INPUT_FILE, dtype=str, low_memory=False)
    print(f"Loaded {len(df)} leads from {INPUT_FILE.name}")
    return df


# ---------------------------------------------------------------------------
# Sheet 1: Frank's Call Sheet
# ---------------------------------------------------------------------------
CALL_HEADERS = [
    'Priority', 'Contact Name', 'Phone', 'Entity / LLC', 'Properties',
    'Portfolio Value', 'Current Lender', 'Equity Available', 'DSCR',
    'ICP Segment', 'Financing Angle', 'Talking Points', 'Email',
    'County', 'Call Status', 'Call Date', 'Follow-Up', 'Notes',
]

def build_call_sheet(df):
    cdf = df[df['phone_1'].notna() & (df['phone_1'] != '')].copy()
    cdf['_priority'] = cdf.apply(calculate_priority, axis=1)
    cdf['_priority_sort'] = cdf['_priority'].map(PRIORITY_ORDER)
    cdf['_financing_angle'] = cdf.apply(generate_financing_angle, axis=1)
    cdf['_sort_score'] = pd.to_numeric(cdf['_score'], errors='coerce').fillna(0)
    cdf['_sort_value'] = pd.to_numeric(cdf['total_portfolio_value'], errors='coerce').fillna(0)
    cdf = cdf.sort_values(['_priority_sort', '_sort_score', '_sort_value'], ascending=[True, False, False])

    rows = []
    for _, r in cdf.iterrows():
        entity = r['OWN_NAME'] if s(r.get('is_entity')) == 'True' else ''
        dscr = safe_float(r.get('est_dscr'))
        rows.append(sanitize_row([
            r['_priority'],
            s(r.get('contact_name')),
            fmt_phone(r.get('phone_1')),
            entity,
            safe_int(r.get('property_count')),
            safe_float(r.get('total_portfolio_value'), 0),
            s(r.get('clean_lender')),
            safe_float(r.get('estimated_equity'), 0),
            dscr if dscr is not None else '',
            s(r.get('_icp')),
            r['_financing_angle'],
            s(r.get('talking_points')),
            s(r.get('email_1')),
            county_name(r.get('CO_NO')),
            '', '', '', '',  # Call Status, Call Date, Follow-Up, Notes
        ]))

    print(f"  Call Sheet: {len(rows)} callable leads")
    return rows


# ---------------------------------------------------------------------------
# Sheet 2: Battlecards
# ---------------------------------------------------------------------------
BC_SECTIONS = [
    ('IDENTITY', ['Contact Name', 'Entity', 'ICP Segment', 'Score', 'Email', 'Phone', 'LinkedIn', 'County']),
    ('PORTFOLIO OVERVIEW', ['Properties', 'Portfolio Value', 'Avg Property Value', 'Property Types', 'Equity', 'Equity %', 'Max Cashout 75%']),
    ('FINANCING INTELLIGENCE', ['Current Lender', 'Lender Type', 'Est. Balance', 'Monthly Payment', 'Interest Rate', 'DSCR', 'Monthly Rent', 'Annual Rent']),
    ('ACQUISITION BEHAVIOR', ['Total Acquisitions', 'Purchases 12mo', 'Purchases 36mo', 'Avg Purchase Price', 'Flips', 'Holds', 'Avg Hold (mo)', 'Cash Purchase %']),
    ('REFI OPPORTUNITY', ['Refi Priority', 'Refi Signals', 'Cash Buyer?', 'BRRRR Exit?', 'Rate Refi?', 'Equity Harvest?', 'Financing Angle']),
    ('WEALTH & NETWORK', ['FEC Donations', 'FEC Recipients', 'SunBiz Entities', 'Wealth Score', 'Officers', 'Registered Agent', 'Entity Status', 'STR Licensed']),
    ('TALKING POINTS & OUTREACH', ['Talking Points', 'Personalization Notes', 'Email Template Used', 'Email Sent?']),
]
BC_COLORS = [BC_IDENTITY, BC_PORTFOLIO, BC_FINANCING, BC_ACQUISITION, BC_REFI, BC_WEALTH, BC_OUTREACH]

def _bc_section_ranges():
    ranges = []
    idx = 0
    for name, cols in BC_SECTIONS:
        ranges.append((name, idx, idx + len(cols)))
        idx += len(cols)
    return ranges

BC_ALL_HEADERS = [col for _, cols in BC_SECTIONS for col in cols]
BC_SECTION_RANGES = _bc_section_ranges()
BC_NUM_COLS = len(BC_ALL_HEADERS)

def build_battlecards(df):
    df2 = df.copy()
    df2['_sort_score'] = pd.to_numeric(df2['_score'], errors='coerce').fillna(0)
    df2 = df2.sort_values('_sort_score', ascending=False)

    rows = []
    for _, r in df2.iterrows():
        dscr = safe_float(r.get('est_dscr'))
        angle = generate_financing_angle(r)
        officers = s(r.get('officer_names')) or s(r.get('entity_officers'))

        rows.append(sanitize_row([
            # IDENTITY
            s(r.get('contact_name')), s(r.get('OWN_NAME')), s(r.get('_icp')),
            safe_int(r.get('_score')), s(r.get('email_1')), fmt_phone(r.get('phone_1')),
            s(r.get('apollo_linkedin')), county_name(r.get('CO_NO')),
            # PORTFOLIO
            safe_int(r.get('property_count')),
            safe_float(r.get('total_portfolio_value'), 0),
            safe_float(r.get('avg_property_value'), 0),
            s(r.get('property_types')),
            safe_float(r.get('estimated_equity'), 0),
            safe_float(r.get('equity_ratio'), 0),
            safe_float(r.get('max_cashout_75'), 0),
            # FINANCING
            s(r.get('clean_lender')), s(r.get('best_lender_type')),
            safe_float(r.get('est_remaining_balance')) or '',
            safe_float(r.get('est_monthly_payment')) or '',
            safe_float(r.get('attom_interest_rate')) or '',
            dscr if dscr is not None else '',
            safe_float(r.get('est_monthly_rent'), 0),
            safe_float(r.get('est_annual_rent'), 0),
            # ACQUISITION
            safe_int(r.get('total_acquisitions')), safe_int(r.get('purchases_last_12mo')),
            safe_int(r.get('purchases_last_36mo')),
            safe_float(r.get('avg_purchase_price'), 0),
            safe_int(r.get('flip_count')), safe_int(r.get('hold_count')),
            safe_int(r.get('avg_hold_period_months')),
            safe_float(r.get('cash_purchase_pct'), 0),
            # REFI
            s(r.get('refi_priority')), s(r.get('refi_signals')),
            check(r.get('probable_cash_buyer')), check(r.get('brrrr_exit_candidate')),
            check(r.get('rate_refi_candidate')), check(r.get('equity_harvest_candidate')),
            angle,
            # WEALTH
            safe_float(r.get('fec_total_donated'), 0), s(r.get('fec_recipients')),
            safe_int(r.get('sunbiz_entity_count')), safe_int(r.get('wealth_signal_score')),
            officers, s(r.get('registered_agent')),
            s(r.get('sunbiz_status')), check(r.get('str_licensed')),
            # OUTREACH
            s(r.get('talking_points')), '', '', '',
        ]))

    print(f"  Battlecards: {len(rows)} leads")
    return rows


# ---------------------------------------------------------------------------
# Sheet 3: Performance
# ---------------------------------------------------------------------------
PF_SECTIONS = [
    ('LEAD INFO', ['Contact Name', 'Entity', 'ICP Segment', 'Score', 'Properties', 'Portfolio Value', 'Lender', 'County']),
    ('OUTREACH TRACKING', ['Channel', 'First Contact', 'Total Touches', 'Last Contact']),
    ('SALES FUNNEL', ['Reached?', 'Conversation?', 'Interest Level', 'Objection', 'Appointment Date',
                       'Application?', 'Loan Amount', 'Loan Type', 'Closed?', 'Close Date', 'Revenue']),
    ('LEARNING', ['What Worked', "What Didn't", 'Referral Given?', 'Notes']),
]
PF_COLORS = [PF_LEAD, PF_OUTREACH, PF_FUNNEL, PF_LEARNING]

def _pf_section_ranges():
    ranges = []
    idx = 0
    for name, cols in PF_SECTIONS:
        ranges.append((name, idx, idx + len(cols)))
        idx += len(cols)
    return ranges

PF_ALL_HEADERS = [col for _, cols in PF_SECTIONS for col in cols]
PF_SECTION_RANGES = _pf_section_ranges()
PF_NUM_COLS = len(PF_ALL_HEADERS)

def build_performance(df):
    df2 = df.copy()
    df2['_sort_score'] = pd.to_numeric(df2['_score'], errors='coerce').fillna(0)
    df2 = df2.sort_values('_sort_score', ascending=False)

    rows = []
    for _, r in df2.iterrows():
        rows.append(sanitize_row([
            s(r.get('contact_name')), s(r.get('OWN_NAME')), s(r.get('_icp')),
            safe_int(r.get('_score')), safe_int(r.get('property_count')),
            safe_float(r.get('total_portfolio_value'), 0),
            s(r.get('clean_lender')), county_name(r.get('CO_NO')),
            # OUTREACH (empty)
            '', '', '', '',
            # FUNNEL (empty)
            '', '', '', '', '', '', '', '', '', '', '',
            # LEARNING (empty)
            '', '', '', '',
        ]))

    print(f"  Performance: {len(rows)} leads")
    return rows


# ---------------------------------------------------------------------------
# Sheet 4: Dashboard data
# ---------------------------------------------------------------------------
def build_dashboard_data(df):
    """Build KPI formulas and ICP breakdown for Dashboard sheet."""
    num = len(df) + 2  # data starts at row 3

    # Row 1: KPI labels
    kpi_labels = ['Total Leads', '', 'Calls Made', '', 'Connect Rate', '',
                  'Appointments', '', 'Pipeline $', '', 'Closed $', '']
    # Row 2: KPI formulas
    kpi_formulas = [
        f"=COUNTA(Performance!A3:A{num})", '',
        f'=COUNTIF(Performance!I3:I{num},"<>")', '',
        f'=IFERROR(COUNTIF(Performance!M3:M{num},"Yes")/COUNTIF(Performance!I3:I{num},"<>"),0)', '',
        f"=COUNTA(Performance!Q3:Q{num})", '',
        f"=SUM(Performance!S3:S{num})", '',
        f"=SUM(Performance!W3:W{num})", '',
    ]

    # ICP breakdown
    icp_counts = df['_icp'].value_counts()
    icp_names = icp_counts.index.tolist()

    # Row 4: section header
    # Row 5: column headers
    # Row 6+: ICP data
    icp_header = ['ICP Segment', 'Leads', '% of Total', 'Contacted', 'Connected', 'Appointments', 'Close Rate']
    total_leads = len(df)

    icp_rows = []
    for i, icp in enumerate(icp_names):
        data_row = 6 + i  # 0-indexed row in the sheet
        icp_rows.append([
            icp,
            int(icp_counts[icp]),
            icp_counts[icp] / total_leads,
            f'=COUNTIFS(Performance!C3:C{num},A{data_row + 1},Performance!I3:I{num},"<>")',
            f'=COUNTIFS(Performance!C3:C{num},A{data_row + 1},Performance!M3:M{num},"Yes")',
            f'=COUNTIFS(Performance!C3:C{num},A{data_row + 1},Performance!Q3:Q{num},"<>")',
            f'=IFERROR(COUNTIFS(Performance!C3:C{num},A{data_row + 1},Performance!U3:U{num},"Yes")/COUNTIF(Performance!C3:C{num},A{data_row + 1}),0)',
        ])

    # County breakdown
    county_header_row = 6 + len(icp_names) + 2  # row index (1-based in sheet)
    county_rows = []
    for county_label in ['Palm Beach', 'Broward']:
        r = county_header_row + 2 + len(county_rows)
        county_rows.append([
            county_label,
            f'=COUNTIF(Performance!H3:H{num},A{r + 1})',
            f'=COUNTIFS(Performance!H3:H{num},A{r + 1},Performance!I3:I{num},"<>")',
            f'=IFERROR(COUNTIFS(Performance!H3:H{num},A{r + 1},Performance!U3:U{num},"Yes")/COUNTIF(Performance!H3:H{num},A{r + 1}),0)',
        ])

    # Objections breakdown
    obj_header_row = county_header_row + 2 + len(county_rows) + 2
    objections = ['Happy with lender', 'Not buying', 'Bad timing', 'Rate too high', "Don't qualify", 'Other']
    obj_header = ['Objection', 'Count', '% of Objections']
    obj_rows = []
    for obj in objections:
        r = obj_header_row + 2 + len(obj_rows)
        obj_rows.append([
            obj,
            f'=COUNTIF(Performance!P3:P{num},A{r + 1})',
            f'=IFERROR(COUNTIF(Performance!P3:P{num},A{r + 1})/COUNTIF(Performance!P3:P{num},"<>"),0)',
        ])

    return {
        'kpi_labels': kpi_labels,
        'kpi_formulas': kpi_formulas,
        'icp_header': icp_header,
        'icp_rows': icp_rows,
        'icp_names': icp_names,
        'county_header_row': county_header_row,
        'county_header': ['County', 'Leads', 'Contacted', 'Close Rate'],
        'county_rows': county_rows,
        'obj_header_row': obj_header_row,
        'obj_header': obj_header,
        'obj_rows': obj_rows,
    }


# ---------------------------------------------------------------------------
# Formatting: Call Sheet
# ---------------------------------------------------------------------------
def fmt_call_sheet(sid, num_rows):
    """Return batchUpdate requests for Call Sheet formatting."""
    nr = num_rows + 1  # total rows including header
    nc = len(CALL_HEADERS)
    reqs = []

    # Header: navy bg, white bold 13pt, centered
    reqs.append({'repeatCell': {
        'range': _r(sid, 0, 1, 0, nc),
        'cell': {'userEnteredFormat': {
            'backgroundColor': NAVY,
            'textFormat': {'bold': True, 'fontSize': 13, 'foregroundColor': WHITE, 'fontFamily': 'Arial'},
            'horizontalAlignment': 'CENTER', 'verticalAlignment': 'MIDDLE',
        }},
        'fields': 'userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment)',
    }})

    # Freeze row 1 + cols A-C, tab color green
    reqs.append({'updateSheetProperties': {'properties': {
        'sheetId': sid,
        'gridProperties': {'frozenRowCount': 1, 'frozenColumnCount': 3},
        'tabColor': GREEN,
    }, 'fields': 'gridProperties.frozenRowCount,gridProperties.frozenColumnCount,tabColor'}})

    # Alternating row colors via conditional formatting
    reqs.append({'addConditionalFormatRule': {'rule': {
        'ranges': [_r(sid, 1, nr, 0, nc)],
        'booleanRule': {
            'condition': {'type': 'CUSTOM_FORMULA',
                          'values': [{'userEnteredValue': '=ISEVEN(ROW())'}]},
            'format': {'backgroundColor': LIGHT_GRAY},
        }}, 'index': 50}})

    # Row height 42px for data
    reqs.append({'updateDimensionProperties': {
        'range': {'sheetId': sid, 'dimension': 'ROWS', 'startIndex': 1, 'endIndex': nr},
        'properties': {'pixelSize': 42}, 'fields': 'pixelSize',
    }})
    # Header row height 48px
    reqs.append({'updateDimensionProperties': {
        'range': {'sheetId': sid, 'dimension': 'ROWS', 'startIndex': 0, 'endIndex': 1},
        'properties': {'pixelSize': 48}, 'fields': 'pixelSize',
    }})

    # Column widths — generous sizing so all text is visible
    # A=Priority, B=Name, C=Phone, D=Entity, E=Props, F=Value, G=Lender,
    # H=Equity, I=DSCR, J=ICP, K=Financing Angle, L=Talking Points,
    # M=Email, N=County, O=Status, P=Call Date, Q=Follow-Up, R=Notes
    widths = [110, 200, 150, 240, 55, 120, 240, 120, 70, 200, 340, 380, 230, 110, 130, 110, 110, 280]
    for i, w in enumerate(widths):
        reqs.append({'updateDimensionProperties': {
            'range': {'sheetId': sid, 'dimension': 'COLUMNS', 'startIndex': i, 'endIndex': i + 1},
            'properties': {'pixelSize': w}, 'fields': 'pixelSize',
        }})

    # Bold + 12pt Contact Name (col 1)
    reqs.append({'repeatCell': {
        'range': _r(sid, 1, nr, 1, 2),
        'cell': {'userEnteredFormat': {'textFormat': {'bold': True, 'fontSize': 12}}},
        'fields': 'userEnteredFormat.textFormat(bold,fontSize)',
    }})

    # Wrap text on ALL data cells so nothing clips
    reqs.append({'repeatCell': {
        'range': _r(sid, 1, nr, 0, nc),
        'cell': {'userEnteredFormat': {'wrapStrategy': 'WRAP'}},
        'fields': 'userEnteredFormat.wrapStrategy',
    }})

    # Currency format: Portfolio Value (5), Equity (7)
    for col in [5, 7]:
        reqs.append({'repeatCell': {
            'range': _r(sid, 1, nr, col, col + 1),
            'cell': {'userEnteredFormat': {'numberFormat': {'type': 'CURRENCY', 'pattern': '$#,##0'}}},
            'fields': 'userEnteredFormat.numberFormat',
        }})

    # DSCR format (col 8): 0.00"x"
    reqs.append({'repeatCell': {
        'range': _r(sid, 1, nr, 8, 9),
        'cell': {'userEnteredFormat': {'numberFormat': {'type': 'NUMBER', 'pattern': '0.00"x"'}}},
        'fields': 'userEnteredFormat.numberFormat',
    }})

    # Right-align numeric cols: Properties (4), Value (5), Equity (7), DSCR (8)
    for col in [4, 5, 7, 8]:
        reqs.append({'repeatCell': {
            'range': _r(sid, 1, nr, col, col + 1),
            'cell': {'userEnteredFormat': {'horizontalAlignment': 'RIGHT'}},
            'fields': 'userEnteredFormat.horizontalAlignment',
        }})

    # Date format: Call Date (15), Follow-Up (16)
    for col in [15, 16]:
        reqs.append({'repeatCell': {
            'range': _r(sid, 1, nr, col, col + 1),
            'cell': {'userEnteredFormat': {'numberFormat': {'type': 'DATE', 'pattern': 'MMM dd, yyyy'}}},
            'fields': 'userEnteredFormat.numberFormat',
        }})

    # Dropdown: Call Status (col 14)
    statuses = ['Not Called', 'VM Left', 'Connected', 'Interested', 'Scheduled', 'Not Qualified', 'DNC']
    reqs.append({'setDataValidation': {
        'range': _r(sid, 1, nr, 14, 15),
        'rule': {'condition': {'type': 'ONE_OF_LIST',
                 'values': [{'userEnteredValue': v} for v in statuses]},
                 'showCustomUi': True, 'strict': False},
    }})

    # Conditional formatting: Equity (col 7)
    for val, op, color, idx in [
        ('200000', 'NUMBER_GREATER_THAN_EQ', LIGHT_GREEN, 0),
        ('0', 'NUMBER_LESS', LIGHT_RED, 2),
    ]:
        reqs.append({'addConditionalFormatRule': {'rule': {
            'ranges': [_r(sid, 1, nr, 7, 8)],
            'booleanRule': {
                'condition': {'type': op, 'values': [{'userEnteredValue': val}]},
                'format': {'backgroundColor': color},
            }}, 'index': idx}})
    # Yellow between 50K-200K
    reqs.append({'addConditionalFormatRule': {'rule': {
        'ranges': [_r(sid, 1, nr, 7, 8)],
        'booleanRule': {
            'condition': {'type': 'CUSTOM_FORMULA',
                          'values': [{'userEnteredValue': f'=AND(H2>=50000,H2<200000)'}]},
            'format': {'backgroundColor': LIGHT_YELLOW},
        }}, 'index': 1}})

    # Conditional formatting: DSCR (col 8)
    reqs.append({'addConditionalFormatRule': {'rule': {
        'ranges': [_r(sid, 1, nr, 8, 9)],
        'booleanRule': {
            'condition': {'type': 'NUMBER_GREATER_THAN_EQ', 'values': [{'userEnteredValue': '1.0'}]},
            'format': {'backgroundColor': LIGHT_GREEN},
        }}, 'index': 10}})
    reqs.append({'addConditionalFormatRule': {'rule': {
        'ranges': [_r(sid, 1, nr, 8, 9)],
        'booleanRule': {
            'condition': {'type': 'CUSTOM_FORMULA',
                          'values': [{'userEnteredValue': '=AND(I2<1,I2<>"")'}]},
            'format': {'backgroundColor': LIGHT_RED},
        }}, 'index': 11}})

    # Conditional formatting: Call Status (col 14) background colors
    status_colors = [
        ('VM Left', LIGHT_YELLOW), ('Connected', LIGHT_BLUE), ('Interested', LIGHT_GREEN),
        ('Scheduled', SCHED_GREEN), ('Not Qualified', MED_GRAY), ('DNC', LIGHT_RED),
    ]
    for i, (label, color) in enumerate(status_colors):
        reqs.append({'addConditionalFormatRule': {'rule': {
            'ranges': [_r(sid, 1, nr, 14, 15)],
            'booleanRule': {
                'condition': {'type': 'TEXT_EQ', 'values': [{'userEnteredValue': label}]},
                'format': {'backgroundColor': color},
            }}, 'index': 20 + i}})

    # Auto-filter
    reqs.append({'setBasicFilter': {'filter': {
        'range': _r(sid, 0, nr, 0, nc),
    }}})

    # Borders: light gray grid + medium bottom on header
    reqs.append({'updateBorders': {
        'range': _r(sid, 0, nr, 0, nc),
        'innerHorizontal': {'style': 'SOLID', 'color': MED_GRAY},
        'innerVertical': {'style': 'SOLID', 'color': MED_GRAY},
    }})
    reqs.append({'updateBorders': {
        'range': _r(sid, 0, 1, 0, nc),
        'bottom': {'style': 'SOLID_MEDIUM', 'color': BLACK},
    }})

    return reqs


# ---------------------------------------------------------------------------
# Formatting: Battlecards
# ---------------------------------------------------------------------------
def fmt_battlecards(sid, num_rows):
    nr = num_rows + 2  # section header row + column header row + data
    reqs = []

    # Section header row (row 0): merge + color
    for (name, sc, ec), color in zip(BC_SECTION_RANGES, BC_COLORS):
        reqs.append({'mergeCells': {
            'range': _r(sid, 0, 1, sc, ec), 'mergeType': 'MERGE_ALL',
        }})
        reqs.append({'repeatCell': {
            'range': _r(sid, 0, 1, sc, ec),
            'cell': {'userEnteredFormat': {
                'backgroundColor': color,
                'textFormat': {'bold': True, 'fontSize': 12, 'foregroundColor': WHITE, 'fontFamily': 'Arial'},
                'horizontalAlignment': 'CENTER', 'verticalAlignment': 'MIDDLE',
            }},
            'fields': 'userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment)',
        }})

    # Column header row (row 1): gray bg, bold
    reqs.append({'repeatCell': {
        'range': _r(sid, 1, 2, 0, BC_NUM_COLS),
        'cell': {'userEnteredFormat': {
            'backgroundColor': MED_GRAY,
            'textFormat': {'bold': True, 'fontSize': 10, 'fontFamily': 'Arial'},
            'horizontalAlignment': 'CENTER',
        }},
        'fields': 'userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)',
    }})

    # Freeze 2 header rows, tab color blue (no col freeze — conflicts with merged section headers)
    reqs.append({'updateSheetProperties': {'properties': {
        'sheetId': sid,
        'gridProperties': {'frozenRowCount': 2},
        'tabColor': BLUE,
    }, 'fields': 'gridProperties.frozenRowCount,tabColor'}})

    # Alternating row colors
    reqs.append({'addConditionalFormatRule': {'rule': {
        'ranges': [_r(sid, 2, nr, 0, BC_NUM_COLS)],
        'booleanRule': {
            'condition': {'type': 'CUSTOM_FORMULA',
                          'values': [{'userEnteredValue': '=ISEVEN(ROW())'}]},
            'format': {'backgroundColor': LIGHT_GRAY},
        }}, 'index': 50}})

    # Data font 10pt
    reqs.append({'repeatCell': {
        'range': _r(sid, 2, nr, 0, BC_NUM_COLS),
        'cell': {'userEnteredFormat': {'textFormat': {'fontSize': 10}}},
        'fields': 'userEnteredFormat.textFormat.fontSize',
    }})

    # Section header row height
    reqs.append({'updateDimensionProperties': {
        'range': {'sheetId': sid, 'dimension': 'ROWS', 'startIndex': 0, 'endIndex': 1},
        'properties': {'pixelSize': 36}, 'fields': 'pixelSize',
    }})

    # Wrap text on all data cells
    reqs.append({'repeatCell': {
        'range': _r(sid, 2, nr, 0, BC_NUM_COLS),
        'cell': {'userEnteredFormat': {'wrapStrategy': 'WRAP'}},
        'fields': 'userEnteredFormat.wrapStrategy',
    }})
    # Column widths — fit content generously across all 50 cols
    # IDENTITY (0-7): Name, Entity, ICP, Score, Email, Phone, LinkedIn, County
    # PORTFOLIO (8-14): Props, Value, AvgVal, Types, Equity, Eq%, Cashout
    # FINANCING (15-22): Lender, Type, Balance, Payment, Rate, DSCR, MoRent, AnnRent
    # ACQUISITION (23-30): Acq, 12mo, 36mo, AvgPrice, Flips, Holds, HoldMo, Cash%
    # REFI (31-37): Priority, Signals, Cash?, BRRRR?, Rate?, Equity?, Angle
    # WEALTH (38-45): FEC$, FECRecip, SunBiz, Wealth, Officers, RegAgent, Status, STR
    # OUTREACH (46-49): Talking, Notes, Template, Sent?
    bc_widths = [
        180, 220, 180, 55, 210, 140, 180, 110,   # IDENTITY
        55, 120, 110, 160, 110, 65, 110,           # PORTFOLIO
        220, 120, 100, 100, 65, 65, 100, 100,      # FINANCING
        55, 55, 55, 120, 55, 55, 65, 65,           # ACQUISITION
        80, 260, 55, 55, 55, 55, 300,              # REFI
        80, 200, 55, 65, 200, 180, 80, 55,         # WEALTH
        360, 200, 140, 65,                          # OUTREACH
    ]
    for i, w in enumerate(bc_widths[:BC_NUM_COLS]):
        reqs.append({'updateDimensionProperties': {
            'range': {'sheetId': sid, 'dimension': 'COLUMNS', 'startIndex': i, 'endIndex': i + 1},
            'properties': {'pixelSize': w}, 'fields': 'pixelSize',
        }})

    # Currency columns: Portfolio Value (col 9), Avg Property Value (10), Equity (12),
    # Max Cashout (14), Est. Balance (17), Monthly Payment (18), Monthly Rent (21), Annual Rent (22),
    # Avg Purchase Price (28), FEC Donations (36)
    currency_cols = [9, 10, 12, 14, 17, 18, 21, 22, 28, 36]
    for col in currency_cols:
        reqs.append({'repeatCell': {
            'range': _r(sid, 2, nr, col, col + 1),
            'cell': {'userEnteredFormat': {'numberFormat': {'type': 'CURRENCY', 'pattern': '$#,##0'}}},
            'fields': 'userEnteredFormat.numberFormat',
        }})

    # Percentage: Equity % (col 13), Cash Purchase % (31)
    for col in [13, 31]:
        reqs.append({'repeatCell': {
            'range': _r(sid, 2, nr, col, col + 1),
            'cell': {'userEnteredFormat': {'numberFormat': {'type': 'PERCENT', 'pattern': '0%'}}},
            'fields': 'userEnteredFormat.numberFormat',
        }})

    # DSCR format (col 20)
    reqs.append({'repeatCell': {
        'range': _r(sid, 2, nr, 20, 21),
        'cell': {'userEnteredFormat': {'numberFormat': {'type': 'NUMBER', 'pattern': '0.00"x"'}}},
        'fields': 'userEnteredFormat.numberFormat',
    }})

    # Interest Rate % (col 19)
    reqs.append({'repeatCell': {
        'range': _r(sid, 2, nr, 19, 20),
        'cell': {'userEnteredFormat': {'numberFormat': {'type': 'NUMBER', 'pattern': '0.00"%"'}}},
        'fields': 'userEnteredFormat.numberFormat',
    }})

    # Borders
    reqs.append({'updateBorders': {
        'range': _r(sid, 1, nr, 0, BC_NUM_COLS),
        'innerHorizontal': {'style': 'SOLID', 'color': MED_GRAY},
        'innerVertical': {'style': 'SOLID', 'color': MED_GRAY},
    }})
    reqs.append({'updateBorders': {
        'range': _r(sid, 1, 2, 0, BC_NUM_COLS),
        'bottom': {'style': 'SOLID_MEDIUM', 'color': BLACK},
    }})

    return reqs


# ---------------------------------------------------------------------------
# Formatting: Performance
# ---------------------------------------------------------------------------
def fmt_performance(sid, num_rows):
    nr = num_rows + 2
    reqs = []

    # Section header row (row 0): merge + color
    for (name, sc, ec), color in zip(PF_SECTION_RANGES, PF_COLORS):
        reqs.append({'mergeCells': {
            'range': _r(sid, 0, 1, sc, ec), 'mergeType': 'MERGE_ALL',
        }})
        reqs.append({'repeatCell': {
            'range': _r(sid, 0, 1, sc, ec),
            'cell': {'userEnteredFormat': {
                'backgroundColor': color,
                'textFormat': {'bold': True, 'fontSize': 12, 'foregroundColor': WHITE, 'fontFamily': 'Arial'},
                'horizontalAlignment': 'CENTER', 'verticalAlignment': 'MIDDLE',
            }},
            'fields': 'userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment)',
        }})

    # Column header row (row 1)
    reqs.append({'repeatCell': {
        'range': _r(sid, 1, 2, 0, PF_NUM_COLS),
        'cell': {'userEnteredFormat': {
            'backgroundColor': MED_GRAY,
            'textFormat': {'bold': True, 'fontSize': 10, 'fontFamily': 'Arial'},
            'horizontalAlignment': 'CENTER',
        }},
        'fields': 'userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)',
    }})

    # Freeze 2 header rows, tab color orange (no col freeze — conflicts with merged section headers)
    reqs.append({'updateSheetProperties': {'properties': {
        'sheetId': sid,
        'gridProperties': {'frozenRowCount': 2},
        'tabColor': ORANGE,
    }, 'fields': 'gridProperties.frozenRowCount,tabColor'}})

    # Alternating row colors
    reqs.append({'addConditionalFormatRule': {'rule': {
        'ranges': [_r(sid, 2, nr, 0, PF_NUM_COLS)],
        'booleanRule': {
            'condition': {'type': 'CUSTOM_FORMULA',
                          'values': [{'userEnteredValue': '=ISEVEN(ROW())'}]},
            'format': {'backgroundColor': LIGHT_GRAY},
        }}, 'index': 50}})

    # Section header row height
    reqs.append({'updateDimensionProperties': {
        'range': {'sheetId': sid, 'dimension': 'ROWS', 'startIndex': 0, 'endIndex': 1},
        'properties': {'pixelSize': 36}, 'fields': 'pixelSize',
    }})

    # Currency: Portfolio Value (col 5), Loan Amount (col 18), Revenue (col 22)
    for col in [5, 18, 22]:
        reqs.append({'repeatCell': {
            'range': _r(sid, 2, nr, col, col + 1),
            'cell': {'userEnteredFormat': {'numberFormat': {'type': 'CURRENCY', 'pattern': '$#,##0'}}},
            'fields': 'userEnteredFormat.numberFormat',
        }})

    # Date format: First Contact (9), Last Contact (11), Appointment Date (16), Close Date (21)
    for col in [9, 11, 16, 21]:
        reqs.append({'repeatCell': {
            'range': _r(sid, 2, nr, col, col + 1),
            'cell': {'userEnteredFormat': {'numberFormat': {'type': 'DATE', 'pattern': 'MMM dd, yyyy'}}},
            'fields': 'userEnteredFormat.numberFormat',
        }})

    # Revenue column: bold green when filled
    reqs.append({'addConditionalFormatRule': {'rule': {
        'ranges': [_r(sid, 2, nr, 22, 23)],
        'booleanRule': {
            'condition': {'type': 'NUMBER_GREATER', 'values': [{'userEnteredValue': '0'}]},
            'format': {'textFormat': {'bold': True, 'foregroundColor': rgb('#1b5e20')}},
        }}, 'index': 0}})

    # Conditional formatting: Closed? (col 20) green if Yes
    reqs.append({'addConditionalFormatRule': {'rule': {
        'ranges': [_r(sid, 2, nr, 20, 21)],
        'booleanRule': {
            'condition': {'type': 'TEXT_EQ', 'values': [{'userEnteredValue': 'Yes'}]},
            'format': {'backgroundColor': LIGHT_GREEN},
        }}, 'index': 1}})

    # Interest Level coloring (col 14)
    interest_colors = [('Hot', LIGHT_RED), ('Warm', LIGHT_YELLOW), ('Cold', LIGHT_BLUE)]
    for i, (label, color) in enumerate(interest_colors):
        reqs.append({'addConditionalFormatRule': {'rule': {
            'ranges': [_r(sid, 2, nr, 14, 15)],
            'booleanRule': {
                'condition': {'type': 'TEXT_EQ', 'values': [{'userEnteredValue': label}]},
                'format': {'backgroundColor': color},
            }}, 'index': 10 + i}})

    # Dropdowns
    dropdowns = {
        8: ['Phone', 'Email', 'LinkedIn', 'Referral', 'Direct Mail'],  # Channel
        12: ['Yes', 'No', 'Voicemail'],  # Reached?
        13: ['Yes', 'No'],  # Conversation?
        14: ['Hot', 'Warm', 'Cold', 'Not Interested'],  # Interest Level
        15: ['Happy with lender', 'Not buying', 'Bad timing', 'Rate too high', "Don't qualify", 'Other'],  # Objection
        17: ['Yes', 'No', 'Pending'],  # Application?
        19: ['Cash-out Refi', 'Rate-Term Refi', 'Purchase', 'Bridge', 'Portfolio'],  # Loan Type
        20: ['Yes', 'No', 'In Progress'],  # Closed?
    }
    for col, values in dropdowns.items():
        reqs.append({'setDataValidation': {
            'range': _r(sid, 2, nr, col, col + 1),
            'rule': {'condition': {'type': 'ONE_OF_LIST',
                     'values': [{'userEnteredValue': v} for v in values]},
                     'showCustomUi': True, 'strict': False},
        }})

    # Wrap text on all data cells
    reqs.append({'repeatCell': {
        'range': _r(sid, 2, nr, 0, PF_NUM_COLS),
        'cell': {'userEnteredFormat': {'wrapStrategy': 'WRAP'}},
        'fields': 'userEnteredFormat.wrapStrategy',
    }})
    # Column widths
    # LEAD INFO (0-7): Name, Entity, ICP, Score, Props, Value, Lender, County
    # OUTREACH (8-11): Channel, First, Touches, Last
    # FUNNEL (12-22): Reached, Convo, Interest, Objection, Appt, App, LoanAmt, Type, Closed, CloseDate, Revenue
    # LEARNING (23-26): Worked, Didn't, Referral, Notes
    pf_widths = [
        180, 220, 180, 55, 55, 120, 200, 110,
        90, 110, 70, 110,
        75, 85, 100, 170, 110, 85, 120, 120, 75, 110, 100,
        220, 220, 85, 280,
    ]
    for i, w in enumerate(pf_widths[:PF_NUM_COLS]):
        reqs.append({'updateDimensionProperties': {
            'range': {'sheetId': sid, 'dimension': 'COLUMNS', 'startIndex': i, 'endIndex': i + 1},
            'properties': {'pixelSize': w}, 'fields': 'pixelSize',
        }})

    # Borders
    reqs.append({'updateBorders': {
        'range': _r(sid, 1, nr, 0, PF_NUM_COLS),
        'innerHorizontal': {'style': 'SOLID', 'color': MED_GRAY},
        'innerVertical': {'style': 'SOLID', 'color': MED_GRAY},
    }})
    reqs.append({'updateBorders': {
        'range': _r(sid, 1, 2, 0, PF_NUM_COLS),
        'bottom': {'style': 'SOLID_MEDIUM', 'color': BLACK},
    }})

    return reqs


# ---------------------------------------------------------------------------
# Formatting: Dashboard
# ---------------------------------------------------------------------------
def fmt_dashboard(sid, dash_data):
    reqs = []
    num_icps = len(dash_data['icp_names'])

    # Tab color purple
    reqs.append({'updateSheetProperties': {'properties': {
        'sheetId': sid, 'tabColor': PURPLE,
    }, 'fields': 'tabColor'}})

    # KPI label row (row 0): merge pairs, navy bg, white text
    for i in range(6):
        col = i * 2
        reqs.append({'mergeCells': {
            'range': _r(sid, 0, 1, col, col + 2), 'mergeType': 'MERGE_ALL',
        }})
        reqs.append({'repeatCell': {
            'range': _r(sid, 0, 1, col, col + 2),
            'cell': {'userEnteredFormat': {
                'backgroundColor': NAVY,
                'textFormat': {'bold': True, 'fontSize': 11, 'foregroundColor': WHITE, 'fontFamily': 'Arial'},
                'horizontalAlignment': 'CENTER', 'verticalAlignment': 'MIDDLE',
            }},
            'fields': 'userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment)',
        }})

    # KPI value row (row 1): merge pairs, large font
    for i in range(6):
        col = i * 2
        reqs.append({'mergeCells': {
            'range': _r(sid, 1, 2, col, col + 2), 'mergeType': 'MERGE_ALL',
        }})
        reqs.append({'repeatCell': {
            'range': _r(sid, 1, 2, col, col + 2),
            'cell': {'userEnteredFormat': {
                'backgroundColor': rgb('#e8eaf6'),
                'textFormat': {'bold': True, 'fontSize': 24, 'fontFamily': 'Arial'},
                'horizontalAlignment': 'CENTER', 'verticalAlignment': 'MIDDLE',
            }},
            'fields': 'userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment)',
        }})

    # Connect Rate format as %
    reqs.append({'repeatCell': {
        'range': _r(sid, 1, 2, 4, 6),
        'cell': {'userEnteredFormat': {'numberFormat': {'type': 'PERCENT', 'pattern': '0%'}}},
        'fields': 'userEnteredFormat.numberFormat',
    }})
    # Pipeline $ and Closed $ as currency
    for col in [8, 10]:
        reqs.append({'repeatCell': {
            'range': _r(sid, 1, 2, col, col + 2),
            'cell': {'userEnteredFormat': {'numberFormat': {'type': 'CURRENCY', 'pattern': '$#,##0'}}},
            'fields': 'userEnteredFormat.numberFormat',
        }})

    # KPI row heights
    for row in [0, 1]:
        reqs.append({'updateDimensionProperties': {
            'range': {'sheetId': sid, 'dimension': 'ROWS', 'startIndex': row, 'endIndex': row + 1},
            'properties': {'pixelSize': 50}, 'fields': 'pixelSize',
        }})

    # ICP section header (row 3) — spans 7 columns
    reqs.append({'mergeCells': {'range': _r(sid, 3, 4, 0, 7), 'mergeType': 'MERGE_ALL'}})
    reqs.append({'repeatCell': {
        'range': _r(sid, 3, 4, 0, 7),
        'cell': {'userEnteredFormat': {
            'backgroundColor': TEAL,
            'textFormat': {'bold': True, 'fontSize': 12, 'foregroundColor': WHITE, 'fontFamily': 'Arial'},
            'horizontalAlignment': 'LEFT', 'verticalAlignment': 'MIDDLE',
        }},
        'fields': 'userEnteredFormat(backgroundColor,textFormat,horizontalAlignment,verticalAlignment)',
    }})

    # ICP column headers (row 4): bold gray bg
    reqs.append({'repeatCell': {
        'range': _r(sid, 4, 5, 0, 7),
        'cell': {'userEnteredFormat': {
            'backgroundColor': MED_GRAY,
            'textFormat': {'bold': True, 'fontSize': 10, 'fontFamily': 'Arial'},
        }},
        'fields': 'userEnteredFormat(backgroundColor,textFormat)',
    }})

    # ICP data rows: % format for col 2 (% of Total) and col 6 (Close Rate)
    icp_end = 5 + num_icps
    reqs.append({'repeatCell': {
        'range': _r(sid, 5, icp_end, 2, 3),
        'cell': {'userEnteredFormat': {'numberFormat': {'type': 'PERCENT', 'pattern': '0.0%'}}},
        'fields': 'userEnteredFormat.numberFormat',
    }})
    reqs.append({'repeatCell': {
        'range': _r(sid, 5, icp_end, 6, 7),
        'cell': {'userEnteredFormat': {'numberFormat': {'type': 'PERCENT', 'pattern': '0%'}}},
        'fields': 'userEnteredFormat.numberFormat',
    }})

    # Alternating rows for ICP table
    reqs.append({'addConditionalFormatRule': {'rule': {
        'ranges': [_r(sid, 5, icp_end, 0, 7)],
        'booleanRule': {
            'condition': {'type': 'CUSTOM_FORMULA',
                          'values': [{'userEnteredValue': '=ISEVEN(ROW())'}]},
            'format': {'backgroundColor': LIGHT_GRAY},
        }}, 'index': 50}})

    # County section header
    ch = dash_data['county_header_row']
    reqs.append({'mergeCells': {'range': _r(sid, ch, ch + 1, 0, 4), 'mergeType': 'MERGE_ALL'}})
    reqs.append({'repeatCell': {
        'range': _r(sid, ch, ch + 1, 0, 4),
        'cell': {'userEnteredFormat': {
            'backgroundColor': BLUE,
            'textFormat': {'bold': True, 'fontSize': 12, 'foregroundColor': WHITE, 'fontFamily': 'Arial'},
            'horizontalAlignment': 'LEFT',
        }},
        'fields': 'userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)',
    }})
    reqs.append({'repeatCell': {
        'range': _r(sid, ch + 1, ch + 2, 0, 4),
        'cell': {'userEnteredFormat': {'backgroundColor': MED_GRAY, 'textFormat': {'bold': True}}},
        'fields': 'userEnteredFormat(backgroundColor,textFormat)',
    }})
    # Close Rate % format
    county_data_end = ch + 2 + len(dash_data['county_rows'])
    reqs.append({'repeatCell': {
        'range': _r(sid, ch + 2, county_data_end, 3, 4),
        'cell': {'userEnteredFormat': {'numberFormat': {'type': 'PERCENT', 'pattern': '0%'}}},
        'fields': 'userEnteredFormat.numberFormat',
    }})

    # Objections section header
    oh = dash_data['obj_header_row']
    reqs.append({'mergeCells': {'range': _r(sid, oh, oh + 1, 0, 3), 'mergeType': 'MERGE_ALL'}})
    reqs.append({'repeatCell': {
        'range': _r(sid, oh, oh + 1, 0, 3),
        'cell': {'userEnteredFormat': {
            'backgroundColor': ORANGE,
            'textFormat': {'bold': True, 'fontSize': 12, 'foregroundColor': WHITE, 'fontFamily': 'Arial'},
            'horizontalAlignment': 'LEFT',
        }},
        'fields': 'userEnteredFormat(backgroundColor,textFormat,horizontalAlignment)',
    }})
    reqs.append({'repeatCell': {
        'range': _r(sid, oh + 1, oh + 2, 0, 3),
        'cell': {'userEnteredFormat': {'backgroundColor': MED_GRAY, 'textFormat': {'bold': True}}},
        'fields': 'userEnteredFormat(backgroundColor,textFormat)',
    }})
    # % format for objection %
    obj_data_end = oh + 2 + len(dash_data['obj_rows'])
    reqs.append({'repeatCell': {
        'range': _r(sid, oh + 2, obj_data_end, 2, 3),
        'cell': {'userEnteredFormat': {'numberFormat': {'type': 'PERCENT', 'pattern': '0%'}}},
        'fields': 'userEnteredFormat.numberFormat',
    }})

    # Column widths
    for col, w in {0: 220, 1: 80, 2: 90, 3: 90, 4: 90, 5: 110, 6: 90}.items():
        reqs.append({'updateDimensionProperties': {
            'range': {'sheetId': sid, 'dimension': 'COLUMNS', 'startIndex': col, 'endIndex': col + 1},
            'properties': {'pixelSize': w}, 'fields': 'pixelSize',
        }})

    return reqs


# ---------------------------------------------------------------------------
# Range helper
# ---------------------------------------------------------------------------
def _r(sheet_id, r1, r2, c1=None, c2=None):
    rng = {'sheetId': sheet_id, 'startRowIndex': r1, 'endRowIndex': r2}
    if c1 is not None:
        rng['startColumnIndex'] = c1
    if c2 is not None:
        rng['endColumnIndex'] = c2
    return rng


# ---------------------------------------------------------------------------
# Build pie chart request
# ---------------------------------------------------------------------------
def build_pie_chart(sid, num_icps):
    return {'addChart': {'chart': {
        'spec': {
            'title': 'Leads by ICP Segment',
            'pieChart': {
                'legendPosition': 'RIGHT_LEGEND',
                'domain': {'sourceRange': {'sources': [_r(sid, 5, 5 + num_icps, 0, 1)]}},
                'series': {'sourceRange': {'sources': [_r(sid, 5, 5 + num_icps, 1, 2)]}},
            },
        },
        'position': {'overlayPosition': {
            'anchorCell': {'sheetId': sid, 'rowIndex': 3, 'columnIndex': 8},
            'widthPixels': 480, 'heightPixels': 320,
        }},
    }}}


# ---------------------------------------------------------------------------
# Upload to Google Sheets
# ---------------------------------------------------------------------------
def upload_to_google_sheets(df, call_rows, bc_rows, pf_rows, dash_data):
    try:
        from googleapiclient.discovery import build as gapi_build
    except ImportError:
        print("ERROR: pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib")
        return None

    from dotenv import load_dotenv
    load_dotenv()

    creds = get_google_creds()
    service = gapi_build("sheets", "v4", credentials=creds)

    # Create spreadsheet with 4 sheets
    spreadsheet = service.spreadsheets().create(body={
        "properties": {"title": "DSCR Lead Gen - Frank's Pipeline"},
        "sheets": [
            {"properties": {"title": "Frank's Call Sheet", "index": 0}},
            {"properties": {"title": "Battlecards", "index": 1}},
            {"properties": {"title": "Performance", "index": 2}},
            {"properties": {"title": "Dashboard", "index": 3}},
        ],
    }).execute()

    sid_map = {sh['properties']['title']: sh['properties']['sheetId'] for sh in spreadsheet['sheets']}
    spreadsheet_id = spreadsheet['spreadsheetId']
    url = spreadsheet['spreadsheetUrl']
    print(f"\nCreated: {url}")

    # --- Write data ---
    def write(tab, values, input_option="USER_ENTERED"):
        service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id, range=f"'{tab}'!A1",
            valueInputOption=input_option, body={"values": values},
        ).execute()

    # Call Sheet: header + data
    print("  Writing Call Sheet...")
    write("Frank's Call Sheet", [CALL_HEADERS] + call_rows)

    # Battlecards: section header row + column header row + data
    print("  Writing Battlecards...")
    sec_row = [''] * BC_NUM_COLS
    for name, sc, ec in BC_SECTION_RANGES:
        sec_row[sc] = name
    write("Battlecards", [sec_row, BC_ALL_HEADERS] + bc_rows)

    # Performance: section header row + column header row + data
    print("  Writing Performance...")
    pf_sec_row = [''] * PF_NUM_COLS
    for name, sc, ec in PF_SECTION_RANGES:
        pf_sec_row[sc] = name
    write("Performance", [pf_sec_row, PF_ALL_HEADERS] + pf_rows)

    # Dashboard
    print("  Writing Dashboard...")
    dash_values = [
        dash_data['kpi_labels'],
        dash_data['kpi_formulas'],
        [],  # blank row
        ['LEADS BY ICP SEGMENT'],
        dash_data['icp_header'],
    ]
    dash_values.extend(dash_data['icp_rows'])
    dash_values.append([])  # blank
    dash_values.append(['CONVERSION BY COUNTY'])
    dash_values.append(dash_data['county_header'])
    dash_values.extend(dash_data['county_rows'])
    dash_values.append([])  # blank
    dash_values.append(['TOP OBJECTIONS'])
    dash_values.append(dash_data['obj_header'])
    dash_values.extend(dash_data['obj_rows'])
    write("Dashboard", dash_values)

    # --- Apply formatting ---
    print("  Applying formatting...")
    all_reqs = []
    all_reqs.extend(fmt_call_sheet(sid_map["Frank's Call Sheet"], len(call_rows)))
    all_reqs.extend(fmt_battlecards(sid_map["Battlecards"], len(bc_rows)))
    all_reqs.extend(fmt_performance(sid_map["Performance"], len(pf_rows)))
    all_reqs.extend(fmt_dashboard(sid_map["Dashboard"], dash_data))

    # Pie chart
    all_reqs.append(build_pie_chart(sid_map["Dashboard"], len(dash_data['icp_names'])))

    # Send all formatting in one batch
    service.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": all_reqs},
    ).execute()

    print(f"  Applied {len(all_reqs)} formatting rules")
    return url


# ---------------------------------------------------------------------------
# Export formatted xlsx
# ---------------------------------------------------------------------------
def export_xlsx(call_rows, bc_rows, pf_rows, dash_data):
    """Export fully formatted Excel workbook matching Google Sheets styling."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule, FormulaRule

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Shared styles
    navy_fill = PatternFill('solid', fgColor='1a237e')
    white_font = Font(bold=True, color='ffffff', size=13, name='Arial')
    white_font_12 = Font(bold=True, color='ffffff', size=12, name='Arial')
    gray_fill = PatternFill('solid', fgColor='e0e0e0')
    alt_fill = PatternFill('solid', fgColor='f5f5f5')
    bold_12 = Font(bold=True, size=12, name='Arial')
    bold_10 = Font(bold=True, size=10, name='Arial')
    normal_10 = Font(size=10, name='Arial')
    wrap = Alignment(wrap_text=True, vertical='top')
    wrap_center = Alignment(wrap_text=True, vertical='center', horizontal='center')
    wrap_right = Alignment(wrap_text=True, vertical='top', horizontal='right')
    thin_border = Border(
        left=Side(style='thin', color='e0e0e0'), right=Side(style='thin', color='e0e0e0'),
        top=Side(style='thin', color='e0e0e0'), bottom=Side(style='thin', color='e0e0e0'),
    )
    header_border = Border(
        left=Side(style='thin', color='e0e0e0'), right=Side(style='thin', color='e0e0e0'),
        top=Side(style='thin', color='e0e0e0'), bottom=Side(style='medium', color='000000'),
    )
    # Conditional format fills
    green_fill = PatternFill('solid', fgColor='c8e6c9')
    yellow_fill = PatternFill('solid', fgColor='fff9c4')
    red_fill = PatternFill('solid', fgColor='ffcdd2')
    blue_fill = PatternFill('solid', fgColor='bbdefb')
    sched_fill = PatternFill('solid', fgColor='81c784')

    def _write_rows(ws, headers, rows, start_row=1):
        """Write headers + data rows starting at start_row (1-based)."""
        for c, h in enumerate(headers, 1):
            ws.cell(row=start_row, column=c, value=h)
        for r, row in enumerate(rows, start_row + 1):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)

    def _style_header(ws, row, num_cols, fill, font, border_style=header_border):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = fill
            cell.font = font
            cell.alignment = wrap_center
            cell.border = border_style

    def _alt_rows(ws, start_row, end_row, num_cols):
        for r in range(start_row, end_row + 1):
            if r % 2 == 0:
                for c in range(1, num_cols + 1):
                    ws.cell(row=r, column=c).fill = alt_fill

    def _borders(ws, start_row, end_row, num_cols):
        for r in range(start_row, end_row + 1):
            for c in range(1, num_cols + 1):
                ws.cell(row=r, column=c).border = thin_border

    # ===================================================================
    # SHEET 1: Frank's Call Sheet
    # ===================================================================
    ws1 = wb.active
    ws1.title = "Frank's Call Sheet"
    ws1.sheet_properties.tabColor = '4caf50'
    nc = len(CALL_HEADERS)
    nr = len(call_rows)

    _write_rows(ws1, CALL_HEADERS, call_rows)
    _style_header(ws1, 1, nc, navy_fill, white_font)
    ws1.freeze_panes = 'D2'  # freeze row 1 + cols A-C
    ws1.row_dimensions[1].height = 40

    # Column widths (chars) and data formatting
    cs_widths = [15, 28, 21, 33, 8, 17, 33, 17, 10, 28, 47, 53, 32, 15, 18, 15, 15, 39]
    for i, w in enumerate(cs_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    for r in range(2, nr + 2):
        ws1.row_dimensions[r].height = 38
        # Wrap all cells
        for c in range(1, nc + 1):
            cell = ws1.cell(row=r, column=c)
            cell.alignment = wrap
            cell.font = normal_10
        # Bold contact name
        ws1.cell(row=r, column=2).font = bold_12
        # Currency: Portfolio Value (F), Equity (H)
        ws1.cell(row=r, column=6).number_format = '$#,##0'
        ws1.cell(row=r, column=8).number_format = '$#,##0'
        # Right-align numbers
        for c in [5, 6, 8, 9]:
            ws1.cell(row=r, column=c).alignment = Alignment(wrap_text=True, horizontal='right', vertical='top')
        # DSCR format (col I=9)
        dscr_val = ws1.cell(row=r, column=9).value
        if dscr_val and dscr_val != '':
            ws1.cell(row=r, column=9).number_format = '0.00"x"'
        # Date cols (P=16, Q=17)
        for c in [16, 17]:
            ws1.cell(row=r, column=c).number_format = 'MMM DD, YYYY'

    # Alternating rows + borders
    _alt_rows(ws1, 2, nr + 1, nc)
    _borders(ws1, 1, nr + 1, nc)
    _style_header(ws1, 1, nc, navy_fill, white_font, header_border)

    # Data validation: Call Status (col O=15)
    dv_status = DataValidation(type="list",
        formula1='"Not Called,VM Left,Connected,Interested,Scheduled,Not Qualified,DNC"',
        allow_blank=True)
    dv_status.error = "Select from dropdown"
    ws1.add_data_validation(dv_status)
    dv_status.add(f'O2:O{nr + 1}')

    # Conditional formatting: Equity (col H)
    ws1.conditional_formatting.add(f'H2:H{nr+1}',
        CellIsRule(operator='greaterThanOrEqual', formula=['200000'], fill=green_fill))
    ws1.conditional_formatting.add(f'H2:H{nr+1}',
        CellIsRule(operator='between', formula=['50000', '200000'], fill=yellow_fill))
    ws1.conditional_formatting.add(f'H2:H{nr+1}',
        CellIsRule(operator='lessThan', formula=['0'], fill=red_fill))

    # Conditional formatting: DSCR (col I)
    ws1.conditional_formatting.add(f'I2:I{nr+1}',
        CellIsRule(operator='greaterThanOrEqual', formula=['1.0'], fill=green_fill))
    ws1.conditional_formatting.add(f'I2:I{nr+1}',
        FormulaRule(formula=[f'AND(I2<1,I2<>"")'], fill=red_fill))

    # Conditional formatting: Call Status (col O)
    for label, fill in [('VM Left', yellow_fill), ('Connected', blue_fill),
                         ('Interested', green_fill), ('Scheduled', sched_fill),
                         ('Not Qualified', gray_fill), ('DNC', red_fill)]:
        ws1.conditional_formatting.add(f'O2:O{nr+1}',
            CellIsRule(operator='equal', formula=[f'"{label}"'], fill=fill))

    # Auto-filter
    ws1.auto_filter.ref = f'A1:{get_column_letter(nc)}{nr+1}'

    # ===================================================================
    # SHEET 2: Battlecards
    # ===================================================================
    ws2 = wb.create_sheet("Battlecards")
    ws2.sheet_properties.tabColor = '2196f3'
    bc_nc = BC_NUM_COLS
    bc_nr = len(bc_rows)

    # Row 1: section headers (merged + colored)
    section_colors_hex = ['1565c0', '00796b', 'e65100', '4a148c', 'b71c1c', '2e7d32', '616161']
    for (name, sc, ec), color_hex in zip(BC_SECTION_RANGES, section_colors_hex):
        ws2.merge_cells(start_row=1, start_column=sc + 1, end_row=1, end_column=ec)
        cell = ws2.cell(row=1, column=sc + 1, value=name)
        cell.fill = PatternFill('solid', fgColor=color_hex)
        cell.font = white_font_12
        cell.alignment = wrap_center
    ws2.row_dimensions[1].height = 30

    # Row 2: column headers
    for c, h in enumerate(BC_ALL_HEADERS, 1):
        cell = ws2.cell(row=2, column=c, value=h)
        cell.fill = gray_fill
        cell.font = bold_10
        cell.alignment = wrap_center
        cell.border = header_border
    ws2.row_dimensions[2].height = 28

    # Data rows (start at row 3)
    for r_idx, row in enumerate(bc_rows, 3):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c, value=val)
            cell.alignment = wrap
            cell.font = normal_10

    ws2.freeze_panes = 'A3'

    # Column widths
    bc_char_widths = [
        25, 30, 25, 8, 29, 20, 25, 15,     # IDENTITY
        8, 17, 15, 22, 15, 9, 15,            # PORTFOLIO
        30, 17, 14, 14, 9, 9, 14, 14,        # FINANCING
        8, 8, 8, 17, 8, 8, 9, 9,             # ACQUISITION
        11, 36, 8, 8, 8, 8, 42,              # REFI
        11, 28, 8, 9, 28, 25, 11, 8,         # WEALTH
        50, 28, 20, 9,                        # OUTREACH
    ]
    for i, w in enumerate(bc_char_widths[:bc_nc], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Currency columns (0-indexed: 9,10,12,14,17,18,21,22,28,36 -> 1-indexed: 10,11,13,15,18,19,22,23,29,37)
    currency_cols_1 = [10, 11, 13, 15, 18, 19, 22, 23, 29, 37]
    for r in range(3, bc_nr + 3):
        for c in currency_cols_1:
            if c <= bc_nc:
                ws2.cell(row=r, column=c).number_format = '$#,##0'
        # Percentage: Equity % (14), Cash Purchase % (32)
        for c in [14, 32]:
            if c <= bc_nc:
                ws2.cell(row=r, column=c).number_format = '0%'
        # DSCR (21)
        dv = ws2.cell(row=r, column=21).value
        if dv and dv != '':
            ws2.cell(row=r, column=21).number_format = '0.00"x"'

    _alt_rows(ws2, 3, bc_nr + 2, bc_nc)
    _borders(ws2, 2, bc_nr + 2, bc_nc)

    # ===================================================================
    # SHEET 3: Performance
    # ===================================================================
    ws3 = wb.create_sheet("Performance")
    ws3.sheet_properties.tabColor = 'ff9800'
    pf_nc = PF_NUM_COLS
    pf_nr = len(pf_rows)

    # Row 1: section headers
    pf_colors_hex = ['546e7a', '1565c0', '2e7d32', 'f9a825']
    for (name, sc, ec), color_hex in zip(PF_SECTION_RANGES, pf_colors_hex):
        ws3.merge_cells(start_row=1, start_column=sc + 1, end_row=1, end_column=ec)
        cell = ws3.cell(row=1, column=sc + 1, value=name)
        cell.fill = PatternFill('solid', fgColor=color_hex)
        cell.font = white_font_12
        cell.alignment = wrap_center
    ws3.row_dimensions[1].height = 30

    # Row 2: column headers
    for c, h in enumerate(PF_ALL_HEADERS, 1):
        cell = ws3.cell(row=2, column=c, value=h)
        cell.fill = gray_fill
        cell.font = bold_10
        cell.alignment = wrap_center
        cell.border = header_border

    # Data rows (start at row 3)
    for r_idx, row in enumerate(pf_rows, 3):
        for c, val in enumerate(row, 1):
            cell = ws3.cell(row=r_idx, column=c, value=val)
            cell.alignment = wrap
            cell.font = normal_10

    ws3.freeze_panes = 'A3'

    # Column widths
    pf_char_widths = [
        25, 30, 25, 8, 8, 17, 28, 15,
        13, 15, 10, 15,
        10, 12, 14, 24, 15, 12, 17, 17, 10, 15, 14,
        30, 30, 12, 39,
    ]
    for i, w in enumerate(pf_char_widths[:pf_nc], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # Currency: Portfolio Value (6), Loan Amount (19), Revenue (23)
    for r in range(3, pf_nr + 3):
        ws3.cell(row=r, column=6).number_format = '$#,##0'
        ws3.cell(row=r, column=19).number_format = '$#,##0'
        ws3.cell(row=r, column=23).number_format = '$#,##0'
        # Date cols: First Contact (10), Last Contact (12), Appt (17), Close Date (22)
        for c in [10, 12, 17, 22]:
            ws3.cell(row=r, column=c).number_format = 'MMM DD, YYYY'

    # Dropdowns
    pf_dropdowns = {
        9: 'Phone,Email,LinkedIn,Referral,Direct Mail',
        13: 'Yes,No,Voicemail',
        14: 'Yes,No',
        15: 'Hot,Warm,Cold,Not Interested',
        16: 'Happy with lender,Not buying,Bad timing,Rate too high,Don\'t qualify,Other',
        18: 'Yes,No,Pending',
        20: 'Cash-out Refi,Rate-Term Refi,Purchase,Bridge,Portfolio',
        21: 'Yes,No,In Progress',
    }
    for col_1, vals in pf_dropdowns.items():
        dv = DataValidation(type="list", formula1=f'"{vals}"', allow_blank=True)
        ws3.add_data_validation(dv)
        dv.add(f'{get_column_letter(col_1)}3:{get_column_letter(col_1)}{pf_nr + 2}')

    # Conditional formatting: Closed=Yes (col U=21) green, Revenue>0 (col W=23) bold green
    ws3.conditional_formatting.add(f'U3:U{pf_nr+2}',
        CellIsRule(operator='equal', formula=['"Yes"'], fill=green_fill))
    ws3.conditional_formatting.add(f'W3:W{pf_nr+2}',
        CellIsRule(operator='greaterThan', formula=['0'],
                   fill=green_fill, font=Font(bold=True, color='1b5e20')))
    # Interest Level coloring (col O=15)
    for label, fill in [('Hot', red_fill), ('Warm', yellow_fill), ('Cold', blue_fill)]:
        ws3.conditional_formatting.add(f'O3:O{pf_nr+2}',
            CellIsRule(operator='equal', formula=[f'"{label}"'], fill=fill))

    _alt_rows(ws3, 3, pf_nr + 2, pf_nc)
    _borders(ws3, 2, pf_nr + 2, pf_nc)

    # ===================================================================
    # SHEET 4: Dashboard
    # ===================================================================
    ws4 = wb.create_sheet("Dashboard")
    ws4.sheet_properties.tabColor = '9c27b0'

    num = len(pf_rows) + 2  # performance data ends at this row

    # KPI Banner — row 1 labels, row 2 formulas
    kpi_items = [
        ('Total Leads', f'=COUNTA(Performance!A3:A{num})'),
        ('Calls Made', f'=COUNTIF(Performance!I3:I{num},"<>")'),
        ('Connect Rate', f'=IFERROR(COUNTIF(Performance!M3:M{num},"Yes")/COUNTIF(Performance!I3:I{num},"<>"),0)'),
        ('Appointments', f'=COUNTA(Performance!Q3:Q{num})'),
        ('Pipeline $', f'=SUM(Performance!S3:S{num})'),
        ('Closed $', f'=SUM(Performance!W3:W{num})'),
    ]
    kpi_fill = PatternFill('solid', fgColor='e8eaf6')
    kpi_font = Font(bold=True, size=22, name='Arial')
    for i, (label, formula) in enumerate(kpi_items):
        col = i * 2 + 1
        # Merge label cells
        ws4.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        cell = ws4.cell(row=1, column=col, value=label)
        cell.fill = navy_fill
        cell.font = Font(bold=True, color='ffffff', size=11, name='Arial')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # Merge value cells
        ws4.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + 1)
        cell = ws4.cell(row=2, column=col, value=formula)
        cell.fill = kpi_fill
        cell.font = kpi_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    # Format Connect Rate as %
    ws4.cell(row=2, column=5).number_format = '0%'
    # Format Pipeline $ and Closed $ as currency
    ws4.cell(row=2, column=9).number_format = '$#,##0'
    ws4.cell(row=2, column=11).number_format = '$#,##0'
    ws4.row_dimensions[1].height = 35
    ws4.row_dimensions[2].height = 45

    # ICP breakdown table (row 4 header, row 5 col headers, row 6+ data)
    r = 4
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    cell = ws4.cell(row=r, column=1, value='LEADS BY ICP SEGMENT')
    cell.fill = PatternFill('solid', fgColor='009688')
    cell.font = white_font_12
    cell.alignment = Alignment(horizontal='left', vertical='center')

    r = 5
    for c, h in enumerate(dash_data['icp_header'], 1):
        cell = ws4.cell(row=r, column=c, value=h)
        cell.fill = gray_fill
        cell.font = bold_10

    for i, icp_row in enumerate(dash_data['icp_rows']):
        r = 6 + i
        for c, val in enumerate(icp_row, 1):
            cell = ws4.cell(row=r, column=c, value=val)
            cell.font = normal_10
        ws4.cell(row=r, column=3).number_format = '0.0%'
        ws4.cell(row=r, column=7).number_format = '0%'

    # County breakdown
    ch = dash_data['county_header_row'] + 1  # convert 0-indexed to 1-indexed
    ws4.merge_cells(start_row=ch, start_column=1, end_row=ch, end_column=4)
    cell = ws4.cell(row=ch, column=1, value='CONVERSION BY COUNTY')
    cell.fill = PatternFill('solid', fgColor='2196f3')
    cell.font = white_font_12
    for c, h in enumerate(dash_data['county_header'], 1):
        cell = ws4.cell(row=ch + 1, column=c, value=h)
        cell.fill = gray_fill
        cell.font = bold_10
    for i, crow in enumerate(dash_data['county_rows']):
        for c, val in enumerate(crow, 1):
            ws4.cell(row=ch + 2 + i, column=c, value=val).font = normal_10
        ws4.cell(row=ch + 2 + i, column=4).number_format = '0%'

    # Objections breakdown
    oh = dash_data['obj_header_row'] + 1
    ws4.merge_cells(start_row=oh, start_column=1, end_row=oh, end_column=3)
    cell = ws4.cell(row=oh, column=1, value='TOP OBJECTIONS')
    cell.fill = PatternFill('solid', fgColor='ff9800')
    cell.font = white_font_12
    for c, h in enumerate(dash_data['obj_header'], 1):
        cell = ws4.cell(row=oh + 1, column=c, value=h)
        cell.fill = gray_fill
        cell.font = bold_10
    for i, orow in enumerate(dash_data['obj_rows']):
        for c, val in enumerate(orow, 1):
            ws4.cell(row=oh + 2 + i, column=c, value=val).font = normal_10
        ws4.cell(row=oh + 2 + i, column=3).number_format = '0%'

    # Dashboard column widths
    for i, w in enumerate([32, 12, 13, 13, 13, 16, 13], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    wb.save(XLSX_OUTPUT)
    print(f"\nFormatted Excel: {XLSX_OUTPUT} ({XLSX_OUTPUT.stat().st_size / 1024:.0f} KB)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Build Professional DSCR MVP Google Sheets")
    parser.add_argument("--xlsx-only", action="store_true", help="Export xlsx only, skip Google upload")
    args = parser.parse_args()

    from dotenv import load_dotenv
    load_dotenv()

    print("=" * 60)
    print("DSCR Lead Gen - Professional MVP Sheet Builder")
    print("=" * 60)

    df = load_master_data()

    print("\nBuilding sheets...")
    call_rows = build_call_sheet(df)
    bc_rows = build_battlecards(df)
    pf_rows = build_performance(df)
    dash_data = build_dashboard_data(df)

    export_xlsx(call_rows, bc_rows, pf_rows, dash_data)

    if not args.xlsx_only:
        print("\nUploading to Google Sheets...")
        url = upload_to_google_sheets(df, call_rows, bc_rows, pf_rows, dash_data)
        if url:
            print(f"\n{'=' * 60}")
            print(f"DONE! Share this with Frank:")
            print(f"  {url}")
            print(f"{'=' * 60}")
    else:
        print("\nSkipped Google upload (--xlsx-only)")


if __name__ == "__main__":
    main()
