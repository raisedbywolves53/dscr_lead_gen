"""
Step 5: Enrich Contacts — Test Mode (25-Record Proof of Concept)
=================================================================

Takes your hottest leads and enriches them with phone numbers and
email addresses using multiple sources, cheapest first.

What this script does:
  1. Reads your existing pipeline output (06_enriched.csv)
  2. Scores and ranks leads, pulls the top N hottest
  3. For leads with a resolved person name, generates email patterns
  4. Re-resolves unresolved LLCs via SunBiz (if needed)
  5. Exports a Datazapp-formatted CSV for batch skip trace ($0.03/record)
  6. Generates a "research tracker" Excel with:
     - Lead details + what we know so far
     - People search URLs (TruePeopleSearch, FastPeopleSearch) for manual lookup
     - Columns to paste in phone/email you find manually
  7. Saves enriched results to data/enriched/

The idea: for 25 records, a mix of automated + 5-min manual lookups
gives you a realistic enrichment rate to project costs at scale.

Usage:
    python scripts/05_enrich_contacts.py --limit 25
    python scripts/05_enrich_contacts.py --limit 25 --input pipeline/output/06_enriched.csv
    python scripts/05_enrich_contacts.py --limit 10 --skip-sunbiz
"""

import argparse
import json
import os
import re
import time
import urllib.parse
from pathlib import Path

import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import requests as req_lib
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

try:
    from bs4 import BeautifulSoup
    HAS_BS4 = True
except ImportError:
    HAS_BS4 = False

try:
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).resolve().parent.parent / ".env")
except ImportError:
    pass

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
ENRICHED_DIR = PROJECT_DIR / "data" / "enriched"
CACHE_DIR = PROJECT_DIR / "data" / "raw"

# Default input: existing pipeline output
DEFAULT_INPUT = Path("pipeline/output/06_enriched.csv")

# Common email domains for real estate investors/LLCs
COMMON_DOMAINS = ["gmail.com", "yahoo.com", "outlook.com", "hotmail.com", "aol.com"]

# SunBiz config (reused from script 04)
SUNBIZ_SEARCH_URL = "https://search.sunbiz.org/Inquiry/CorporationSearch/ByName"
SUNBIZ_BASE = "https://search.sunbiz.org"
SUNBIZ_DELAY = 3.0

PRIORITY_TITLES = [
    "MGR", "MGRM", "MANAGER", "MANAGING MEMBER", "MEMBER",
    "PRESIDENT", "PRES", "CEO", "OWNER", "PRINCIPAL",
]

AGENT_SERVICE_KEYWORDS = [
    " LLC", " INC", " CORP", " SERVICE", " AGENT", " REGISTERED",
    " SOLUTIONS", " FILING", " STATUTORY",
]


# ---------------------------------------------------------------------------
# Scoring — rank leads to find the hottest 25
# ---------------------------------------------------------------------------

def score_lead(row) -> int:
    """Quick scoring to find the hottest leads from existing pipeline data."""
    score = 0

    def flag_true(val):
        return str(val).strip().lower() in ("true", "1", "yes")

    # Entity / LLC ownership
    if flag_true(row.get("is_entity", "")):
        score += 10

    # Absentee / out-of-state
    if flag_true(row.get("out_of_state", "")):
        score += 15
    elif flag_true(row.get("is_absentee", "")):
        score += 10

    # Portfolio size
    prop_count = 1
    try:
        prop_count = int(float(str(row.get("property_count", 1))))
    except (ValueError, TypeError):
        pass
    if prop_count >= 5:
        score += 20
    elif prop_count >= 2:
        score += 10

    # Cash buyer / refi signals
    if flag_true(row.get("probable_cash_buyer", "")):
        score += 15

    # STR licensed
    if flag_true(row.get("str_licensed", "")):
        score += 5

    # Already has a resolved person (easier to enrich)
    resolved = str(row.get("resolved_person", "")).strip()
    if resolved and resolved.upper() not in ("NAN", "NONE", ""):
        score += 5

    # Already has a phone (even better)
    phone = str(row.get("phone", "")).strip()
    if phone and phone.upper() not in ("NAN", "NONE", ""):
        score += 5

    return score


# ---------------------------------------------------------------------------
# Name parsing
# ---------------------------------------------------------------------------

def parse_person_name(name: str) -> dict:
    """Parse a person name into first, last, middle components."""
    if not name or name.upper() in ("NAN", "NONE", ""):
        return {"first": "", "last": "", "full": ""}

    name = name.strip()

    # "LAST, FIRST MIDDLE" format
    if "," in name:
        parts = name.split(",", 1)
        last = parts[0].strip()
        rest = parts[1].strip() if len(parts) > 1 else ""
        first = rest.split()[0] if rest else ""
        return {"first": first.title(), "last": last.title(), "full": f"{first} {last}".title().strip()}

    # "FIRST LAST" format
    parts = name.split()
    if len(parts) >= 2:
        return {"first": parts[0].title(), "last": parts[-1].title(), "full": name.title()}
    elif len(parts) == 1:
        return {"first": parts[0].title(), "last": "", "full": parts[0].title()}

    return {"first": "", "last": "", "full": name.title()}


# ---------------------------------------------------------------------------
# Email pattern generation
# ---------------------------------------------------------------------------

def generate_email_patterns(first: str, last: str, domain: str = None) -> list:
    """
    Generate likely email patterns from a person's name.
    Returns list of candidate emails to validate.
    """
    if not first or not last:
        return []

    first = first.lower().strip()
    last = last.lower().strip()

    patterns = []

    # If we have a domain (from LLC website), use it
    domains = [domain] if domain else COMMON_DOMAINS

    for d in domains:
        patterns.extend([
            f"{first}.{last}@{d}",
            f"{first}{last}@{d}",
            f"{first[0]}{last}@{d}",
            f"{first}{last[0]}@{d}",
            f"{first}@{d}",
        ])

    return patterns


# ---------------------------------------------------------------------------
# People search URL generation (for manual lookup)
# ---------------------------------------------------------------------------

def truepeoplesearch_url(first: str, last: str, city: str, state: str) -> str:
    """Generate a TruePeopleSearch URL for manual lookup."""
    name = f"{first} {last}".strip()
    location = f"{city}, {state}".strip(", ")
    q = urllib.parse.quote(f"{name} {location}")
    return f"https://www.truepeoplesearch.com/results?name={urllib.parse.quote(name)}&citystatezip={urllib.parse.quote(location)}"


def fastpeoplesearch_url(first: str, last: str, city: str, state: str) -> str:
    """Generate a FastPeopleSearch URL for manual lookup."""
    name_slug = f"{first}-{last}".strip("-").replace(" ", "-")
    location_slug = f"{city}-{state}".strip("-").replace(" ", "-")
    return f"https://www.fastpeoplesearch.com/name/{name_slug}_{location_slug}"


def linkedin_search_url(first: str, last: str, company: str = "") -> str:
    """Generate a LinkedIn search URL."""
    q = f"{first} {last}"
    if company:
        q += f" {company}"
    return f"https://www.linkedin.com/search/results/all/?keywords={urllib.parse.quote(q)}"


# ---------------------------------------------------------------------------
# SunBiz re-resolution (for leads missing resolved_person)
# ---------------------------------------------------------------------------

def resolve_sunbiz(entity_name: str, session) -> dict:
    """Resolve an LLC via SunBiz. Returns dict with resolved_person, officers, etc."""
    result = {
        "resolved_person": "",
        "registered_agent_name": "",
        "officer_names": "",
        "status": "",
    }

    if not HAS_REQUESTS or not HAS_BS4 or not session:
        return result

    search_name = entity_name.strip()
    for suffix in [" LLC", " L.L.C.", " L.L.C", " INC", " INC.",
                   " CORP", " CORP.", " LP", " LTD", " LTD."]:
        if search_name.upper().endswith(suffix):
            search_name = search_name[: len(search_name) - len(suffix)].strip()

    try:
        resp = session.post(
            SUNBIZ_SEARCH_URL,
            data={"SearchTerm": search_name, "InquiryType": "EntityName", "SearchNameOrder": ""},
            timeout=30,
        )
        if resp.status_code != 200:
            result["status"] = f"HTTP {resp.status_code}"
            return result

        soup = BeautifulSoup(resp.text, "html.parser")
        links = soup.find_all("a", href=re.compile(r"SearchResultDetail"))
        if not links:
            result["status"] = "NO RESULTS"
            return result

        # Find best match
        detail_href = None
        entity_upper = entity_name.upper().strip()
        for link in links:
            if link.get_text(strip=True).upper() == entity_upper:
                detail_href = link["href"]
                break
        if not detail_href:
            detail_href = links[0]["href"]

        time.sleep(1)

        detail_resp = session.get(SUNBIZ_BASE + detail_href, timeout=30)
        if detail_resp.status_code != 200:
            return result

        detail_soup = BeautifulSoup(detail_resp.text, "html.parser")
        sections = detail_soup.find_all("div", class_="detailSection")
        officers = []

        for section in sections:
            lines = [l.strip() for l in section.get_text(separator="\n", strip=True).split("\n") if l.strip()]
            if not lines:
                continue
            header = lines[0]

            if "Registered Agent" in header:
                agent_lines = [l for l in lines[1:] if not l.startswith("Name Changed:") and not l.startswith("Address Changed:")]
                if agent_lines:
                    result["registered_agent_name"] = agent_lines[0]

            elif "Officer/Director" in header or "Authorized Person" in header:
                i = 1
                while i < len(lines):
                    if lines[i].startswith("Title "):
                        title_val = lines[i].replace("Title ", "").strip()
                        officer = {"title": title_val, "name": ""}
                        if i + 1 < len(lines) and not lines[i + 1].startswith("Title "):
                            officer["name"] = lines[i + 1]
                            i += 1
                        officers.append(officer)
                    i += 1

        if officers:
            result["officer_names"] = "; ".join(f"{o['name']} ({o['title']})" for o in officers if o.get("name"))

        # Pick the human
        for officer in officers:
            if officer.get("title", "").upper() in PRIORITY_TITLES and officer.get("name"):
                result["resolved_person"] = officer["name"]
                break
        if not result["resolved_person"] and officers:
            for o in officers:
                if o.get("name"):
                    result["resolved_person"] = o["name"]
                    break
        if not result["resolved_person"] and result["registered_agent_name"]:
            agent = result["registered_agent_name"].upper()
            if not any(kw in agent for kw in AGENT_SERVICE_KEYWORDS):
                result["resolved_person"] = result["registered_agent_name"]

        result["status"] = "OK"

    except Exception as e:
        result["status"] = f"ERROR: {e}"

    return result


# ---------------------------------------------------------------------------
# Datazapp export
# ---------------------------------------------------------------------------

def export_datazapp_csv(df: pd.DataFrame, output_path: Path):
    """
    Export a CSV formatted for Datazapp batch skip trace upload.
    Datazapp needs: First Name, Last Name, Address, City, State, Zip
    Cost: $0.03/record — for 25 records that's $0.75.
    """
    rows = []
    for _, row in df.iterrows():
        person = parse_person_name(str(row.get("resolved_person", "") or row.get("OWN_NAME", "")))

        rows.append({
            "First Name": person["first"],
            "Last Name": person["last"],
            "Address": str(row.get("OWN_ADDR1", "")).strip(),
            "City": str(row.get("OWN_CITY", "")).strip(),
            "State": str(row.get("OWN_STATE_DOM", row.get("OWN_STATE", ""))).strip(),
            "Zip": str(row.get("OWN_ZIPCD", "")).strip()[:5],
        })

    out_df = pd.DataFrame(rows)
    out_df.to_csv(output_path, index=False)
    return len(out_df)


# ---------------------------------------------------------------------------
# Research tracker Excel
# ---------------------------------------------------------------------------

def build_research_tracker(df: pd.DataFrame, output_path: Path):
    """
    Build an Excel workbook with everything you need to manually
    research 25 leads: details, search URLs, and blank columns
    to paste in what you find.
    """
    if not HAS_OPENPYXL:
        print("  WARNING: openpyxl not installed. Skipping Excel tracker.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Research Tracker"

    # Styles
    header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    link_font = Font(color="0563C1", underline="single", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Headers
    headers = [
        "Rank", "Score", "LLC / Owner Name", "Resolved Person",
        "Properties", "Portfolio Value", "Mailing Address",
        "Existing Phone", "Existing Email",
        "TruePeopleSearch Link", "FastPeopleSearch Link", "LinkedIn Link",
        "FOUND: Phone", "FOUND: Email", "FOUND: Source", "Notes",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = wrap
        cell.border = thin_border

    # Data rows
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        owner = str(row.get("OWN_NAME", ""))
        resolved = str(row.get("resolved_person", ""))
        if resolved.upper() in ("NAN", "NONE", ""):
            resolved = ""

        person = parse_person_name(resolved or owner)
        city = str(row.get("OWN_CITY", "")).strip()
        state = str(row.get("OWN_STATE_DOM", row.get("OWN_STATE", ""))).strip()

        existing_phone = str(row.get("phone", ""))
        if existing_phone.upper() in ("NAN", "NONE", ""):
            existing_phone = ""
        existing_email = str(row.get("email", ""))
        if existing_email.upper() in ("NAN", "NONE", ""):
            existing_email = ""

        prop_count = str(row.get("property_count", "1"))
        portfolio_val = str(row.get("total_portfolio_value", ""))
        address = f"{row.get('OWN_ADDR1', '')}, {city}, {state} {str(row.get('OWN_ZIPCD', ''))[:5]}"

        score = row.get("_enrich_score", 0)

        values = [
            row_idx - 1,           # Rank
            score,                 # Score
            owner,                 # LLC / Owner
            resolved,              # Resolved Person
            prop_count,            # Properties
            portfolio_val,         # Portfolio Value
            address,               # Mailing Address
            existing_phone,        # Existing Phone
            existing_email,        # Existing Email
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = wrap
            cell.border = thin_border

        # Search URL links
        if person["first"] and person["last"]:
            tps_url = truepeoplesearch_url(person["first"], person["last"], city, state)
            fps_url = fastpeoplesearch_url(person["first"], person["last"], city, state)
            li_url = linkedin_search_url(person["first"], person["last"], owner if str(row.get("is_entity", "")).lower() in ("true", "1") else "")

            ws.cell(row=row_idx, column=10, value="Search").hyperlink = tps_url
            ws.cell(row=row_idx, column=10).font = link_font
            ws.cell(row=row_idx, column=10).border = thin_border

            ws.cell(row=row_idx, column=11, value="Search").hyperlink = fps_url
            ws.cell(row=row_idx, column=11).font = link_font
            ws.cell(row=row_idx, column=11).border = thin_border

            ws.cell(row=row_idx, column=12, value="Search").hyperlink = li_url
            ws.cell(row=row_idx, column=12).font = link_font
            ws.cell(row=row_idx, column=12).border = thin_border
        else:
            for c in (10, 11, 12):
                cell = ws.cell(row=row_idx, column=c, value="(need person name)")
                cell.border = thin_border

        # Yellow input cells for manual entry
        for c in (13, 14, 15, 16):
            cell = ws.cell(row=row_idx, column=c, value="")
            cell.fill = input_fill
            cell.border = thin_border

    # Column widths
    widths = [5, 6, 35, 25, 8, 15, 35, 15, 25, 12, 12, 12, 15, 25, 15, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # Freeze top row
    ws.freeze_panes = "A2"

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Enrich contacts for top leads — test mode (Step 5)"
    )
    parser.add_argument(
        "--input",
        type=str,
        default=str(DEFAULT_INPUT),
        help=f"Input CSV (default: {DEFAULT_INPUT})",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Number of top leads to enrich (default: 0 = all)",
    )
    parser.add_argument(
        "--skip-sunbiz",
        action="store_true",
        help="Skip SunBiz re-resolution (faster, use cached data only)",
    )
    parser.add_argument(
        "--counties",
        type=str,
        default="",
        help=(
            'Filter to leads with mailing address in these counties. '
            'Comma-separated city keywords, e.g. "palm beach,broward". '
            'Leads in other counties are excluded unless they have <25%% '
            'of portfolio outside target area.'
        ),
    )
    args = parser.parse_args()

    ENRICHED_DIR.mkdir(parents=True, exist_ok=True)
    CACHE_DIR.mkdir(parents=True, exist_ok=True)

    # -------------------------------------------------------------------
    # 1. Load existing pipeline data
    # -------------------------------------------------------------------
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"\nInput file not found: {input_path}")
        print("Point --input to your pipeline output CSV.")
        return

    print(f"\n  Loading: {input_path}")
    df = pd.read_csv(input_path, dtype=str, low_memory=False)
    print(f"  Total records: {len(df):,}")

    # -------------------------------------------------------------------
    # 1b. Filter by county if requested
    # -------------------------------------------------------------------
    if args.counties:
        county_keywords = [c.strip().upper() for c in args.counties.split(",") if c.strip()]
        print(f"  Filtering to counties: {', '.join(county_keywords)}")

        # Palm Beach county cities / Broward county cities for matching
        # We match on mailing city (OWN_CITY) since that's where the owner is
        PB_CITIES = {
            "WEST PALM BEACH", "PALM BEACH", "PALM BCH", "PALM BEACH GARDENS",
            "PALM BCH GDNS", "LAKE WORTH", "BOYNTON BEACH", "DELRAY BEACH",
            "BOCA RATON", "JUPITER", "WELLINGTON", "ROYAL PALM BEACH",
            "GREENACRES", "RIVIERA BEACH", "N PALM BEACH", "NORTH PALM BEACH",
            "LANTANA", "PAHOKEE", "BELLE GLADE", "LOXAHATCHEE", "TEQUESTA",
            "JUNO BEACH", "SINGER ISLAND", "MANGONIA PARK", "GLEN RIDGE",
            "HAVERHILL", "CLOUD LAKE", "PALM SPRINGS", "LAKE CLARKE SHORES",
            "HYPOLUXO", "MANALAPAN", "OCEAN RIDGE", "BRINY BREEZES",
            "GULF STREAM", "HIGHLAND BEACH", "SOUTH PALM BEACH",
        }
        BROWARD_CITIES = {
            "FORT LAUDERDALE", "FT LAUDERDALE", "HOLLYWOOD", "PEMBROKE PINES",
            "CORAL SPRINGS", "MIRAMAR", "SUNRISE", "PLANTATION", "DAVIE",
            "DEERFIELD BEACH", "POMPANO BEACH", "COCONUT CREEK", "TAMARAC",
            "MARGATE", "WESTON", "DANIA", "DANIA BEACH", "HALLANDALE",
            "HALLANDALE BEACH", "LAUDERHILL", "LAUDERDALE LAKES",
            "OAKLAND PARK", "NORTH LAUDERDALE", "LIGHTHOUSE POINT",
            "WILTON MANORS", "COOPER CITY", "PARKLAND", "SOUTHWEST RANCHES",
            "LAZY LAKE", "SEA RANCH LAKES", "HILLSBORO BEACH",
        }

        target_cities = set()
        for kw in county_keywords:
            if "PALM" in kw or "PB" in kw:
                target_cities.update(PB_CITIES)
            elif "BROWARD" in kw or "FT LAUDERDALE" in kw or "FORT LAUDERDALE" in kw:
                target_cities.update(BROWARD_CITIES)
            else:
                # Generic: just match as a city name substring
                target_cities.add(kw)

        mail_city = df["OWN_CITY"].astype(str).str.strip().str.upper()
        in_target = mail_city.isin(target_cities)

        # Also accept leads whose city contains the keyword
        for kw in county_keywords:
            in_target = in_target | mail_city.str.contains(kw, na=False)

        before = len(df)
        df = df[in_target].copy()
        print(f"  Filtered: {len(df):,} leads in target counties (dropped {before - len(df):,})")

    # -------------------------------------------------------------------
    # 2. Score and rank leads
    # -------------------------------------------------------------------
    print("  Scoring leads to find the hottest...")
    df["_enrich_score"] = df.apply(score_lead, axis=1)
    df = df.sort_values("_enrich_score", ascending=False)

    if args.limit > 0:
        top = df.head(args.limit).copy()
    else:
        top = df.copy()
    print(f"  Selected {len(top)} leads (scores {top['_enrich_score'].max()}-{top['_enrich_score'].min()})")

    # -------------------------------------------------------------------
    # 3. SunBiz re-resolution for unresolved LLCs
    # -------------------------------------------------------------------
    needs_resolve = top[
        (top["is_entity"].astype(str).str.lower().isin(["true", "1", "yes"])) &
        (top["resolved_person"].fillna("").isin(["", "nan", "NaN", "NONE", "None"]))
    ]

    if len(needs_resolve) > 0 and not args.skip_sunbiz and HAS_REQUESTS and HAS_BS4:
        print(f"\n  SunBiz: {len(needs_resolve)} LLCs need person resolution...")

        session = req_lib.Session()
        session.headers.update({
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        })
        try:
            session.get(SUNBIZ_SEARCH_URL, timeout=30)
        except Exception:
            pass

        resolved_count = 0
        for idx, row in needs_resolve.iterrows():
            entity_name = str(row["OWN_NAME"]).strip()
            print(f"    Looking up: {entity_name[:50]}")

            result = resolve_sunbiz(entity_name, session)

            if result["resolved_person"]:
                top.at[idx, "resolved_person"] = result["resolved_person"]
                top.at[idx, "entity_officers"] = result.get("officer_names", "")
                resolved_count += 1
                print(f"      → {result['resolved_person']}")
            else:
                print(f"      → (no person found: {result['status']})")

            time.sleep(SUNBIZ_DELAY)

        print(f"  SunBiz resolved: {resolved_count}/{len(needs_resolve)}")
    elif len(needs_resolve) > 0:
        print(f"\n  {len(needs_resolve)} LLCs still unresolved (use --skip-sunbiz=false to retry)")

    # -------------------------------------------------------------------
    # 4. Generate email patterns for resolved leads
    # -------------------------------------------------------------------
    print("\n  Generating email patterns for resolved names...")
    email_candidates = []
    for idx, row in top.iterrows():
        resolved = str(row.get("resolved_person", "")).strip()
        if resolved.upper() in ("NAN", "NONE", ""):
            email_candidates.append("")
            continue

        person = parse_person_name(resolved)
        if person["first"] and person["last"]:
            patterns = generate_email_patterns(person["first"], person["last"])
            # Store the most likely patterns (gmail is most common for individuals)
            top.at[idx, "email_patterns"] = "; ".join(patterns[:5])
            email_candidates.append(patterns[0] if patterns else "")
        else:
            email_candidates.append("")

    has_patterns = sum(1 for e in email_candidates if e)
    print(f"  Generated email patterns for {has_patterns}/{len(top)} leads")

    # -------------------------------------------------------------------
    # 5. Export Datazapp CSV
    # -------------------------------------------------------------------
    datazapp_path = ENRICHED_DIR / "datazapp_upload.csv"
    datazapp_count = export_datazapp_csv(top, datazapp_path)
    cost = datazapp_count * 0.03
    print(f"\n  Datazapp CSV: {datazapp_path}")
    print(f"    {datazapp_count} records — estimated cost: ${cost:.2f}")
    print(f"    Upload at: https://www.datazapp.com/")
    print(f"    Returns: cell phone, landline, email for each match")

    # -------------------------------------------------------------------
    # 6. Build research tracker Excel
    # -------------------------------------------------------------------
    tracker_path = ENRICHED_DIR / "research_tracker.xlsx"
    build_research_tracker(top, tracker_path)
    print(f"\n  Research tracker: {tracker_path}")
    print(f"    Open this file and use the search links to manually")
    print(f"    look up phone/email. Paste findings in the yellow columns.")

    # -------------------------------------------------------------------
    # 7. Save enriched CSV
    # -------------------------------------------------------------------
    output_csv = ENRICHED_DIR / "top_leads_enriched.csv"
    top.to_csv(output_csv, index=False)

    # -------------------------------------------------------------------
    # Summary
    # -------------------------------------------------------------------
    has_phone = top["phone"].fillna("").ne("").sum()
    has_email = top["email"].fillna("").ne("").sum()
    has_resolved = top["resolved_person"].fillna("").apply(
        lambda x: x.upper() not in ("", "NAN", "NONE")
    ).sum()
    has_entity = top["is_entity"].astype(str).str.lower().isin(["true", "1"]).sum()

    print()
    print("=" * 60)
    print("  ENRICHMENT SUMMARY")
    print("=" * 60)
    print(f"  Top leads selected:    {len(top)}")
    print(f"  LLC/Entity owned:      {has_entity}")
    print(f"  Resolved to person:    {has_resolved}")
    print(f"  Has phone already:     {has_phone}")
    print(f"  Has email already:     {has_email}")
    print(f"  Email patterns made:   {has_patterns}")
    print()
    print("  NEXT STEPS:")
    print("  " + "-" * 50)
    print(f"  1. Open research_tracker.xlsx")
    print(f"     Click the search links for each lead.")
    print(f"     Paste phone/email into the yellow columns.")
    print(f"     (~5 min per lead, ~2 hours for all {len(top)})")
    print()
    print(f"  2. Upload datazapp_upload.csv to datazapp.com")
    print(f"     Cost: ${cost:.2f} for {datazapp_count} records")
    print(f"     Download results, save as data/enriched/datazapp_results.csv")
    print()
    print(f"  3. After manual research + Datazapp, run:")
    print(f"     python scripts/05b_merge_enrichment.py")
    print(f"     (merges all sources into one enriched file)")
    print()
    print(f"  4. Then validate:")
    print(f"     python scripts/06_validate_contacts.py --county top_leads")
    print()
    print(f"  FILES CREATED:")
    print(f"    {output_csv}")
    print(f"    {tracker_path}")
    print(f"    {datazapp_path}")
    print()


if __name__ == "__main__":
    main()
